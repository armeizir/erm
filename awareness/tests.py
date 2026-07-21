from datetime import date, timedelta
from io import BytesIO, StringIO
from smtplib import SMTPAuthenticationError

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.contrib.messages import get_messages
from django.core.management import call_command
from django.core import mail
from django.core.mail.backends.locmem import EmailBackend as LocMemEmailBackend
from django.test import RequestFactory, override_settings
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    AwarenessAnswer,
    AwarenessAttempt,
    AwarenessCampaign,
    AwarenessQuestion,
    AwarenessUnitTarget,
)
from .admin import AwarenessAttemptAdmin, AwarenessAttemptGroupFilter
from .notifications import send_awareness_notification
from risk.models import AppSetting, PenugasanUnitBisnis


class DummySMTPBackend(LocMemEmailBackend):
    init_kwargs = None

    def __init__(self, *args, **kwargs):
        type(self).init_kwargs = kwargs.copy()
        super().__init__(*args, **kwargs)


class AwarenessAttemptAdminTests(TestCase):
    def test_group_column_is_after_user_and_sortable(self):
        User = get_user_model()
        admin_user = User.objects.create_superuser(username="admin", password="secret")
        user = User.objects.create_user(username="risk.user", password="secret")
        group = Group.objects.create(name="BID RISIKO")
        user.groups.add(group)
        campaign = AwarenessCampaign.objects.create(
            title="Awareness 2026",
            description="Kuis",
            topic="manajemen_risiko",
            start_date=timezone.localdate(),
            end_date=timezone.localdate() + timedelta(days=1),
        )
        attempt = AwarenessAttempt.objects.create(campaign=campaign, user=user, attempt_number=1)
        request = RequestFactory().get("/admin/awareness/awarenessattempt/")
        request.user = admin_user
        attempt_admin = AwarenessAttemptAdmin(AwarenessAttempt, AdminSite())

        queryset = attempt_admin.get_queryset(request).filter(pk=attempt.pk)
        row = queryset.order_by("user_group_order").first()

        self.assertEqual(attempt_admin.list_display[1:3], ("user", "user_groups"))
        self.assertEqual(AwarenessAttemptAdmin.user_groups.admin_order_field, "user_group_order")
        self.assertEqual(row.user_group_order, "BID RISIKO")
        self.assertEqual(attempt_admin.user_groups(row), "BID RISIKO")

    def test_group_filter_limits_attempts_by_user_group(self):
        User = get_user_model()
        admin_user = User.objects.create_superuser(username="admin", password="secret")
        group_a = Group.objects.create(name="SEKPER")
        group_b = Group.objects.create(name="UB BES")
        user_a = User.objects.create_user(username="sekper.user", password="secret")
        user_b = User.objects.create_user(username="ubbes.user", password="secret")
        user_a.groups.add(group_a)
        user_b.groups.add(group_b)
        campaign = AwarenessCampaign.objects.create(
            title="Awareness 2026",
            description="Kuis",
            topic="manajemen_risiko",
            start_date=timezone.localdate(),
            end_date=timezone.localdate() + timedelta(days=1),
        )
        attempt_a = AwarenessAttempt.objects.create(campaign=campaign, user=user_a, attempt_number=1)
        AwarenessAttempt.objects.create(campaign=campaign, user=user_b, attempt_number=1)
        request = RequestFactory().get("/admin/awareness/awarenessattempt/", {"user_group": str(group_a.pk)})
        request.user = admin_user
        attempt_admin = AwarenessAttemptAdmin(AwarenessAttempt, AdminSite())
        group_filter = AwarenessAttemptGroupFilter(
            request,
            {"user_group": str(group_a.pk)},
            AwarenessAttempt,
            attempt_admin,
        )

        queryset = group_filter.queryset(request, attempt_admin.get_queryset(request))

        self.assertEqual(list(queryset), [attempt_a])

    def test_export_attempts_xlsx_includes_group_column(self):
        User = get_user_model()
        admin_user = User.objects.create_superuser(username="admin", password="secret")
        user = User.objects.create_user(username="risk.user", email="risk@example.com", password="secret")
        group = Group.objects.create(name="BID RISIKO")
        user.groups.add(group)
        campaign = AwarenessCampaign.objects.create(
            title="Awareness 2026",
            description="Kuis",
            topic="manajemen_risiko",
            start_date=timezone.localdate(),
            end_date=timezone.localdate() + timedelta(days=1),
        )
        attempt = AwarenessAttempt.objects.create(
            campaign=campaign,
            user=user,
            attempt_number=1,
            status=AwarenessAttempt.STATUS_PASSED,
            score=100,
            correct_count=10,
        )
        request = RequestFactory().post("/admin/awareness/awarenessattempt/")
        request.user = admin_user
        attempt_admin = AwarenessAttemptAdmin(AwarenessAttempt, AdminSite())

        response = attempt_admin.export_attempts_xlsx(
            request,
            attempt_admin.get_queryset(request).filter(pk=attempt.pk),
        )
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("group", rows[0])
        self.assertEqual(rows[1][3], "BID RISIKO")


class AwarenessFlowTests(TestCase):
    def setUp(self):
        self.client = Client()
        User = get_user_model()
        self.user = User.objects.create_user(username="risk.user", email="risk@example.com", password="secret")
        self.other = User.objects.create_user(username="other.user", email="other@example.com", password="secret")
        self.admin = User.objects.create_superuser(username="admin", email="admin@example.com", password="secret")
        self.staff = User.objects.create_user(
            username="staff.awareness",
            email="staff-awareness@example.com",
            password="secret",
            is_staff=True,
        )
        today = timezone.localdate()
        self.campaign = AwarenessCampaign.objects.create(
            title="Awareness Manajemen Risiko Dasar 2026",
            description="Kuis dasar",
            topic="manajemen_risiko",
            start_date=today - timedelta(days=1),
            end_date=today + timedelta(days=30),
            passing_score=70,
            max_attempts=1,
            is_active=True,
        )
        self.q1 = AwarenessQuestion.objects.create(
            campaign=self.campaign,
            order=1,
            question_text="Apa tujuan utama manajemen risiko?",
            option_a="Menghilangkan risiko",
            option_b="Mengidentifikasi, menilai, mengendalikan, dan memantau risiko",
            option_c="Menghindari bisnis",
            option_d="Menunda keputusan",
            correct_answer="B",
            explanation="Manajemen risiko mengelola ketidakpastian.",
        )
        self.q2 = AwarenessQuestion.objects.create(
            campaign=self.campaign,
            order=2,
            question_text="Apa arti KRI?",
            option_a="Key Risk Indicator",
            option_b="Key Revenue Income",
            option_c="Knowledge Risk Input",
            option_d="Key Review Instruction",
            correct_answer="A",
            explanation="KRI adalah indikator risiko utama.",
        )

    def test_user_login_can_view_active_campaign(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("awareness:campaign_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.campaign.title)

    def test_anonymous_user_redirected_to_login(self):
        response = self.client.get(reverse("awareness:campaign_list"))

        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login/", response["Location"])

    def test_user_can_start_quiz(self):
        self.client.force_login(self.user)

        response = self.client.post(reverse("awareness:start_campaign", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 302)
        attempt = AwarenessAttempt.objects.get(user=self.user, campaign=self.campaign)
        self.assertEqual(attempt.status, AwarenessAttempt.STATUS_IN_PROGRESS)
        self.assertIn(reverse("awareness:quiz_attempt", args=[attempt.pk]), response["Location"])

    def test_campaign_list_links_to_material_before_quiz(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("awareness:campaign_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Baca Materi")
        self.assertContains(response, reverse("awareness:campaign_material", args=[self.campaign.pk]))
        self.assertFalse(AwarenessAttempt.objects.filter(user=self.user, campaign=self.campaign).exists())

    def test_material_page_shows_uploaded_image_and_posts_start(self):
        self.campaign.material_image = "awareness/materials/materi.png"
        self.campaign.save(update_fields=["material_image"])
        self.client.force_login(self.user)

        response = self.client.get(reverse("awareness:campaign_material", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Materi Sebelum Kuis")
        self.assertContains(response, self.campaign.material_image.url)
        self.assertContains(response, reverse("awareness:start_campaign", args=[self.campaign.pk]))
        self.assertFalse(AwarenessAttempt.objects.filter(user=self.user, campaign=self.campaign).exists())

    def test_campaign_participants_shows_awareness_and_group_only(self):
        unit = Group.objects.create(name="BID RISIKO")
        earlier_unit = Group.objects.create(name="BID AGA")
        self.user.first_name = "Risk"
        self.user.last_name = "User"
        self.user.save(update_fields=["first_name", "last_name"])
        self.user.groups.add(unit)
        self.other.first_name = "Other"
        self.other.last_name = "User"
        self.other.save(update_fields=["first_name", "last_name"])
        self.other.groups.add(earlier_unit)
        AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
            correct_count=2,
            wrong_count=0,
            score=100,
            status=AwarenessAttempt.STATUS_PASSED,
            submitted_at=timezone.now(),
        )
        AwarenessAttempt.objects.create(
            user=self.other,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
            correct_count=2,
            wrong_count=0,
            score=100,
            status=AwarenessAttempt.STATUS_PASSED,
            submitted_at=timezone.now(),
        )
        self.client.force_login(self.other)

        response = self.client.get(reverse("awareness:campaign_participants", args=[self.campaign.pk]))
        html = response.content.decode()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AWARENESS")
        self.assertContains(response, "GROUP")
        self.assertContains(response, "Other User")
        self.assertContains(response, "BID AGA")
        self.assertContains(response, "Risk User")
        self.assertContains(response, "BID RISIKO")
        self.assertLess(html.index("BID AGA"), html.index("BID RISIKO"))
        self.assertNotContains(response, "SCORE")
        self.assertNotContains(response, "STATUS")
        self.assertNotContains(response, "Lulus")

    def test_user_can_submit_answers_and_score_is_calculated(self):
        self.client.force_login(self.user)
        attempt = AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
        )

        response = self.client.post(reverse("awareness:submit_attempt", args=[attempt.pk]), {
            f"question_{self.q1.pk}": "B",
            f"question_{self.q2.pk}": "D",
        })

        attempt.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(attempt.correct_count, 1)
        self.assertEqual(attempt.wrong_count, 1)
        self.assertEqual(float(attempt.score), 50.0)
        self.assertEqual(attempt.status, AwarenessAttempt.STATUS_FAILED)
        self.assertEqual(AwarenessAnswer.objects.filter(attempt=attempt).count(), 2)

    def test_passing_score_marks_passed(self):
        self.client.force_login(self.user)
        attempt = AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
        )

        self.client.post(reverse("awareness:submit_attempt", args=[attempt.pk]), {
            f"question_{self.q1.pk}": "B",
            f"question_{self.q2.pk}": "A",
        })

        attempt.refresh_from_db()
        self.assertEqual(float(attempt.score), 100.0)
        self.assertEqual(attempt.status, AwarenessAttempt.STATUS_PASSED)

    def test_user_cannot_view_other_user_attempt(self):
        self.client.force_login(self.user)
        attempt = AwarenessAttempt.objects.create(
            user=self.other,
            campaign=self.campaign,
            attempt_number=1,
        )

        response = self.client.get(reverse("awareness:quiz_attempt", args=[attempt.pk]))

        self.assertEqual(response.status_code, 403)

    def test_max_attempts_blocks_new_attempt(self):
        self.client.force_login(self.user)
        AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            status=AwarenessAttempt.STATUS_FAILED,
            submitted_at=timezone.now(),
        )

        response = self.client.post(reverse("awareness:start_campaign", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(AwarenessAttempt.objects.filter(user=self.user, campaign=self.campaign).count(), 1)

    def test_start_campaign_requires_post(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("awareness:start_campaign", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 405)
        self.assertFalse(AwarenessAttempt.objects.filter(user=self.user, campaign=self.campaign).exists())

    def test_admin_export_excel_generates_xlsx(self):
        self.client.force_login(self.admin)
        AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            status=AwarenessAttempt.STATUS_PASSED,
            score=100,
            correct_count=2,
            wrong_count=0,
            submitted_at=timezone.now(),
        )

        response = self.client.get(reverse("risk_admin:awareness_campaign_export_xlsx", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertGreater(len(response.content), 1000)

    def _grant_awareness_permissions(self, user, *codenames):
        permissions = Permission.objects.filter(
            content_type__app_label="awareness",
            codename__in=codenames,
        )
        user.user_permissions.add(*permissions)

    def test_staff_cannot_open_awareness_admin_without_explicit_permission(self):
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:awareness_awarenesscampaign_changelist"))

        self.assertEqual(response.status_code, 403)

    def test_staff_can_open_awareness_campaign_admin_with_view_permission(self):
        self._grant_awareness_permissions(self.staff, "view_awarenesscampaign")
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:awareness_awarenesscampaign_changelist"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Kirim Report")
        self.assertNotContains(response, "Report Awareness")
        self.assertNotContains(response, "Kirim Test")

    def test_awareness_campaign_admin_labels_report_actions(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("risk_admin:awareness_awarenesscampaign_changelist"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Kirim Report")
        self.assertContains(response, "Report Awareness")
        self.assertNotContains(response, "Kirim Test")

    def test_admin_can_open_campaign_report_detail(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("risk_admin:awareness_campaign_report_detail", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Report Awareness")
        self.assertContains(response, self.campaign.title)
        self.assertContains(response, "Progress Responden per Bidang / Unit")

    def test_staff_can_open_campaign_report_detail_with_report_permission(self):
        self._grant_awareness_permissions(self.staff, "view_campaign_report")
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:awareness_campaign_report_detail", args=[self.campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Report Awareness")

    def test_staff_cannot_open_awareness_question_admin_without_explicit_permission(self):
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:awareness_awarenessquestion_changelist"))

        self.assertEqual(response.status_code, 403)

    def test_staff_can_open_awareness_question_admin_with_view_permission(self):
        self._grant_awareness_permissions(self.staff, "view_awarenessquestion")
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:awareness_awarenessquestion_changelist"))

        self.assertEqual(response.status_code, 200)

    def test_staff_dashboard_hides_risk_awareness_menu_without_permission(self):
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:index"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Risk Awareness")

    def test_staff_dashboard_shows_risk_awareness_menu_with_permission(self):
        self._grant_awareness_permissions(self.staff, "view_awarenesscampaign")
        self.client.force_login(self.staff)

        response = self.client.get(reverse("risk_admin:index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Risk Awareness")

    def test_seed_command_creates_campaign_and_questions(self):
        out = StringIO()

        call_command("seed_awareness_risk_management", stdout=out)

        campaign = AwarenessCampaign.objects.get(title="Awareness Manajemen Risiko Dasar 2026")
        self.assertGreaterEqual(campaign.questions.count(), 10)
        self.assertIn("Seed awareness selesai", out.getvalue())

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM CSIRT <noreply@plnbatam.com>",
        ALLOWED_HOSTS=["127.0.0.1"],
    )
    def test_send_awareness_reminder_command_sends_html_email_to_test_address(self):
        out = StringIO()

        call_command(
            "send_awareness_reminders",
            "--campaign-id",
            str(self.campaign.pk),
            "--email",
            "armeizir@plnbatam.com",
            "--base-url",
            "http://127.0.0.1:8001",
            stdout=out,
        )

        self.assertEqual(len(mail.outbox), 1)
        message = mail.outbox[0]
        self.assertEqual(message.to, ["armeizir@plnbatam.com"])
        self.assertIn("Pelaksanaan Awareness Manajemen Risiko Dasar 2026", message.subject)
        self.assertIn("http://127.0.0.1:8001/awareness/", message.body)
        self.assertTrue(message.alternatives)
        self.assertIn("Isi Awareness Sekarang", message.alternatives[0].content)
        self.assertIn("Email awareness terkirim", out.getvalue())

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM CSIRT <noreply@plnbatam.com>",
    )
    def test_admin_can_send_awareness_test_notification(self):
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("risk_admin:awareness_campaign_send_test", args=[self.campaign.pk]),
            {"email": "armeizir@plnbatam.com"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["armeizir@plnbatam.com"])

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM ERM <erm@plnbatam.com>",
    )
    def test_admin_send_awareness_test_uses_campaign_test_email(self):
        self.campaign.notification_test_email = "risk.admin@plnbatam.com"
        self.campaign.save(update_fields=["notification_test_email"])
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("risk_admin:awareness_campaign_send_test", args=[self.campaign.pk]),
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["risk.admin@plnbatam.com"])

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM ERM <erm@plnbatam.com>",
    )
    def test_awareness_email_uses_campaign_heading_and_topic(self):
        self.campaign.email_header_title = "Awareness Manajemen Risiko"
        self.campaign.email_header_subtitle = "Manajemen Risiko"
        self.campaign.save(update_fields=["email_header_title", "email_header_subtitle"])

        send_awareness_notification(
            self.campaign,
            ["risk.admin@plnbatam.com"],
            base_url="https://erm.plnbatam.com",
        )

        html_body = mail.outbox[0].alternatives[0].content
        self.assertIn("AWARENESS MANAJEMEN RISIKO", html_body)
        self.assertIn("Manajemen Risiko", html_body)
        self.assertNotIn("CYBER SECURITY", html_body)
        self.assertNotIn("Keamanan Teknologi Informasi", html_body)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM ERM <erm@plnbatam.com>",
    )
    def test_awareness_email_includes_unit_progress_table(self):
        unit = Group.objects.create(name="Bidang Manajemen Risiko dan Kepatuhan")
        self.user.groups.add(unit)
        self.other.groups.add(unit)
        AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
            correct_count=2,
            wrong_count=0,
            score=100,
            status=AwarenessAttempt.STATUS_PASSED,
            submitted_at=timezone.now(),
        )

        send_awareness_notification(
            self.campaign,
            ["risk.admin@plnbatam.com"],
            base_url="https://erm.plnbatam.com",
        )

        html_body = mail.outbox[0].alternatives[0].content
        text_body = mail.outbox[0].body
        self.assertIn("Progress Responden per Bidang / Unit", html_body)
        self.assertIn("Bidang Manajemen Risiko dan Kepatuhan", html_body)
        self.assertIn("50%", html_body)
        self.assertIn("TOTAL", html_body)
        self.assertIn("1/2 responden (50%)", text_body)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM ERM <erm@plnbatam.com>",
    )
    def test_awareness_email_includes_participants_link(self):
        send_awareness_notification(
            self.campaign,
            ["risk.admin@plnbatam.com"],
            base_url="https://erm.plnbatam.com",
        )

        html_body = mail.outbox[0].alternatives[0].content
        text_body = mail.outbox[0].body
        participants_path = reverse("awareness:campaign_participants", args=[self.campaign.pk])
        self.assertIn("Lihat Pengisi Awareness", html_body)
        self.assertIn(f"https://erm.plnbatam.com{participants_path}", html_body)
        self.assertIn(f"https://erm.plnbatam.com{participants_path}", text_body)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="PLNBATAM ERM <erm@plnbatam.com>",
    )
    def test_awareness_email_uses_manual_unit_employee_targets(self):
        unit = Group.objects.create(name="Unit Bisnis Infrastruktur Teknologi Informasi")
        self.user.groups.add(unit)
        PenugasanUnitBisnis.objects.create(
            user=self.other,
            unit_bisnis=unit,
            peran=PenugasanUnitBisnis.ROLE_PAIRING_OFFICER,
        )
        AwarenessUnitTarget.objects.create(
            campaign=self.campaign,
            unit_name="Unit Bisnis Infrastruktur Teknologi Informasi",
            employee_count=23,
            order=1,
        )
        AwarenessAttempt.objects.create(
            user=self.user,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
            correct_count=2,
            wrong_count=0,
            score=100,
            status=AwarenessAttempt.STATUS_PASSED,
            submitted_at=timezone.now(),
        )
        AwarenessAttempt.objects.create(
            user=self.other,
            campaign=self.campaign,
            attempt_number=1,
            total_questions=2,
            correct_count=2,
            wrong_count=0,
            score=100,
            status=AwarenessAttempt.STATUS_PASSED,
            submitted_at=timezone.now(),
        )

        send_awareness_notification(
            self.campaign,
            ["risk.admin@plnbatam.com"],
            base_url="https://erm.plnbatam.com",
        )

        html_body = mail.outbox[0].alternatives[0].content
        text_body = mail.outbox[0].body
        self.assertIn("Unit Bisnis Infrastruktur Teknologi Informasi", html_body)
        self.assertIn(">23<", html_body)
        self.assertIn(">1<", html_body)
        self.assertIn(">22<", html_body)
        self.assertIn("1/23 responden (4%)", text_body)

    def test_admin_send_awareness_test_handles_smtp_auth_error(self):
        self.client.force_login(self.admin)

        import awareness.admin as awareness_admin
        old_send = awareness_admin.send_awareness_notification

        def failing_send(*args, **kwargs):
            raise SMTPAuthenticationError(535, b"5.7.3 Authentication unsuccessful")

        awareness_admin.send_awareness_notification = failing_send
        try:
            response = self.client.get(
                reverse("risk_admin:awareness_campaign_send_test", args=[self.campaign.pk]),
                {"email": "armeizir@plnbatam.com"},
                follow=True,
            )
        finally:
            awareness_admin.send_awareness_notification = old_send

        self.assertEqual(response.status_code, 200)
        message_texts = [str(message) for message in get_messages(response.wsgi_request)]
        self.assertTrue(any("Authentication unsuccessful" in text for text in message_texts))

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.console.EmailBackend",
        DEFAULT_FROM_EMAIL="webmaster@localhost",
    )
    def test_awareness_notification_uses_smtp_from_app_setting_when_active(self):
        app_setting = AppSetting.get_solo()
        app_setting.email_smtp_aktif = True
        app_setting.email_host = "smtp.plnbatam.com"
        app_setting.email_port = 587
        app_setting.email_host_user = "csirt"
        app_setting.email_host_password = "secret"
        app_setting.email_use_tls = True
        app_setting.email_use_ssl = False
        app_setting.default_from_email = "PLNBATAM CSIRT <noreply@plnbatam.com>"
        app_setting.save()

        import awareness.notifications as notifications
        old_get_connection = notifications.get_connection

        def fake_connection(**kwargs):
            kwargs.pop("backend", None)
            return DummySMTPBackend(**kwargs)

        notifications.get_connection = fake_connection
        try:
            sent = send_awareness_notification(
                self.campaign,
                ["armeizir@plnbatam.com"],
                base_url="https://erm.plnbatam.com",
            )
        finally:
            notifications.get_connection = old_get_connection

        self.assertEqual(sent, 1)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].from_email, "PLNBATAM CSIRT <noreply@plnbatam.com>")
        self.assertEqual(DummySMTPBackend.init_kwargs["host"], "smtp.plnbatam.com")
        self.assertEqual(DummySMTPBackend.init_kwargs["username"], "csirt")
