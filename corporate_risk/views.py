import json
from datetime import datetime
from decimal import Decimal
from statistics import mean, median, pstdev

from .models import MonteCarloKorporatResult

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.forms import modelformset_factory
from openpyxl import load_workbook

from .models import RiskMetric, MonteCarloMetricHistory
from masterdata.models import PeriodeLaporan

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, permission_required
from django import forms
from django.utils import timezone


class AssignedMetricHistoryForm(forms.ModelForm):
    class Meta:
        model = MonteCarloMetricHistory
        fields = ("metric_value", "target_value", "keterangan")
        widgets = {"keterangan": forms.Textarea(attrs={"rows": 5})}


def _can_manage_metric_history(user):
    return (
        user.is_active
        and user.is_staff
        and (
            user.has_perm("corporate_risk.add_montecarlometrichistory")
            or user.has_perm("corporate_risk.change_montecarlometrichistory")
        )
    )


def _require_metric_history_permission(request):
    if not _can_manage_metric_history(request.user):
        raise PermissionDenied


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_excel_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %B %Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(value):
    if value in (None, ""):
        return None
    cleaned = str(value).strip().replace(",", "")
    try:
        return Decimal(cleaned)
    except Exception:
        return None


def _resolve_period(value, tanggal_data=None):
    if value not in (None, ""):
        text = str(value).strip()
        periode = (
            PeriodeLaporan.objects.filter(kode_periode__iexact=text).first()
            or PeriodeLaporan.objects.filter(nama_periode__iexact=text).first()
        )
        if periode:
            return periode
        if text.isdigit():
            periode = PeriodeLaporan.objects.filter(pk=int(text)).first()
            if periode:
                return periode

    if tanggal_data:
        return PeriodeLaporan.objects.filter(
            tanggal_mulai__lte=tanggal_data,
            tanggal_selesai__gte=tanggal_data,
        ).first()

    return None


def _import_metric_history_excel(metric, uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel kosong.")

    headers = [_normalize_header(value) for value in rows[0]]
    aliases = {
        "periode": {"periode", "period", "bulan", "month", "kode_periode"},
        "tanggal_data": {"tanggal_data", "tanggal", "date", "tgl_data"},
        "metric_value": {"nilai_aktual", "actual", "realisasi", "metric_value", "nilai"},
        "target_value": {"target", "target_value", "target_rkap"},
        "keterangan": {"keterangan", "catatan", "note", "notes"},
    }

    column_map = {}
    for field, names in aliases.items():
        for idx, header in enumerate(headers):
            if header in names:
                column_map[field] = idx
                break

    if "metric_value" not in column_map:
        raise ValueError("Kolom nilai aktual tidak ditemukan. Gunakan header: Nilai Aktual atau Realisasi.")

    imported = 0
    skipped = 0

    for raw_row in rows[1:]:
        if not raw_row or all(value in (None, "") for value in raw_row):
            continue

        def cell(field):
            idx = column_map.get(field)
            if idx is None or idx >= len(raw_row):
                return None
            return raw_row[idx]

        tanggal_data = _parse_excel_date(cell("tanggal_data"))
        periode = _resolve_period(cell("periode"), tanggal_data=tanggal_data)
        metric_value = _parse_decimal(cell("metric_value"))
        target_value = _parse_decimal(cell("target_value"))
        keterangan = str(cell("keterangan") or "").strip()

        if not periode or metric_value is None:
            skipped += 1
            continue

        if tanggal_data is None:
            tanggal_data = periode.tanggal_selesai

        MonteCarloMetricHistory.objects.update_or_create(
            metric=metric,
            periode=periode,
            defaults={
                "tanggal_data": tanggal_data,
                "metric_value": metric_value,
                "target_value": target_value,
                "keterangan": keterangan,
                "status": MonteCarloMetricHistory.STATUS_UPDATED,
            },
        )
        imported += 1

    return imported, skipped


def _metric_statistics(values):
    values = [float(value) for value in values if value not in (None, "")]
    if not values:
        return None

    growth_rates = []
    for idx in range(1, len(values)):
        previous = values[idx - 1]
        current = values[idx]
        if previous:
            growth_rates.append((current - previous) / previous * 100)

    avg_value = mean(values)
    std_value = pstdev(values) if len(values) > 1 else 0
    coefficient_variation = (std_value / avg_value * 100) if avg_value else 0
    trend = "Stabil"
    if len(values) > 1:
        if values[-1] > values[0] * 1.05:
            trend = "Meningkat"
        elif values[-1] < values[0] * 0.95:
            trend = "Menurun"

    volatility = "Rendah"
    if coefficient_variation > 20:
        volatility = "Tinggi"
    elif coefficient_variation > 10:
        volatility = "Sedang"

    return {
        "count": len(values),
        "min": min(values),
        "max": max(values),
        "mean": avg_value,
        "median": median(values),
        "std_dev": std_value,
        "coefficient_variation": coefficient_variation,
        "avg_growth": mean(growth_rates) if growth_rates else 0,
        "min_growth": min(growth_rates) if growth_rates else 0,
        "max_growth": max(growth_rates) if growth_rates else 0,
        "trend": trend,
        "volatility": volatility,
    }

@login_required
@permission_required("corporate_risk.view_montecarlokorporatresult", raise_exception=True)
def monte_carlo_result_chart(request, pk):
    result = get_object_or_404(MonteCarloKorporatResult, pk=pk)

    history_labels = [x["periode"] for x in result.history_snapshot]
    history_values = [x["value"] for x in result.history_snapshot]

    context = {
        "result": result,
        "history_labels_json": json.dumps(history_labels),
        "history_values_json": json.dumps(history_values),
        "simulation_values_json": json.dumps(result.simulation_snapshot),
    }
    return render(request, "corporate_risk/monte_carlo_chart.html", context)

@staff_member_required
def bulk_metric_input(request, metric_id):
    _require_metric_history_permission(request)
    metric = get_object_or_404(RiskMetric, id=metric_id)
    messages.info(
        request,
        "Input histori metric sekarang tersedia di panel Monte Carlo halaman Profil Risiko Korporat.",
    )
    return redirect(
        f"{reverse('risk_admin:risk_profilrisikokorporatsummary_change', args=[metric.corporate_risk_item.summary_id])}"
        f"#metric-{metric.pk}"
    )

    HistoryFormSet = modelformset_factory(
        MonteCarloMetricHistory,
        fields=(
            "periode",
            "tanggal_data",
            "metric_value",
            "target_value",
            "keterangan",
        ),
        extra=12,
        can_delete=True,
    )

    queryset = MonteCarloMetricHistory.objects.filter(
        metric=metric
    ).order_by("tanggal_data")
    statistics = _metric_statistics(
        queryset.exclude(status=MonteCarloMetricHistory.STATUS_UNUPDATED).values_list(
            "metric_value", flat=True
        )
    )

    if request.method == "POST":
        if "upload_excel" in request.POST:
            uploaded_file = request.FILES.get("excel_file")
            if not uploaded_file:
                messages.error(request, "Pilih file Excel terlebih dahulu.")
            else:
                try:
                    imported, skipped = _import_metric_history_excel(metric, uploaded_file)
                    messages.success(
                        request,
                        f"Upload Excel berhasil. {imported} baris tersimpan, {skipped} baris dilewati.",
                    )
                    return redirect("bulk_metric_input", metric_id=metric.id)
                except Exception as exc:
                    messages.error(request, f"Gagal upload Excel: {exc}")

        formset = HistoryFormSet(
            request.POST,
            queryset=queryset,
        )

        if formset.is_valid():
            instances = formset.save(commit=False)

            for obj in instances:
                obj.metric = metric
                obj.status = MonteCarloMetricHistory.STATUS_UPDATED
                obj.save()

            for obj in formset.deleted_objects:
                obj.delete()

            return redirect(
                "bulk_metric_input",
                metric_id=metric.id,
            )

    else:
        formset = HistoryFormSet(queryset=queryset)

    return render(
        request,
        "corporate_risk/bulk_metric_input.html",
        {
            "metric": metric,
            "formset": formset,
            "statistics": statistics,
        },
    )

    metric = get_object_or_404(RiskMetric, id=metric_id)

    rows = []
    histories = MonteCarloMetricHistory.objects.filter(
        metric=metric
    ).select_related("periode").order_by("tanggal_data")

    for h in histories:
        rows.append({
            "id": h.id,
            "periode_id": h.periode_id,
            "periode": str(h.periode),
            "tanggal_data": h.tanggal_data.isoformat() if h.tanggal_data else "",
            "metric_value": float(h.metric_value) if h.metric_value is not None else None,
            "target_value": float(h.target_value) if h.target_value is not None else None,
            "keterangan": h.keterangan or "",
        })

    return JsonResponse({
        "metric": {
            "id": metric.id,
            "name": metric.name,
            "unit": metric.unit,
            "direction": metric.direction,
            "weight": float(metric.weight),
        },
        "rows": rows,
    })


@staff_member_required
def metric_history_input_menu(request):
    _require_metric_history_permission(request)
    messages.info(
        request,
        "Input Histori / Upload Excel sekarang tersedia di halaman Profil Risiko Korporat.",
    )
    return redirect("risk_admin:risk_profilrisikokorporatsummary_changelist")


@login_required
def assigned_metric_history_input(request, pk):
    history = get_object_or_404(
        MonteCarloMetricHistory.objects.select_related(
            "assigned_to", "metric__corporate_risk_item", "periode"
        ),
        pk=pk,
    )
    if request.user != history.assigned_to and not request.user.is_superuser:
        raise PermissionDenied("Data ini ditugaskan kepada user lain.")
    form = AssignedMetricHistoryForm(request.POST or None, instance=history)
    if request.method == "POST" and form.is_valid():
        history = form.save(commit=False)
        history.status = MonteCarloMetricHistory.STATUS_UPDATED
        history.completed_by = request.user
        history.completed_at = timezone.now()
        history.save()
        messages.success(request, "Data histori berhasil disimpan.")
        return redirect("metric_history_assigned_input", pk=history.pk)
    return render(
        request,
        "corporate_risk/assigned_metric_history_input.html",
        {"history": history, "form": form},
    )
