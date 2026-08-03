from datetime import date, timedelta
from decimal import Decimal
import json
from io import BytesIO
import tempfile

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core import mail
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from masterdata.models import (
    OrganizationUnit,
    OrganizationUnitUserAssignment,
    PeriodeLaporan,
    TahunBuku,
)
from risk.admin import ReAssessmentItemAdmin
from risk.models import (
    AppSetting,
    BagianKontrakManajemen,
    ItemKontrakManajemen,
    KontrakManajemen,
    KPMRIndikatorResmi,
    KPMRPeriode,
    KnowledgeBaseArticle,
    KnowledgeBaseCategory,
    MasterBagianKM,
    MasterTemplateKM,
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    PenugasanUnitBisnis,
    ReAssessmentItem,
    ReAssessmentSummary,
)

from .admin import MonthlyRiskReportAdmin, MonthlyRiskReportAdminForm, MonthlyRiskReportGroupFilter, MonthlyRiskReportItemInline, _monthly_risk_item_label
from .models import (
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
    MonthlyRiskReportImportBatch,
)
from .import_services import (
    IMPORT_PARSER_VERSION,
    _match_item,
    _parse_workbook,
    analyze_import_batch,
    apply_import_batch,
    batch_analysis_is_current,
    extract_risk_number,
    normalize_cause_code,
    normalize_risk_code,
    normalize_treatment_code,
)
from .notifications import send_monthly_report_notification
from .recipient_services import build_approved_report_recipients
from .services import (
    duplicate_approved_report_to_next_month,
    initialize_monthly_report_structure_from_profile,
    initialize_monthly_report_structure_from_reference,
    initialize_monthly_report_structure_from_previous,
    recommended_structure_reference,
    refresh_monthly_report_summary,
    structure_reference_reports,
)


class MonthlyRiskReportAdminTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin_user = User.objects.create_superuser(username="admin", password="secret")
        self.prepared_by = User.objects.create_user(username="prepared", password="secret")
        self.tahun_buku = TahunBuku.objects.create(tahun=2026)
        self.periode = PeriodeLaporan.objects.create(
            tahun_buku=self.tahun_buku,
            kode_periode="2026-02",
            nama_periode="Februari 2026",
            jenis_periode="bulanan",
            tanggal_mulai=date(2026, 2, 1),
            tanggal_selesai=date(2026, 2, 28),
        )

    def _report(self, group_name):
        group = Group.objects.create(name=group_name)
        kontrak = KontrakManajemen.objects.create(
            judul=f"KM {group_name}",
            tahun=2026,
            unit_bisnis=group,
        )
        reassessment = ReAssessmentSummary.objects.create(
            judul=f"Profil Risiko {group_name}",
            tahun=2026,
            unit_bisnis=group,
            kontrak_manajemen=kontrak,
        )
        return MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=self.periode,
            reassessment=reassessment,
            prepared_by=self.prepared_by,
        )

    def test_report_syncs_canonical_km_from_reassessment(self):
        report = self._report("BID SYNC KM")
        self.assertEqual(
            report.kontrak_manajemen_id,
            report.reassessment.kontrak_manajemen_id,
        )

    def test_report_item_syncs_canonical_km_item_from_risk_event(self):
        report = self._report("BID SYNC ITEM")
        risk_event = self._risk_item(report)
        report_item = MonthlyRiskReportItem.objects.create(
            report=report,
            risk_event=risk_event,
        )
        self.assertEqual(report_item.km_item_id, risk_event.km_item_id)

    def _assign_pairing_officer(self, report, username="pairing", email="pairing@example.com"):
        User = get_user_model()
        pairing = User.objects.create_user(username=username, email=email)
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report.reassessment.unit_bisnis,
            user=pairing,
            peran=PenugasanUnitBisnis.ROLE_PAIRING_OFFICER,
        )
        return pairing

    def test_duplicate_approved_report_creates_next_month_draft_with_lineage(self):
        User = get_user_model()
        source = self._report("BID BIS")
        source.kode = "MRR-BIS-2026-02"
        source.status = "approved"
        source.approved_by = self.admin_user
        source.approved_at = timezone.now()
        source.is_locked = True
        source.save()

        risk_officer = User.objects.create_user(username="bis.ro")
        risk_champion = User.objects.create_user(username="bis.rc")
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=source.reassessment.unit_bisnis,
            user=risk_officer,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=source.reassessment.unit_bisnis,
            user=risk_champion,
            peran=PenugasanUnitBisnis.ROLE_RISK_CHAMPION,
        )
        risk_event = self._risk_item(source, no_item=1, no_risiko=1)
        source_item = MonthlyRiskReportItem.objects.create(
            report=source,
            risk_event=risk_event,
            realisasi_nilai_dampak=Decimal("1000"),
            realisasi_nilai_probabilitas=Decimal("50"),
            realisasi_rencana_perlakuan="Realisasi Februari",
            next_action="Lanjutkan mitigasi",
        )
        MonthlyRiskReportChange.objects.create(
            report=source,
            jenis_perubahan=MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
            penjelasan="Perubahan Februari",
        )
        MonthlyRiskReportLossEvent.objects.create(
            report=source,
            nama_kejadian="Kejadian Februari",
        )

        target = duplicate_approved_report_to_next_month(source, self.admin_user)

        self.assertEqual(target.status, "draft")
        self.assertEqual(target.periode.kode_periode, "2026-03")
        self.assertEqual(target.kode, "MRR-BIS-2026-03")
        self.assertEqual(target.copied_from, source)
        self.assertEqual(target.copied_by, self.admin_user)
        self.assertIsNotNone(target.copied_at)
        self.assertEqual(target.prepared_by, risk_officer)
        self.assertEqual(target.reviewed_by, risk_champion)
        self.assertEqual(target.approved_by, self.admin_user)
        self.assertIsNone(target.submitted_at)
        self.assertIsNone(target.approved_at)
        self.assertFalse(target.is_locked)
        self.assertFalse(target.is_aggregated_to_corporate)
        self.assertEqual(target.display_profile_name, "Profil Risiko BID BIS (copy bulan Februari)")

        copied_item = target.items.get()
        self.assertEqual(copied_item.risk_event, source_item.risk_event)
        self.assertEqual(copied_item.realisasi_nilai_dampak, Decimal("1000"))
        self.assertEqual(copied_item.realisasi_nilai_probabilitas, Decimal("50"))
        self.assertEqual(copied_item.realisasi_rencana_perlakuan, "Realisasi Februari")
        self.assertEqual(copied_item.next_action, "Lanjutkan mitigasi")
        self.assertEqual(target.changes.count(), 1)
        self.assertEqual(target.loss_events.count(), 1)
        self.assertEqual(target.submission_logs.get().action, "duplicate")

        with self.assertRaises(ValidationError):
            duplicate_approved_report_to_next_month(source, self.admin_user)

    def test_duplicate_rejects_report_that_is_not_approved(self):
        source = self._report("BID OPS")

        with self.assertRaisesMessage(
            ValidationError,
            "Hanya laporan berstatus Approved yang dapat disalin.",
        ):
            duplicate_approved_report_to_next_month(source, self.admin_user)

    def _risk_item(
        self,
        report,
        no_item=1,
        no_risiko=None,
        no_penyebab_risiko=None,
        peristiwa_risiko=None,
    ):
        item_suffix = f"{no_item}-{no_risiko or no_item}-{no_penyebab_risiko or 'x'}"
        template, _ = MasterTemplateKM.objects.get_or_create(
            tahun=2026,
            defaults={"nama": "Template 2026"},
        )
        master_bagian = MasterBagianKM.objects.create(
            template=template,
            kode_bagian=f"B{report.pk}-{item_suffix}",
            nama_bagian="Keuangan",
            urutan=1,
        )
        bagian = BagianKontrakManajemen.objects.create(
            kontrak=report.reassessment.kontrak_manajemen,
            kode_bagian=f"B{report.pk}-{item_suffix}",
            nama_bagian="Keuangan",
        )
        km_item = ItemKontrakManajemen.objects.create(
            kontrak=report.reassessment.kontrak_manajemen,
            bagian=bagian,
            master_bagian=master_bagian,
            no_urut=1,
            indikator_kinerja_kunci=f"KPI {report.reassessment.unit_bisnis.name}",
            satuan="%",
            bobot=10,
            target="100",
        )
        return ReAssessmentItem.objects.create(
            summary=report.reassessment,
            no_item=no_item,
            km_item=km_item,
            no_risiko=no_risiko or no_item,
            no_penyebab_risiko=no_penyebab_risiko,
            peristiwa_risiko=peristiwa_risiko or f"Risiko {report.reassessment.unit_bisnis.name}",
            deskripsi_peristiwa_risiko="Deskripsi risiko",
            penyebab_risiko="Penyebab",
            rencana_perlakuan_risiko="Mitigasi",
            output_perlakuan_risiko="Output",
        )

    def _import_workbook(self, probability=40, progress=75):
        workbook = Workbook()
        iiia = workbook.active
        iiia.title = "III.A"
        iiib = workbook.create_sheet("III.B")
        iiia.cell(9, 1, "Start pengisian")
        iiib.cell(9, 1, "Start pengisian")
        iiia.cell(10, 2, 1)
        iiia.cell(10, 3, "Risiko BID BIS")
        iiia.cell(10, 12, "Asumsi dampak Februari")
        iiia.cell(10, 13, 1500000)
        iiia.cell(10, 17, 3)
        iiia.cell(10, 25, probability)
        iiia.cell(10, 29, 2)
        iiia.cell(10, 57, "Efektif")
        iiib.cell(10, 2, 1)
        iiib.cell(10, 3, "Risiko BID BIS")
        iiib.cell(10, 6, "BID BIS-1-a")
        iiib.cell(10, 11, "Realisasi mitigasi")
        iiib.cell(10, 12, "Output mitigasi")
        iiib.cell(10, 13, 250000)
        iiib.cell(10, 15, "PIC BIS")
        iiib.cell(10, 28, "Continue")
        iiib.cell(10, 29, "Sesuai jadwal")
        iiib.cell(10, 30, progress)
        iiib.cell(10, 41, "<= 10 hari")
        iiib.cell(10, 42, "Aman")
        stream = BytesIO()
        workbook.save(stream)
        return SimpleUploadedFile(
            "profil_bis_februari.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_excel_import_is_staged_then_applied_after_confirmation(self):
        report = self._report("BID BIS")
        risk_event = self._risk_item(
            report,
            no_item=1,
            no_risiko=1,
            no_penyebab_risiko="a",
            peristiwa_risiko="Risiko BID BIS",
        )
        item = MonthlyRiskReportItem.objects.create(report=report, risk_event=risk_event)
        MasterSkalaDampak.objects.create(nama="Menengah", urutan=3)
        MasterSkalaProbabilitas.objects.create(nama="Jarang", urutan=2)

        with tempfile.TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=self._import_workbook(),
                original_filename="profil_bis_februari.xlsx",
                file_sha256="a" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            row = batch.rows.get(source_reference="III.A:10")
            self.assertEqual(row.validation_level, row.LEVEL_GREEN)
            self.assertEqual(row.user_decision, row.DECISION_IMPORT)
            item.refresh_from_db()
            self.assertIsNone(item.realisasi_nilai_dampak)

            apply_import_batch(batch, self.admin_user)
            item.refresh_from_db()
            batch.refresh_from_db()

        self.assertEqual(batch.status, batch.STATUS_IMPORTED)
        self.assertEqual(item.realisasi_nilai_dampak, Decimal("1500000"))
        self.assertEqual(item.realisasi_nilai_probabilitas, Decimal("40"))
        self.assertEqual(item.progress_pelaksanaan_percent, Decimal("75"))
        self.assertEqual(item.realisasi_rencana_perlakuan, "Realisasi mitigasi")
        self.assertEqual(report.submission_logs.filter(action="import").count(), 1)

    def test_excel_import_rejects_invalid_value_until_user_skips_row(self):
        report = self._report("BID BIS INVALID")
        risk_event = self._risk_item(
            report, no_item=1, no_risiko=1, no_penyebab_risiko="a",
            peristiwa_risiko="Risiko BID BIS",
        )
        MonthlyRiskReportItem.objects.create(report=report, risk_event=risk_event)
        with tempfile.TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=self._import_workbook(probability=140),
                original_filename="invalid.xlsx",
                file_sha256="b" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            row = batch.rows.get(source_reference="III.A:10")
            self.assertEqual(row.validation_level, row.LEVEL_RED)
            with self.assertRaisesMessage(ValidationError, "belum dikonfirmasi"):
                apply_import_batch(batch, self.admin_user)
            row.user_decision = row.DECISION_SKIP
            row.save(update_fields=["user_decision"])
            apply_import_batch(batch, self.admin_user)

    def _anchored_spi_workbook(self, risks=15, treatments=22):
        workbook = Workbook()
        iiia = workbook.active
        iiia.title = "III.A"
        iiib = workbook.create_sheet("III.B")
        iiia.cell(
            10,
            3,
            "Nama peristiwa risiko harus sama persis dengan nama peristiwa risiko "
            "yang ada pada tabel profil risiko",
        )
        iiia.cell(11, 1, " Start Pengisian ")
        for offset in range(risks):
            row = 13 + offset
            iiia.cell(row, 2, offset + 1)
            iiia.cell(row, 3, f"Risiko SPI {offset + 1}")
        iiia.cell(29, 2, "=SUM(B13:B27)")
        iiia.cell(29, 3, "TOTAL")

        iiib.cell(10, 1, "START PENGISIAN")
        for offset in range(treatments):
            row = 11 + offset
            risk_number = (offset % risks) + 1
            iiib.cell(row, 2, risk_number)
            iiib.cell(row, 3, f"Risiko SPI {risk_number}")
            iiib.cell(row, 6, f"SPI {risk_number}{chr(97 + (offset // risks))}")
            iiib.cell(row, 11, f"Rencana {offset + 1}")
        stream = BytesIO()
        workbook.save(stream)
        return SimpleUploadedFile(
            "spi_mei_2026.xlsx",
            stream.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def _bis_header_workbook(self):
        workbook = Workbook()
        iiia = workbook.active
        iiia.title = "III.A"
        iiia["A4"] = "Data Item"
        iiia["B4"] = "No. Risiko"
        iiia["C4"] = "Peristiwa Risiko"
        iiia["L4"] = "Realisasi Residual Risk"
        iiia["L5"] = "Asumsi Perhitungan Dampak"
        iiia.merge_cells("L4:O4")
        iiia["A11"] = "Start pengisian"
        iiia["B11"] = "1-A.2-KEU-1"
        iiia["C11"] = "Risiko BIS Pertama"
        iiia["B12"] = "BID BIS-02-a"
        iiia["C12"] = "Risiko BIS Kedua"
        iiia["B13"] = "TOTAL"
        iiia["C13"] = "TOTAL RISIKO"

        iiib = workbook.create_sheet("III.B")
        iiib["A4"] = "Data Item"
        iiib["B4"] = "Nomor Risiko"
        iiib["C4"] = "Nama Peristiwa Risiko"
        iiib["F4"] = "Kode Penyebab"
        iiib["H4"] = "Rencana Perlakuan"
        iiib["K4"] = "Realisasi Rencana Perlakuan Risiko"
        iiib["A10"] = "Start pengisian"
        iiib["B10"] = "1-A.2-KEU-1"
        iiib["C10"] = "Risiko BIS Pertama"
        iiib["F10"] = "BIS-1a"
        iiib["H10"] = "Mitigasi pertama"
        iiib.merge_cells("B10:B11")
        iiib.merge_cells("C10:C11")
        iiib["F11"] = "BIS-1b"
        iiib["H11"] = "Mitigasi kedua"
        iiib["B12"] = "TOTAL"
        iiib["C12"] = "TOTAL"

        stream = BytesIO()
        workbook.save(stream)
        return SimpleUploadedFile(
            "bis_mei_2026.xlsx",
            stream.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def test_bis_header_layout_parses_codes_merges_and_diagnostics(self):
        report = self._report("BIS HEADER PARSER")
        for number, name in (
            (1, "Risiko BIS Pertama"),
            (2, "Risiko BIS Kedua"),
        ):
            risk_event = self._risk_item(
                report,
                no_item=number,
                no_risiko=number,
                peristiwa_risiko=name,
            )
            MonthlyRiskReportItem.objects.create(
                report=report, risk_event=risk_event
            )
        self.assertEqual(extract_risk_number(1), 1)
        self.assertEqual(extract_risk_number(1.0), 1)
        self.assertEqual(extract_risk_number("1."), 1)
        self.assertEqual(extract_risk_number("01"), 1)
        self.assertEqual(extract_risk_number("BID BIS-02-a"), 2)
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            upload = self._bis_header_workbook()
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=upload,
                original_filename=upload.name,
                file_sha256="b" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            batch.refresh_from_db()

        summary = batch.analysis_summary
        self.assertEqual(summary["source_risks"], 2)
        self.assertEqual(summary["source_treatments"], 2)
        self.assertEqual(summary["source_total"], 4)
        self.assertEqual(summary["parser_diagnostics"]["III.A"]["anchor"], "A11")
        self.assertEqual(
            summary["parser_diagnostics"]["III.A"]["columns"]["risk_number"],
            {"index": 2, "letter": "B"},
        )
        self.assertEqual(
            summary["parser_diagnostics"]["III.A"]["skip_reasons"],
            {"total_or_footer": 1},
        )
        self.assertEqual(
            summary["parser_diagnostics"]["III.B"]["accepted_rows"], 2
        )
        self.assertEqual(
            summary["parser_diagnostics"]["III.B"]["skip_reasons"],
            {"total_or_footer": 1},
        )
        self.assertEqual(batch.rows.count(), 4)
        first_row = batch.rows.order_by("pk").first()
        self.assertEqual(first_row.raw_data["source_row"], 11)
        self.assertEqual(
            first_row.raw_data["source_columns"]["risk_number"]["letter"],
            "B",
        )

    def test_bis_scope_uses_17_profile_items_and_audits_continuations(self):
        report = self._report("BID BIS")
        for number in range(1, 18):
            risk_event = self._risk_item(
                report,
                no_item=number,
                no_risiko=number,
                no_penyebab_risiko="a",
                peristiwa_risiko=f"Risiko BIS {number}",
            )
            MonthlyRiskReportItem.objects.create(
                report=report, risk_event=risk_event
            )

        workbook = Workbook()
        iiia = workbook.active
        iiia.title = "III.A"
        iiia["B4"], iiia["C4"] = "Kode Risiko", "Peristiwa Risiko"
        iiia["A11"] = "Start pengisian"
        iiia["B11"], iiia["C11"] = "1-A.2-KEU-1", "Risiko KEU"
        for number in range(1, 18):
            row = 23 + number
            iiia.cell(row, 2, f"BID BIS-{number}-a")
            iiia.cell(row, 3, f"Risiko BIS {number}")
        iiia["B41"], iiia["C41"] = "BID BIS--e", "Risiko kode salah"
        iiia["B42"], iiia["C42"] = "TOTAL", "TOTAL"

        iiib = workbook.create_sheet("III.B")
        iiib["B4"], iiib["C4"] = "No Risiko", "Nama Peristiwa Risiko"
        iiib["F4"], iiib["H4"] = "Kode Penyebab", "Rencana Perlakuan"
        iiib["A10"] = "Start pengisian"
        iiib["B10"], iiib["C10"] = "1-A.2-KEU-1", "Risiko KEU"
        iiib["F10"], iiib["H10"] = "KEU-1-a", "Mitigasi KEU"
        for number in range(1, 18):
            row = 31 + number
            iiib.cell(row, 2, f"BID BIS-{number}-a")
            iiib.cell(row, 3, f"Risiko BIS {number}")
            iiib.cell(row, 6, f"BIS-{number}-a")
            iiib.cell(row, 8, f"Mitigasi BIS {number}")
        # Merged parent cells reproduce the continuation layout.
        iiib.merge_cells("B32:B33")
        iiib.merge_cells("C32:C33")
        iiib["F33"], iiib["H33"] = "BIS-1-a", "Mitigasi lanjutan BIS 1"

        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        upload = SimpleUploadedFile("bis-actual-layout.xlsx", payload)
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=upload,
                original_filename=upload.name,
                file_sha256="8" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            batch.refresh_from_db()
            second_upload = SimpleUploadedFile(
                "bis-actual-layout.xlsx", payload
            )
            second_batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=second_upload,
                original_filename=second_upload.name,
                file_sha256="9" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(second_batch)

        summary = batch.analysis_summary
        self.assertEqual(summary["target_risks"], 17)
        self.assertEqual(summary["matched_risks"], 17)
        self.assertEqual(summary["outside_target_profile"], 2)
        self.assertEqual(summary["malformed"], 1)
        self.assertEqual(summary["continuation_rows"], 1)
        self.assertEqual(
            summary["parser_diagnostics"]["III.B"]["context_reasons"],
            {"inherited_parent_context": 1},
        )
        self.assertEqual(
            batch.rows.filter(
                match_method="outside_target_profile",
                user_decision="skip",
            ).count(),
            2,
        )
        malformed = batch.rows.get(match_method="malformed_code")
        self.assertEqual(malformed.risk_code, "BID BIS--e")
        self.assertEqual(malformed.raw_data["source_row"], 41)
        self.assertEqual(malformed.raw_data["source_risk_code"], "BID BIS--e")
        self.assertIsNone(malformed.raw_data["source_risk_sequence"])
        self.assertEqual(report.items.count(), 17)
        self.assertEqual(second_batch.analysis_summary["matched_risks"], 17)

    def test_missing_required_bis_headers_has_specific_error(self):
        report = self._report("BIS MISSING HEADER")
        workbook = Workbook()
        workbook.active.title = "III.A"
        workbook["III.A"]["A10"] = "Start pengisian"
        workbook["III.A"]["D10"] = "BID BIS-1-a"
        workbook["III.A"]["E10"] = "Risiko tanpa header"
        workbook.create_sheet("III.B")
        workbook["III.B"]["A10"] = "Start pengisian"
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("missing-bis-header.xlsx", stream.getvalue())
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=upload,
                original_filename=upload.name,
                file_sha256="7" * 64,
                uploaded_by=self.admin_user,
            )
            with self.assertRaisesMessage(
                ValidationError,
                "Kolom Nomor Risiko dan Peristiwa Risiko tidak ditemukan pada sheet III.A.",
            ):
                analyze_import_batch(batch)

    def test_parser_uses_anchor_and_counts_15_risks_plus_22_treatments(self):
        report = self._report("SPI EMPTY IMPORT")
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=self._anchored_spi_workbook(),
                original_filename="spi_mei_2026.xlsx",
                file_sha256="c" * 64,
                uploaded_by=self.admin_user,
            )

            analyze_import_batch(batch)
            batch.refresh_from_db()

        self.assertEqual(batch.parser_version, IMPORT_PARSER_VERSION)
        self.assertEqual(batch.analysis_summary["source_risks"], 15)
        self.assertEqual(batch.analysis_summary["source_treatments"], 22)
        self.assertEqual(batch.analysis_summary["source_total"], 37)
        self.assertIn("belum memiliki item risiko", batch.blocking_reason)
        self.assertFalse(batch.rows.exists())
        self.assertNotIn("III.A:10", batch.ai_summary)

    def test_empty_target_review_is_blocked_and_has_safe_navigation(self):
        report = self._report("SPI EMPTY REVIEW")
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=self._anchored_spi_workbook(),
                original_filename="spi_mei_2026.xlsx",
                file_sha256="1" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            self.client.force_login(self.admin_user)
            response = self.client.get(
                reverse(
                    "risk_admin:"
                    "monthly_report_monthlyriskreport_import_profile_review",
                    args=[report.pk, batch.pk],
                )
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "belum memiliki item risiko")
        self.assertContains(response, "Risiko sumber: 15")
        self.assertContains(response, "Rencana perlakuan: 22")
        self.assertContains(response, "Total sumber: 37")
        self.assertContains(response, "Konfirmasi dan Import")
        self.assertContains(response, "disabled aria-disabled")
        self.assertContains(response, "Kembali ke laporan")
        self.assertNotContains(response, "'realisasi_nilai_dampak':")

    def test_parser_requires_anchor_and_required_sheets(self):
        report = self._report("SPI INVALID TEMPLATE")
        workbook = Workbook()
        workbook.active.title = "III.A"
        workbook.create_sheet("III.B")
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("no-anchor.xlsx", stream.getvalue())
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=upload,
                original_filename=upload.name,
                file_sha256="d" * 64,
                uploaded_by=self.admin_user,
            )
            with self.assertRaisesMessage(
                ValidationError, "Anchor 'Start pengisian' tidak ditemukan"
            ):
                analyze_import_batch(batch)

        workbook = Workbook()
        workbook.active.title = "III.A"
        workbook["III.A"].cell(1, 1, "Start pengisian")
        stream = BytesIO()
        workbook.save(stream)
        upload = SimpleUploadedFile("missing-sheet.xlsx", stream.getvalue())
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=upload,
                original_filename=upload.name,
                file_sha256="e" * 64,
                uploaded_by=self.admin_user,
            )
            with self.assertRaisesMessage(
                ValidationError, "sheet III.A dan III.B"
            ):
                analyze_import_batch(batch)

    def test_risk_number_prevents_duplicate_names_from_cross_matching(self):
        report = self._report("SPI DUPLICATE NAMES")
        risk_five = self._risk_item(
            report,
            no_item=5,
            no_risiko=5,
            peristiwa_risiko="Tim SPI tidak menerima reminder penugasan HCR OCR",
        )
        risk_six = self._risk_item(
            report,
            no_item=6,
            no_risiko=6,
            peristiwa_risiko="Tim SPI tidak menerima reminder penugasan HCR OCR",
        )
        MonthlyRiskReportItem.objects.create(report=report, risk_event=risk_five)
        item_six = MonthlyRiskReportItem.objects.create(
            report=report, risk_event=risk_six
        )
        entry = {
            "no_risiko": 6,
            "normalized_code": "",
            "cause": "",
            "risk_event_text": risk_six.peristiwa_risiko,
        }

        matched, strategy, confidence, candidates = _match_item(report, entry)

        self.assertEqual(matched, item_six)
        self.assertEqual(strategy, "exact_risk_number")
        self.assertEqual(confidence, Decimal("100"))
        self.assertEqual(candidates, [])

    def test_treatment_code_normalization_is_stable(self):
        keys = {
            normalize_treatment_code(value)
            for value in ("SPI-14a", "SPI 14a", "SPI14a", "spi-14A")
        }
        self.assertEqual(keys, {"SPI14A"})
        self.assertNotEqual(
            normalize_treatment_code("SPI-14a"),
            normalize_treatment_code("SPI14b"),
        )

    def test_canonical_bis_codes_preserve_cause_suffix(self):
        equivalent = {
            normalize_risk_code(value)
            for value in ("BID BIS-1-a", "bid bis 1 a", "BID-BIS-1-A")
        }
        self.assertEqual(equivalent, {"BIDBIS1A"})
        self.assertEqual(normalize_cause_code("BIS - 1 - a"), "BIS1A")
        self.assertNotEqual(
            normalize_risk_code("BID BIS-1-a"),
            normalize_risk_code("BID BIS-1-b"),
        )

    def test_treatment_matching_uses_parent_number_and_normalized_code(self):
        report = self._report("SPI TREATMENT MATCH")
        risk_five = self._risk_item(
            report, no_item=5, no_risiko=5, no_penyebab_risiko="a"
        )
        risk_six = self._risk_item(
            report, no_item=6, no_risiko=6, no_penyebab_risiko="a"
        )
        item_five = MonthlyRiskReportItem.objects.create(
            report=report, risk_event=risk_five
        )
        item_six = MonthlyRiskReportItem.objects.create(
            report=report, risk_event=risk_six
        )
        common = {"cause": "a", "risk_event_text": ""}

        matched_five, method_five, _, _ = _match_item(
            report,
            {
                **common,
                "no_risiko": 5,
                "normalized_code": normalize_treatment_code("SPI-5a"),
            },
        )
        matched_six, method_six, _, _ = _match_item(
            report,
            {
                **common,
                "no_risiko": 6,
                "normalized_code": normalize_treatment_code("SPI 6a"),
            },
        )

        self.assertEqual(matched_five, item_five)
        self.assertEqual(matched_six, item_six)
        self.assertEqual(method_five, "exact_risk_number_and_code")
        self.assertEqual(method_six, "exact_risk_number_and_code")

    def test_duplicate_normalized_name_without_number_is_ambiguous(self):
        report = self._report("SPI AMBIGUOUS NAME")
        for number in (11, 12):
            risk = self._risk_item(
                report,
                no_item=number,
                no_risiko=number,
                peristiwa_risiko="Ketidaksesuaian perencanaan (PACA)",
            )
            MonthlyRiskReportItem.objects.create(report=report, risk_event=risk)

        matched, strategy, _, candidates = _match_item(
            report,
            {
                "no_risiko": None,
                "normalized_code": "",
                "cause": "",
                "risk_event_text": "Ketidaksesuaian perencanaan PACA",
            },
        )

        self.assertIsNone(matched)
        self.assertEqual(strategy, "ambiguous")
        self.assertEqual(len(candidates), 2)

    def test_empty_report_can_copy_structure_without_monthly_realization_or_logs(self):
        report = self._report("SPI COPY STRUCTURE")
        january = PeriodeLaporan.objects.create(
            tahun_buku=self.tahun_buku,
            kode_periode="2026-01",
            nama_periode="Januari 2026",
            jenis_periode="bulanan",
            tanggal_mulai=date(2026, 1, 1),
            tanggal_selesai=date(2026, 1, 31),
        )
        source = MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=january,
            reassessment=report.reassessment,
            prepared_by=self.prepared_by,
            status="approved",
            evidence_url="https://example.com/evidence-january",
        )
        risk_event = self._risk_item(report, no_item=1, no_risiko=1)
        source_item = MonthlyRiskReportItem.objects.create(
            report=source,
            risk_event=risk_event,
            km_item=risk_event.km_item,
            realisasi_nilai_dampak=Decimal("999"),
            progress_pelaksanaan_percent=Decimal("100"),
        )

        copied_from, copied_count = (
            initialize_monthly_report_structure_from_previous(report)
        )
        report.refresh_from_db()
        target_item = report.items.get()

        self.assertEqual(copied_from, source)
        self.assertEqual(copied_count, 1)
        self.assertEqual(target_item.risk_event, source_item.risk_event)
        self.assertEqual(target_item.km_item, source_item.km_item)
        self.assertIsNone(target_item.realisasi_nilai_dampak)
        self.assertIsNone(target_item.progress_pelaksanaan_percent)
        self.assertEqual(report.status, "draft")
        self.assertEqual(report.evidence_url, "")
        self.assertFalse(report.submission_logs.exists())

    def test_copy_structure_without_previous_report_fails_safely(self):
        report = self._report("SPI NO PREVIOUS STRUCTURE")

        with self.assertRaisesMessage(
            ValidationError, "Tidak ada laporan bulan sebelumnya"
        ):
            initialize_monthly_report_structure_from_previous(report)

        self.assertFalse(report.items.exists())

    def test_profile_initialization_is_primary_idempotent_and_recovers_failed_batch(self):
        report = self._report("BIS PROFILE INITIALIZATION")
        for number, name in (
            (1, "Risiko BIS Pertama"),
            (2, "Risiko BIS Kedua"),
        ):
            self._risk_item(
                report,
                no_item=number,
                no_risiko=number,
                peristiwa_risiko=name,
            )
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            upload = self._bis_header_workbook()
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=upload,
                original_filename=upload.name,
                file_sha256="6" * 64,
                uploaded_by=self.admin_user,
                status=MonthlyRiskReportImportBatch.STATUS_FAILED,
                error_message="Parser lama gagal.",
            )
            self.client.force_login(self.admin_user)
            upload_url = reverse(
                "risk_admin:"
                "monthly_report_monthlyriskreport_import_profile",
                args=[report.pk],
            )
            review_url = reverse(
                "risk_admin:"
                "monthly_report_monthlyriskreport_import_profile_review",
                args=[report.pk, batch.pk],
            )
            upload_response = self.client.get(upload_url)
            review_response = self.client.get(review_url)
            self.assertContains(
                upload_response, "Ambil Struktur dari Profil Risiko"
            )
            self.assertContains(
                review_response, "Ambil Struktur dari Profil Risiko"
            )

            response = self.client.post(
                review_url,
                {"action": "initialize_profile"},
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        report.refresh_from_db()
        batch.refresh_from_db()
        self.assertEqual(report.status, "draft")
        self.assertEqual(report.items.count(), 2)
        self.assertEqual(batch.status, batch.STATUS_REVIEW)
        self.assertEqual(batch.analysis_summary["source_risks"], 2)
        self.assertEqual(batch.analysis_summary["source_treatments"], 2)
        self.assertEqual(batch.rows.count(), 4)
        self.assertEqual(report.evidence_url, "")
        self.assertIsNone(report.submitted_at)
        self.assertIsNone(report.approved_at)
        self.assertFalse(report.submission_logs.exclude(action="import").exists())

        _, total, created = initialize_monthly_report_structure_from_profile(
            report
        )
        self.assertEqual(total, 2)
        self.assertEqual(created, 0)
        self.assertEqual(report.items.count(), 2)

    def test_approved_future_reference_is_recommended_and_copies_final_structure(self):
        target = self._report("SPI REFERENCE SELECTOR")

        def period(month, name):
            return PeriodeLaporan.objects.create(
                tahun_buku=self.tahun_buku,
                kode_periode=f"2026-{month:02d}",
                nama_periode=f"{name} 2026",
                jenis_periode="bulanan",
                tanggal_mulai=date(2026, month, 1),
                tanggal_selesai=date(2026, month, 28),
            )

        april = period(4, "April")
        may = period(5, "Mei")
        june = period(6, "Juni")
        target.periode = may
        target.save(update_fields=["periode", "updated_at"])
        april_report = MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=april,
            reassessment=target.reassessment,
            prepared_by=self.prepared_by,
            status="draft",
        )
        june_report = MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=june,
            reassessment=target.reassessment,
            prepared_by=self.prepared_by,
            reviewed_by=self.admin_user,
            approved_by=self.admin_user,
            status="approved",
            submitted_at=timezone.now(),
            approved_at=timezone.now(),
            evidence_url="https://example.com/june-evidence",
        )
        risk_events = []
        for number in range(1, 23):
            risk_events.append(
                self._risk_item(
                    target,
                    no_item=number,
                    no_risiko=number,
                    peristiwa_risiko=f"Risiko SPI {number}",
                )
            )
        for risk_event in risk_events[:15]:
            MonthlyRiskReportItem.objects.create(
                report=april_report, risk_event=risk_event
            )
        for risk_event in risk_events:
            MonthlyRiskReportItem.objects.create(
                report=june_report,
                risk_event=risk_event,
                realisasi_nilai_dampak=Decimal("999"),
                progress_pelaksanaan_percent=Decimal("100"),
            )

        references = list(structure_reference_reports(target))
        self.assertEqual([report.pk for report in references], [
            june_report.pk,
            april_report.pk,
        ])
        self.assertEqual(recommended_structure_reference(target), june_report)
        self.assertNotIn(target.pk, [report.pk for report in references])

        copied_from, copied_count = initialize_monthly_report_structure_from_reference(
            target, june_report
        )
        self.assertEqual(copied_from, june_report)
        self.assertEqual(copied_count, 22)
        self.assertEqual(target.items.count(), 22)
        self.assertFalse(
            target.items.filter(realisasi_nilai_dampak__isnull=False).exists()
        )
        self.assertFalse(
            target.items.filter(
                progress_pelaksanaan_percent__isnull=False
            ).exists()
        )
        target.refresh_from_db()
        self.assertEqual(target.status, "draft")
        self.assertIsNone(target.reviewed_by)
        self.assertIsNone(target.approved_by)
        self.assertIsNone(target.submitted_at)
        self.assertIsNone(target.approved_at)
        self.assertEqual(target.evidence_url, "")
        self.assertFalse(target.submission_logs.exists())

        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=target,
                source_file=self._anchored_spi_workbook(),
                original_filename="spi_mei_2026.xlsx",
                file_sha256="f" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            batch.refresh_from_db()

        self.assertEqual(batch.analysis_summary["source_risks"], 15)
        self.assertEqual(batch.analysis_summary["source_treatments"], 22)
        self.assertEqual(batch.analysis_summary["source_total"], 37)
        self.assertEqual(batch.analysis_summary["target_risks"], 22)
        self.assertEqual(batch.analysis_summary["matched_risks"], 15)
        self.assertEqual(batch.analysis_summary["target_only"], 7)
        self.assertEqual(batch.analysis_summary["ambiguous"], 0)
        self.assertEqual(batch.analysis_summary["invalid"], 0)
        self.assertEqual(len(batch.analysis_summary["target_only_ids"]), 7)
        target_only_ids = batch.analysis_summary["target_only_ids"]
        target_only_timestamps = dict(
            target.items.filter(pk__in=target_only_ids).values_list(
                "pk", "updated_at"
            )
        )

        apply_import_batch(batch, self.admin_user)

        self.assertEqual(target.items.count(), 22)
        target.refresh_from_db()
        self.assertEqual(target.status, "draft")
        self.assertEqual(
            dict(
                target.items.filter(pk__in=target_only_ids).values_list(
                    "pk", "updated_at"
                )
            ),
            target_only_timestamps,
        )
        self.assertFalse(
            target.items.filter(
                pk__in=target_only_ids,
                realisasi_nilai_dampak__isnull=False,
            ).exists()
        )

    def test_reference_copy_requires_explicit_confirmation_and_shows_future_warning(self):
        target = self._report("SPI REFERENCE UI")
        may = PeriodeLaporan.objects.create(
            tahun_buku=self.tahun_buku,
            kode_periode="2026-05-ui",
            nama_periode="Mei 2026",
            jenis_periode="bulanan",
            tanggal_mulai=date(2026, 5, 1),
            tanggal_selesai=date(2026, 5, 31),
        )
        june = PeriodeLaporan.objects.create(
            tahun_buku=self.tahun_buku,
            kode_periode="2026-06-ui",
            nama_periode="Juni 2026",
            jenis_periode="bulanan",
            tanggal_mulai=date(2026, 6, 1),
            tanggal_selesai=date(2026, 6, 30),
        )
        target.periode = may
        target.save(update_fields=["periode", "updated_at"])
        source = MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=june,
            reassessment=target.reassessment,
            prepared_by=self.prepared_by,
            status="approved",
        )
        risk_event = self._risk_item(target, no_item=1, no_risiko=1)
        MonthlyRiskReportItem.objects.create(report=source, risk_event=risk_event)
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=target,
                source_file=self._anchored_spi_workbook(),
                original_filename="spi_mei_2026.xlsx",
                file_sha256="9" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            self.client.force_login(self.admin_user)
            url = reverse(
                "risk_admin:"
                "monthly_report_monthlyriskreport_import_profile_review",
                args=[target.pk, batch.pk],
            )
            response = self.client.get(url)
            self.assertContains(response, "Pilih dan Salin Struktur Referensi")
            self.assertContains(response, "Juni 2026")
            self.assertContains(response, "Direkomendasikan")
            self.assertContains(response, "setelah Mei 2026")

            response = self.client.post(
                url,
                {"action": "copy_structure", "reference_report": source.pk},
                follow=True,
            )

        self.assertContains(
            response, "Konfirmasi pilihan laporan referensi wajib diberikan"
        )
        self.assertFalse(target.items.exists())

    def test_parser_version_or_target_change_invalidates_cached_analysis(self):
        report = self._report("SPI CACHE INVALIDATION")
        risk_event = self._risk_item(report, no_item=1, no_risiko=1)
        MonthlyRiskReportItem.objects.create(report=report, risk_event=risk_event)
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=self._import_workbook(),
                original_filename="cache.xlsx",
                file_sha256="2" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            batch.refresh_from_db()
            self.assertTrue(batch_analysis_is_current(batch))

            batch.parser_version = IMPORT_PARSER_VERSION - 1
            batch.save(update_fields=["parser_version"])
            self.assertFalse(batch_analysis_is_current(batch))

            batch.parser_version = IMPORT_PARSER_VERSION
            batch.save(update_fields=["parser_version"])
            another_risk = self._risk_item(report, no_item=2, no_risiko=2)
            MonthlyRiskReportItem.objects.create(
                report=report, risk_event=another_risk
            )
            self.assertFalse(batch_analysis_is_current(batch))

    def test_apply_rolls_back_all_rows_when_a_critical_save_fails(self):
        from unittest.mock import patch

        report = self._report("BID BIS")
        risk_event = self._risk_item(
            report,
            no_item=1,
            no_risiko=1,
            peristiwa_risiko="Risiko BID BIS",
        )
        item = MonthlyRiskReportItem.objects.create(
            report=report, risk_event=risk_event
        )
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            batch = MonthlyRiskReportImportBatch.objects.create(
                report=report,
                source_file=self._import_workbook(),
                original_filename="atomic.xlsx",
                file_sha256="f" * 64,
                uploaded_by=self.admin_user,
            )
            analyze_import_batch(batch)
            original_save = MonthlyRiskReportItem.save

            def fail_on_treatment(instance, *args, **kwargs):
                if instance.realisasi_rencana_perlakuan:
                    raise RuntimeError("simulated critical failure")
                return original_save(instance, *args, **kwargs)

            with patch.object(
                MonthlyRiskReportItem, "save", new=fail_on_treatment
            ), self.assertRaisesMessage(RuntimeError, "simulated critical failure"):
                apply_import_batch(batch, self.admin_user)

        item.refresh_from_db()
        batch.refresh_from_db()
        self.assertIsNone(item.realisasi_nilai_dampak)
        self.assertEqual(batch.status, batch.STATUS_REVIEW)

    def test_group_filter_limits_monthly_reports_by_reassessment_group(self):
        report_aga = self._report("BID AGA")
        self._report("SETPER")
        request = RequestFactory().get(
            "/admin/monthly_report/monthlyriskreport/",
            {"group": str(report_aga.reassessment.unit_bisnis_id)},
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())
        group_filter = MonthlyRiskReportGroupFilter(
            request,
            {"group": str(report_aga.reassessment.unit_bisnis_id)},
            MonthlyRiskReport,
            report_admin,
        )

        queryset = group_filter.queryset(request, report_admin.get_queryset(request))

        self.assertEqual(list(queryset), [report_aga])
        self.assertIn(
            (str(report_aga.reassessment.unit_bisnis_id), "BID AGA"),
            list(group_filter.lookups(request, report_admin)),
        )

    def test_monthly_report_form_uses_searchable_autocomplete_fields(self):
        self.assertEqual(getattr(MonthlyRiskReportItemInline, "autocomplete_fields", ()), ())
        self.assertEqual(
            MonthlyRiskReportAdmin.autocomplete_fields,
            ("reassessment",),
        )
        self.assertIn("web_button", MonthlyRiskReportAdmin.list_display)
        self.assertIn("excel_button", MonthlyRiskReportAdmin.list_display)
        self.assertNotIn("pdf_button", MonthlyRiskReportAdmin.list_display)

    def test_monthly_report_excel_export_uses_official_template_and_kpmr_sheet(self):
        report = self._report("BID XLSX")
        report.status = "approved"
        report.approved_by = self.admin_user
        report.approved_at = timezone.now()
        report.is_locked = True
        report.save(
            update_fields=[
                "status",
                "approved_by",
                "approved_at",
                "is_locked",
            ]
        )

        risk_event = self._risk_item(
            report,
            no_item=1,
            no_risiko=1,
            no_penyebab_risiko="a",
            peristiwa_risiko="Risiko ekspor Excel",
        )
        risk_event.key_risk_indicators = "KRI ekspor"
        risk_event.unit_satuan_kri = "%"
        risk_event.threshold_aman = ">= 95%"
        risk_event.threshold_hati_hati = "90%-94%"
        risk_event.threshold_bahaya = "< 90%"
        risk_event.timeline_2 = 1
        risk_event.save()
        MonthlyRiskReportItem.objects.create(
            report=report,
            risk_event=risk_event,
            realisasi_nilai_dampak=Decimal("1000"),
            realisasi_nilai_probabilitas=Decimal("50"),
            realisasi_rencana_perlakuan="Realisasi mitigasi",
            realisasi_output_perlakuan="Output mitigasi",
            realisasi_biaya_perlakuan=Decimal("100"),
            progress_pelaksanaan_percent=Decimal("100"),
            realisasi_threshold_kri="3. Hijau",
            realisasi_threshold_kri_skor="100",
            status_rencana_perlakuan="continue",
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse(
                "risk_admin:monthly_report_monthlyriskreport_excel",
                args=[report.pk],
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertIn("III.A", workbook.sheetnames)
        self.assertIn("III.B", workbook.sheetnames)
        self.assertIn("KPMR", workbook.sheetnames)
        self.assertEqual(workbook["III.B"]["B11"].value, 1)
        self.assertEqual(workbook["III.B"]["C11"].value, "Risiko ekspor Excel")
        self.assertEqual(workbook["III.B"]["AO11"].value, "3. Hijau")
        self.assertEqual(workbook["KPMR"]["I7"].value, "=H8+H12+H18+(30%*(H22+H25+H28+H31))")

    def test_monthly_report_preparer_is_automatic_for_report_unit(self):
        User = get_user_model()
        report_infra = self._report("UB INFRA")
        report_bes = self._report("UB BES")
        infra_ro = User.objects.create_user(username="infra_ro")
        infra_rc = User.objects.create_user(username="infra_rc")
        infra_member = User.objects.create_user(username="infra_member")
        bes_ro = User.objects.create_user(username="bes_ro")
        bes_rc = User.objects.create_user(username="bes_rc")
        outsider = User.objects.create_user(username="outsider")
        report_infra.reassessment.unit_bisnis.user_set.add(infra_member)
        report_bes.reassessment.unit_bisnis.user_set.add(outsider)
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report_infra.reassessment.unit_bisnis,
            user=infra_ro,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report_infra.reassessment.unit_bisnis,
            user=infra_rc,
            peran=PenugasanUnitBisnis.ROLE_RISK_CHAMPION,
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report_bes.reassessment.unit_bisnis,
            user=bes_ro,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report_bes.reassessment.unit_bisnis,
            user=bes_rc,
            peran=PenugasanUnitBisnis.ROLE_RISK_CHAMPION,
        )

        form = MonthlyRiskReportAdminForm(instance=report_infra)
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        self.assertNotIn("prepared_by", form.fields)
        self.assertEqual(report_admin.prepared_by_display(report_infra), "infra_ro")
        self.assertEqual(list(form.fields["reviewed_by"].queryset), [infra_rc])
        self.assertEqual(form.fields["reviewed_by"].initial, infra_rc.pk)
        self.assertEqual(form.initial["reviewed_by"], infra_rc.pk)
        self.assertTrue(form.fields["reviewed_by"].disabled)
        self.assertEqual(list(form.fields["approved_by"].queryset), [infra_member])

    def test_monthly_report_form_handles_signer_fields_excluded_by_admin(self):
        report_infra = self._report("UB INFRA")

        class ReadonlySignerForm(MonthlyRiskReportAdminForm):
            class Meta(MonthlyRiskReportAdminForm.Meta):
                exclude = MonthlyRiskReportAdminForm.Meta.exclude + (
                    "prepared_by",
                    "reviewed_by",
                    "approved_by",
                )

        form = ReadonlySignerForm(instance=report_infra)

        self.assertNotIn("prepared_by", form.fields)
        self.assertNotIn("reviewed_by", form.fields)
        self.assertNotIn("approved_by", form.fields)

    def test_monthly_report_status_is_readonly_after_saved(self):
        report_infra = self._report("INFRA")
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/change/"
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        readonly_fields = report_admin.get_readonly_fields(request, report_infra)

        self.assertIn("status", readonly_fields)

    def test_monthly_report_flow_moves_draft_to_submitted_to_under_review_to_approved(self):
        User = get_user_model()
        reviewer = User.objects.create_user(username="reviewer")
        approver = User.objects.create_user(username="approver")
        report_infra = self._report("INFRA")
        report_infra.reviewed_by = reviewer
        report_infra.approved_by = approver
        report_infra.save(update_fields=["reviewed_by", "approved_by"])
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report_infra.reassessment.unit_bisnis,
            user=self.prepared_by,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())
        report_infra.evidence_url = (
            "https://brightbox.plnbatam.com/drive/d/f/test-evidence"
        )
        report_infra.save(update_fields=["evidence_url"])
        report_admin._apply_flow_action(report_infra, "submit", self.admin_user)
        report_infra.refresh_from_db()
        self.assertEqual(report_infra.status, "submitted")
        self.assertIsNotNone(report_infra.submitted_at)

        report_admin._apply_flow_action(report_infra, "review", reviewer)
        report_infra.refresh_from_db()
        self.assertEqual(report_infra.status, "under_review")

        report_admin._apply_flow_action(report_infra, "approve", approver)
        report_infra.refresh_from_db()
        self.assertEqual(report_infra.status, "approved")
        self.assertIsNotNone(report_infra.approved_at)
        self.assertEqual(
            list(report_infra.submission_logs.order_by("action_at").values_list("action", flat=True)),
            ["submit", "review", "approve"],
        )

    def test_reviewer_can_return_submitted_report_to_drafter_with_required_comment(self):
        User = get_user_model()
        reviewer = User.objects.create_user(
            username="revision.reviewer",
            email="reviewer@example.com",
        )
        report = self._report("INFRA REVISION")
        report.status = "submitted"
        report.reviewed_by = reviewer
        report.save(update_fields=["status", "reviewed_by"])
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        with self.assertRaisesMessage(ValidationError, "Komentar koreksi wajib diisi"):
            report_admin._apply_flow_action(report, "revise", reviewer)

        report_admin._apply_flow_action(
            report,
            "revise",
            reviewer,
            note="Perbaiki nilai residual dan tambahkan dasar perhitungan.",
        )
        report.refresh_from_db()

        self.assertEqual(report.status, "revision")
        revision_log = report.submission_logs.get(action="revise")
        self.assertEqual(revision_log.action_by, reviewer)
        self.assertIn("Perbaiki nilai residual", revision_log.note)
        rendered_comment = str(report_admin.latest_revision_comment(report))
        self.assertIn("revision.reviewer", rendered_comment)
        self.assertIn("Perbaiki nilai residual", rendered_comment)

    def test_approver_can_return_under_review_report_to_drafter(self):
        User = get_user_model()
        approver = User.objects.create_user(username="revision.approver")
        report = self._report("INFRA APPROVER REVISION")
        report.status = "under_review"
        report.approved_by = approver
        report.approved_at = timezone.now()
        report.save(update_fields=["status", "approved_by", "approved_at"])
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        report_admin._apply_flow_action(
            report,
            "revise",
            approver,
            note="Lengkapi eviden perlakuan risiko.",
        )
        report.refresh_from_db()

        self.assertEqual(report.status, "revision")
        self.assertIsNone(report.approved_at)

    def test_monthly_report_submit_requires_evidence_on_nas(self):
        report = self._report("BID AGA EVIDENCE")
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report.reassessment.unit_bisnis,
            user=self.prepared_by,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        with self.assertRaisesMessage(ValidationError, "minimal satu Link Eviden"):
            report_admin._apply_flow_action(report, "submit", self.admin_user)

    def test_monthly_report_evidence_accepts_any_https_domain(self):
        report = self._report("BID AGA INVALID EVIDENCE")
        report.evidence_url = "https://example.com/evidence.pdf"

        report.full_clean()

    def test_monthly_report_evidence_rejects_http_link(self):
        report = self._report("BID AGA HTTP EVIDENCE")
        report.evidence_url = "http://example.com/evidence.pdf"

        with self.assertRaisesMessage(ValidationError, "menggunakan HTTPS"):
            report.full_clean()

    def test_monthly_report_flow_button_matches_current_status(self):
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        self.assertEqual(report_admin._flow_action_for_status("draft"), ("submit", "Submit Laporan"))
        self.assertEqual(report_admin._flow_action_for_status("revision"), ("submit", "Submit Ulang"))
        self.assertEqual(report_admin._flow_action_for_status("submitted"), ("review", "Review & Paraf"))
        self.assertEqual(report_admin._flow_action_for_status("under_review"), ("approve", "Approve"))
        self.assertIsNone(report_admin._flow_action_for_status("approved"))

    def test_monthly_report_admin_loads_select2_for_inline_risk_event_dropdown(self):
        media = str(MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite()).media)

        self.assertIn("admin/js/vendor/select2/select2.full.js", media)
        self.assertIn("monthly_report_items_searchable.js", media)

    def test_monthly_report_web_view_returns_report_context(self):
        report_infra = self._report("INFRA")
        risk_item = self._risk_item(
            report_infra,
            no_item=1,
            no_risiko=1,
            no_penyebab_risiko="a",
        )
        MonthlyRiskReportItem.objects.create(report=report_infra, risk_event=risk_item)
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/web/"
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.web_report_view(request, str(report_infra.pk))
        response.render()

        self.assertEqual(response.context_data["report"], report_infra)
        self.assertEqual(len(response.context_data["iiia_rows"]), 1)
        self.assertEqual(len(response.context_data["iiib_rows"]), 1)
        self.assertIn(b"LAPORAN REALISASI MANAJEMEN RISIKO", response.content)
        self.assertIn(b"III.A. FORMAT TABEL REALISASI RISIKO RESIDUAL BULANAN", response.content)

    def test_refresh_summary_counts_only_high_scores_from_20(self):
        report_infra = self._report("INFRA")
        moderate_event = self._risk_item(report_infra, no_item=1, no_risiko=1)
        high_event = self._risk_item(report_infra, no_item=2, no_risiko=2)
        moderate_report_item = MonthlyRiskReportItem.objects.create(
            report=report_infra,
            risk_event=moderate_event,
        )
        high_report_item = MonthlyRiskReportItem.objects.create(
            report=report_infra,
            risk_event=high_event,
        )
        MonthlyRiskReportItem.objects.filter(pk=moderate_report_item.pk).update(
            realisasi_skor_risiko=19,
            realisasi_level_risiko="Moderate to High",
        )
        MonthlyRiskReportItem.objects.filter(pk=high_report_item.pk).update(
            realisasi_skor_risiko=20,
            realisasi_level_risiko="High",
        )

        refresh_monthly_report_summary(report_infra)
        report_infra.refresh_from_db()

        self.assertEqual(report_infra.total_risiko, 2)
        self.assertEqual(report_infra.total_high, 1)

    def test_monthly_report_item_inline_cost_absorption_is_read_only(self):
        inline = MonthlyRiskReportItemInline(MonthlyRiskReport, AdminSite())

        self.assertIn("persentase_serapan_biaya", inline.readonly_fields)

    def test_monthly_report_item_calculates_cost_absorption_from_budget_formula(self):
        report_infra = self._report("INFRA")
        risk_item = self._risk_item(report_infra, no_item=1, no_risiko=1)
        risk_item.biaya_perlakuan_risiko = Decimal("100.00")
        risk_item.save()

        report_item = MonthlyRiskReportItem.objects.create(
            report=report_infra,
            risk_event=risk_item,
            realisasi_biaya_perlakuan=Decimal("150.00"),
            persentase_serapan_biaya=Decimal("12.00"),
        )

        self.assertEqual(report_item.persentase_serapan_biaya, Decimal("150.00"))

    def test_peta_risiko_iiic_includes_automatic_kpmr_calculation(self):
        report_infra = self._report("INFRA")
        report_infra.status = "approved"
        report_infra.save(update_fields=["status"])
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/peta-risiko-iiic/"
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.peta_risiko_iiic_view(request, str(report_infra.pk))

        self.assertEqual(response.context_data["kpmr_quarter"], 1)
        self.assertEqual(response.context_data["kpmr_calculation"].unit, report_infra.reassessment.unit_bisnis)
        self.assertEqual(response.context_data["kpmr_calculation"].report_count, 1)

    def test_peta_risiko_iiic_hides_kpmr_for_draft_report(self):
        report_infra = self._report("INFRA DRAFT")
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/peta-risiko-iiic/"
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(
            MonthlyRiskReport,
            AdminSite(),
        )

        response = report_admin.peta_risiko_iiic_view(
            request,
            str(report_infra.pk),
        )
        response.render()

        self.assertIsNone(response.context_data["kpmr_calculation"])
        self.assertNotIn(b"KPMR Otomatis Bulanan", response.content)

    def test_peta_risiko_iiic_shows_kpmr_after_report_is_submitted(self):
        for status in ("submitted", "under_review"):
            with self.subTest(status=status):
                report_infra = self._report(f"INFRA {status}")
                report_infra.status = status
                report_infra.save(update_fields=["status"])

                request = RequestFactory().get(
                    f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/peta-risiko-iiic/"
                )
                request.user = self.admin_user
                report_admin = MonthlyRiskReportAdmin(
                    MonthlyRiskReport,
                    AdminSite(),
                )

                response = report_admin.peta_risiko_iiic_view(
                    request,
                    str(report_infra.pk),
                )
                response.render()

                self.assertIsNotNone(
                    response.context_data["kpmr_calculation"]
                )
                self.assertIn(
                    b"KPMR Otomatis Bulanan",
                    response.content,
                )

    def test_peta_risiko_iiic_exposes_previous_and_next_month_for_same_profile(self):
        report_infra = self._report("INFRA")
        january = PeriodeLaporan.objects.create(
            tahun_buku=self.tahun_buku,
            kode_periode="2026-01",
            nama_periode="Januari 2026",
            jenis_periode="bulanan",
            tanggal_mulai=date(2026, 1, 1),
            tanggal_selesai=date(2026, 1, 31),
        )
        march = PeriodeLaporan.objects.create(
            tahun_buku=self.tahun_buku,
            kode_periode="2026-03",
            nama_periode="Maret 2026",
            jenis_periode="bulanan",
            tanggal_mulai=date(2026, 3, 1),
            tanggal_selesai=date(2026, 3, 31),
        )
        previous_report = MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=january,
            reassessment=report_infra.reassessment,
            prepared_by=self.prepared_by,
        )
        next_report = MonthlyRiskReport.objects.create(
            tahun_buku=self.tahun_buku,
            periode=march,
            reassessment=report_infra.reassessment,
            prepared_by=self.prepared_by,
        )
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/peta-risiko-iiic/"
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.peta_risiko_iiic_view(request, str(report_infra.pk))

        self.assertEqual(response.context_data["previous_report"], previous_report)
        self.assertEqual(response.context_data["next_report"], next_report)

    def test_peta_risiko_iiic_calculates_kpmr_from_monthly_data_even_when_saved_result_exists(self):
        report_infra = self._report("INFRA")
        report_infra.status = "approved"
        report_infra.save(update_fields=["status"])
        period = KPMRPeriode.objects.create(
            tahun=2026,
            triwulan=1,
            unit_bisnis=report_infra.reassessment.unit_bisnis,
            skor_total=Decimal("81.00"),
            rating="FAIR",
        )
        KPMRIndikatorResmi.objects.create(
            periode=period,
            kode="I1",
            nama="Pencapaian Nilai Eksposur Risiko dibandingkan target Risiko Residual",
            bobot=Decimal("30.00"),
            jawaban="b",
            hasil=Decimal("60.00"),
            skor=Decimal("18.00"),
        )
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/peta-risiko-iiic/"
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.peta_risiko_iiic_view(request, str(report_infra.pk))

        self.assertNotEqual(response.context_data["kpmr_calculation"].score_total, Decimal("81.00"))
        self.assertIn(
            "Belum ada data eksposur lengkap maupun pasangan skor residual-target yang dapat dihitung.",
            response.context_data["kpmr_calculation"].notes,
        )

    def test_monthly_report_notification_sends_prepare_stage_to_risk_office_and_cc_pairing(self):
        report_infra = self._report("INFRA")
        User = get_user_model()
        first_officer = User.objects.create_user(username="risk.office.1", email="risk.office.1@example.com")
        second_officer = User.objects.create_user(username="risk.office.2", email="risk.office.2@example.com")
        first_officer.first_name = "Risk"
        first_officer.last_name = "Officer Satu"
        first_officer.save(update_fields=["first_name", "last_name"])
        second_officer.first_name = "Risk"
        second_officer.last_name = "Officer Dua"
        second_officer.save(update_fields=["first_name", "last_name"])
        for officer in (first_officer, second_officer):
            PenugasanUnitBisnis.objects.create(
                unit_bisnis=report_infra.reassessment.unit_bisnis,
                user=officer,
                peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
            )
        self._assign_pairing_officer(report_infra)

        sent = send_monthly_report_notification(report_infra, base_url="https://erm.plnbatam.com")

        self.assertEqual(sent, 1)
        self.assertCountEqual(
            mail.outbox[0].to,
            ["risk.office.1@example.com", "risk.office.2@example.com"],
        )
        self.assertEqual(mail.outbox[0].cc, [])
        self.assertEqual(mail.outbox[0].bcc, ["pairing@example.com"])
        self.assertNotIn("[MODE UJI COBA]", mail.outbox[0].body)
        self.assertIn("Yth. Risk Officer Dua; Risk Officer Satu,", mail.outbox[0].body)
        self.assertIn(
            "Yth. Risk Officer Dua; Risk Officer Satu,",
            mail.outbox[0].alternatives[0].content,
        )
        self.assertIn("Pairing Officer", mail.outbox[0].body)
        self.assertIn("Input Laporan Risiko Bulanan", mail.outbox[0].subject)
        self.assertIn("Februari 2026", mail.outbox[0].body)
        self.assertIn("5 Maret 2026", mail.outbox[0].body)
        self.assertIn(
            "https://erm.plnbatam.com/admin/monthly_report/monthlyriskreport/",
            mail.outbox[0].body,
        )

    def test_monthly_report_review_notification_still_uses_test_email_when_configured(self):
        app_setting = AppSetting.get_solo()
        app_setting.monthly_report_notification_test_email = "armeizir@plnbatam.com"
        app_setting.save(update_fields=["monthly_report_notification_test_email"])
        User = get_user_model()
        reviewer = User.objects.create_user(username="reviewer", email="reviewer@example.com")
        report_infra = self._report("INFRA")
        report_infra.status = "submitted"
        report_infra.reviewed_by = reviewer
        report_infra.save(update_fields=["status", "reviewed_by"])

        sent = send_monthly_report_notification(report_infra, base_url="https://erm.plnbatam.com")

        self.assertEqual(sent, 1)
        self.assertEqual(mail.outbox[0].to, ["armeizir@plnbatam.com"])
        self.assertIn("[MODE UJI COBA]", mail.outbox[0].body)
        self.assertIn("Paraf / Review Laporan Risiko Bulanan", mail.outbox[0].subject)
        self.assertIn("Februari 2026", mail.outbox[0].body)
        self.assertIn(
            "https://erm.plnbatam.com/admin/monthly_report/monthlyriskreport/",
            mail.outbox[0].body,
        )

    def test_revision_notification_sends_comment_to_all_drafters_and_bcc_pairing(self):
        app_setting = AppSetting.get_solo()
        app_setting.monthly_report_notification_test_email = ""
        app_setting.save(update_fields=["monthly_report_notification_test_email"])
        report = self._report("INFRA CORRECTION EMAIL")
        report.status = "revision"
        report.save(update_fields=["status"])
        User = get_user_model()
        first_officer = User.objects.create_user(
            username="revision.officer.1",
            email="revision.officer.1@example.com",
        )
        second_officer = User.objects.create_user(
            username="revision.officer.2",
            email="revision.officer.2@example.com",
        )
        for officer in (first_officer, second_officer):
            PenugasanUnitBisnis.objects.create(
                unit_bisnis=report.reassessment.unit_bisnis,
                user=officer,
                peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
            )
        pairing = self._assign_pairing_officer(
            report,
            username="revision.pairing",
            email="revision.pairing@example.com",
        )

        sent = send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            correction_note="Perbaiki nilai residual dan unggah eviden pendukung.",
        )

        self.assertEqual(sent, 1)
        self.assertCountEqual(
            mail.outbox[-1].to,
            [
                "revision.officer.1@example.com",
                "revision.officer.2@example.com",
            ],
        )
        self.assertEqual(mail.outbox[-1].bcc, [pairing.email])
        self.assertIn("Koreksi Laporan Risiko Bulanan", mail.outbox[-1].subject)
        self.assertIn("Perbaiki nilai residual", mail.outbox[-1].body)
        self.assertIn("Submit Ulang", mail.outbox[-1].body)
        self.assertIn(
            "Perbaiki nilai residual",
            mail.outbox[-1].alternatives[0].content,
        )

    def test_monthly_report_workflow_notifications_include_pairing_and_superior_chain(self):
        User = get_user_model()
        app_setting = AppSetting.get_solo()
        app_setting.monthly_report_notification_test_email = ""
        app_setting.save(update_fields=["monthly_report_notification_test_email"])
        report = self._report("INFRA WORKFLOW")
        reviewer = User.objects.create_user(
            username="workflow.reviewer", email="reviewer@example.com"
        )
        approver = User.objects.create_user(
            username="workflow.approver", email="approver@example.com"
        )
        pairing = self._assign_pairing_officer(
            report, username="workflow.pairing", email="pairing.workflow@example.com"
        )
        directorate = OrganizationUnit.objects.create(
            code="ORG-DIR",
            name="DIREKTORAT",
        )
        mrk_unit = OrganizationUnit.objects.create(
            code="ORG-MRK",
            name="BID MRK",
            parent=directorate,
        )
        risk_subunit = OrganizationUnit.objects.create(
            code="ORG-RISK",
            name="SBID RIS",
            parent=mrk_unit,
        )
        subunit_head = User.objects.create_user(
            username="workflow.subunit.head",
            email="subunit.head@example.com",
        )
        mrk_head = User.objects.create_user(
            username="workflow.mrk.head",
            email="mrk.head@example.com",
        )
        director = User.objects.create_user(
            username="workflow.director",
            email="director@example.com",
        )
        unrelated_mrk_user = User.objects.create_user(
            username="workflow.unrelated.mrk",
            email="unrelated.mrk@example.com",
        )
        mrk_group = Group.objects.create(name="BID MRK")
        unrelated_mrk_user.groups.add(mrk_group)
        OrganizationUnitUserAssignment.objects.create(
            user=pairing,
            organization_unit=risk_subunit,
        )
        OrganizationUnitUserAssignment.objects.create(
            user=subunit_head,
            organization_unit=risk_subunit,
            is_unit_head=True,
        )
        OrganizationUnitUserAssignment.objects.create(
            user=mrk_head,
            organization_unit=mrk_unit,
            is_unit_head=True,
        )
        OrganizationUnitUserAssignment.objects.create(
            user=director,
            organization_unit=directorate,
            is_unit_head=True,
        )
        report.reviewed_by = reviewer
        report.approved_by = approver

        report.status = "submitted"
        report.save(update_fields=["status", "reviewed_by", "approved_by"])
        send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
        )
        self.assertEqual(mail.outbox[-1].to, [reviewer.email])
        self.assertEqual(mail.outbox[-1].cc, [approver.email])
        self.assertEqual(
            mail.outbox[-1].bcc,
            [
                pairing.email,
                subunit_head.email,
                mrk_head.email,
                director.email,
            ],
        )
        self.assertIn("Total KPMR", mail.outbox[-1].body)

        report.status = "under_review"
        report.save(update_fields=["status"])
        send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
        )
        self.assertEqual(mail.outbox[-1].to, [approver.email])
        self.assertEqual(mail.outbox[-1].cc, [reviewer.email])
        self.assertEqual(
            mail.outbox[-1].bcc,
            [
                pairing.email,
                subunit_head.email,
                mrk_head.email,
                director.email,
            ],
        )
        self.assertIn("Total KPMR", mail.outbox[-1].body)

        report.status = "approved"
        report.save(update_fields=["status"])
        send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            approved_transition=True,
        )
        self.assertEqual(mail.outbox[-1].to, [pairing.email])
        self.assertEqual(
            mail.outbox[-1].cc,
            [reviewer.email, approver.email],
        )
        self.assertEqual(
            mail.outbox[-1].bcc,
            [subunit_head.email, mrk_head.email, director.email],
        )
        self.assertNotIn(unrelated_mrk_user.email, mail.outbox[-1].to)
        self.assertNotIn(unrelated_mrk_user.email, mail.outbox[-1].cc)
        self.assertIn("Total KPMR", mail.outbox[-1].body)
        self.assertIn("telah disetujui", mail.outbox[-1].body)

    def test_approved_notification_falls_back_to_pairing_when_org_assignment_missing(self):
        app_setting = AppSetting.get_solo()
        app_setting.monthly_report_notification_test_email = ""
        app_setting.save(update_fields=["monthly_report_notification_test_email"])
        report = self._report("INFRA APPROVED PAIRING ONLY")
        pairing = self._assign_pairing_officer(
            report,
            username="approved.pairing.only",
            email="approved.pairing.only@example.com",
        )
        report.status = "approved"
        report.save(update_fields=["status"])

        send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            approved_transition=True,
        )

        self.assertEqual(mail.outbox[-1].to, [pairing.email])
        self.assertEqual(mail.outbox[-1].cc, [])
        self.assertEqual(mail.outbox[-1].bcc, [])

    def test_approved_recipient_resolution_filters_assignments_and_invalid_emails(self):
        User = get_user_model()
        report = self._report("APPROVED RECIPIENT FILTERS")
        pairing = self._assign_pairing_officer(
            report, username="filter.pairing", email="Pairing@Example.com"
        )
        today = timezone.localdate()
        root = OrganizationUnit.objects.create(code="FILTER-ROOT", name="Root")
        expired_level = OrganizationUnit.objects.create(
            code="FILTER-EXPIRED", name="Expired", parent=root
        )
        future_level = OrganizationUnit.objects.create(
            code="FILTER-FUTURE", name="Future", parent=expired_level
        )
        inactive_level = OrganizationUnit.objects.create(
            code="FILTER-INACTIVE", name="Inactive", parent=future_level
        )
        current = OrganizationUnit.objects.create(
            code="FILTER-CURRENT", name="Current", parent=inactive_level
        )
        OrganizationUnitUserAssignment.objects.create(
            user=pairing, organization_unit=current
        )
        ordinary = User.objects.create_user(
            username="filter.member", email="member@example.com"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=ordinary, organization_unit=current, utama=False
        )
        blank_head = User.objects.create_user(username="filter.blank", email="")
        OrganizationUnitUserAssignment.objects.create(
            user=blank_head,
            organization_unit=current,
            is_unit_head=True,
            utama=False,
        )
        inactive_head = User.objects.create_user(
            username="filter.inactive", email="inactive@example.com"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=inactive_head,
            organization_unit=inactive_level,
            is_unit_head=True,
            utama=False,
            aktif=False,
        )
        future_head = User.objects.create_user(
            username="filter.future", email="future@example.com"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=future_head,
            organization_unit=future_level,
            is_unit_head=True,
            utama=False,
            tanggal_mulai=today + timedelta(days=1),
        )
        expired_head = User.objects.create_user(
            username="filter.expired", email="expired@example.com"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=expired_head,
            organization_unit=expired_level,
            is_unit_head=True,
            utama=False,
            tanggal_mulai=today - timedelta(days=2),
            tanggal_selesai=today - timedelta(days=1),
        )
        invalid_head = User.objects.create_user(
            username="filter.invalid", email="bukan-email"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=invalid_head,
            organization_unit=root,
            is_unit_head=True,
            utama=False,
        )

        recipients = build_approved_report_recipients(report, on_date=today)

        self.assertEqual(recipients.to, [pairing.email])
        self.assertEqual(recipients.cc, [])
        excluded = " ".join(recipients.excluded)
        self.assertIn("filter.blank", excluded)
        self.assertIn("filter.invalid", excluded)
        for email in (
            ordinary.email,
            inactive_head.email,
            future_head.email,
            expired_head.email,
        ):
            self.assertNotIn(email, recipients.cc)

    def test_approved_recipient_resolution_deduplicates_case_insensitively_and_excludes_to(self):
        User = get_user_model()
        report = self._report("APPROVED RECIPIENT DEDUP")
        pairing = self._assign_pairing_officer(
            report, username="dedup.pairing", email="same@example.com"
        )
        root = OrganizationUnit.objects.create(code="DEDUP-ROOT", name="Root")
        child = OrganizationUnit.objects.create(
            code="DEDUP-CHILD", name="Child", parent=root
        )
        OrganizationUnitUserAssignment.objects.create(
            user=pairing,
            organization_unit=child,
            is_unit_head=True,
        )
        root_head = User.objects.create_user(
            username="dedup.root", email="SAME@EXAMPLE.COM"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=root_head,
            organization_unit=root,
            is_unit_head=True,
            utama=False,
        )

        recipients = build_approved_report_recipients(report)

        self.assertEqual(recipients.to, ["same@example.com"])
        self.assertEqual(recipients.cc, [])
        self.assertEqual(len(recipients.excluded), 2)

    def test_approved_recipient_resolution_skips_headless_unit_and_handles_cycle(self):
        User = get_user_model()
        report = self._report("APPROVED RECIPIENT CYCLE")
        pairing = self._assign_pairing_officer(
            report, username="cycle.pairing", email="cycle.pairing@example.com"
        )
        first = OrganizationUnit.objects.create(code="CYCLE-A", name="A")
        second = OrganizationUnit.objects.create(code="CYCLE-B", name="B", parent=first)
        OrganizationUnit.objects.filter(pk=first.pk).update(parent=second)
        first.refresh_from_db()
        head = User.objects.create_user(
            username="cycle.head", email="cycle.head@example.com"
        )
        OrganizationUnitUserAssignment.objects.create(
            user=pairing, organization_unit=first
        )
        OrganizationUnitUserAssignment.objects.create(
            user=head,
            organization_unit=second,
            is_unit_head=True,
            utama=False,
        )

        recipients = build_approved_report_recipients(report)

        self.assertEqual(recipients.to, [pairing.email])
        self.assertEqual(recipients.cc, [head.email])

    def test_approved_notification_is_transition_only_and_keeps_subject_body(self):
        app_setting = AppSetting.get_solo()
        app_setting.monthly_report_notification_test_email = ""
        app_setting.save(update_fields=["monthly_report_notification_test_email"])
        report = self._report("APPROVED IDEMPOTENT")
        pairing = self._assign_pairing_officer(
            report, username="idempotent.pairing", email="idempotent@example.com"
        )
        report.status = "approved"
        report.save(update_fields=["status"])

        sent = send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            approved_transition=True,
        )
        report.save()

        self.assertEqual(sent, 1)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, [pairing.email])
        self.assertIn("Laporan Risiko Bulanan Telah Disetujui", mail.outbox[0].subject)
        self.assertIn("telah disetujui", mail.outbox[0].body)
        with self.assertRaisesMessage(
            ValidationError,
            "Notifikasi Approved hanya dikirim saat transisi status menjadi Approved.",
        ):
            send_monthly_report_notification(
                report, base_url="https://erm.plnbatam.com"
            )
        self.assertEqual(len(mail.outbox), 1)

    def test_risk_item_autocomplete_is_limited_by_selected_reassessment(self):
        report_infra = self._report("INFRA")
        report_bes = self._report("BES")
        infra_item = self._risk_item(report_infra, no_item=1)
        self._risk_item(report_bes, no_item=2)
        request = RequestFactory().get(
            "/admin/autocomplete/",
            {"reassessment": str(report_infra.reassessment_id), "term": "Risiko"},
        )
        request.user = self.admin_user
        item_admin = ReAssessmentItemAdmin(ReAssessmentItem, AdminSite())

        queryset = item_admin.get_queryset(request)

        self.assertEqual(list(queryset), [infra_item])

    def test_inline_risk_event_field_is_limited_by_parent_report(self):
        report_infra = self._report("INFRA")
        report_bes = self._report("BES")
        infra_item = self._risk_item(report_infra, no_item=1)
        self._risk_item(report_bes, no_item=2)
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/change/"
        )
        request.user = self.admin_user
        request._monthly_report_reassessment_id = report_infra.reassessment_id
        inline = MonthlyRiskReportItemInline(MonthlyRiskReport, AdminSite())

        formfield = inline.formfield_for_foreignkey(
            MonthlyRiskReportItem._meta.get_field("risk_event"),
            request,
        )

        self.assertEqual(list(formfield.queryset), [infra_item])

    def test_monthly_risk_item_label_uses_excel_risk_number_and_cause_code(self):
        report_infra = self._report("UB INFRA")
        item = self._risk_item(
            report_infra,
            no_item=42,
            no_risiko=25,
            no_penyebab_risiko="ae",
            peristiwa_risiko="Tidak tercapai KPI HCR, OCR dan Produktifitas",
        )

        label = _monthly_risk_item_label(item)

        self.assertIn("UB INFRA-25.ae", label)
        self.assertIn("Risiko 25", label)
        self.assertNotIn("Item 42", label)

    def test_inline_risk_event_label_includes_item_number_and_event(self):
        report_infra = self._report("INFRA")
        infra_item = self._risk_item(report_infra, no_item=1)
        request = RequestFactory().get(
            f"/admin/monthly_report/monthlyriskreport/{report_infra.pk}/change/"
        )
        request.user = self.admin_user
        request._monthly_report_reassessment_id = report_infra.reassessment_id
        inline = MonthlyRiskReportItemInline(MonthlyRiskReport, AdminSite())

        formfield = inline.formfield_for_foreignkey(
            MonthlyRiskReportItem._meta.get_field("risk_event"),
            request,
        )

        label = formfield.label_from_instance(infra_item)
        self.assertIn("INFRA-1.a", label)
        self.assertIn("Risiko 1", label)
        self.assertIn("Penyebab a", label)
        self.assertIn("Risiko INFRA", label)

    def test_risk_items_endpoint_uses_informative_labels(self):
        report_infra = self._report("INFRA")
        self._risk_item(report_infra, no_item=1, no_risiko=1, no_penyebab_risiko="a")
        request = RequestFactory().get(
            "/admin/monthly_report/monthlyriskreport/risk-items/",
            {"reassessment": str(report_infra.reassessment_id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.risk_items_for_reassessment(request)
        payload = json.loads(response.content)
        label = payload["items"][0]["text"]

        self.assertIn("INFRA-1.a", label)
        self.assertIn("Risiko 1", label)
        self.assertIn("Penyebab a", label)
        self.assertIn("Risiko INFRA", label)

    def test_risk_items_endpoint_orders_by_item_then_cause_code(self):
        report_infra = self._report("INFRA")
        self._risk_item(
            report_infra,
            no_item=1,
            no_risiko=3,
            no_penyebab_risiko="c",
            peristiwa_risiko="Risiko pertama",
        )
        self._risk_item(
            report_infra,
            no_item=1,
            no_risiko=1,
            no_penyebab_risiko="a",
            peristiwa_risiko="Risiko pertama",
        )
        self._risk_item(
            report_infra,
            no_item=1,
            no_risiko=2,
            no_penyebab_risiko="b",
            peristiwa_risiko="Risiko pertama",
        )
        request = RequestFactory().get(
            "/admin/monthly_report/monthlyriskreport/risk-items/",
            {"reassessment": str(report_infra.reassessment_id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.risk_items_for_reassessment(request)
        payload = json.loads(response.content)
        labels = [item["text"] for item in payload["items"]]

        self.assertIn("INFRA-1.a", labels[0])
        self.assertIn("INFRA-2.b", labels[1])
        self.assertIn("INFRA-3.c", labels[2])

    def test_risk_items_endpoint_uses_excel_risk_number_when_internal_item_number_jumps(self):
        report_infra = self._report("INFRA")
        self._risk_item(
            report_infra,
            no_item=10,
            no_risiko=26,
            no_penyebab_risiko="p",
            peristiwa_risiko="Risiko urutan kesatu",
        )
        self._risk_item(
            report_infra,
            no_item=28,
            no_risiko=11,
            no_penyebab_risiko="q",
            peristiwa_risiko="Risiko urutan kedua",
        )
        request = RequestFactory().get(
            "/admin/monthly_report/monthlyriskreport/risk-items/",
            {"reassessment": str(report_infra.reassessment_id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        request.user = self.admin_user
        report_admin = MonthlyRiskReportAdmin(MonthlyRiskReport, AdminSite())

        response = report_admin.risk_items_for_reassessment(request)
        payload = json.loads(response.content)
        labels = [item["text"] for item in payload["items"]]

        self.assertIn("INFRA-26.p", labels[0])
        self.assertIn("INFRA-11.q", labels[1])
        self.assertNotIn("INFRA-28.q", labels[1])


    def test_notification_test_mode_only_uses_explicit_test_email(self):
        report = self._report("INFRA TEST DELIVERY")
        User = get_user_model()
        officer = User.objects.create_user(
            username="delivery.officer",
            email="delivery.officer@example.com",
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report.reassessment.unit_bisnis,
            user=officer,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        self._assign_pairing_officer(
            report,
            username="delivery.pairing",
            email="delivery.pairing@example.com",
        )

        sent = send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            delivery_mode="test",
            test_email_override="admin.test@example.com",
            subject_override="Test Notifikasi ERM",
            instruction_override="Periksa tampilan email ini.",
        )

        self.assertEqual(sent, 1)
        self.assertEqual(mail.outbox[-1].to, ["admin.test@example.com"])
        self.assertEqual(mail.outbox[-1].cc, [])
        self.assertEqual(mail.outbox[-1].bcc, [])
        self.assertEqual(mail.outbox[-1].subject, "Test Notifikasi ERM")
        self.assertIn("[MODE UJI COBA]", mail.outbox[-1].body)
        self.assertIn("Periksa tampilan email ini.", mail.outbox[-1].body)

    def test_notification_email_includes_published_youtube_tutorial(self):
        category = KnowledgeBaseCategory.objects.create(
            nama="Tutorial ERM",
        )
        KnowledgeBaseArticle.objects.create(
            kategori=category,
            judul="Cara Mengisi Laporan Risiko Bulanan",
            ringkasan="Panduan singkat pengisian dan pengiriman laporan.",
            konten="Konten tutorial.",
            status=KnowledgeBaseArticle.STATUS_PUBLISHED,
            tutorial_placement=(
                KnowledgeBaseArticle.TUTORIAL_PLACEMENT_MONTHLY_REPORT_EMAIL
            ),
            video_youtube_url="https://youtu.be/tutorial-erm",
            dipublikasikan_pada=timezone.now(),
        )
        report = self._report("INFRA TUTORIAL EMAIL")
        User = get_user_model()
        officer = User.objects.create_user(
            username="tutorial.officer",
            email="tutorial.officer@example.com",
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report.reassessment.unit_bisnis,
            user=officer,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )

        sent = send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            delivery_mode="test",
            test_email_override="admin.test@example.com",
        )

        self.assertEqual(sent, 1)
        self.assertIn(
            "Tutorial Penggunaan: Cara Mengisi Laporan Risiko Bulanan",
            mail.outbox[-1].body,
        )
        self.assertIn(
            "https://youtu.be/tutorial-erm",
            mail.outbox[-1].body,
        )
        html = mail.outbox[-1].alternatives[0].content
        self.assertIn("Tutorial Penggunaan", html)
        self.assertIn("Tonton Video Tutorial", html)
        self.assertIn("https://youtu.be/tutorial-erm", html)

    def test_notification_email_hides_unpublished_tutorial(self):
        category = KnowledgeBaseCategory.objects.create(
            nama="Tutorial Draft",
        )
        KnowledgeBaseArticle.objects.create(
            kategori=category,
            judul="Tutorial Belum Terbit",
            konten="Konten tutorial.",
            status=KnowledgeBaseArticle.STATUS_DRAFT,
            tutorial_placement=(
                KnowledgeBaseArticle.TUTORIAL_PLACEMENT_MONTHLY_REPORT_EMAIL
            ),
            video_youtube_url="https://youtu.be/tutorial-draft",
        )
        report = self._report("INFRA TUTORIAL DRAFT")
        User = get_user_model()
        officer = User.objects.create_user(
            username="tutorial.draft.officer",
            email="tutorial.draft.officer@example.com",
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report.reassessment.unit_bisnis,
            user=officer,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )

        send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            delivery_mode="test",
            test_email_override="admin.test@example.com",
        )

        self.assertNotIn("Tutorial Belum Terbit", mail.outbox[-1].body)
        self.assertNotIn(
            "https://youtu.be/tutorial-draft",
            mail.outbox[-1].alternatives[0].content,
        )

    def test_notification_final_mode_ignores_global_test_email(self):
        app_setting = AppSetting.get_solo()
        app_setting.monthly_report_notification_test_email = (
            "global.test@example.com"
        )
        app_setting.save(
            update_fields=["monthly_report_notification_test_email"]
        )
        User = get_user_model()
        reviewer = User.objects.create_user(
            username="final.reviewer",
            email="final.reviewer@example.com",
        )
        approver = User.objects.create_user(
            username="final.approver",
            email="final.approver@example.com",
        )
        report = self._report("INFRA FINAL DELIVERY")
        report.status = "submitted"
        report.reviewed_by = reviewer
        report.approved_by = approver
        report.save(
            update_fields=["status", "reviewed_by", "approved_by"]
        )
        self._assign_pairing_officer(
            report,
            username="final.pairing",
            email="final.pairing@example.com",
        )

        sent = send_monthly_report_notification(
            report,
            base_url="https://erm.plnbatam.com",
            delivery_mode="final",
        )

        self.assertEqual(sent, 1)
        self.assertEqual(mail.outbox[-1].to, [reviewer.email])
        self.assertEqual(mail.outbox[-1].cc, [approver.email])
        self.assertEqual(mail.outbox[-1].bcc, ["final.pairing@example.com"])
        self.assertNotIn("global.test@example.com", mail.outbox[-1].to)
        self.assertNotIn("[MODE UJI COBA]", mail.outbox[-1].body)

    def test_notification_configuration_get_does_not_send_email(self):
        report = self._report("INFRA CONFIG PAGE")
        User = get_user_model()
        officer = User.objects.create_user(
            username="config.officer",
            email="config.officer@example.com",
        )
        PenugasanUnitBisnis.objects.create(
            unit_bisnis=report.reassessment.unit_bisnis,
            user=officer,
            peran=PenugasanUnitBisnis.ROLE_RISK_OFFICER,
        )
        self._assign_pairing_officer(
            report,
            username="config.pairing",
            email="config.pairing@example.com",
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse(
                "risk_admin:monthly_report_monthlyriskreport_send_notification",
                args=[report.pk],
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Konfigurasi Notifikasi")
        self.assertContains(response, "Kirim Test")
        self.assertContains(response, "Kirim Notifikasi Final")
        self.assertContains(response, officer.email)
        self.assertEqual(len(mail.outbox), 0)
