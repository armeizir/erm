from __future__ import annotations

from collections import defaultdict
from copy import copy
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from django.utils.text import get_valid_filename
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage

from risk.services.kpmr_automation import calculate_kpmr_for_report

from .models import MonthlyRiskReport


TEMPLATE_DIR = Path(__file__).resolve().parent / "excel_templates"
MONTHLY_REPORT_TEMPLATE = TEMPLATE_DIR / "monthly_risk_report_template.xlsx"
KPMR_SHEET_TEMPLATE = TEMPLATE_DIR / "kpmr_sheet_template.xlsx"

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


MONTH_NAMES = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


def _safe_text(value):
    if value in (None, ""):
        return None
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _number(value):
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, TypeError, ValueError):
        return _safe_text(value)


def _integer(value):
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
        if number == number.to_integral_value():
            return int(number)
        return float(number)
    except (InvalidOperation, TypeError, ValueError):
        return _safe_text(value)


def _probability(value):
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return _safe_text(value)
    if abs(number) > 1:
        number /= Decimal("100")
    return float(number)


def _fraction(value):
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return _safe_text(value)
    if abs(number) > 1:
        number /= Decimal("100")
    return float(number)


def _scale_value(value):
    if value in (None, ""):
        return None
    return getattr(value, "urutan", None) or _integer(value)


def _display(instance, field_name):
    method = getattr(instance, f"get_{field_name}_display", None)
    if callable(method):
        value = method()
        if value:
            return value
    return getattr(instance, field_name, None)


def _quarter(month):
    return ((int(month) - 1) // 3) + 1


def _copy_row_style(ws, source_row, target_row, min_col, max_col):
    if source_row == target_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
    for column in range(min_col, max_col + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _clear_values(ws, start_row, end_row, min_col, max_col):
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _ordered_items(report):
    return list(
        report.items.select_related(
            "risk_event",
            "risk_event__summary",
            "risk_event__summary__unit_bisnis",
            "risk_event__km_item",
            "risk_event__skala_dampak_q1",
            "risk_event__skala_dampak_q2",
            "risk_event__skala_dampak_q3",
            "risk_event__skala_dampak_q4",
            "risk_event__skala_probabilitas_q1",
            "risk_event__skala_probabilitas_q2",
            "risk_event__skala_probabilitas_q3",
            "risk_event__skala_probabilitas_q4",
            "realisasi_skala_dampak",
            "realisasi_skala_probabilitas",
        ).order_by(
            "risk_event__no_risiko",
            "risk_event__no_penyebab_risiko",
            "risk_event__no_item",
            "pk",
        )
    )


def _item_completeness(item):
    field_names = (
        "realisasi_asumsi_dampak",
        "realisasi_nilai_dampak",
        "realisasi_skala_dampak_id",
        "realisasi_nilai_probabilitas",
        "realisasi_skala_probabilitas_id",
        "realisasi_eksposur",
        "realisasi_skor_risiko",
        "realisasi_level_risiko",
        "efektivitas_perlakuan_risiko",
    )
    return sum(getattr(item, field_name, None) not in (None, "") for field_name in field_names)


def _representative_items(items):
    selected = {}
    for item in items:
        risk_number = item.risk_event.no_risiko or item.risk_event.no_item or item.risk_event_id
        current = selected.get(risk_number)
        if current is None or _item_completeness(item) > _item_completeness(current):
            selected[risk_number] = item
    return dict(sorted(selected.items(), key=lambda pair: (pair[0], pair[1].pk)))


def _history(report):
    reports = (
        MonthlyRiskReport.objects.filter(
            reassessment_id=report.reassessment_id,
            tahun_buku_id=report.tahun_buku_id,
            periode__tanggal_mulai__lte=report.periode.tanggal_mulai,
        )
        .select_related("periode", "reassessment", "reassessment__unit_bisnis")
        .prefetch_related(
            "items__risk_event",
            "items__realisasi_skala_dampak",
            "items__realisasi_skala_probabilitas",
        )
        .order_by("periode__tanggal_mulai", "versi", "pk")
    )
    by_month = {}
    for history_report in reports:
        by_month[history_report.periode.tanggal_mulai.month] = history_report
    return by_month


def _history_item_maps(report):
    by_month = _history(report)
    items_by_month = {}
    representative_by_quarter = {}
    items_by_quarter = {}
    for month, history_report in by_month.items():
        items = list(history_report.items.all())
        items_by_month[month] = {item.risk_event_id: item for item in items}
        quarter = _quarter(month)
        representative_by_quarter[quarter] = _representative_items(items)
        items_by_quarter[quarter] = {item.risk_event_id: item for item in items}
    return by_month, items_by_month, representative_by_quarter, items_by_quarter


def _risk_type(risk):
    explicit = getattr(risk, "jenis_risiko", None)
    if explicit == "kuantitatif":
        return "Kuantitatif"
    if explicit == "kualitatif":
        return "Kualitatif"

    # Fallback hanya untuk data legacy sebelum field Jenis Risiko tersedia.
    if any(
        getattr(risk, f"nilai_dampak_q{quarter}", None) not in (None, "", 0)
        for quarter in range(1, 5)
    ):
        return "Kuantitatif"
    return "Kualitatif"


def _fill_iiia(workbook, report, items, representative_by_quarter):
    ws = workbook["III.A"]
    data_start = 13
    _clear_values(ws, data_start, ws.max_row, 2, 59)
    representatives = _representative_items(items)
    report_quarter = _quarter(report.periode.tanggal_mulai.month)

    quarter_columns = {
        1: {"impact": 15, "impact_scale": 19, "probability": 27, "probability_scale": 31, "exposure": 39, "score": 43, "level": 51},
        2: {"impact": 16, "impact_scale": 20, "probability": 28, "probability_scale": 32, "exposure": 40, "score": 44, "level": 52},
        3: {"impact": 17, "impact_scale": 21, "probability": 29, "probability_scale": 33, "exposure": 41, "score": 45, "level": 53},
        4: {"impact": 18, "impact_scale": 22, "probability": 30, "probability_scale": 34, "exposure": 42, "score": 46, "level": 54},
    }

    for offset, (risk_number, current_item) in enumerate(representatives.items()):
        row = data_start + offset
        _copy_row_style(ws, data_start, row, 2, 59)
        risk = current_item.risk_event
        target_quarter = report_quarter

        ws.cell(row, 2, _integer(risk_number))
        ws.cell(row, 3, _safe_text(risk.peristiwa_risiko))
        ws.cell(row, 4, _risk_type(risk))
        ws.cell(row, 5, _safe_text(current_item.realisasi_asumsi_dampak or risk.asumsi_perhitungan_dampak))
        ws.cell(row, 6, _number(getattr(risk, f"nilai_dampak_q{target_quarter}", None) or risk.nilai_dampak))
        ws.cell(row, 7, _scale_value(getattr(risk, f"skala_dampak_q{target_quarter}", None)))
        ws.cell(row, 9, _probability(getattr(risk, f"nilai_probabilitas_q{target_quarter}", None) or risk.nilai_probabilitas))
        ws.cell(row, 10, _scale_value(getattr(risk, f"skala_probabilitas_q{target_quarter}", None) or risk.skala_probabilitas))

        target_exposure = getattr(risk, f"eksposur_risiko_q{target_quarter}", None)
        if target_exposure not in (None, ""):
            ws.cell(row, 12, _number(target_exposure))
        else:
            ws.cell(row, 12, f'=IF(OR(F{row}="",I{row}=""),"",F{row}*I{row})')
        ws.cell(row, 13, _integer(getattr(risk, f"skala_risiko_q{target_quarter}", None)))
        ws.cell(row, 14, _safe_text(current_item.realisasi_asumsi_dampak or risk.asumsi_perhitungan_dampak))

        for quarter in range(1, 5):
            historical_item = representative_by_quarter.get(quarter, {}).get(risk_number)
            if historical_item is None or quarter > report_quarter:
                continue
            columns = quarter_columns[quarter]
            ws.cell(row, columns["impact"], _number(historical_item.realisasi_nilai_dampak))
            ws.cell(row, columns["impact_scale"], _scale_value(historical_item.realisasi_skala_dampak))
            ws.cell(row, columns["probability"], _probability(historical_item.realisasi_nilai_probabilitas))
            ws.cell(row, columns["probability_scale"], _scale_value(historical_item.realisasi_skala_probabilitas))
            impact_letter = ws.cell(1, columns["impact"]).column_letter
            probability_letter = ws.cell(1, columns["probability"]).column_letter
            ws.cell(
                row,
                columns["exposure"],
                f'=IF(OR({impact_letter}{row}="",{probability_letter}{row}=""),"",{impact_letter}{row}*{probability_letter}{row})',
            )
            ws.cell(row, columns["score"], _integer(historical_item.realisasi_skor_risiko))
            ws.cell(row, columns["level"], _safe_text(historical_item.realisasi_level_risiko))

        ws.cell(row, 59, _safe_text(_display(current_item, "efektivitas_perlakuan_risiko")))


def _status_text(item):
    value = item.status_rencana_perlakuan
    display = _display(item, "status_rencana_perlakuan")
    if value == "continue":
        return "2. Continue"
    if value == "discontinue":
        return "1. Discontinue"
    return _safe_text(display)


def _fill_iiib(workbook, report, items, items_by_month, items_by_quarter):
    ws = workbook["III.B"]
    data_start = 11
    _clear_values(ws, data_start, ws.max_row, 2, 54)
    current_month = report.periode.tanggal_mulai.month
    current_quarter = _quarter(current_month)
    quarter_progress_columns = {1: 30, 2: 31, 3: 32, 4: 33}  # AD:AG

    for offset, item in enumerate(items):
        row = data_start + offset
        _copy_row_style(ws, data_start, row, 2, 54)
        risk = item.risk_event
        risk_number = risk.no_risiko or risk.no_item

        ws.cell(row, 2, _integer(risk_number))
        ws.cell(row, 3, _safe_text(risk.peristiwa_risiko))
        ws.cell(row, 4, _safe_text(risk.deskripsi_peristiwa_risiko))
        ws.cell(row, 5, _safe_text(risk.no_penyebab_risiko))
        ws.cell(row, 6, _safe_text(risk.kode_penyebab_risiko))
        ws.cell(row, 7, _safe_text(risk.penyebab_risiko))
        ws.cell(row, 8, _safe_text(risk.rencana_perlakuan_risiko))
        ws.cell(row, 9, _safe_text(risk.output_perlakuan_risiko))
        ws.cell(row, 10, _number(risk.biaya_perlakuan_risiko))
        ws.cell(row, 11, _safe_text(item.realisasi_rencana_perlakuan))
        ws.cell(row, 12, _safe_text(item.realisasi_output_perlakuan))
        ws.cell(row, 13, _number(item.realisasi_biaya_perlakuan))
        ws.cell(row, 14, f'=IFERROR(M{row}/J{row},"n/a")')
        ws.cell(row, 15, _safe_text(item.realisasi_pic or getattr(risk, "pic_display", None) or risk.pic))

        # MONTHLY_ACTUAL_TIMELINE_EXPORT_V1
        for month in range(1, 13):
            ws.cell(
                row,
                15 + month,
                _integer(getattr(item, f"realisasi_timeline_{month}", 0)) or None,
            )

        ws.cell(row, 28, _status_text(item))
        ws.cell(row, 29, _safe_text(item.penjelasan_status_rencana))

        for quarter in range(1, current_quarter + 1):
            quarter_item = items_by_quarter.get(quarter, {}).get(risk.id)
            if quarter_item:
                ws.cell(
                    row,
                    quarter_progress_columns[quarter],
                    _fraction(quarter_item.progress_pelaksanaan_percent),
                )

        ws.cell(row, 34, _safe_text(risk.key_risk_indicators))
        ws.cell(row, 35, _safe_text(risk.unit_satuan_kri))
        ws.cell(row, 36, _safe_text(risk.threshold_aman))
        ws.cell(row, 37, _safe_text(risk.threshold_hati_hati))
        ws.cell(row, 38, _safe_text(risk.threshold_bahaya))

        for month in range(1, 13):
            history_item = items_by_month.get(month, {}).get(risk.id)
            if history_item is None:
                continue
            threshold_col = 39 + ((month - 1) * 2)
            score_col = threshold_col + 1
            ws.cell(row, threshold_col, _safe_text(history_item.realisasi_threshold_kri))
            actual_kri = history_item.realisasi_nilai_kri
            if actual_kri is None:
                actual_kri = history_item.realisasi_threshold_kri_skor
            ws.cell(row, score_col, _number(actual_kri))


def _fill_iiic(workbook, report, items):
    ws = workbook["III.C"]
    # AE:AR adalah area tabel data heatmap. AS:AT adalah legenda template
    # dan memuat merged cell AT8:AT9, sehingga tidak boleh dibersihkan/ditimpa.
    _clear_values(ws, 8, 37, 31, 44)
    representatives = _representative_items(items)
    quarter = _quarter(report.periode.tanggal_mulai.month)

    for offset, (risk_number, item) in enumerate(representatives.items()):
        if offset >= 30:
            break
        row = 8 + offset
        risk = item.risk_event
        ws.cell(row, 31, _integer(risk_number))  # AE
        ws.cell(row, 32, _safe_text(risk.peristiwa_risiko))  # AF
        ws.cell(row, 33, _scale_value(getattr(risk, f"skala_dampak_q{quarter}", None)))
        ws.cell(row, 34, _scale_value(getattr(risk, f"skala_probabilitas_q{quarter}", None)))
        ws.cell(row, 35, _integer(getattr(risk, f"skala_risiko_q{quarter}", None)))
        ws.cell(row, 39, _scale_value(item.realisasi_skala_dampak))
        ws.cell(row, 40, _scale_value(item.realisasi_skala_probabilitas))
        ws.cell(row, 41, _integer(item.realisasi_skor_risiko))


def _fill_iiid(workbook, report):
    ws = workbook["III.D"]
    _clear_values(ws, 8, ws.max_row, 2, 4)
    changes = list(report.changes.all().order_by("pk"))
    if not changes:
        ws["B8"] = "Tidak ada perubahan"
        ws["C8"] = "Tidak ada perubahan"
        ws["D8"] = "Tidak ada perubahan"
        return
    for offset, change in enumerate(changes):
        row = 8 + offset
        _copy_row_style(ws, 8, row, 1, 4)
        ws.cell(row, 2, _safe_text(change.get_jenis_perubahan_display()))
        ws.cell(row, 3, _safe_text(change.peristiwa_risiko_terdampak))
        ws.cell(row, 4, _safe_text(change.penjelasan))


def _yes_no(value):
    if value == "ya":
        return "1. Ya"
    if value == "tidak":
        return "2. Tidak"
    return None


def _loss_source(value):
    if value == "internal":
        return "1. Internal"
    if value == "external":
        return "2. Eksternal"
    return None


def _fill_iiie(workbook, report):
    ws = workbook["III.E"]
    _clear_values(ws, 8, ws.max_row, 2, 21)
    events = list(report.loss_events.all().order_by("pk"))
    if not events:
        for column in range(2, 22):
            ws.cell(8, column, "Tidak terjadi loss event")
        return

    for offset, event in enumerate(events):
        row = 8 + offset
        _copy_row_style(ws, 8, row, 1, 21)
        values = [
            event.nama_kejadian,
            event.identifikasi_kejadian,
            event.kategori_kejadian,
            _loss_source(event.sumber_penyebab_kejadian),
            event.penyebab_kejadian,
            event.penanganan_saat_kejadian,
            event.deskripsi_kejadian_risk_event,
            event.kategori_risiko_bumn,
            event.kategori_risiko_t2_t3_kbumn,
            event.penjelasan_kerugian,
            _number(event.nilai_kerugian),
            _yes_no(event.kejadian_berulang),
            event.frekuensi_kejadian,
            event.mitigasi_direncanakan,
            event.realisasi_mitigasi,
            event.perbaikan_mendatang,
            event.pihak_terkait,
            _yes_no(event.status_asuransi),
            _number(event.nilai_premi),
            _number(event.nilai_klaim),
        ]
        for column, value in enumerate(values, start=2):
            ws.cell(row, column, _safe_text(value) if isinstance(value, str) else value)


def _copy_kpmr_sheet(workbook):
    if "KPMR" in workbook.sheetnames:
        workbook.remove(workbook["KPMR"])

    source_workbook = load_workbook(KPMR_SHEET_TEMPLATE, data_only=False)
    source = source_workbook["KPMR"]
    target = workbook.create_sheet("KPMR")

    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)
    target.freeze_panes = source.freeze_panes
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.protection = copy(source.protection)
    target.auto_filter = copy(source.auto_filter)

    for key, dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[key]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit
        target_dimension.outlineLevel = dimension.outlineLevel
        target_dimension.collapsed = dimension.collapsed
    for key, dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[key]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel
        target_dimension.collapsed = dimension.collapsed

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    merged_non_start = set()
    for merged_range in source.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                if row == min_row and column == min_col:
                    continue
                merged_non_start.add((row, column))

    for row in source.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            if (source_cell.row, source_cell.column) in merged_non_start:
                continue
            target_cell = target[source_cell.coordinate]
            target_cell.value = source_cell.value
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    for validation in source.data_validations.dataValidation:
        target.add_data_validation(copy(validation))

    for conditional_format, rules in source.conditional_formatting._cf_rules.items():
        for rule in rules:
            target.conditional_formatting.add(str(conditional_format.sqref), copy(rule))

    for image in getattr(source, "_images", []):
        try:
            copied_image = ExcelImage(BytesIO(image._data()))
            copied_image.width = image.width
            copied_image.height = image.height
            copied_image.anchor = copy(image.anchor)
            target.add_image(copied_image)
        except Exception:
            continue

    try:
        target.print_area = source.print_area
        target.print_title_rows = source.print_title_rows
        target.print_title_cols = source.print_title_cols
    except Exception:
        pass
    return target


def _fill_kpmr(workbook, report):
    ws = _copy_kpmr_sheet(workbook)
    ws["B3"] = f"Unit/Bidang: {report.reassessment.unit_bisnis.name} | Periode: {report.periode.nama_periode}"

    try:
        calculation = calculate_kpmr_for_report(report)
    except Exception as exc:
        ws["K8"] = f"KPMR belum dapat dihitung otomatis: {exc}"
        return

    indicators = {item.get("kode"): item for item in calculation.indicators}
    indicator_cells = {
        "I1": "8",
        "I2": "12",
        "I3": "18",
    }
    for code, row in indicator_cells.items():
        indicator = indicators.get(code, {})
        ws[f"D{row}"] = _safe_text(indicator.get("jawaban"))
        ws[f"J{row}"] = _safe_text(indicator.get("dokumen_referensi"))
        ws[f"K{row}"] = _safe_text(indicator.get("keterangan"))

    i4 = indicators.get("I4", {})
    subindicators = {item.get("kode"): item for item in i4.get("subindikator", [])}
    subindicator_cells = {
        "IDENTIFIKASI": "22",
        "KUANTIFIKASI": "25",
        "RENCANA": "28",
        "PRIORITISASI": "31",
    }
    for code, row in subindicator_cells.items():
        indicator = subindicators.get(code, {})
        ws[f"D{row}"] = _safe_text(indicator.get("jawaban"))
        ws[f"J{row}"] = _safe_text(indicator.get("dokumen_referensi") or i4.get("dokumen_referensi"))
        ws[f"K{row}"] = _safe_text(indicator.get("keterangan"))


def build_monthly_risk_report_workbook(report):
    if not MONTHLY_REPORT_TEMPLATE.exists():
        raise FileNotFoundError(f"Template laporan bulanan tidak ditemukan: {MONTHLY_REPORT_TEMPLATE}")
    if not KPMR_SHEET_TEMPLATE.exists():
        raise FileNotFoundError(f"Template KPMR tidak ditemukan: {KPMR_SHEET_TEMPLATE}")

    workbook = load_workbook(MONTHLY_REPORT_TEMPLATE, data_only=False, keep_links=True)
    items = _ordered_items(report)
    _, items_by_month, representative_by_quarter, items_by_quarter = _history_item_maps(report)

    _fill_iiia(workbook, report, items, representative_by_quarter)
    _fill_iiib(workbook, report, items, items_by_month, items_by_quarter)
    _fill_iiic(workbook, report, items)
    _fill_iiid(workbook, report)
    _fill_iiie(workbook, report)
    _fill_kpmr(workbook, report)

    workbook.properties.title = f"Laporan Risiko {report.reassessment} - {report.periode.nama_periode}"
    workbook.properties.subject = "Laporan Realisasi Manajemen Risiko dan KPMR"
    workbook.properties.description = (
        f"Dihasilkan dari ERM untuk {report.reassessment.unit_bisnis.name}, "
        f"periode {report.periode.nama_periode}."
    )
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass
    return workbook


def build_monthly_risk_report_excel(report):
    workbook = build_monthly_risk_report_workbook(report)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    unit = get_valid_filename(report.reassessment.unit_bisnis.name or "UNIT")
    period = get_valid_filename(report.periode.kode_periode or report.periode.nama_periode)
    filename = f"Laporan_Risiko_{unit}_{period}.xlsx"
    return output, filename
