import hashlib
import json
import re
from decimal import Decimal, InvalidOperation

import httpx
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from risk.models import AppSetting, MasterSkalaDampak, MasterSkalaProbabilitas

from .models import (
    MonthlyRiskReportImportBatch,
    MonthlyRiskReportImportRow,
    MonthlyRiskReportSubmissionLog,
)
from .services import refresh_monthly_report_summary


IMPORTABLE_FIELDS = (
    "realisasi_asumsi_dampak",
    "realisasi_nilai_dampak",
    "realisasi_skala_dampak_id",
    "realisasi_nilai_probabilitas",
    "realisasi_skala_probabilitas_id",
    "jenis_risiko",
    "realisasi_skala_dampak_kbumn",
    "realisasi_skala_probabilitas_kbumn",
    "realisasi_skala_nilai_risiko_kbumn",
    "realisasi_eksposur",
    "realisasi_skor_risiko",
    "realisasi_level_risiko_bumn",
    "realisasi_level_risiko_kbumn",
    "efektivitas_perlakuan_risiko",
    "realisasi_rencana_perlakuan",
    "realisasi_output_perlakuan",
    "realisasi_biaya_perlakuan",
    "realisasi_pic",
    "status_rencana_perlakuan",
    "penjelasan_status_rencana",
    "progress_pelaksanaan_percent",
    "realisasi_threshold_kri",
    "realisasi_threshold_kri_skor",
    "realisasi_nilai_kri",
)
IMPORT_PARSER_VERSION = 5

IIIA_QUARTER_COLUMNS = {
    "realisasi_nilai_dampak": {1: 14, 2: 15, 3: 16, 4: 17},       # O:R
    "realisasi_skala_dampak_id": {1: 18, 2: 19, 3: 20, 4: 21},    # S:V
    "realisasi_nilai_probabilitas": {1: 26, 2: 27, 3: 28, 4: 29}, # AA:AD
    "realisasi_skala_probabilitas_id": {1: 30, 2: 31, 3: 32, 4: 33},
    "realisasi_skala_dampak_kbumn": {1: 22, 2: 23, 3: 24, 4: 25},       # W:Z
    "realisasi_skala_probabilitas_kbumn": {1: 34, 2: 35, 3: 36, 4: 37}, # AI:AL
    "source_realisasi_eksposur": {1: 38, 2: 39, 3: 40, 4: 41},          # AM:AP
    "source_realisasi_skor_risiko": {1: 42, 2: 43, 3: 44, 4: 45},       # AQ:AT
    "realisasi_skala_nilai_risiko_kbumn": {1: 46, 2: 47, 3: 48, 4: 49},# AU:AX
    "realisasi_level_risiko_bumn": {1: 50, 2: 51, 3: 52, 4: 53},        # AY:BB
    "realisasi_level_risiko_kbumn": {1: 54, 2: 55, 3: 56, 4: 57},       # BC:BF
}


def report_quarter(month):
    if month not in range(1, 13):
        raise ValueError("Bulan harus berada antara 1 dan 12.")
    return ((month - 1) // 3) + 1


def _is_standard_iiia_layout(worksheet):
    """Recognize the official III.A layout before using fixed quarter columns."""
    return (
        normalize_header_name(worksheet.cell(5, 14).value)
        == "asumsi perhitungan dampak"
        and normalize_header_name(worksheet.cell(5, 15).value) == "nilai dampak"
        and normalize_header_name(worksheet.cell(5, 19).value) == "skala dampak"
        and normalize_header_name(worksheet.cell(5, 27).value) == "nilai probabilitas"
        and normalize_header_name(worksheet.cell(5, 31).value) == "skala probabilitas"
    )
FIELD_LABELS = {
    "realisasi_asumsi_dampak": "Asumsi perhitungan dampak",
    "realisasi_nilai_dampak": "Nilai dampak",
    "realisasi_skala_dampak_id": "Skala dampak",
    "realisasi_nilai_probabilitas": "Nilai probabilitas",
    "realisasi_skala_probabilitas_id": "Skala probabilitas",
    "realisasi_skala_dampak_kbumn": "Skala dampak KBUMN",
    "realisasi_skala_probabilitas_kbumn": "Skala probabilitas KBUMN",
    "realisasi_skala_nilai_risiko_kbumn": "Skala nilai risiko KBUMN",
    "realisasi_level_risiko_bumn": "Level risiko BUMN",
    "realisasi_level_risiko_kbumn": "Level risiko KBUMN",
    "realisasi_eksposur": "Nilai eksposur risiko",
    "realisasi_skor_risiko": "Skala nilai risiko BUMN",
    "efektivitas_perlakuan_risiko": "Efektivitas perlakuan",
    "realisasi_rencana_perlakuan": "Realisasi rencana",
    "realisasi_output_perlakuan": "Realisasi output",
    "realisasi_biaya_perlakuan": "Realisasi biaya",
    "realisasi_pic": "PIC",
    "status_rencana_perlakuan": "Status rencana",
    "penjelasan_status_rencana": "Penjelasan status",
    "progress_pelaksanaan_percent": "Progress",
    "realisasi_threshold_kri": "Nilai KRI",
    "realisasi_threshold_kri_skor": "Threshold KRI",
    "realisasi_nilai_kri": "Nilai realisasi KRI",
}


def file_sha256(uploaded_file):
    digest = hashlib.sha256()
    for chunk in uploaded_file.chunks():
        digest.update(chunk)
    uploaded_file.seek(0)
    return digest.hexdigest()


def _text(value):
    return str(value or "").strip()


def _safe_import_text(value):
    text = _text(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _normalize(value):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", _text(value).lower())).strip()


def normalize_header_name(value):
    text = _text(value).casefold().replace("\n", " ")
    text = re.sub(r"[./]+", " ", text)
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9%]+", " ", text)).strip()


HEADER_ALIASES = {
    "risk_number": {
        "no risiko",
        "nomor risiko",
        "kode risiko",
    },
    "risk_event": {
        "peristiwa risiko",
        "nama peristiwa risiko",
    },
    "cause_code": {
        "kode penyebab",
        "kode penyebab risiko",
    },
    "treatment_plan": {
        "rencana perlakuan",
        "rencana perlakuan risiko",
    },
    "impact_assumption": {
        "asumsi perhitungan dampak",
        "asumsi perhitungan dampak kuantitatif penjelasana dampak kualitatif",
    },
    "impact_value": {"nilai dampak"},
    "impact_scale": {"skala dampak"},
    "probability_value": {"nilai probabilitas"},
    "probability_scale": {"skala probabilitas"},
    "effectiveness": {
        "efektifitas perlakuan risiko",
        "efektivitas perlakuan risiko",
    },
    "actual_treatment": {"realisasi rencana perlakuan risiko"},
    "actual_output": {
        "realisasi output atas masing masing breakdown perlakuan risiko",
        "realisasi output perlakuan risiko",
    },
    "actual_cost": {
        "realisasi biaya perlakuan risiko rp usd",
        "realisasi biaya perlakuan risiko",
    },
    "actual_pic": {"realisasi pic"},
    "treatment_status": {"status rencana perlakuan risiko"},
    "treatment_status_note": {"penjelasan status rencana perlakuan"},
    "treatment_progress": {"progress pelaksanaan rencana perlakuan"},
}


def _header_key(value):
    normalized = normalize_header_name(value)
    for key, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return key
    return None


def find_header_columns(worksheet, anchor_row):
    """Map semantic columns from the multi-row header surrounding the anchor."""
    columns = {}
    header_rows = set()
    start = max(1, anchor_row - 10)
    end = min(worksheet.max_row, anchor_row + 5)
    for row in worksheet.iter_rows(min_row=start, max_row=end):
        for cell in row:
            key = _header_key(cell.value)
            if key and key not in columns:
                columns[key] = cell.column - 1
                header_rows.add(cell.row)
    return columns, sorted(header_rows)


def _normalize_business_code(value):
    text = _text(value).replace("–", "-").replace("—", "-").upper()
    return re.sub(r"[^A-Z0-9]+", "", text)


def normalize_risk_code(value):
    """Canonical comparison key without discarding meaningful suffixes."""
    return _normalize_business_code(value)


def normalize_cause_code(value):
    return _normalize_business_code(value)


def normalize_treatment_code(value):
    # Backward-compatible public name used by existing imports/tests.
    return normalize_cause_code(value)


def is_malformed_risk_code(value):
    text = _text(value).replace("–", "-").replace("—", "-")
    return bool(not text or re.search(r"-\s*-", text))


def _positive_integer(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        text = value.strip()
        match = re.fullmatch(r"0*(\d+)(?:\.0+|\.)?", text)
        if match:
            number = int(match.group(1))
            return number if number > 0 else None
    number = _decimal(value)
    if number is None or number <= 0 or number != number.to_integral_value():
        return None
    return int(number)


def extract_risk_number(value):
    direct = _positive_integer(value)
    if direct:
        return direct
    text = _text(value)
    if not text:
        return None
    match = re.match(r"^0*(\d+)(?=[.\s-]|$)", text)
    if not match:
        match = re.search(r"(?:^|[\s-])0*(\d+)(?=[.\s-]|$)", text)
    if not match:
        return None
    number = int(match.group(1))
    return number if number > 0 else None


def _profile_item_codes(item):
    unit_name = _text(item.summary.unit_bisnis.name)
    number = item.no_risiko
    cause = _text(item.no_penyebab_risiko)
    values = {
        item.kode_penyebab_risiko,
        f"{unit_name}-{number}-{cause}",
        f"{unit_name} {number} {cause}",
    }
    return {normalize_risk_code(value) for value in values if _text(value)}


def find_start_anchor(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if _text(cell.value).casefold() == "start pengisian":
                return cell.row, cell.coordinate
    raise ValidationError(
        f"Anchor 'Start pengisian' tidak ditemukan pada sheet {worksheet.title}."
    )


def find_start_row(worksheet):
    return find_start_anchor(worksheet)[0]


def iter_data_rows_after_anchor(worksheet):
    start_row = find_start_row(worksheet)
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=start_row + 1, values_only=True),
        start=start_row + 1,
    ):
        yield row_number, list(row)


def iter_candidate_rows(worksheet, max_row=None):
    """The anchor may occupy the first data row, as in the BIS workbook."""
    start_row = find_start_row(worksheet)
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=start_row,
            max_row=max_row or worksheet.max_row,
            values_only=True,
        ),
        start=start_row,
    ):
        yield row_number, list(row)


def is_valid_risk_row(row):
    return bool(
        _positive_integer(row[1] if len(row) > 1 else None)
        and _text(row[2] if len(row) > 2 else None)
    )


def is_valid_treatment_row(row):
    number = _positive_integer(row[1] if len(row) > 1 else None)
    code = (row[5] if len(row) > 5 else None) or (
        row[1] if len(row) > 1 else None
    )
    event = row[2] if len(row) > 2 else None
    treatment = row[10] if len(row) > 10 else None
    return bool(number and _text(code) and (_text(event) or _text(treatment)))


def target_item_fingerprint(report):
    payload = "|".join(
        f"{pk}:{risk_id}:{number}:{cause}:{updated.isoformat()}"
        for pk, risk_id, number, cause, updated in report.items.order_by("pk").values_list(
            "pk",
            "risk_event_id",
            "risk_event__no_risiko",
            "risk_event__no_penyebab_risiko",
            "updated_at",
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def batch_analysis_is_current(batch):
    return (
        batch.parser_version == IMPORT_PARSER_VERSION
        and batch.target_fingerprint == target_item_fingerprint(batch.report)
    )


def build_display_changes(row):
    changes = []
    item = row.matched_report_item
    for field_name, after in row.proposed_data.items():
        if field_name not in IMPORTABLE_FIELDS:
            continue
        before = _field_value(item, field_name) if item else None
        changes.append(
            {
                "label": FIELD_LABELS.get(field_name, field_name),
                "before": "-" if before in (None, "") else before,
                "after": "-" if after in (None, "") else after,
            }
        )
    return changes


def _decimal(value):
    if value in (None, ""):
        return None
    text = str(value).strip().replace("Rp", "").replace(" ", "")
    if text.endswith("%"):
        text = text[:-1]
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 else text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _percent(value):
    number = _decimal(value)
    if number is None:
        return None
    if 0 < number <= 1:
        number *= 100
    return number


def _risk_identity(code):
    text = _text(code)
    match = re.search(r"(?:^|[-\s])(\d+)[-\s]*([A-Za-z]+)\s*$", text)
    if not match:
        return None, ""
    return int(match.group(1)), match.group(2).lower()


def _scale_id(model, value):
    if value in (None, ""):
        return None
    number = _decimal(value)
    if number is not None:
        candidate = model.objects.filter(urutan=int(number), aktif=True).first()
        if candidate:
            return candidate.pk
    normalized = _normalize(value)
    for scale in model.objects.filter(aktif=True):
        if normalized in {_normalize(scale.nama), _normalize(getattr(scale, "label", ""))}:
            return scale.pk
    return None


def _effectiveness(value):
    normalized = _normalize(value)
    if "tidak" in normalized:
        return "tidak_efektif"
    if "cukup" in normalized:
        return "cukup_efektif"
    if "efektif" in normalized:
        return "efektif"
    return None


def _treatment_status(value):
    normalized = _normalize(value)
    if any(token in normalized for token in ("discontinue", "dihentikan", "stop")):
        return "discontinue"
    if any(token in normalized for token in ("continue", "dilanjut", "lanjut")):
        return "continue"
    return None


def _json_value(value):
    if isinstance(value, Decimal):
        return str(value)
    return value


def _is_total_or_footer(*values):
    normalized = " ".join(normalize_header_name(value) for value in values if value)
    return any(token in normalized.split() for token in ("total", "jumlah", "footer"))


def _legacy_layout_available(worksheet, anchor_row):
    for row in worksheet.iter_rows(
        min_row=anchor_row,
        max_row=min(worksheet.max_row, anchor_row + 50),
        values_only=True,
    ):
        if extract_risk_number(row[1] if len(row) > 1 else None) and _text(
            row[2] if len(row) > 2 else None
        ):
            return True
    return False


def _parser_diagnostic(anchor_coordinate, header_rows, columns):
    return {
        "anchor": anchor_coordinate,
        "header_rows": header_rows,
        "columns": {
            key: {"index": index + 1, "letter": get_column_letter(index + 1)}
            for key, index in columns.items()
        },
        "candidate_rows": 0,
        "accepted_rows": 0,
        "skipped_rows": 0,
        "skip_reasons": {},
        "context_reasons": {},
        "continuation_rows": 0,
        "skipped": [],
    }


def _record_skip(diagnostic, row_number, reason):
    diagnostic["skipped_rows"] += 1
    diagnostic["skip_reasons"][reason] = (
        diagnostic["skip_reasons"].get(reason, 0) + 1
    )
    if len(diagnostic["skipped"]) < 100:
        diagnostic["skipped"].append({"row": row_number, "reason": reason})


def _column_value(row, columns, key, fallback=None):
    index = columns.get(key, fallback)
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _last_relevant_row(worksheet, column_indexes, start_row):
    last_row = start_row
    for row_number in range(start_row, worksheet.max_row + 1):
        if any(
            _text(worksheet.cell(row_number, index + 1).value)
            for index in column_indexes
            if index is not None
        ):
            last_row = row_number
    return last_row


def _parse_workbook(batch, include_diagnostics=False):
    batch.source_file.open("rb")
    workbook = load_workbook(batch.source_file, data_only=True, read_only=False)
    if "III.A" not in workbook.sheetnames or "III.B" not in workbook.sheetnames:
        raise ValidationError(
            "File harus memiliki sheet III.A dan III.B sesuai template laporan ERM."
        )

    month = batch.report.periode.tanggal_mulai.month
    quarter = report_quarter(month)
    iiia_schemas = (
        (11 + quarter, 15 + quarter, 23 + quarter, 27 + quarter, 56, 11),
        (13 + quarter, 17 + quarter, 21 + quarter, 25 + quarter, 46, 13),
        (13 + quarter, 17 + quarter, 25 + quarter, 29 + quarter, 58, 13),
        (11 + quarter, 15 + quarter, 19 + quarter, 23 + quarter, 40, 11),
    )
    entries = []
    diagnostics = {}

    iiia = workbook["III.A"]
    standard_iiia_layout = _is_standard_iiia_layout(iiia)
    anchor_row, anchor_coordinate = find_start_anchor(iiia)
    columns, header_rows = find_header_columns(iiia, anchor_row)
    if not {"risk_number", "risk_event"}.issubset(columns):
        if _legacy_layout_available(iiia, anchor_row):
            columns.setdefault("risk_number", 1)
            columns.setdefault("risk_event", 2)
        else:
            raise ValidationError(
                "Kolom Nomor Risiko dan Peristiwa Risiko tidak ditemukan pada sheet III.A."
            )
    diagnostic = _parser_diagnostic(anchor_coordinate, header_rows, columns)
    diagnostics["III.A"] = diagnostic
    iiia_last_row = _last_relevant_row(
        iiia,
        [columns["risk_number"], columns["risk_event"]],
        anchor_row,
    )
    for row_number, row in iter_candidate_rows(iiia, iiia_last_row):
        code = _column_value(row, columns, "risk_number")
        event = _column_value(row, columns, "risk_event")
        if row_number in header_rows:
            diagnostic["candidate_rows"] += 1
            _record_skip(diagnostic, row_number, "header_row")
            continue
        if code in (None, "") and event in (None, ""):
            diagnostic["candidate_rows"] += 1
            _record_skip(diagnostic, row_number, "blank_row")
            continue
        diagnostic["candidate_rows"] += 1
        if _is_total_or_footer(code, event):
            _record_skip(diagnostic, row_number, "total_or_footer")
            continue
        if not _text(event):
            _record_skip(diagnostic, row_number, "missing_risk_event")
            continue
        no_risiko = extract_risk_number(code)
        malformed = is_malformed_risk_code(code)

        def schema_score(schema):
            _, impact_scale, _, probability_scale, _, _ = schema
            return int(
                bool(
                    len(row) > impact_scale
                    and _scale_id(MasterSkalaDampak, row[impact_scale])
                )
            ) + int(
                bool(
                    len(row) > probability_scale
                    and _scale_id(
                        MasterSkalaProbabilitas, row[probability_scale]
                    )
                )
            )

        if standard_iiia_layout:
            impact = IIIA_QUARTER_COLUMNS["realisasi_nilai_dampak"][quarter]
            impact_scale = IIIA_QUARTER_COLUMNS["realisasi_skala_dampak_id"][quarter]
            probability = IIIA_QUARTER_COLUMNS["realisasi_nilai_probabilitas"][quarter]
            probability_scale = IIIA_QUARTER_COLUMNS[
                "realisasi_skala_probabilitas_id"
            ][quarter]
            effectiveness, assumption = 58, 13
        else:
            impact, impact_scale, probability, probability_scale, effectiveness, assumption = max(
                iiia_schemas, key=schema_score
            )
        source_cells = {
            "realisasi_nilai_dampak": f"III.A!{get_column_letter(impact + 1)}{row_number}",
            "realisasi_skala_dampak_id": f"III.A!{get_column_letter(impact_scale + 1)}{row_number}",
            "realisasi_nilai_probabilitas": f"III.A!{get_column_letter(probability + 1)}{row_number}",
            "realisasi_skala_probabilitas_id": f"III.A!{get_column_letter(probability_scale + 1)}{row_number}",
        }
        standard_values = {}
        if standard_iiia_layout:
            for field_name in (
                "realisasi_skala_dampak_kbumn",
                "realisasi_skala_probabilitas_kbumn",
                "source_realisasi_eksposur",
                "source_realisasi_skor_risiko",
                "realisasi_skala_nilai_risiko_kbumn",
                "realisasi_level_risiko_bumn",
                "realisasi_level_risiko_kbumn",
            ):
                column_index = IIIA_QUARTER_COLUMNS[field_name][quarter]
                source_cells[field_name] = (
                    f"III.A!{get_column_letter(column_index + 1)}{row_number}"
                )
                standard_values[field_name] = (
                    row[column_index] if len(row) > column_index else None
                )
            source_cells["realisasi_eksposur"] = source_cells["source_realisasi_eksposur"]
            source_cells["realisasi_skor_risiko"] = source_cells["source_realisasi_skor_risiko"]
        empty_quarter_fields = (
            [
                field_name
                for field_name, column_index in (
                    ("realisasi_nilai_dampak", impact),
                    ("realisasi_skala_dampak_id", impact_scale),
                    ("realisasi_nilai_probabilitas", probability),
                    ("realisasi_skala_probabilitas_id", probability_scale),
                    (
                        "realisasi_skala_dampak_kbumn",
                        IIIA_QUARTER_COLUMNS["realisasi_skala_dampak_kbumn"][quarter],
                    ),
                    (
                        "realisasi_skala_probabilitas_kbumn",
                        IIIA_QUARTER_COLUMNS["realisasi_skala_probabilitas_kbumn"][quarter],
                    ),
                    (
                        "realisasi_skala_nilai_risiko_kbumn",
                        IIIA_QUARTER_COLUMNS["realisasi_skala_nilai_risiko_kbumn"][quarter],
                    ),
                    (
                        "realisasi_level_risiko_bumn",
                        IIIA_QUARTER_COLUMNS["realisasi_level_risiko_bumn"][quarter],
                    ),
                    (
                        "realisasi_level_risiko_kbumn",
                        IIIA_QUARTER_COLUMNS["realisasi_level_risiko_kbumn"][quarter],
                    ),
                    (
                        "realisasi_eksposur",
                        IIIA_QUARTER_COLUMNS["source_realisasi_eksposur"][quarter],
                    ),
                    (
                        "realisasi_skor_risiko",
                        IIIA_QUARTER_COLUMNS["source_realisasi_skor_risiko"][quarter],
                    ),
                )
                if len(row) <= column_index or row[column_index] in (None, "")
            ]
            if standard_iiia_layout
            else []
        )
        entries.append(
            {
                "source_reference": f"III.A:{row_number}",
                "source_sheet": "III.A",
                "source_row": row_number,
                "risk_code": _text(code),
                "source_risk_code": _text(code),
                "source_risk_sequence": no_risiko,
                "source_risk_event": _text(event),
                "risk_event_text": _text(event),
                "no_risiko": no_risiko,
                "cause": "",
                "normalized_code": normalize_risk_code(code),
                "scope_hint": "malformed_code" if malformed else "",
                "raw_data": {
                    "source_sheet": "III.A",
                    "source_row": row_number,
                    "source_columns": diagnostic["columns"],
                    "original_code": _text(code),
                    "sequence_warning": (
                        "sequence_not_extracted" if no_risiko is None else ""
                    ),
                    "selected_quarter": f"Q{quarter}",
                    "source_cells": source_cells,
                    "empty_quarter_fields": empty_quarter_fields,
                },
                "proposed_data": {
                    "jenis_risiko": (
                        _text(row[3]).casefold()
                        if len(row) > 3
                        and _text(row[3]).casefold() in {"kuantitatif", "kualitatif"}
                        else None
                    ),
                    "realisasi_asumsi_dampak": _safe_import_text(
                        row[assumption] if len(row) > assumption else None
                    ),
                    "realisasi_nilai_dampak": _decimal(row[impact])
                    if len(row) > impact
                    else None,
                    "realisasi_skala_dampak_id": _scale_id(
                        MasterSkalaDampak,
                        row[impact_scale] if len(row) > impact_scale else None,
                    ),
                    "realisasi_nilai_probabilitas": _percent(row[probability])
                    if len(row) > probability
                    else None,
                    "realisasi_skala_probabilitas_id": _scale_id(
                        MasterSkalaProbabilitas,
                        row[probability_scale]
                        if len(row) > probability_scale
                        else None,
                    ),
                    "realisasi_skala_dampak_kbumn": _positive_integer(
                        standard_values.get("realisasi_skala_dampak_kbumn")
                    ),
                    "realisasi_skala_probabilitas_kbumn": _positive_integer(
                        standard_values.get("realisasi_skala_probabilitas_kbumn")
                    ),
                    "realisasi_skala_nilai_risiko_kbumn": _positive_integer(
                        standard_values.get("realisasi_skala_nilai_risiko_kbumn")
                    ),
                    "realisasi_level_risiko_bumn": _safe_import_text(
                        standard_values.get("realisasi_level_risiko_bumn")
                    ),
                    "realisasi_level_risiko_kbumn": _safe_import_text(
                        standard_values.get("realisasi_level_risiko_kbumn")
                    ),
                    "realisasi_eksposur": _decimal(
                        standard_values.get("source_realisasi_eksposur")
                    ),
                    "realisasi_skor_risiko": _positive_integer(
                        standard_values.get("source_realisasi_skor_risiko")
                    ),
                    "efektivitas_perlakuan_risiko": _effectiveness(
                        row[effectiveness]
                        if len(row) > effectiveness
                        else None
                    ),
                },
            }
        )
        diagnostic["accepted_rows"] += 1
    if not diagnostic["accepted_rows"]:
        raise ValidationError(
            "Tidak ada data risiko valid setelah anchor pada sheet III.A. "
            f"Alasan: {diagnostic['skip_reasons']}."
        )

    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[quarter] - 1
    threshold_col = 39 + ((month - 1) * 2) - 1
    iiib = workbook["III.B"]
    anchor_row, anchor_coordinate = find_start_anchor(iiib)
    columns, header_rows = find_header_columns(iiib, anchor_row)
    if not {"risk_number", "risk_event"}.issubset(columns):
        if _legacy_layout_available(iiib, anchor_row):
            columns.setdefault("risk_number", 1)
            columns.setdefault("risk_event", 2)
            columns.setdefault("cause_code", 5)
            columns.setdefault("actual_treatment", 10)
        else:
            raise ValidationError(
                "Kolom Nomor Risiko dan Peristiwa Risiko tidak ditemukan pada sheet III.B."
            )
    diagnostic = _parser_diagnostic(anchor_coordinate, header_rows, columns)
    diagnostics["III.B"] = diagnostic
    iiib_last_row = _last_relevant_row(
        iiib,
        [
            columns.get("risk_number"),
            columns.get("risk_event"),
            columns.get("cause_code", 5),
            columns.get("treatment_plan", 7),
            columns.get("actual_treatment", 10),
        ],
        anchor_row,
    )
    last_number = None
    last_event = ""
    last_code = ""
    for row_number, row in iter_candidate_rows(iiib, iiib_last_row):
        raw_number = _column_value(row, columns, "risk_number")
        raw_event = _column_value(row, columns, "risk_event")
        raw_code = _column_value(row, columns, "cause_code", 5)
        treatment = _column_value(row, columns, "treatment_plan", 7)
        actual_treatment = _column_value(row, columns, "actual_treatment", 10)
        if row_number in header_rows:
            diagnostic["candidate_rows"] += 1
            _record_skip(diagnostic, row_number, "header_row")
            continue
        if not any(
            _text(value)
            for value in (
                raw_number,
                raw_event,
                raw_code,
                treatment,
                actual_treatment,
            )
        ):
            diagnostic["candidate_rows"] += 1
            _record_skip(diagnostic, row_number, "blank_row")
            continue
        diagnostic["candidate_rows"] += 1
        if _is_total_or_footer(raw_number, raw_event):
            _record_skip(diagnostic, row_number, "total_or_footer")
            continue
        explicit_number = extract_risk_number(raw_number)
        code_number = extract_risk_number(raw_code)
        no_risiko = explicit_number or code_number or last_number
        event = _text(raw_event) or last_event
        code = _text(raw_code) or last_code or _text(raw_number)
        continuation = not _text(raw_number) and not _text(raw_event)
        context_reason = ""
        if continuation:
            diagnostic["continuation_rows"] += 1
            if last_event or last_number:
                context_reason = "inherited_parent_context"
            elif raw_code:
                context_reason = "derived_parent_from_code"
            else:
                _record_skip(diagnostic, row_number, "missing_parent_context")
                continue
        if not event:
            # Cause-only rows can still be resolved deterministically against
            # the canonical target profile during matching.
            if raw_code:
                context_reason = "derived_parent_from_code"
            else:
                _record_skip(diagnostic, row_number, "missing_parent_context")
                continue
        if context_reason:
            diagnostic["context_reasons"][context_reason] = (
                diagnostic["context_reasons"].get(context_reason, 0) + 1
            )
        if not (
            _text(treatment) or _text(actual_treatment) or _text(raw_event)
        ):
            _record_skip(diagnostic, row_number, "unsupported_layout")
            continue

        _, cause = _risk_identity(code)
        entries.append(
            {
                "source_reference": f"III.B:{row_number}",
                "source_sheet": "III.B",
                "source_row": row_number,
                "risk_code": _text(code),
                "source_risk_code": _text(code),
                "source_risk_sequence": no_risiko,
                "source_risk_event": event,
                "risk_event_text": event,
                "no_risiko": no_risiko,
                "cause": cause,
                "normalized_code": normalize_cause_code(code),
                "scope_hint": (
                    "malformed_code" if is_malformed_risk_code(code) else ""
                ),
                "raw_data": {
                    "source_sheet": "III.B",
                    "source_row": row_number,
                    "normalized_treatment_code": normalize_cause_code(code),
                    "source_columns": diagnostic["columns"],
                    "parent_context": context_reason,
                    "continuation_row": continuation,
                    "original_code": _text(code),
                },
                "proposed_data": {
                    "realisasi_rencana_perlakuan": _safe_import_text(
                        _column_value(row, columns, "actual_treatment", 10)
                    ),
                    "realisasi_output_perlakuan": _safe_import_text(
                        _column_value(row, columns, "actual_output", 11)
                    ),
                    "realisasi_biaya_perlakuan": _decimal(
                        _column_value(row, columns, "actual_cost", 12)
                    ),
                    "realisasi_pic": _safe_import_text(
                        _column_value(row, columns, "actual_pic", 14)
                    ),
                    "status_rencana_perlakuan": _treatment_status(
                        _column_value(row, columns, "treatment_status", 27)
                    ),
                    "penjelasan_status_rencana": _safe_import_text(
                        _column_value(
                            row, columns, "treatment_status_note", 28
                        )
                    ),
                    "progress_pelaksanaan_percent": _percent(
                        _column_value(
                            row,
                            columns,
                            "treatment_progress",
                            progress_col,
                        )
                    ),
                    "realisasi_threshold_kri": _safe_import_text(
                        row[threshold_col] if len(row) > threshold_col else ""
                    ),
                    "realisasi_threshold_kri_skor": _safe_import_text(
                        row[threshold_col + 1]
                        if len(row) > threshold_col + 1
                        else ""
                    ),
                    "realisasi_nilai_kri": _decimal(
                        row[threshold_col + 1]
                        if len(row) > threshold_col + 1
                        else None
                    ),
                },
            }
        )
        diagnostic["accepted_rows"] += 1
        if explicit_number:
            last_number = explicit_number
        elif code_number:
            last_number = code_number
        if raw_event:
            last_event = _text(raw_event)
        if raw_code:
            last_code = _text(raw_code)
    if not diagnostic["accepted_rows"]:
        raise ValidationError(
            "Tidak ada data rencana perlakuan valid setelah anchor pada sheet III.B. "
            f"Alasan: {diagnostic['skip_reasons']}."
        )
    if include_diagnostics:
        return entries, diagnostics
    return entries


def _match_item(report, entry):
    items = list(
        report.items.select_related(
            "risk_event", "risk_event__summary__unit_bisnis"
        )
    )
    if not items:
        return None, "target_report_empty", Decimal("0"), []
    if entry.get("scope_hint") == "malformed_code":
        source_name = _normalize(entry.get("risk_event_text"))
        likely = [
            item.pk
            for item in items
            if source_name
            and _normalize(item.risk_event.peristiwa_risiko) == source_name
        ]
        return None, "malformed_code", Decimal("0"), likely

    normalized_code = normalize_risk_code(
        entry.get("risk_code") or entry.get("normalized_code")
    )
    if normalized_code:
        exact_code = [
            item
            for item in items
            if normalized_code in _profile_item_codes(item.risk_event)
        ]
        if len(exact_code) == 1:
            return exact_code[0], "matched_target_profile", Decimal("100"), []
        if len(exact_code) > 1:
            return (
                None,
                "ambiguous_target_profile",
                Decimal("0"),
                [item.pk for item in exact_code],
            )
        unit_name = _text(report.reassessment.unit_bisnis.name)
        unit_key = normalize_risk_code(unit_name)
        unit_short_key = normalize_risk_code(unit_name.split()[-1]) if unit_name else ""
        contains_letters = bool(re.search(r"[A-Z]", normalized_code))
        if (
            entry.get("source_sheet")
            and unit_key
            and contains_letters
            and not normalized_code.startswith(unit_key)
            and not (
                unit_short_key
                and normalized_code.startswith(unit_short_key)
            )
        ):
            return None, "outside_target_profile", Decimal("0"), []

    same_number = [
        item
        for item in items
        if entry["no_risiko"] is not None
        and item.risk_event.no_risiko == entry["no_risiko"]
    ]
    if normalized_code:
        exact_code = [
            item
            for item in same_number
            if normalize_cause_code(
                f"{item.risk_event.no_risiko}"
                f"{item.risk_event.no_penyebab_risiko or ''}"
            ) == normalized_code
            or normalized_code.endswith(
                normalize_cause_code(
                    f"{item.risk_event.no_risiko}"
                    f"{item.risk_event.no_penyebab_risiko or ''}"
                )
            )
        ]
        if len(exact_code) == 1:
            return exact_code[0], "exact_risk_number_and_code", Decimal("100"), []
    if len(same_number) == 1:
        item = same_number[0]
        return item, "exact_risk_number", Decimal("100"), []
    source_event = _normalize(entry["risk_event_text"])
    same_name = [
        item
        for item in items
        if source_event
        and _normalize(item.risk_event.peristiwa_risiko) == source_event
    ]
    if len(same_name) == 1:
        return same_name[0], "unique_normalized_name", Decimal("85"), []
    if len(same_number) > 1 and entry.get("source_sheet") == "III.A":
        return (
            None,
            "ambiguous_target_profile",
            Decimal("0"),
            [item.pk for item in same_number],
        )
    if len(same_number) > 1 or len(same_name) > 1:
        candidates = sorted({item.pk for item in same_number + same_name})
        return None, "ambiguous", Decimal("0"), candidates

    return None, "unmatched", Decimal("0"), []


def _validate_entry(entry, item, method, confidence, candidates):
    issues = []
    fatal = False
    proposed = entry["proposed_data"]
    probability = proposed.get("realisasi_nilai_probabilitas")
    progress = proposed.get("progress_pelaksanaan_percent")
    for field_name in entry.get("raw_data", {}).get("empty_quarter_fields", []):
        source_cell = entry["raw_data"].get("source_cells", {}).get(field_name, "")
        issues.append(
            f"{FIELD_LABELS[field_name]} untuk quarter terpilih kosong ({source_cell}); "
            "nilai quarter sebelumnya tidak digunakan."
        )
    if probability is not None and not 0 <= probability <= 100:
        issues.append("Nilai probabilitas harus berada antara 0 dan 100%.")
        fatal = True
    if progress is not None and not 0 <= progress <= 100:
        issues.append("Progress pelaksanaan harus berada antara 0 dan 100%.")
        fatal = True
    if proposed.get("realisasi_biaya_perlakuan") is not None and proposed["realisasi_biaya_perlakuan"] < 0:
        issues.append("Realisasi biaya tidak boleh negatif.")
        fatal = True
    for field_name in (
        "realisasi_skala_dampak_kbumn",
        "realisasi_skala_probabilitas_kbumn",
    ):
        value = proposed.get(field_name)
        if value is not None and not 1 <= int(value) <= 5:
            issues.append(f"{FIELD_LABELS.get(field_name, field_name)} harus 1 sampai 5.")
            fatal = True
    risk_scale = proposed.get("realisasi_skala_nilai_risiko_kbumn")
    if risk_scale is not None and not 1 <= int(risk_scale) <= 25:
        issues.append("Skala nilai risiko KBUMN harus 1 sampai 25.")
        fatal = True
    bumn_risk_scale = proposed.get("realisasi_skor_risiko")
    if bumn_risk_scale is not None and not 1 <= int(bumn_risk_scale) <= 25:
        issues.append("Skala nilai risiko BUMN harus 1 sampai 25.")
        fatal = True
    exposure = proposed.get("realisasi_eksposur")
    if exposure is not None and Decimal(str(exposure)) < 0:
        issues.append("Nilai eksposur risiko tidak boleh negatif.")
        fatal = True
    effectiveness = proposed.get("efektivitas_perlakuan_risiko")
    if effectiveness not in (None, "", "efektif", "tidak_efektif"):
        issues.append("Efektivitas harus Efektif atau Tidak Efektif.")
        fatal = True
    if not item:
        if method == "outside_target_profile":
            issues.append(
                "Baris berada di luar Profil Risiko target dan akan diabaikan."
            )
            return MonthlyRiskReportImportRow.LEVEL_GREEN, issues
        if method == "malformed_code":
            candidate_text = (
                ", kandidat target: " + ", ".join(str(value) for value in candidates)
                if candidates
                else ""
            )
            issues.append(
                "Kode risiko malformed; perlu konfirmasi manual"
                f"{candidate_text}."
            )
        elif method in {"ambiguous", "ambiguous_target_profile"}:
            issues.append(
                "Pencocokan ambigu; kandidat target: "
                + ", ".join(str(value) for value in candidates)
                + "."
            )
        else:
            issues.append("Risiko pada Excel belum dapat dicocokkan dengan item laporan ERM.")
        fatal = True
    elif entry["risk_event_text"] and _normalize(entry["risk_event_text"]) != _normalize(
        item.risk_event.peristiwa_risiko
    ):
        issues.append(
            "Nomor/kode cocok, tetapi nama peristiwa berbeda; mohon konfirmasi."
        )
    if fatal:
        return MonthlyRiskReportImportRow.LEVEL_RED, issues
    if issues:
        return MonthlyRiskReportImportRow.LEVEL_YELLOW, issues
    return MonthlyRiskReportImportRow.LEVEL_GREEN, issues


def _extract_json(text):
    cleaned = (text or "").strip().strip("`")
    if cleaned.lower().startswith("json"):
        cleaned = cleaned[4:].strip()
    start, end = cleaned.find("{"), cleaned.rfind("}")
    if start >= 0 and end > start:
        cleaned = cleaned[start:end + 1]
    return json.loads(cleaned)


def _run_ai_review(batch, rows):
    setting = AppSetting.get_solo()
    if not setting.ai_aktif or not setting.runtime_ai_api_key:
        return False, "AI tidak aktif; analisis deterministik digunakan."
    ambiguous = [row for row in rows if row.validation_level != row.LEVEL_GREEN]
    if not ambiguous:
        return False, "Semua baris cocok secara deterministik; AI tidak diperlukan."
    payload = [
        {
            "row_id": row.pk,
            "risk_code": row.risk_code,
            "risk_event": row.risk_event_text,
            "matched_item_id": row.matched_report_item_id,
            "confidence": float(row.confidence),
            "issues": row.issues,
        }
        for row in ambiguous[:100]
    ]
    prompt = (
        "Anda adalah analis ERM PLN Batam. Tinjau hasil validasi import berikut tanpa "
        "menciptakan fakta baru. Kembalikan JSON valid: {summary:string, rows:[{row_id:int," 
        "analysis:string}]}. Jelaskan singkat hal yang perlu dikonfirmasi user. Data: "
        + json.dumps(payload, ensure_ascii=False)
    )
    try:
        if setting.ai_provider == AppSetting.AI_PROVIDER_GEMINI:
            base = (setting.ai_base_url or "https://generativelanguage.googleapis.com/v1beta").rstrip("/")
            if "api.openai.com" in base:
                base = "https://generativelanguage.googleapis.com/v1beta"
            model = setting.ai_model if not setting.ai_model.startswith("gpt-") else "gemini-3.1-flash-lite"
            response = httpx.post(
                f"{base}/models/{model}:generateContent",
                headers={"x-goog-api-key": setting.runtime_ai_api_key},
                json={"contents": [{"parts": [{"text": prompt}]}], "generationConfig": {"responseMimeType": "application/json"}},
                timeout=30,
            )
            response.raise_for_status()
            data = response.json()
            text = data["candidates"][0]["content"]["parts"][0]["text"]
        else:
            base = (setting.ai_base_url or "https://api.openai.com/v1").rstrip("/")
            response = httpx.post(
                f"{base}/chat/completions",
                headers={"Authorization": f"Bearer {setting.runtime_ai_api_key}"},
                json={"model": setting.ai_model or "gpt-4.1-mini", "temperature": float(setting.ai_temperature), "messages": [{"role": "user", "content": prompt}], "response_format": {"type": "json_object"}},
                timeout=30,
            )
            response.raise_for_status()
            text = response.json()["choices"][0]["message"]["content"]
        result = _extract_json(text)
        by_id = {row.pk: row for row in ambiguous}
        for ai_row in result.get("rows", []):
            row = by_id.get(ai_row.get("row_id"))
            if row:
                row.ai_analysis = _text(ai_row.get("analysis"))
                row.save(update_fields=["ai_analysis", "updated_at"])
        return True, _text(result.get("summary")) or "Analisis AI selesai."
    except Exception as exc:
        return False, f"AI tidak tersedia ({exc}); analisis deterministik tetap dapat direview."


@transaction.atomic
def analyze_import_batch(batch):
    if batch.report.status not in {"draft", "revision"}:
        raise ValidationError("Import hanya dapat dilakukan pada laporan Draft atau Revision.")
    batch.rows.all().delete()
    entries, parser_diagnostics = _parse_workbook(
        batch, include_diagnostics=True
    )
    if not entries:
        raise ValidationError("Tidak ditemukan baris risiko valid pada sheet III.A/III.B.")
    source_risks = sum(entry["source_sheet"] == "III.A" for entry in entries)
    source_treatments = sum(entry["source_sheet"] == "III.B" for entry in entries)
    summary = {
        "source_risks": source_risks,
        "source_treatments": source_treatments,
        "source_total": len(entries),
        "target_risks": batch.report.items.count(),
        "matched_risks": 0,
        "target_only": 0,
        "target_only_ids": [],
        "parser_diagnostics": parser_diagnostics,
        "matched": 0,
        "ambiguous": 0,
        "outside_target_profile": 0,
        "malformed": 0,
        "unmatched": 0,
        "continuation_rows": parser_diagnostics.get("III.B", {}).get(
            "continuation_rows", 0
        ),
        "invalid": 0,
        "skipped": 0,
    }
    fingerprint = target_item_fingerprint(batch.report)
    if not batch.report.items.exists():
        batch.status = batch.STATUS_REVIEW
        batch.parser_version = IMPORT_PARSER_VERSION
        batch.target_fingerprint = fingerprint
        batch.analysis_summary = summary
        batch.blocking_reason = (
            f"Laporan {batch.report.periode.nama_periode} belum memiliki item risiko "
            "sebagai target import."
        )
        batch.ai_used = False
        batch.ai_summary = (
            f"Risiko sumber: {source_risks}; rencana perlakuan sumber: "
            f"{source_treatments}; total sumber: {len(entries)}. "
            "Matching tidak dijalankan karena target kosong."
        )
        batch.analyzed_at = timezone.now()
        batch.error_message = ""
        batch.save(
            update_fields=[
                "status",
                "parser_version",
                "target_fingerprint",
                "analysis_summary",
                "blocking_reason",
                "ai_used",
                "ai_summary",
                "analyzed_at",
                "error_message",
                "updated_at",
            ]
        )
        return batch
    rows = []
    for entry in entries:
        item, method, confidence, candidates = _match_item(batch.report, entry)
        level, issues = _validate_entry(
            entry, item, method, confidence, candidates
        )
        proposed = {
            key: _json_value(value)
            for key, value in entry["proposed_data"].items()
            if value not in (None, "")
            or (
                entry["source_sheet"] == "III.A"
                and key in entry["raw_data"].get("empty_quarter_fields", [])
            )
        }
        rows.append(
            MonthlyRiskReportImportRow.objects.create(
                batch=batch,
                source_reference=entry["source_reference"],
                risk_code=entry["risk_code"],
                risk_event_text=entry["risk_event_text"],
                matched_report_item=item,
                match_method=method,
                confidence=confidence,
                validation_level=level,
                issues=issues,
                raw_data={
                    **entry["raw_data"],
                    "parser_version": IMPORT_PARSER_VERSION,
                    "source_risk_code": entry.get("source_risk_code", ""),
                    "source_risk_sequence": entry.get("source_risk_sequence"),
                    "source_risk_event": entry.get("source_risk_event", ""),
                    "normalized_risk_name": _normalize(entry["risk_event_text"]),
                    "ambiguity_candidates": candidates,
                },
                proposed_data=proposed,
                user_decision=(
                    "skip"
                    if method == "outside_target_profile"
                    else "import"
                    if level == "green"
                    else "pending"
                ),
            )
        )
        if item:
            summary["matched"] += 1
        if method in {"ambiguous", "ambiguous_target_profile"}:
            summary["ambiguous"] += 1
        if method == "outside_target_profile":
            summary["outside_target_profile"] += 1
        if method == "malformed_code":
            summary["malformed"] += 1
        if method == "unmatched":
            summary["unmatched"] += 1
        if level == MonthlyRiskReportImportRow.LEVEL_RED:
            summary["invalid"] += 1
    matched_risk_ids = {
        row.matched_report_item_id
        for row in rows
        if row.raw_data.get("source_sheet") == "III.A"
        and row.matched_report_item_id
    }
    target_ids = set(batch.report.items.values_list("pk", flat=True))
    target_only_ids = sorted(target_ids - matched_risk_ids)
    summary["matched_risks"] = len(matched_risk_ids)
    summary["target_only"] = len(target_only_ids)
    summary["target_only_ids"] = target_only_ids
    ai_used, ai_summary = _run_ai_review(batch, rows)
    batch.status = batch.STATUS_REVIEW
    batch.ai_used = ai_used
    batch.ai_summary = ai_summary
    batch.analyzed_at = timezone.now()
    batch.error_message = ""
    batch.parser_version = IMPORT_PARSER_VERSION
    batch.target_fingerprint = fingerprint
    batch.analysis_summary = summary
    batch.blocking_reason = ""
    batch.save(update_fields=[
        "status",
        "ai_used",
        "ai_summary",
        "analyzed_at",
        "error_message",
        "parser_version",
        "target_fingerprint",
        "analysis_summary",
        "blocking_reason",
        "updated_at",
    ])
    return batch


def _field_value(item, field_name):
    value = getattr(item, field_name)
    return _json_value(value)


@transaction.atomic
def apply_import_batch(batch, user):
    batch = MonthlyRiskReportImportBatch.objects.select_for_update().select_related("report").get(pk=batch.pk)
    if batch.status != batch.STATUS_REVIEW:
        raise ValidationError("Batch import ini tidak lagi menunggu konfirmasi.")
    if batch.report.status not in {"draft", "revision"}:
        raise ValidationError("Laporan bukan lagi Draft/Revision sehingga import dibatalkan.")
    if batch.blocking_reason:
        raise ValidationError(batch.blocking_reason)
    if not batch_analysis_is_current(batch):
        raise ValidationError(
            "Struktur target laporan berubah setelah preview. Jalankan analisis ulang."
        )
    rows = list(batch.rows.select_related("matched_report_item"))
    unresolved = [
        row for row in rows
        if row.validation_level in {row.LEVEL_YELLOW, row.LEVEL_RED}
        and row.user_decision == row.DECISION_PENDING
    ]
    if unresolved:
        raise ValidationError("Masih ada baris kuning/merah yang belum dikonfirmasi.")
    for row in rows:
        if row.user_decision == row.DECISION_SKIP:
            continue
        if row.validation_level == row.LEVEL_RED or not row.matched_report_item_id:
            raise ValidationError(f"{row.source_reference} berstatus merah dan tidak dapat diimpor.")
        item = row.matched_report_item
        previous = {}
        applied = {}
        for field_name, raw_value in row.proposed_data.items():
            if field_name not in IMPORTABLE_FIELDS:
                continue
            if (
                field_name in {
                    "realisasi_nilai_kri",
                    "realisasi_threshold_kri",
                    "realisasi_threshold_kri_skor",
                }
                and raw_value in (None, "")
            ):
                continue
            previous[field_name] = _field_value(item, field_name)
            model_field = field_name[:-3] if field_name.endswith("_id") else field_name
            field = item._meta.get_field(model_field)
            value = raw_value
            if field.get_internal_type() == "DecimalField" and raw_value not in (None, ""):
                value = Decimal(str(raw_value))
            setattr(item, field_name, value)
            applied[field_name] = _json_value(value)
        item.full_clean()
        item.save(update_fields=[*applied.keys(), "updated_at"])
        row.previous_data = previous
        row.applied_data = applied
        row.save(update_fields=["previous_data", "applied_data", "updated_at"])
    refresh_monthly_report_summary(batch.report)
    batch.status = batch.STATUS_IMPORTED
    batch.imported_by = user
    batch.imported_at = timezone.now()
    batch.save(update_fields=["status", "imported_by", "imported_at", "updated_at"])
    MonthlyRiskReportSubmissionLog.objects.create(
        report=batch.report,
        action="import",
        action_by=user,
        note=f"Import Excel {batch.original_filename}; batch ID {batch.pk}.",
    )
    return batch
