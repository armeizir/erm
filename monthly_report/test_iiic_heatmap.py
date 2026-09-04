from datetime import date
from types import SimpleNamespace

from django.test import SimpleTestCase
from openpyxl import load_workbook

from .excel_reports import MONTHLY_REPORT_TEMPLATE, _fill_iiic


class MonthlyRiskReportIIICHeatmapTests(SimpleTestCase):
    @staticmethod
    def _scale(number):
        return SimpleNamespace(urutan=number)

    def _item(
        self,
        *,
        pk,
        risk_number,
        impact,
        likelihood,
        score,
        residual_impact=None,
        residual_likelihood=None,
        residual_score=None,
    ):
        risk = SimpleNamespace(
            id=1000 + pk,
            no_risiko=risk_number,
            no_item=risk_number,
            peristiwa_risiko=f"Risiko {risk_number}",
            skala_dampak_q2=self._scale(impact),
            skala_probabilitas_q2=self._scale(likelihood),
            skala_risiko_q2=score,
        )
        return SimpleNamespace(
            pk=pk,
            risk_event=risk,
            risk_event_id=risk.id,
            realisasi_skala_dampak=(
                self._scale(residual_impact) if residual_impact else None
            ),
            realisasi_skala_probabilitas=(
                self._scale(residual_likelihood) if residual_likelihood else None
            ),
            realisasi_skor_risiko=residual_score,
        )

    def test_iiic_plots_inherent_and_residual_markers_on_matrix(self):
        workbook = load_workbook(MONTHLY_REPORT_TEMPLATE, data_only=False)
        report = SimpleNamespace(
            periode=SimpleNamespace(tanggal_mulai=date(2026, 4, 1))
        )
        items = [
            self._item(
                pk=1,
                risk_number=1,
                impact=5,
                likelihood=4,
                score=24,
                residual_impact=3,
                residual_likelihood=2,
                residual_score=11,
            ),
            self._item(
                pk=2,
                risk_number=8,
                impact=3,
                likelihood=4,
                score=14,
            ),
        ]

        _fill_iiic(workbook, report, items)
        ws = workbook["III.C"]

        self.assertEqual(ws["AE8"].value, 1)
        self.assertEqual(ws["AG8"].value, 5)
        self.assertEqual(ws["AH8"].value, 4)
        self.assertEqual(ws["AI8"].value, 24)
        self.assertIn("●I: 1", ws["Y10"].value)
        self.assertIn("●I: 8", ws["O10"].value)
        self.assertIn("○R: 1", ws["O20"].value)
        self.assertEqual(ws["B4"].value, "1")
        self.assertEqual(ws["C4"].value, "Risiko Inheren")
        self.assertEqual(ws["F4"].value, "1")
        self.assertEqual(ws["G4"].value, "Risiko Residual Triwulan/Tahun Berjalan")
        self.assertIn("C4:E4", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertIn("G4:M4", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertEqual(ws["B4"].fill.fgColor.rgb, "004E73C4")
        self.assertEqual(ws["AT6"].value, "●I = Risiko Inheren")
        self.assertEqual(
            ws["AT8"].value,
            "○R = Risiko Residual Triwulan/Tahun Berjalan",
        )

    def test_iiic_plot_is_idempotent_and_groups_same_cell(self):
        workbook = load_workbook(MONTHLY_REPORT_TEMPLATE, data_only=False)
        report = SimpleNamespace(
            periode=SimpleNamespace(tanggal_mulai=date(2026, 4, 1))
        )
        items = [
            self._item(pk=1, risk_number=1, impact=5, likelihood=4, score=24),
            self._item(pk=2, risk_number=8, impact=5, likelihood=4, score=24),
        ]

        _fill_iiic(workbook, report, items)
        _fill_iiic(workbook, report, items)
        value = workbook["III.C"]["Y10"].value

        self.assertEqual(value.count("●I:"), 1)
        self.assertIn("●I: 1, 8", value)
