#!/usr/bin/env python3
"""
Safe importer Monthly Risk Report BID AGA — Mei, Juni, Juli 2026.

Sumber:
- Laporan Realisasi Mei 2026 BID AGA (1).xlsx
- Laporan Realisasi Juni 2026 BID AGA (2).xlsx
- Laporan Realisasi Juli 2026 BID AGA.xlsx

Kondisi master yang DIHARAPKAN sebelum import:
- ReAssessmentSummary ID=1
- Judul = Profil Risiko NIAGA
- Unit = BID AGA
- Tahun = 2026
- Tepat 14 ReAssessmentItem, no_item 1..14
- Pairing KM final sudah benar (A.01 ... F.14)

Prinsip keselamatan:
- Default DRY RUN: seluruh perubahan database di-rollback.
- --apply wajib untuk commit.
- Seluruh Mei+Juni+Juli diproses dalam satu transaction.atomic().
- Pada --apply, SQLite dibackup lebih dulu dan PRAGMA quick_check diverifikasi.
- Hash ketiga workbook harus sama dengan sumber yang direview.
- Tidak mengubah KM, pairing km_item, event, cause, KRI master, atau struktur profil.
- Mapping source -> master = nomor item 1..14 + exact normalized event validation.
- Existing report hanya boleh status draft/revision dan tidak locked.
- Jika ada >1 report untuk profil/periode/versi yang sama, import dibatalkan.
- Tidak menghapus MonthlyRiskReportItem valid. Missing item dibuat; item asing/extras menyebabkan abort.
- Mei dan Juni memakai residual Q2; Juli memakai residual Q3. TANPA fallback quarter.
- Nilai kosong pada source-owned monthly fields akan mengosongkan nilai lama saat re-import.
- Untuk risiko kualitatif, nilai dampak numerik dibiarkan None meski template berisi 0.
- KRI aktual memakai kolom bulan berjalan saja; TANPA fallback bulan sebelumnya.
- Threshold + skor KRI dari workbook dipertahankan sebagai historical source-of-truth.
- Numeric KRI hanya disimpan bila cell murni numerik dan lolos evaluator threshold aktif;
  jika tidak kompatibel, numeric dibiarkan None dan warning dicetak (tidak bypass model).
- III.D disinkronkan sesuai sumber.
- III.E: dua baris legacy/template 2024/2025 diabaikan. Unknown existing loss event
  dilindungi: jika source tidak punya loss event valid tetapi DB punya loss event non-legacy,
  import dibatalkan agar tidak menghapus data manual tanpa review.
- Budget rencana pada master TIDAK diubah. Script hanya mengaudit apakah total budget
  master berbeda dari workbook. Monthly actual cost tetap diimpor.
- Tidak memanggil full_clean() pada ReAssessmentItem master, sehingga legacy choices
  seperti "Low to Moderate" pada field profil tidak menghambat import bulanan.

Run dari project root:
  export DJANGO_SETTINGS_MODULE=riskproject.settings.prod

  python monthly_report/scripts/import_aga_may_jun_jul_2026.py \
    --may /tmp/aga_mei_2026.xlsx \
    --june /tmp/aga_juni_2026.xlsx \
    --july /tmp/aga_juli_2026.xlsx

Apply hanya setelah dry-run direview:
  python monthly_report/scripts/import_aga_may_jun_jul_2026.py \
    --may /tmp/aga_mei_2026.xlsx \
    --june /tmp/aga_juni_2026.xlsx \
    --july /tmp/aga_juli_2026.xlsx \
    --apply
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import os
import re
import sqlite3
import sys
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

if not os.environ.get("DJANGO_SETTINGS_MODULE"):
    raise RuntimeError(
        "DJANGO_SETTINGS_MODULE belum diset. Untuk PROD jalankan: "
        "export DJANGO_SETTINGS_MODULE=riskproject.settings.prod"
    )

import django  # noqa: E402

django.setup()

from django.conf import settings  # noqa: E402
from django.core.exceptions import ValidationError  # noqa: E402
from django.contrib.auth import get_user_model  # noqa: E402
from django.db import transaction  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from masterdata.models import PeriodeLaporan, TahunBuku  # noqa: E402
from monthly_report.models import (  # noqa: E402
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from monthly_report.services import refresh_monthly_report_summary  # noqa: E402
from risk.models import (  # noqa: E402
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)


YEAR = 2026
PROFILE_ID = 1
PROFILE_TITLE = "Profil Risiko NIAGA"
UNIT_NAME = "BID AGA"
EXPECTED_SOURCE_ITEMS = 14

MONTH_NAMES = {5: "Mei", 6: "Juni", 7: "Juli"}
MONTH_QUARTER = {5: 2, 6: 2, 7: 3}

EXPECTED_SHA256 = {
    5: "0377c28a934e0b16cc4d576af3ffeaf1badc1e594d9a54e07895e137e0d2d301",
    6: "ec36d32e63e65101bb56b70adb5294c2d758701bddea554e8b168f7179f9ecb0",
    7: "034fdfb77a9d29d3e6ca6bb6df87ac5702d39b0e79cda65f8998ebb0db756548",
}

REPORT_CODES = {
    5: "MRR-BIDAGA-2026-05",
    6: "MRR-BIDAGA-2026-06",
    7: "MRR-BIDAGA-2026-07",
}

# III.A: B=nomor risk; C=event; D=kategori; N=asumsi residual.
IIIA_RISK_NO_INDEX = 1
IIIA_EVENT_INDEX = 2
IIIA_CATEGORY_INDEX = 3
IIIA_ASSUMPTION_INDEX = 13

# III.A official residual Q1..Q4 — zero based indexes.
IIIA_QUARTER_COLUMNS = {
    "realisasi_nilai_dampak": {1: 14, 2: 15, 3: 16, 4: 17},               # O:R
    "realisasi_skala_dampak": {1: 18, 2: 19, 3: 20, 4: 21},               # S:V
    "realisasi_skala_dampak_kbumn": {1: 22, 2: 23, 3: 24, 4: 25},         # W:Z
    "realisasi_nilai_probabilitas": {1: 26, 2: 27, 3: 28, 4: 29},         # AA:AD
    "realisasi_skala_probabilitas": {1: 30, 2: 31, 3: 32, 4: 33},         # AE:AH
    "realisasi_skala_probabilitas_kbumn": {1: 34, 2: 35, 3: 36, 4: 37},   # AI:AL
    "realisasi_eksposur": {1: 38, 2: 39, 3: 40, 4: 41},                    # AM:AP
    "realisasi_skor_risiko": {1: 42, 2: 43, 3: 44, 4: 45},                # AQ:AT
    "realisasi_skala_nilai_risiko_kbumn": {1: 46, 2: 47, 3: 48, 4: 49},  # AU:AX
    "realisasi_level_risiko_bumn": {1: 50, 2: 51, 3: 52, 4: 53},          # AY:BB
    "realisasi_level_risiko_kbumn": {1: 54, 2: 55, 3: 56, 4: 57},         # BC:BF
}
IIIA_EFFECTIVENESS_INDEX = 58  # BG

# III.B: B=nomor risk, C=event, F=kode penyebab BID AGA-N-a.
IIIB_RISK_NO_INDEX = 1
IIIB_EVENT_INDEX = 2
IIIB_CAUSE_NO_INDEX = 4
IIIB_CODE_INDEX = 5

# Rencana (audit only, tidak mengubah master).
IIIB_PLAN_INDEX = 7       # H
IIIB_OUTPUT_PLAN_INDEX = 8  # I
IIIB_BUDGET_INDEX = 9     # J

# Realisasi.
IIIB_ACTUAL_TREATMENT_INDEX = 10  # K
IIIB_ACTUAL_OUTPUT_INDEX = 11     # L
IIIB_ACTUAL_COST_INDEX = 12       # M
IIIB_ACTUAL_PIC_INDEX = 14        # O
IIIB_TIMELINE_FIRST = 15          # P = Jan ... V = Jul
IIIB_STATUS_INDEX = 27            # AB
IIIB_STATUS_NOTE_INDEX = 28       # AC
IIIB_PROGRESS_BY_QUARTER = {1: 29, 2: 30, 3: 31, 4: 32}  # AD:AG

# KRI master/audit.
IIIB_KRI_NAME_INDEX = 33       # AH
IIIB_KRI_UNIT_INDEX = 34       # AI
IIIB_KRI_SAFE_INDEX = 35       # AJ
IIIB_KRI_CAUTION_INDEX = 36    # AK
IIIB_KRI_DANGER_INDEX = 37     # AL

LEGACY_LOSS_EVENT_NAMES = {
    "kerugian keuangan akibat harga gas dan kurs",
    "harga jual rata rata dibawah target rkap",
    "harga jual rata rata di bawah target rkap",
}


@dataclass(frozen=True)
class ImportSpec:
    month: int
    path: Path


@dataclass
class MonthStats:
    month: int
    report_id: int | None = None
    report_created: bool = False
    source_iiia: int = 0
    source_iiib: int = 0
    created_items: int = 0
    updated_items: int = 0
    field_changes: int = 0
    source_change_rows: int = 0
    source_valid_loss_rows: int = 0
    source_legacy_loss_rows: int = 0
    deleted_legacy_db_loss_rows: int = 0
    kri_with_data: int = 0
    kri_blank: list[str] = field(default_factory=list)
    kri_numeric_stored: list[str] = field(default_factory=list)
    kri_numeric_skipped: list[str] = field(default_factory=list)
    residual_incomplete: list[str] = field(default_factory=list)
    category_mismatches: list[str] = field(default_factory=list)
    current_month_timeline_on: list[str] = field(default_factory=list)
    current_month_timeline_off: list[str] = field(default_factory=list)
    master_budget_mismatches: list[str] = field(default_factory=list)
    kri_master_mismatches: list[str] = field(default_factory=list)
    report_item_count: int = 0
    distinct_risk_count: int = 0


def normalize(value) -> str:
    text = str(value or "").casefold().replace("\xa0", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def text_or_none(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.upper() in {"#NAME?", "#REF!", "#N/A", "#DIV/0!"}:
        return None
    return text


def decimal_or_none(value):
    if value in (None, ""):
        return None
    multiplier = Decimal("1")
    if isinstance(value, str):
        stripped = value.strip().casefold()
        if stripped in {
            "-", "n/a", "na", "no data", "none", "null",
            "#div/0!", "#n/a", "#name?", "#ref!",
        }:
            return None
        if re.search(r"miliar|milyar", stripped):
            multiplier = Decimal("1000000000")
        elif "juta" in stripped:
            multiplier = Decimal("1000000")
        cleaned = re.sub(r"[^0-9,.\-]", "", stripped)
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        if not cleaned or cleaned in {"-", ".", "-."}:
            return None
        value = cleaned
    try:
        return Decimal(str(value).strip()) * multiplier
    except (InvalidOperation, ValueError, TypeError):
        return None


def kri_numeric_or_none(value):
    """
    KRI numeric hanya diambil bila cell memang numerik murni.

    Penting: cell seperti "R = 97%\nI = 33%" adalah teks komposit dan TIDAK
    boleh diubah menjadi 9733. Nilai tersebut tetap disimpan pada field skor teks.
    """
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    text = str(value).strip()
    if not text:
        return None

    # Hanya izinkan satu angka murni opsional tanda %, pemisah ribuan/desimal.
    # Adanya huruf, newline, '=', atau lebih dari satu angka berarti skor bersifat teks.
    if re.search(r"[A-Za-z]", text) or "\n" in text or "=" in text:
        return None
    if not re.fullmatch(r"[-+]?\s*[0-9][0-9.,]*\s*%?", text):
        return None

    number = decimal_or_none(text.rstrip("%").strip())
    if number is None:
        return None
    return number


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    # Workbook menyimpan probabilitas/progress sebagai fraction 0..1.
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    if Decimal("0") <= number <= Decimal("100"):
        return number
    return None


def int_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    try:
        return int(number)
    except (ValueError, TypeError, OverflowError):
        return None


def value_at(row, zero_index):
    return row[zero_index] if len(row) > zero_index else None


def file_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_risk_number(value):
    number = int_or_none(value)
    return number if number is not None and 1 <= number <= EXPECTED_SOURCE_ITEMS else None


def cause_code_risk_number(value):
    match = re.search(r"bid\s*aga\s*[- ]?(\d+)\s*[- ]?([a-z])?", normalize(value))
    if not match:
        return None, None
    return int(match.group(1)), (match.group(2) or None)


def get_profile():
    profile = (
        ReAssessmentSummary.objects
        .select_related("unit_bisnis", "kontrak_manajemen")
        .filter(pk=PROFILE_ID)
        .first()
    )
    if profile is None:
        raise RuntimeError(f"Profil AGA expected ID={PROFILE_ID} tidak ditemukan.")
    if profile.tahun != YEAR:
        raise RuntimeError(f"Profile ID={PROFILE_ID} tahun={profile.tahun}, expected {YEAR}.")
    if normalize(profile.judul) != normalize(PROFILE_TITLE):
        raise RuntimeError(
            f"Profile ID={PROFILE_ID} judul={profile.judul!r}, expected {PROFILE_TITLE!r}."
        )
    unit_name = getattr(getattr(profile, "unit_bisnis", None), "name", "")
    if normalize(unit_name) != normalize(UNIT_NAME):
        raise RuntimeError(
            f"Profile ID={PROFILE_ID} unit={unit_name!r}, expected {UNIT_NAME!r}."
        )
    if profile.item.count() != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"Profile ID={PROFILE_ID} harus tepat {EXPECTED_SOURCE_ITEMS} item; "
            f"ditemukan {profile.item.count()}."
        )
    return profile


def build_master_map(profile):
    items = list(
        ReAssessmentItem.objects
        .filter(summary=profile)
        .select_related("kategori_dampak", "km_item", "km_item__master_bagian")
        .order_by("no_item", "id")
    )
    mapping = {}
    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        found = [x for x in items if int(x.no_item or 0) == n]
        if len(found) != 1:
            raise RuntimeError(
                f"Master AGA no_item={n} harus tepat satu; ditemukan {len(found)}: "
                + ", ".join(
                    f"RE={x.pk}/R{x.no_risiko} {x.peristiwa_risiko}" for x in found
                )
            )
        mapping[n] = found[0]

    if len({x.pk for x in mapping.values()}) != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError("Mapping master AGA tidak menghasilkan 14 target unik.")

    print("\nMASTER AGA FINAL 14 RISIKO:")
    for n, item in mapping.items():
        km = getattr(item, "km_item", None)
        km_label = "-"
        if km and getattr(km, "master_bagian", None):
            km_label = f"{km.master_bagian.kode_bagian}.{int(km.no_urut):02d}"
        print(
            f"- {n:02d} -> RE={item.pk:<4} | KM={km_label:<5} "
            f"| R={item.no_risiko} | {item.peristiwa_risiko}"
        )
    return mapping


def is_qualitative_master(item):
    category = normalize(getattr(getattr(item, "kategori_dampak", None), "nama", ""))
    return "kual" in category


def scale_by_level(model, value, label):
    number = int_or_none(value)
    if number is None:
        return None
    if not 1 <= number <= 5:
        raise RuntimeError(f"{label}: nilai skala {value!r} di luar 1..5.")
    obj = model.objects.filter(urutan=number).first()
    if obj is None:
        raise RuntimeError(f"{label}: master skala urutan={number} tidak ditemukan.")
    return obj


def treatment_status(value):
    txt = normalize(value)
    if not txt:
        return None
    if "discontinue" in txt:
        return "discontinue"
    if "continue" in txt:
        return "continue"
    # Jangan membuat choice ilegal bila source memakai label lain.
    return None


def treatment_effectiveness(value):
    txt = normalize(value)
    if not txt:
        return None
    if "tidak efektif" in txt:
        return "tidak_efektif"
    if "efektif" in txt:
        return "efektif"
    return None


def parse_iiia(workbook, month, master_map, stats):
    ws = workbook["III.A"]
    parsed = {}

    for row_no, cells in enumerate(ws.iter_rows(values_only=True), start=1):
        row = list(cells)
        n = source_risk_number(value_at(row, IIIA_RISK_NO_INDEX))
        if n is None:
            continue
        event = text_or_none(value_at(row, IIIA_EVENT_INDEX))
        if not event:
            continue

        # Hanya blok AGA 1..14 yang event-nya sama dengan master final.
        master = master_map[n]
        if normalize(event) != normalize(master.peristiwa_risiko):
            # Bisa saja ada nomor 1..14 pada template lain; jangan salah ambil.
            continue

        if n in parsed:
            raise RuntimeError(
                f"III.A {MONTH_NAMES[month]}: risiko {n} duplikat "
                f"(row {parsed[n][0]} dan {row_no})."
            )
        parsed[n] = (row_no, row)

    stats.source_iiia = len(parsed)
    expected = set(range(1, EXPECTED_SOURCE_ITEMS + 1))
    if set(parsed) != expected:
        missing = sorted(expected - set(parsed))
        raise RuntimeError(
            f"III.A {MONTH_NAMES[month]} harus menghasilkan 14 risiko AGA; "
            f"ditemukan {len(parsed)}. Missing={missing}."
        )
    return parsed


def parse_iiib(workbook, month, master_map, stats):
    ws = workbook["III.B"]
    parsed = {}

    for row_no, cells in enumerate(ws.iter_rows(values_only=True), start=1):
        row = list(cells)
        code = value_at(row, IIIB_CODE_INDEX)
        n_code, cause_letter = cause_code_risk_number(code)
        if n_code is None or not 1 <= n_code <= EXPECTED_SOURCE_ITEMS:
            continue

        n_col = source_risk_number(value_at(row, IIIB_RISK_NO_INDEX))
        if n_col != n_code:
            raise RuntimeError(
                f"III.B {MONTH_NAMES[month]} row {row_no}: "
                f"nomor B={n_col} berbeda dengan kode {code!r}."
            )

        event = text_or_none(value_at(row, IIIB_EVENT_INDEX))
        if not event:
            raise RuntimeError(
                f"III.B {MONTH_NAMES[month]} row {row_no}: event kosong untuk {code!r}."
            )
        master = master_map[n_code]
        if normalize(event) != normalize(master.peristiwa_risiko):
            raise RuntimeError(
                f"III.B {MONTH_NAMES[month]} risiko {n_code}: event source tidak sama "
                f"dengan master RE={master.pk}.\n"
                f"MASTER={master.peristiwa_risiko}\nSOURCE={event}"
            )

        if n_code in parsed:
            raise RuntimeError(
                f"III.B {MONTH_NAMES[month]}: risiko {n_code} duplikat "
                f"(row {parsed[n_code][0]} dan {row_no})."
            )
        parsed[n_code] = (row_no, row, cause_letter)

    stats.source_iiib = len(parsed)
    expected = set(range(1, EXPECTED_SOURCE_ITEMS + 1))
    if set(parsed) != expected:
        missing = sorted(expected - set(parsed))
        raise RuntimeError(
            f"III.B {MONTH_NAMES[month]} harus menghasilkan 14 risiko AGA; "
            f"ditemukan {len(parsed)}. Missing={missing}."
        )
    return parsed


def validate_iiia_vs_iiib(parsed_a, parsed_b, month):
    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        row_a, values_a = parsed_a[n]
        row_b, values_b, _ = parsed_b[n]
        event_a = normalize(value_at(values_a, IIIA_EVENT_INDEX))
        event_b = normalize(value_at(values_b, IIIB_EVENT_INDEX))
        if event_a != event_b:
            raise RuntimeError(
                f"{MONTH_NAMES[month]} risiko {n}: event III.A row {row_a} "
                f"berbeda dengan III.B row {row_b}."
            )


def locate_monthly_kri_columns(ws, month):
    month_name = normalize(MONTH_NAMES[month])
    # Cari cell header yang mengandung "Realisasi Threshold KRI <bulan>".
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, ws.max_column + 1):
            txt = normalize(ws.cell(r, c).value)
            if (
                "realisasi threshold kri" in txt
                and month_name in txt
                and "skor" not in txt
            ):
                return c - 1, c  # zero-based threshold; score ada di kolom berikutnya

    # Validasi eksplisit template AGA yang sudah direview.
    known = {
        5: (46, 47),  # AU/AV zero-based
        6: (48, 49),  # AW/AX
        7: (50, 51),  # AY/AZ
    }
    threshold_idx, score_idx = known[month]
    header_text = normalize(ws.cell(8, threshold_idx + 1).value)
    if month_name in header_text or "realisasi threshold kri" in header_text:
        return threshold_idx, score_idx

    raise RuntimeError(
        f"Kolom KRI bulan {MONTH_NAMES[month]} tidak dapat ditemukan pada III.B."
    )


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir", is_active=True).first()
        or User.objects.filter(is_superuser=True, is_active=True).order_by("id").first()
        or User.objects.filter(is_active=True).order_by("id").first()
    )


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(YEAR, month)
    period, _ = PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{YEAR}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {YEAR}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{YEAR}-{month:02d}-01",
            "tanggal_selesai": f"{YEAR}-{month:02d}-{last_day:02d}",
        },
    )
    return period


def get_or_create_report(profile, period, tahun_buku, prepared_by, month):
    qs = MonthlyRiskReport.objects.filter(
        reassessment=profile,
        periode=period,
        versi=1,
    ).order_by("id")

    count = qs.count()
    if count > 1:
        detail = "; ".join(
            f"id={r.pk}/kode={r.kode}/status={r.status}/items={r.items.count()}"
            for r in qs
        )
        raise RuntimeError(
            f"{MONTH_NAMES[month]} {YEAR}: ditemukan {count} report versi=1. "
            f"Import dibatalkan; review duplicate dulu. {detail}"
        )

    report = qs.first()
    created = False
    canonical_code = REPORT_CODES[month]

    if report is None:
        report = MonthlyRiskReport.objects.create(
            reassessment=profile,
            periode=period,
            tahun_buku=tahun_buku,
            versi=1,
            kode=canonical_code,
            status="draft",
            prepared_by=prepared_by,
        )
        created = True
    else:
        if report.status not in {"draft", "revision"}:
            raise RuntimeError(
                f"Report id={report.pk} {MONTH_NAMES[month]} status={report.status!r}; "
                "hanya Draft/Revision boleh diimpor."
            )
        if getattr(report, "is_locked", False):
            raise RuntimeError(
                f"Report id={report.pk} {MONTH_NAMES[month]} sedang locked."
            )

        # Normalisasi kode hanya bila tidak bentrok.
        if report.kode != canonical_code:
            collision = (
                MonthlyRiskReport.objects
                .exclude(pk=report.pk)
                .filter(kode=canonical_code)
                .exists()
            )
            if collision:
                raise RuntimeError(
                    f"Tidak dapat menormalisasi kode ke {canonical_code}: kode sudah dipakai."
                )
            report.kode = canonical_code

        if not report.prepared_by_id:
            report.prepared_by = prepared_by
        if not report.tahun_buku_id:
            report.tahun_buku = tahun_buku
        report.save()

    return report, created


def _field_changed(old, new):
    if hasattr(old, "pk"):
        old = old.pk
    if hasattr(new, "pk"):
        new = new.pk
    return old != new


def normalize_decimal_for_model_field(item, field_name, value):
    if value is None or not isinstance(value, Decimal):
        return value
    model_field = item._meta.get_field(field_name)
    decimal_places = getattr(model_field, "decimal_places", None)
    if decimal_places is None:
        return value
    quantum = Decimal("1").scaleb(-decimal_places)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def apply_fields(item, values, stats):
    changed_fields = []
    for field_name, new_value in values.items():
        new_value = normalize_decimal_for_model_field(item, field_name, new_value)
        old_value = getattr(item, field_name)
        if _field_changed(old_value, new_value):
            setattr(item, field_name, new_value)
            changed_fields.append(field_name)
            stats.field_changes += 1

    if changed_fields:
        # Sengaja tidak full_clean() global agar unrelated legacy fields/choices
        # tidak menghambat import source-owned monthly fields.
        item.save()
        stats.updated_items += 1
    return changed_fields


def validate_existing_report_items(report, master_map, month):
    target_ids = {x.pk for x in master_map.values()}
    existing = list(report.items.select_related("risk_event").all())

    foreign = [
        x for x in existing
        if x.risk_event_id not in target_ids
    ]
    if foreign:
        detail = "; ".join(
            f"item={x.pk}/RE={x.risk_event_id}/{x.risk_event.peristiwa_risiko}"
            for x in foreign
        )
        raise RuntimeError(
            f"{MONTH_NAMES[month]}: report existing memiliki item yang bukan bagian "
            f"dari master AGA final. Tidak akan dihapus otomatis: {detail}"
        )

    risk_ids = [x.risk_event_id for x in existing]
    if len(risk_ids) != len(set(risk_ids)):
        raise RuntimeError(
            f"{MONTH_NAMES[month]}: terdapat duplicate risk_event pada report existing."
        )

    if len(existing) > EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"{MONTH_NAMES[month]}: report existing memiliki {len(existing)} item > 14."
        )


def audit_master_against_iiib(master, row, n, stats):
    source_budget = decimal_or_none(value_at(row, IIIB_BUDGET_INDEX))
    master_budget = decimal_or_none(getattr(master, "biaya_perlakuan_risiko", None))
    if source_budget != master_budget:
        stats.master_budget_mismatches.append(
            f"R{n:02d} RE={master.pk}: budget master={master_budget} "
            f"| source={source_budget}"
        )

    checks = [
        ("KRI", getattr(master, "key_risk_indicators", None), value_at(row, IIIB_KRI_NAME_INDEX)),
        ("unit", getattr(master, "unit_satuan_kri", None), value_at(row, IIIB_KRI_UNIT_INDEX)),
        ("aman", getattr(master, "threshold_aman", None), value_at(row, IIIB_KRI_SAFE_INDEX)),
        ("hati", getattr(master, "threshold_hati_hati", None), value_at(row, IIIB_KRI_CAUTION_INDEX)),
        ("bahaya", getattr(master, "threshold_bahaya", None), value_at(row, IIIB_KRI_DANGER_INDEX)),
    ]
    for label, master_value, source_value in checks:
        if normalize(master_value) != normalize(source_value):
            stats.kri_master_mismatches.append(
                f"R{n:02d} {label}: master={text_or_none(master_value)!r} "
                f"| source={text_or_none(source_value)!r}"
            )


def import_iiia(report, master_map, parsed, month, stats):
    quarter = MONTH_QUARTER[month]

    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        master = master_map[n]
        row_no, row = parsed[n]

        item, created = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=master,
        )
        if created:
            stats.created_items += 1

        source_category = normalize(value_at(row, IIIA_CATEGORY_INDEX))
        if "kuant" in source_category:
            qualitative = False
        elif "kual" in source_category:
            qualitative = True
        else:
            qualitative = is_qualitative_master(master)

        master_qualitative = is_qualitative_master(master)
        if qualitative != master_qualitative:
            stats.category_mismatches.append(
                f"R{n:02d} RE={master.pk}: "
                f"master={getattr(getattr(master, 'kategori_dampak', None), 'nama', '-')} "
                f"| source={value_at(row, IIIA_CATEGORY_INDEX)!r}; "
                "monthly mengikuti source, master tidak diubah."
            )

        impact_raw = value_at(
            row, IIIA_QUARTER_COLUMNS["realisasi_nilai_dampak"][quarter]
        )
        impact_scale_raw = value_at(
            row, IIIA_QUARTER_COLUMNS["realisasi_skala_dampak"][quarter]
        )
        probability_raw = value_at(
            row, IIIA_QUARTER_COLUMNS["realisasi_nilai_probabilitas"][quarter]
        )
        probability_scale_raw = value_at(
            row, IIIA_QUARTER_COLUMNS["realisasi_skala_probabilitas"][quarter]
        )

        impact_value = None if qualitative else decimal_or_none(impact_raw)
        impact_scale = scale_by_level(
            MasterSkalaDampak,
            impact_scale_raw,
            f"{MONTH_NAMES[month]} R{n} skala dampak Q{quarter}",
        )
        probability_value = percent_or_none(probability_raw)
        probability_scale = scale_by_level(
            MasterSkalaProbabilitas,
            probability_scale_raw,
            f"{MONTH_NAMES[month]} R{n} skala probabilitas Q{quarter}",
        )

        required_snapshot = [
            ("skala_dampak", impact_scale),
            ("nilai_probabilitas", probability_value),
            ("skala_probabilitas", probability_scale),
        ]
        if not qualitative:
            required_snapshot.append(("nilai_dampak", impact_value))

        missing = [name for name, val in required_snapshot if val is None]
        if missing:
            stats.residual_incomplete.append(
                f"R{n:02d} III.A row {row_no} Q{quarter}: " + ", ".join(missing)
            )

        values = {
            "jenis_risiko": "kualitatif" if qualitative else "kuantitatif",
            "realisasi_asumsi_dampak": text_or_none(
                value_at(row, IIIA_ASSUMPTION_INDEX)
            ),
            "realisasi_nilai_dampak": impact_value,
            "realisasi_skala_dampak": impact_scale,
            "realisasi_skala_dampak_kbumn": int_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_skala_dampak_kbumn"][quarter],
                )
            ),
            "realisasi_nilai_probabilitas": probability_value,
            "realisasi_skala_probabilitas": probability_scale,
            "realisasi_skala_probabilitas_kbumn": int_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_skala_probabilitas_kbumn"][quarter],
                )
            ),
            "realisasi_eksposur": decimal_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_eksposur"][quarter],
                )
            ),
            "realisasi_skor_risiko": int_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_skor_risiko"][quarter],
                )
            ),
            "realisasi_skala_nilai_risiko_kbumn": int_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_skala_nilai_risiko_kbumn"][quarter],
                )
            ),
            "realisasi_level_risiko_bumn": text_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_level_risiko_bumn"][quarter],
                )
            ),
            "realisasi_level_risiko_kbumn": text_or_none(
                value_at(
                    row,
                    IIIA_QUARTER_COLUMNS["realisasi_level_risiko_kbumn"][quarter],
                )
            ),
            "efektivitas_perlakuan_risiko": treatment_effectiveness(
                value_at(row, IIIA_EFFECTIVENESS_INDEX)
            ),
        }

        apply_fields(item, values, stats)


def apply_kri_source_values(item, master, kri_status, kri_score_text, kri_numeric, stats, n, month):
    """
    Simpan threshold + skor persis dari workbook sebagai authoritative historical data.

    realisasi_nilai_kri hanya dicoba bila cell skor memang numerik murni dan model
    KRI dapat mengevaluasinya terhadap konfigurasi threshold aktif. Jika konfigurasi
    master tidak kompatibel dengan historical score/unit, nilai numerik TIDAK dipaksa;
    threshold dan skor teks tetap tersimpan, sehingga import tidak gagal dan tidak
    mengakali model.save().
    """
    # Tahap 1: kosongkan numeric lama dan simpan source threshold/score.
    source_values = {
        "realisasi_nilai_kri": None,
        "realisasi_threshold_kri": kri_status,
        "realisasi_threshold_kri_skor": kri_score_text,
    }
    apply_fields(item, source_values, stats)

    if kri_numeric is None:
        return

    # Tahap 2: coba simpan numeric melalui model.save() agar business rule tetap berlaku.
    try:
        with transaction.atomic():
            item.realisasi_nilai_kri = normalize_decimal_for_model_field(
                item, "realisasi_nilai_kri", kri_numeric
            )
            item.save(update_fields=["realisasi_nilai_kri"])
    except ValidationError as exc:
        # Nested atomic rollback hanya membatalkan attempt numeric. Source threshold/score
        # dari tahap 1 tetap ada pada outer transaction. Refresh object supaya state bersih.
        item.refresh_from_db(
            fields=[
                "realisasi_nilai_kri",
                "realisasi_threshold_kri",
                "realisasi_threshold_kri_skor",
            ]
        )
        stats.kri_numeric_skipped.append(
            f"R{n:02d} {MONTH_NAMES[month]}: score={kri_score_text!r} "
            f"tidak kompatibel dengan threshold aktif; numeric tidak disimpan; "
            f"source threshold={kri_status!r} tetap dipertahankan. Error={'; '.join(exc.messages)}"
        )
    else:
        stats.kri_numeric_stored.append(
            f"R{n:02d} {MONTH_NAMES[month]}: numeric={kri_numeric}"
        )
        # Model.save() dapat mengubah atribut threshold di memory; DB source fields
        # tetap authoritative karena update_fields hanya numeric. Sinkronkan object.
        item.refresh_from_db(
            fields=[
                "realisasi_nilai_kri",
                "realisasi_threshold_kri",
                "realisasi_threshold_kri_skor",
            ]
        )


def import_iiib(report, master_map, parsed, workbook, month, stats):
    ws = workbook["III.B"]
    quarter = MONTH_QUARTER[month]
    kri_threshold_idx, kri_score_idx = locate_monthly_kri_columns(ws, month)
    current_timeline_idx = IIIB_TIMELINE_FIRST + (month - 1)

    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        master = master_map[n]
        row_no, row, cause_letter = parsed[n]

        item, created = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=master,
        )
        if created:
            stats.created_items += 1

        audit_master_against_iiib(master, row, n, stats)

        timeline_raw = value_at(row, current_timeline_idx)
        if decimal_or_none(timeline_raw) not in (None, Decimal("0")):
            stats.current_month_timeline_on.append(f"R{n:02d}")
        else:
            stats.current_month_timeline_off.append(f"R{n:02d}")

        actual_values = {
            "realisasi_rencana_perlakuan": text_or_none(
                value_at(row, IIIB_ACTUAL_TREATMENT_INDEX)
            ),
            "realisasi_output_perlakuan": text_or_none(
                value_at(row, IIIB_ACTUAL_OUTPUT_INDEX)
            ),
            "realisasi_biaya_perlakuan": decimal_or_none(
                value_at(row, IIIB_ACTUAL_COST_INDEX)
            ),
            "realisasi_pic": text_or_none(
                value_at(row, IIIB_ACTUAL_PIC_INDEX)
            ),
            "status_rencana_perlakuan": treatment_status(
                value_at(row, IIIB_STATUS_INDEX)
            ),
            "penjelasan_status_rencana": text_or_none(
                value_at(row, IIIB_STATUS_NOTE_INDEX)
            ),
            "progress_pelaksanaan_percent": percent_or_none(
                value_at(row, IIIB_PROGRESS_BY_QUARTER[quarter])
            ),
        }
        apply_fields(item, actual_values, stats)

        kri_status = text_or_none(value_at(row, kri_threshold_idx))
        kri_score_raw = value_at(row, kri_score_idx)
        kri_score_text = text_or_none(kri_score_raw)
        kri_numeric = kri_numeric_or_none(kri_score_raw)

        if kri_status is None and kri_score_text is None:
            stats.kri_blank.append(
                f"R{n:02d} III.B row {row_no}"
            )
        else:
            stats.kri_with_data += 1

        # Threshold dan skor workbook adalah source-of-truth historical. Numeric hanya
        # disimpan bila lolos evaluator threshold model; tidak ada bypass model.save().
        apply_kri_source_values(
            item=item,
            master=master,
            kri_status=kri_status,
            kri_score_text=kri_score_text,
            kri_numeric=kri_numeric,
            stats=stats,
            n=n,
            month=month,
        )


def parse_source_changes(workbook):
    if "III.D" not in workbook.sheetnames:
        return []
    ws = workbook["III.D"]
    rows = []
    started = False
    for cells in ws.iter_rows(values_only=True):
        row = list(cells)
        if normalize(value_at(row, 0)) == "start pengisian":
            started = True
            # Data bisa ada pada baris yang sama mulai kolom B.
        if not started:
            continue
        change_type = text_or_none(value_at(row, 1))
        if not change_type:
            continue
        rows.append(row)
    return rows


def import_changes(workbook, report, stats):
    rows = parse_source_changes(workbook)
    stats.source_change_rows = len(rows)

    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }

    parsed = []
    for row in rows:
        raw_type = normalize(value_at(row, 1))
        change_type = mapping.get(raw_type)
        if change_type is None:
            raise RuntimeError(
                f"III.D: jenis perubahan tidak dikenali: {value_at(row, 1)!r}"
            )
        parsed.append(
            (
                change_type,
                text_or_none(value_at(row, 2)),
                text_or_none(value_at(row, 3)),
            )
        )

    MonthlyRiskReportChange.objects.filter(report=report).delete()
    for change_type, event, explanation in parsed:
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=event,
            penjelasan=explanation,
        )


def parse_source_loss_events(workbook, stats):
    if "III.E" not in workbook.sheetnames:
        return []

    ws = workbook["III.E"]
    valid = []
    started = False

    for cells in ws.iter_rows(values_only=True):
        row = list(cells)
        if normalize(value_at(row, 0)) == "start pengisian":
            started = True
        if not started:
            continue

        name = text_or_none(value_at(row, 1))
        if not name:
            continue

        norm_name = normalize(name)
        if norm_name in LEGACY_LOSS_EVENT_NAMES:
            stats.source_legacy_loss_rows += 1
            continue

        valid.append(row)

    stats.source_valid_loss_rows = len(valid)
    return valid


def import_loss_events(workbook, report, stats):
    valid_rows = parse_source_loss_events(workbook, stats)

    existing = list(
        MonthlyRiskReportLossEvent.objects
        .filter(report=report)
        .order_by("id")
    )

    if not valid_rows:
        unknown_existing = [
            x for x in existing
            if normalize(x.nama_kejadian) not in LEGACY_LOSS_EVENT_NAMES
        ]
        if unknown_existing:
            detail = "; ".join(
                f"id={x.pk}/{x.nama_kejadian}" for x in unknown_existing
            )
            raise RuntimeError(
                "Source III.E tidak mempunyai loss event valid untuk bulan ini, "
                "tetapi DB mempunyai loss event non-legacy. Tidak akan dihapus otomatis: "
                + detail
            )

        legacy_existing = [
            x for x in existing
            if normalize(x.nama_kejadian) in LEGACY_LOSS_EVENT_NAMES
        ]
        stats.deleted_legacy_db_loss_rows = len(legacy_existing)
        if legacy_existing:
            MonthlyRiskReportLossEvent.objects.filter(
                pk__in=[x.pk for x in legacy_existing]
            ).delete()
        return

    # Current source files are expected to have 0 valid loss event. Generic path
    # below exists for safety if a later exact source unexpectedly contains valid rows.
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()

    for row in valid_rows:
        src = normalize(value_at(row, 4))
        repeat = normalize(value_at(row, 12))
        insured = normalize(value_at(row, 18))

        MonthlyRiskReportLossEvent.objects.create(
            report=report,
            nama_kejadian=text_or_none(value_at(row, 1)),
            identifikasi_kejadian=text_or_none(value_at(row, 2)),
            kategori_kejadian=text_or_none(value_at(row, 3)),
            sumber_penyebab_kejadian=(
                "external" if "eksternal" in src
                else "internal" if "internal" in src
                else None
            ),
            penyebab_kejadian=text_or_none(value_at(row, 5)),
            penanganan_saat_kejadian=text_or_none(value_at(row, 6)),
            deskripsi_kejadian_risk_event=text_or_none(value_at(row, 7)),
            kategori_risiko_bumn=text_or_none(value_at(row, 8)),
            kategori_risiko_t2_t3_kbumn=text_or_none(value_at(row, 9)),
            penjelasan_kerugian=text_or_none(value_at(row, 10)),
            nilai_kerugian=decimal_or_none(value_at(row, 11)),
            kejadian_berulang=(
                "ya" if "ya" in repeat
                else "tidak" if "tidak" in repeat
                else None
            ),
            frekuensi_kejadian=text_or_none(value_at(row, 13)),
            mitigasi_direncanakan=text_or_none(value_at(row, 14)),
            realisasi_mitigasi=text_or_none(value_at(row, 15)),
            perbaikan_mendatang=text_or_none(value_at(row, 16)),
            pihak_terkait=text_or_none(value_at(row, 17)),
            status_asuransi=(
                "ya" if "ya" in insured
                else "tidak" if "tidak" in insured
                else None
            ),
            nilai_premi=decimal_or_none(value_at(row, 19)),
            nilai_klaim=decimal_or_none(value_at(row, 20)),
        )


def backup_sqlite():
    engine = settings.DATABASES["default"].get("ENGINE", "")
    if "sqlite" not in engine:
        raise RuntimeError(
            f"APPLY dibatalkan: database aktif bukan SQLite ({engine}). "
            "Siapkan backup DB manual lalu sesuaikan script bila memang diperlukan."
        )

    db_path = Path(str(settings.DATABASES["default"]["NAME"])).resolve()
    if not db_path.exists():
        raise RuntimeError(f"SQLite DB tidak ditemukan: {db_path}")

    backup_dir = PROJECT_ROOT / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = (
        backup_dir
        / f"db_before_aga_may_jun_jul_2026_{datetime.now():%Y%m%d_%H%M%S}.sqlite3"
    )

    src = sqlite3.connect(str(db_path), timeout=30)
    dst = sqlite3.connect(str(backup_path), timeout=30)
    try:
        src.backup(dst)
        result = dst.execute("PRAGMA quick_check").fetchone()[0]
        if result != "ok":
            raise RuntimeError(
                f"Backup SQLite quick_check gagal: {result}"
            )
    finally:
        dst.close()
        src.close()

    return backup_path


def open_source(spec):
    actual_sha = file_sha256(spec.path)
    expected_sha = EXPECTED_SHA256[spec.month]
    if actual_sha != expected_sha:
        raise RuntimeError(
            f"SHA256 {MONTH_NAMES[spec.month]} tidak sama dengan file yang direview.\n"
            f"Expected={expected_sha}\nActual  ={actual_sha}\nFile={spec.path}"
        )

    wb = load_workbook(
        spec.path,
        data_only=True,
        read_only=True,
        keep_links=False,
    )
    required = {"III.A", "III.B", "III.D", "III.E"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise RuntimeError(
            f"{spec.path.name}: sheet wajib tidak ada: {sorted(missing)}"
        )
    return wb, actual_sha


def process_month(spec, profile, master_map, tahun_buku, prepared_by):
    stats = MonthStats(month=spec.month)
    wb, source_sha = open_source(spec)

    print(
        f"\nSOURCE {MONTH_NAMES[spec.month]}: {spec.path}"
        f"\n- SHA256: {source_sha}"
        f"\n- Residual: Q{MONTH_QUARTER[spec.month]}"
    )

    parsed_a = parse_iiia(wb, spec.month, master_map, stats)
    parsed_b = parse_iiib(wb, spec.month, master_map, stats)
    validate_iiia_vs_iiib(parsed_a, parsed_b, spec.month)

    period = get_period(tahun_buku, spec.month)
    report, created = get_or_create_report(
        profile, period, tahun_buku, prepared_by, spec.month
    )
    stats.report_id = report.pk
    stats.report_created = created

    validate_existing_report_items(report, master_map, spec.month)

    import_iiia(report, master_map, parsed_a, spec.month, stats)
    import_iiib(report, master_map, parsed_b, wb, spec.month, stats)
    import_changes(wb, report, stats)
    import_loss_events(wb, report, stats)

    refresh_monthly_report_summary(report)
    report.refresh_from_db()

    stats.report_item_count = report.items.count()
    stats.distinct_risk_count = (
        report.items.values("risk_event_id").distinct().count()
    )

    target_ids = {x.pk for x in master_map.values()}
    actual_ids = set(report.items.values_list("risk_event_id", flat=True))

    if stats.report_item_count != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"{MONTH_NAMES[spec.month]} hasil import item={stats.report_item_count}, "
            f"expected 14."
        )
    if stats.distinct_risk_count != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"{MONTH_NAMES[spec.month]} distinct risk_event={stats.distinct_risk_count}, "
            f"expected 14."
        )
    if actual_ids != target_ids:
        raise RuntimeError(
            f"{MONTH_NAMES[spec.month]} risk_event set tidak sama dengan profile final."
        )
    if report.total_risiko != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"{MONTH_NAMES[spec.month]} total_risiko={report.total_risiko}, expected 14."
        )

    return report, stats


def print_stats(report, stats):
    print("\n" + "=" * 108)
    print(
        f"{MONTH_NAMES[stats.month].upper()} {YEAR} | "
        f"REPORT ID={report.pk} | {report.kode} | status={report.status}"
    )
    print("=" * 108)
    print(
        f"- Report: {'BARU' if stats.report_created else 'EXISTING'} "
        f"| items={stats.report_item_count} "
        f"| distinct risk={stats.distinct_risk_count} "
        f"| total_risiko={report.total_risiko}"
    )
    print(f"- Source III.A: {stats.source_iiia}/14")
    print(f"- Source III.B: {stats.source_iiib}/14")
    print(f"- Residual source: Q{MONTH_QUARTER[stats.month]}")
    print(
        f"- MonthlyRiskReportItem dibuat: {stats.created_items}"
        f" | item tersentuh: {stats.updated_items}"
        f" | perubahan field: {stats.field_changes}"
    )
    print(
        f"- KRI bulan berjalan: ada data={stats.kri_with_data}"
        f" | kosong={len(stats.kri_blank)}"
    )
    if stats.kri_blank:
        print("  KRI kosong: " + ", ".join(stats.kri_blank))
    print(
        f"- KRI numeric tersimpan via evaluator: {len(stats.kri_numeric_stored)}"
        f" | numeric tidak dipaksa: {len(stats.kri_numeric_skipped)}"
    )
    if stats.kri_numeric_stored:
        for msg in stats.kri_numeric_stored:
            print("  +", msg)
    if stats.kri_numeric_skipped:
        print("- WARNING KRI numeric tidak kompatibel dengan threshold aktif:")
        for msg in stats.kri_numeric_skipped:
            print("  *", msg)

    print(
        f"- Timeline bulan berjalan = 1: {len(stats.current_month_timeline_on)}"
        f" | kosong/0: {len(stats.current_month_timeline_off)}"
    )
    if stats.current_month_timeline_off:
        print("  Timeline kosong/0: " + ", ".join(stats.current_month_timeline_off))

    print(f"- III.D perubahan: {stats.source_change_rows}")
    print(
        f"- III.E loss event valid: {stats.source_valid_loss_rows}"
        f" | legacy/template diabaikan: {stats.source_legacy_loss_rows}"
        f" | legacy DB dibersihkan: {stats.deleted_legacy_db_loss_rows}"
    )

    if stats.residual_incomplete:
        print("- Residual incomplete (dipertahankan kosong, TANPA fallback):")
        for msg in stats.residual_incomplete:
            print("  *", msg)

    if stats.category_mismatches:
        print("- Perbedaan kategori master vs source (master TIDAK diubah):")
        for msg in stats.category_mismatches:
            print("  *", msg)

    if stats.master_budget_mismatches:
        print("- AUDIT budget master vs source III.B (TIDAK diubah oleh importer):")
        for msg in stats.master_budget_mismatches:
            print("  *", msg)
    else:
        print("- AUDIT budget master vs source III.B: sama.")

    if stats.kri_master_mismatches:
        print("- AUDIT KRI master vs source III.B (TIDAK diubah oleh importer):")
        for msg in stats.kri_master_mismatches:
            print("  *", msg)
    else:
        print("- AUDIT KRI master vs source III.B: sama.")


def main():
    parser = argparse.ArgumentParser(
        description="Import laporan BID AGA Mei, Juni, Juli 2026 secara aman."
    )
    parser.add_argument("--may", required=True, type=Path)
    parser.add_argument("--june", required=True, type=Path)
    parser.add_argument("--july", required=True, type=Path)
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Commit perubahan. Default DRY RUN / rollback.",
    )
    args = parser.parse_args()

    specs = [
        ImportSpec(5, args.may.expanduser().resolve()),
        ImportSpec(6, args.june.expanduser().resolve()),
        ImportSpec(7, args.july.expanduser().resolve()),
    ]
    for spec in specs:
        if not spec.path.exists():
            raise FileNotFoundError(spec.path)

    mode = "APPLY" if args.apply else "DRY RUN"
    print("=" * 108)
    print(f"IMPORT BID AGA MEI–JULI {YEAR} | MODE={mode}")
    print("=" * 108)

    profile = get_profile()
    master_map = build_master_map(profile)

    prepared_by = get_prepared_by()
    if prepared_by is None:
        raise RuntimeError("Tidak ada user aktif untuk prepared_by.")

    print(
        f"\nPROFILE: id={profile.pk} | {profile.judul} "
        f"| unit={profile.unit_bisnis} | master={profile.item.count()}"
    )
    print(
        f"KM: id={getattr(profile, 'kontrak_manajemen_id', None)} "
        f"| {getattr(profile, 'kontrak_manajemen', None)}"
    )
    print(f"PREPARED BY: id={prepared_by.pk} | {prepared_by.get_username()}")

    if args.apply:
        backup = backup_sqlite()
        print("\nBACKUP DB:", backup)

    results = []

    with transaction.atomic():
        tahun_buku, _ = TahunBuku.objects.get_or_create(
            tahun=YEAR,
            defaults={"aktif": True},
        )

        # Lock profile on APPLY so master relationship cannot change mid-import.
        if args.apply:
            locked_profile = (
                ReAssessmentSummary.objects
                .select_for_update()
                .get(pk=profile.pk)
            )
            if locked_profile.item.count() != EXPECTED_SOURCE_ITEMS:
                raise RuntimeError(
                    "Profile berubah saat lock; item bukan lagi 14. APPLY dibatalkan."
                )

        for spec in specs:
            report, stats = process_month(
                spec, profile, master_map, tahun_buku, prepared_by
            )
            results.append((report, stats))

        if not args.apply:
            transaction.set_rollback(True)

    for report, stats in results:
        print_stats(report, stats)

    print("\n" + "=" * 108)
    if args.apply:
        print("RESULT: APPLY BERHASIL — Mei, Juni, Juli 2026 sudah di-commit dalam satu transaksi.")
    else:
        print("RESULT: DRY RUN BERHASIL — database TIDAK berubah (rollback).")
        print("Review seluruh output di atas. Jika bersih, ulangi command yang sama dengan --apply.")
    print("=" * 108)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        raise
