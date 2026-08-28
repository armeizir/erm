#!/usr/bin/env python3
"""
Import laporan bulanan BID HCGA Juni dan Juli 2026 ke modul monthly_report.

Karakteristik:
- standalone: TIDAK bergantung pada import_hcga_feb_may_2026_v2.py;
- default DRY RUN (transaction di-rollback);
- --apply untuk commit;
- matching target berbasis peristiwa risiko/master HCGA, bukan nomor source;
- menangani 3 penyebab untuk event "Ketidaksiapan organisasi dan SDM...";
- dampak numerik risiko kualitatif dibiarkan kosong, probabilitas tetap dibaca;
- Q2/Q3 dibaca sesuai bulan, tanpa fallback quarter sebelumnya;
- KRI dibaca dari kolom bulan yang berlabel Juni/Juli, tanpa fallback bulan lain.

Contoh:
  python monthly_report/scripts/import_hcga_jun_jul_2026.py \
    --june /tmp/hcga_juni.xlsx \
    --july /tmp/hcga_juli.xlsx

Apply setelah dry-run bersih:
  python monthly_report/scripts/import_hcga_jun_jul_2026.py \
    --june /tmp/hcga_juni.xlsx \
    --july /tmp/hcga_juli.xlsx \
    --apply
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django  # noqa: E402

django.setup()

from django.conf import settings  # noqa: E402
from django.contrib.auth import get_user_model  # noqa: E402
from django.core.exceptions import ValidationError  # noqa: E402
from django.db import transaction  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from masterdata.models import PeriodeLaporan, TahunBuku  # noqa: E402
from monthly_report.models import (  # noqa: E402
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from monthly_report.services import refresh_monthly_report_summary  # noqa: E402
from risk.models import (  # noqa: E402
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)


YEAR = 2026
EXPECTED_SOURCE_ITEMS = 15
MONTH_NAMES = {6: "Juni", 7: "Juli"}

# Official III.A, zero-based column indexes.
IIIA_QUARTER_COLUMNS = {
    "realisasi_nilai_dampak": {1: 14, 2: 15, 3: 16, 4: 17},
    "realisasi_skala_dampak": {1: 18, 2: 19, 3: 20, 4: 21},
    "realisasi_skala_dampak_kbumn": {1: 22, 2: 23, 3: 24, 4: 25},
    "realisasi_nilai_probabilitas": {1: 26, 2: 27, 3: 28, 4: 29},
    "realisasi_skala_probabilitas": {1: 30, 2: 31, 3: 32, 4: 33},
    "realisasi_skala_probabilitas_kbumn": {1: 34, 2: 35, 3: 36, 4: 37},
    "realisasi_eksposur": {1: 38, 2: 39, 3: 40, 4: 41},
    "realisasi_skor_risiko": {1: 42, 2: 43, 3: 44, 4: 45},
    "realisasi_skala_nilai_risiko_kbumn": {1: 46, 2: 47, 3: 48, 4: 49},
    "realisasi_level_risiko_bumn": {1: 50, 2: 51, 3: 52, 4: 53},
    "realisasi_level_risiko_kbumn": {1: 54, 2: 55, 3: 56, 4: 57},
}
IIIA_EFFECTIVENESS_INDEX = 58  # BG
IIIA_ASSUMPTION_INDEX = 13     # N

# Official III.B, zero-based fixed indexes up to KRI header area.
IIIB_ACTUAL_TREATMENT_INDEX = 10   # K
IIIB_ACTUAL_OUTPUT_INDEX = 11      # L
IIIB_ACTUAL_COST_INDEX = 12        # M
IIIB_ACTUAL_PIC_INDEX = 14         # O
IIIB_STATUS_INDEX = 27             # AB
IIIB_STATUS_NOTE_INDEX = 28        # AC
IIIB_PROGRESS_BY_QUARTER = {1: 29, 2: 30, 3: 31, 4: 32}  # AD:AG


@dataclass(frozen=True)
class ImportSpec:
    month: int
    path: Path


@dataclass
class MonthStats:
    month: int
    source_iiia: int = 0
    source_iiib: int = 0
    created_items: int = 0
    updated_items: int = 0
    field_changes: int = 0
    kri_rows_with_current_month_data: int = 0
    kri_rows_blank_current_month: list[str] = field(default_factory=list)
    empty_quarter_fields: list[str] = field(default_factory=list)
    skipped_rows: list[str] = field(default_factory=list)
    changes_count: int = 0
    loss_events_count: int = 0
    report_item_count: int = 0


def normalize(value) -> str:
    text = str(value or "").casefold().replace("\xa0", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def text_or_none(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        stripped = value.strip().casefold()
        if stripped in {"-", "n/a", "no data", "#div/0!", "#n/a", "#name?"}:
            return None
        multiplier = Decimal("1")
        if re.search(r"miliar|milyar", stripped):
            multiplier = Decimal("1000000000")
        elif "juta" in stripped:
            multiplier = Decimal("1000000")
        cleaned = re.sub(r"[^0-9,.\-]", "", stripped)
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        if not cleaned:
            return None
        value = cleaned
    else:
        multiplier = Decimal("1")
    try:
        return Decimal(str(value).strip()) * multiplier
    except (InvalidOperation, ValueError, TypeError):
        return None


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    if Decimal("0") <= number <= Decimal("100"):
        return number
    return None


def int_or_none(value):
    number = decimal_or_none(value)
    return int(number) if number is not None else None


def value_at(row, zero_index):
    return row[zero_index] if len(row) > zero_index else None


def file_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def quarter_for_month(month):
    return ((month - 1) // 3) + 1


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.filter(is_active=True).order_by("id").first()
    )


def get_hcga_profile():
    candidates = list(
        ReAssessmentSummary.objects.filter(tahun=YEAR)
        .select_related("unit_bisnis")
        .order_by("id")
    )
    matched = [
        item
        for item in candidates
        if "hcga" in normalize(item.judul)
        or (item.unit_bisnis_id and "hcga" in normalize(item.unit_bisnis.name))
    ]
    if len(matched) != 1:
        raise RuntimeError(
            f"Profil HCGA tahun {YEAR} harus tepat satu; ditemukan {len(matched)}: "
            + ", ".join(f"id={x.id} {x.judul}" for x in matched)
        )
    return matched[0]


def is_qualitative_master(item: ReAssessmentItem) -> bool:
    category = normalize(getattr(getattr(item, "kategori_dampak", None), "nama", ""))
    return "kual" in category


def _contains_all(text, *parts):
    return all(normalize(part) in text for part in parts)


def build_master_map(profile):
    items = list(
        ReAssessmentItem.objects.filter(summary=profile)
        .select_related("kategori_dampak")
        .order_by("no_item", "no_risiko", "id")
    )

    def unique(key, predicate):
        found = [x for x in items if predicate(normalize(x.peristiwa_risiko), normalize(x.key_risk_indicators))]
        if len(found) != 1:
            raise RuntimeError(
                f"Mapping master '{key}' harus tepat satu, ditemukan {len(found)}: "
                + ", ".join(
                    f"RE={x.id}/R{x.no_risiko}/{x.no_penyebab_risiko or '-'} {x.peristiwa_risiko}"
                    for x in found
                )
            )
        return found[0]

    mapping = {
        "r1": unique("r1", lambda e, k: _contains_all(e, "biaya pemeliharaan", "korektif")),
        "r2": unique("r2", lambda e, k: _contains_all(e, "pendapatan", "non pln group")),
        "r3": unique("r3", lambda e, k: _contains_all(e, "digital", "general affair")),
        "r4": unique("r4", lambda e, k: _contains_all(e, "pengalihan hak kepemilikan", "legalitas")),
        "r5": unique("r5", lambda e, k: _contains_all(e, "keterlambatan proses pengadaan", "jasa konstruksi")),
        "r6": unique("r6", lambda e, k: "formasi tenaga kerja" in e),
        "r7": unique("r7", lambda e, k: "karya inovasi" in e),
        "r8": unique("r8", lambda e, k: "sertifikasi" in e and "pegawai" in e),
        "r9": unique("r9", lambda e, k: "rencana perlakuan risiko tidak terlaksana" in e),
        "wellness": unique("wellness", lambda e, k: "wellness" in e),
        "provider": unique("provider", lambda e, k: "provider" in e and "kesehatan" in e),
        "yanhc": unique("yanhc", lambda e, k: "yanhc" in e),
    }

    org_candidates = [
        x for x in items
        if _contains_all(normalize(x.peristiwa_risiko), "ketidaksiapan", "organisasi", "sdm")
    ]
    if len(org_candidates) < 3:
        raise RuntimeError(
            "Master HCGA harus memiliki minimal 3 item untuk event Ketidaksiapan organisasi dan SDM; "
            f"ditemukan {len(org_candidates)}."
        )

    def resolve_org(letter, kri_keyword):
        by_cause = [x for x in org_candidates if normalize(x.no_penyebab_risiko) == letter]
        if len(by_cause) == 1:
            return by_cause[0]
        by_kri = [x for x in org_candidates if kri_keyword in normalize(x.key_risk_indicators)]
        if len(by_kri) == 1:
            return by_kri[0]
        raise RuntimeError(
            f"Tidak dapat menentukan org_{letter}; kandidat="
            + ", ".join(f"RE={x.id}/R{x.no_risiko}/{x.no_penyebab_risiko}" for x in org_candidates)
        )

    mapping["org_a"] = resolve_org("a", "ees")
    mapping["org_b"] = resolve_org("b", "talent diversity")
    mapping["org_c"] = resolve_org("c", "pendapatan")

    if len({item.pk for item in mapping.values()}) != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError("Mapping master HCGA tidak menghasilkan 15 target unik.")
    return mapping


def source_key_from_event(event):
    e = normalize(event)
    if not e:
        return None
    rules = [
        ("r1", ("biaya pemeliharaan", "korektif")),
        ("r2", ("pendapatan", "non pln group")),
        ("r3", ("digital", "general affair")),
        ("r4", ("pengalihan hak kepemilikan", "legalitas")),
        ("r5", ("keterlambatan proses pengadaan", "jasa konstruksi")),
        ("r6", ("formasi tenaga kerja",)),
        ("r7", ("karya inovasi",)),
        ("r8", ("sertifikasi", "pegawai")),
        ("r9", ("rencana perlakuan risiko tidak terlaksana",)),
        ("wellness", ("wellness",)),
        ("provider", ("provider", "kesehatan")),
        ("yanhc", ("yanhc",)),
        ("org_a", ("ketidaksiapan", "organisasi", "sdm")),
    ]
    for key, parts in rules:
        if all(part in e for part in parts):
            return key
    return None


def find_start_row(ws):
    for row in range(1, min(ws.max_row, 80) + 1):
        if "start pengisian" in normalize(ws.cell(row, 1).value):
            return row
    raise RuntimeError(f"Anchor 'Start pengisian' tidak ditemukan pada sheet {ws.title}.")


def locate_monthly_kri_columns(ws, month):
    month_name = MONTH_NAMES[month]
    needle = normalize(f"Realisasi Threshold KRI {month_name}")
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, ws.max_column + 1):
            if needle == normalize(ws.cell(row, col).value):
                return col - 1, col  # zero-based threshold, zero-based score (next col)
    raise RuntimeError(
        f"Kolom 'Realisasi Threshold KRI {month_name}' tidak ditemukan pada III.B."
    )


def scale_by_level(model, value):
    number = int_or_none(value)
    if number is None:
        return None
    return model.objects.filter(urutan=number).first()


def treatment_status(value):
    txt = normalize(value)
    if not txt:
        return None
    if "discontinue" in txt:
        return "discontinue"
    if "continue" in txt:
        return "continue"
    # model saat ini belum punya pilihan "revisi"; jangan memaksa nilai ilegal.
    return None


def treatment_effectiveness(value):
    txt = normalize(value)
    if not txt:
        return None
    if "tidak efektif" in txt:
        return "tidak_efektif"
    if "efektif" in txt:
        return "efektif"
    return None


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(YEAR, month)
    period, _ = PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{YEAR}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {YEAR}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{YEAR}-{month:02d}-01",
            "tanggal_selesai": f"{YEAR}-{month:02d}-{last_day:02d}",
        },
    )
    return period


def get_or_create_report(profile, period, tahun_buku, prepared_by, month):
    report = MonthlyRiskReport.objects.filter(
        reassessment=profile,
        periode=period,
        versi=1,
    ).first()
    created = False
    if report is None:
        report = MonthlyRiskReport.objects.create(
            reassessment=profile,
            periode=period,
            tahun_buku=tahun_buku,
            versi=1,
            kode=f"MRR-HCGA-{YEAR}-{month:02d}",
            status="draft",
            prepared_by=prepared_by,
        )
        created = True
    else:
        if report.status not in {"draft", "revision"}:
            raise RuntimeError(
                f"{report.kode} berstatus {report.status}; hanya Draft/Revision yang boleh diimpor."
            )
        if getattr(report, "is_locked", False):
            raise RuntimeError(f"{report.kode} sedang locked; import dibatalkan.")
        fields = []
        if not report.prepared_by_id:
            report.prepared_by = prepared_by
            fields.append("prepared_by")
        if not report.tahun_buku_id:
            report.tahun_buku = tahun_buku
            fields.append("tahun_buku")
        if fields:
            report.save(update_fields=fields + ["updated_at"])
    return report, created


def _field_changed(old, new):
    if hasattr(old, "pk"):
        old = old.pk
    if hasattr(new, "pk"):
        new = new.pk
    return old != new


def _normalize_decimal_for_model_field(item, field_name, value):
    """Round Decimal values to the precision declared by the Django model field."""
    if value is None or not isinstance(value, Decimal):
        return value
    model_field = item._meta.get_field(field_name)
    decimal_places = getattr(model_field, "decimal_places", None)
    if decimal_places is None:
        return value
    quantum = Decimal("1").scaleb(-decimal_places)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def apply_fields(item, values, stats: MonthStats, *, overwrite_none_fields=()):
    changed_fields = []
    for field_name, new_value in values.items():
        new_value = _normalize_decimal_for_model_field(item, field_name, new_value)
        if new_value is None and field_name not in overwrite_none_fields:
            continue
        old_value = getattr(item, field_name)
        if _field_changed(old_value, new_value):
            setattr(item, field_name, new_value)
            changed_fields.append(field_name)
            stats.field_changes += 1
    if changed_fields:
        item.full_clean()
        item.save(update_fields=changed_fields + ["updated_at"])
        stats.updated_items += 1
    return changed_fields


def parse_iiia_rows(workbook, month, stats):
    ws = workbook["III.A"]
    quarter = quarter_for_month(month)
    start = find_start_row(ws)
    parsed = {}
    org_continuation = 0

    relevant_indexes = [
        IIIA_QUARTER_COLUMNS[name][quarter]
        for name in IIIA_QUARTER_COLUMNS
    ]
    for row_no, cells in enumerate(
        ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True), start=start
    ):
        row = list(cells)
        event = value_at(row, 2)  # C
        key = source_key_from_event(event)
        if key == "org_a":
            org_continuation = 2
        elif key is None and org_continuation:
            has_quarter_data = any(value_at(row, idx) not in (None, "") for idx in relevant_indexes)
            if has_quarter_data:
                key = "org_b" if org_continuation == 2 else "org_c"
                org_continuation -= 1
            else:
                org_continuation = 0
        elif event not in (None, ""):
            org_continuation = 0

        if not key:
            continue
        if key in parsed:
            raise RuntimeError(f"III.A {MONTH_NAMES[month]}: key {key} duplikat (row {row_no}).")
        parsed[key] = (row_no, row)

    stats.source_iiia = len(parsed)
    if len(parsed) != EXPECTED_SOURCE_ITEMS:
        missing = sorted(set(expected_keys()) - set(parsed))
        raise RuntimeError(
            f"III.A {MONTH_NAMES[month]} harus menghasilkan {EXPECTED_SOURCE_ITEMS} item; "
            f"ditemukan {len(parsed)}. Missing={missing}."
        )
    return parsed


def parse_iiib_rows(workbook, month, stats):
    ws = workbook["III.B"]
    start = find_start_row(ws)
    parsed = {}
    for row_no, cells in enumerate(
        ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True), start=start
    ):
        row = list(cells)
        cause_code = normalize(value_at(row, 5))  # F
        if "hcga" not in cause_code:
            continue
        event = value_at(row, 2)  # C
        key = source_key_from_event(event)
        if key == "org_a":
            match = re.search(r"hcga\s*(\d+)\s*([abc])?", cause_code)
            letter = match.group(2) if match else "a"
            key = f"org_{letter or 'a'}"
        elif key is None:
            match = re.search(r"hcga\s*(\d+)\s*([abc])?", cause_code)
            if match and match.group(1) == "13" and match.group(2) in {"a", "b", "c"}:
                key = f"org_{match.group(2)}"
        if not key:
            stats.skipped_rows.append(f"III.B row {row_no}: tidak dikenali ({cause_code})")
            continue
        if key in parsed:
            raise RuntimeError(f"III.B {MONTH_NAMES[month]}: key {key} duplikat (row {row_no}).")
        parsed[key] = (row_no, row)

    stats.source_iiib = len(parsed)
    if len(parsed) != EXPECTED_SOURCE_ITEMS:
        missing = sorted(set(expected_keys()) - set(parsed))
        raise RuntimeError(
            f"III.B {MONTH_NAMES[month]} harus menghasilkan {EXPECTED_SOURCE_ITEMS} item; "
            f"ditemukan {len(parsed)}. Missing={missing}."
        )
    return parsed


def expected_keys():
    return [
        "r1", "r2", "r3", "r4", "r5", "r6", "r7", "r8", "r9",
        "wellness", "provider", "yanhc", "org_a", "org_b", "org_c",
    ]


def import_iiia(report, master_map, parsed, month, stats):
    quarter = quarter_for_month(month)
    overwrite_none = {
        "realisasi_nilai_dampak",
        "realisasi_skala_dampak",
        "realisasi_nilai_probabilitas",
        "realisasi_skala_probabilitas",
        "realisasi_skala_dampak_kbumn",
        "realisasi_skala_probabilitas_kbumn",
        "realisasi_eksposur",
        "realisasi_skor_risiko",
        "realisasi_skala_nilai_risiko_kbumn",
        "realisasi_level_risiko_bumn",
        "realisasi_level_risiko_kbumn",
    }

    for key in expected_keys():
        master = master_map[key]
        row_no, row = parsed[key]
        item, created = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=master)
        if created:
            stats.created_items += 1

        qualitative = is_qualitative_master(master)
        impact_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_nilai_dampak"][quarter])
        impact_scale_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_dampak"][quarter])
        prob_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_nilai_probabilitas"][quarter])
        prob_scale_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_probabilitas"][quarter])

        impact_value = None if qualitative else decimal_or_none(impact_raw)
        impact_scale = scale_by_level(MasterSkalaDampak, impact_scale_raw)
        probability_value = percent_or_none(prob_raw)
        probability_scale = scale_by_level(MasterSkalaProbabilitas, prob_scale_raw)

        for label, val in (
            ("skala_dampak", impact_scale),
            ("nilai_probabilitas", probability_value),
            ("skala_probabilitas", probability_scale),
        ):
            if val is None:
                stats.empty_quarter_fields.append(
                    f"{key} III.A row {row_no}: {label} Q{quarter} kosong"
                )
        if not qualitative and impact_value is None:
            stats.empty_quarter_fields.append(
                f"{key} III.A row {row_no}: nilai_dampak Q{quarter} kuantitatif kosong"
            )

        values = {
            "jenis_risiko": "kualitatif" if qualitative else "kuantitatif",
            "realisasi_asumsi_dampak": text_or_none(value_at(row, IIIA_ASSUMPTION_INDEX)),
            "realisasi_nilai_dampak": impact_value,
            "realisasi_skala_dampak": impact_scale,
            "realisasi_nilai_probabilitas": probability_value,
            "realisasi_skala_probabilitas": probability_scale,
            "realisasi_skala_dampak_kbumn": int_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_dampak_kbumn"][quarter])
            ),
            "realisasi_skala_probabilitas_kbumn": int_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_probabilitas_kbumn"][quarter])
            ),
            "realisasi_eksposur": decimal_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_eksposur"][quarter])
            ),
            "realisasi_skor_risiko": int_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skor_risiko"][quarter])
            ),
            "realisasi_skala_nilai_risiko_kbumn": int_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_nilai_risiko_kbumn"][quarter])
            ),
            "realisasi_level_risiko_bumn": text_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_level_risiko_bumn"][quarter])
            ),
            "realisasi_level_risiko_kbumn": text_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_level_risiko_kbumn"][quarter])
            ),
            "efektivitas_perlakuan_risiko": treatment_effectiveness(
                value_at(row, IIIA_EFFECTIVENESS_INDEX)
            ),
        }
        apply_fields(item, values, stats, overwrite_none_fields=overwrite_none)


def import_iiib(report, master_map, parsed, workbook, month, stats):
    ws = workbook["III.B"]
    quarter = quarter_for_month(month)
    kri_threshold_idx, kri_score_idx = locate_monthly_kri_columns(ws, month)
    for key in expected_keys():
        master = master_map[key]
        row_no, row = parsed[key]
        item, created = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=master)
        if created:
            stats.created_items += 1

        values = {
            "realisasi_rencana_perlakuan": text_or_none(value_at(row, IIIB_ACTUAL_TREATMENT_INDEX)),
            "realisasi_output_perlakuan": text_or_none(value_at(row, IIIB_ACTUAL_OUTPUT_INDEX)),
            "realisasi_biaya_perlakuan": decimal_or_none(value_at(row, IIIB_ACTUAL_COST_INDEX)),
            "realisasi_pic": text_or_none(value_at(row, IIIB_ACTUAL_PIC_INDEX)),
            "status_rencana_perlakuan": treatment_status(value_at(row, IIIB_STATUS_INDEX)),
            "penjelasan_status_rencana": text_or_none(value_at(row, IIIB_STATUS_NOTE_INDEX)),
            "progress_pelaksanaan_percent": percent_or_none(
                value_at(row, IIIB_PROGRESS_BY_QUARTER[quarter])
            ),
        }
        apply_fields(item, values, stats)

        kri_status = text_or_none(value_at(row, kri_threshold_idx))
        kri_score_raw = value_at(row, kri_score_idx)
        kri_score_text = text_or_none(kri_score_raw)
        kri_numeric = decimal_or_none(kri_score_raw)
        if kri_status is None and kri_score_text is None:
            stats.kri_rows_blank_current_month.append(f"{key} (III.B row {row_no})")
        else:
            stats.kri_rows_with_current_month_data += 1
            kri_values = {}
            if kri_status is not None:
                kri_values["realisasi_threshold_kri"] = kri_status
            if kri_score_text is not None:
                kri_values["realisasi_threshold_kri_skor"] = kri_score_text
            if kri_numeric is not None:
                kri_values["realisasi_nilai_kri"] = kri_numeric
            apply_fields(item, kri_values, stats)


def import_changes(workbook, report):
    if "III.D" not in workbook.sheetnames:
        return 0
    ws = workbook["III.D"]
    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }
    rows = []
    for cells in ws.iter_rows(min_row=find_start_row(ws), values_only=True):
        row = list(cells)
        change_type = mapping.get(normalize(value_at(row, 1)))
        if not change_type:
            continue
        rows.append((change_type, row))
    MonthlyRiskReportChange.objects.filter(report=report).delete()
    for change_type, row in rows:
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=value_at(row, 2),
            penjelasan=value_at(row, 3),
        )
    return len(rows)


def import_loss_events(workbook, report):
    if "III.E" not in workbook.sheetnames:
        return 0
    ws = workbook["III.E"]
    rows = []
    for cells in ws.iter_rows(min_row=find_start_row(ws), values_only=True):
        row = list(cells)
        name = text_or_none(value_at(row, 1))
        if name:
            rows.append(row)
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()
    for row in rows:
        src = normalize(value_at(row, 4))
        repeat = normalize(value_at(row, 12))
        insured = normalize(value_at(row, 18))
        MonthlyRiskReportLossEvent.objects.create(
            report=report,
            nama_kejadian=value_at(row, 1),
            identifikasi_kejadian=value_at(row, 2),
            kategori_kejadian=value_at(row, 3),
            sumber_penyebab_kejadian="external" if "eksternal" in src else "internal" if "internal" in src else None,
            penyebab_kejadian=value_at(row, 5),
            penanganan_saat_kejadian=value_at(row, 6),
            deskripsi_kejadian_risk_event=value_at(row, 7),
            kategori_risiko_bumn=value_at(row, 8),
            kategori_risiko_t2_t3_kbumn=value_at(row, 9),
            penjelasan_kerugian=value_at(row, 10),
            nilai_kerugian=decimal_or_none(value_at(row, 11)),
            kejadian_berulang="ya" if "ya" in repeat else "tidak" if "tidak" in repeat else None,
            frekuensi_kejadian=value_at(row, 13),
            mitigasi_direncanakan=value_at(row, 14),
            realisasi_mitigasi=value_at(row, 15),
            perbaikan_mendatang=value_at(row, 16),
            pihak_terkait=value_at(row, 17),
            status_asuransi="ya" if "ya" in insured else "tidak" if "tidak" in insured else None,
            nilai_premi=decimal_or_none(value_at(row, 19)),
            nilai_klaim=decimal_or_none(value_at(row, 20)),
        )
    return len(rows)


def create_sqlite_backup_if_possible():
    db = settings.DATABASES.get("default", {})
    if db.get("ENGINE") != "django.db.backends.sqlite3":
        print("BACKUP: database bukan SQLite; backup otomatis file tidak dibuat.")
        return None
    source = Path(db.get("NAME"))
    if not source.exists():
        print(f"BACKUP: file SQLite tidak ditemukan: {source}")
        return None
    backup_dir = PROJECT_ROOT / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"db_before_hcga_jun_jul_2026_{stamp}.sqlite3"
    shutil.copy2(source, target)
    print(f"BACKUP: {target}")
    return target


def process_month(spec, profile, master_map, tahun_buku, prepared_by):
    stats = MonthStats(month=spec.month)
    workbook = load_workbook(spec.path, data_only=True, read_only=False, keep_links=False)
    required = {"III.A", "III.B"}
    missing_sheets = required - set(workbook.sheetnames)
    if missing_sheets:
        raise RuntimeError(f"{spec.path.name}: sheet wajib tidak ada: {sorted(missing_sheets)}")

    period = get_period(tahun_buku, spec.month)
    report, report_created = get_or_create_report(
        profile, period, tahun_buku, prepared_by, spec.month
    )
    parsed_iiia = parse_iiia_rows(workbook, spec.month, stats)
    parsed_iiib = parse_iiib_rows(workbook, spec.month, stats)

    import_iiia(report, master_map, parsed_iiia, spec.month, stats)
    import_iiib(report, master_map, parsed_iiib, workbook, spec.month, stats)
    stats.changes_count = import_changes(workbook, report)
    stats.loss_events_count = import_loss_events(workbook, report)

    refresh_monthly_report_summary(report)
    report.refresh_from_db()
    stats.report_item_count = report.items.count()
    return report, report_created, stats


def print_master_map(master_map):
    print("MASTER MAPPING (event-based):")
    for key in expected_keys():
        item = master_map[key]
        category = getattr(getattr(item, "kategori_dampak", None), "nama", "-")
        print(
            f"- {key:8s} -> RE={item.id:>4} | Risk={item.no_risiko} | "
            f"Cause={item.no_penyebab_risiko or '-'} | {category} | {item.peristiwa_risiko}"
        )


def print_stats(report, created, stats):
    print(f"\n{MONTH_NAMES[stats.month].upper()} {YEAR} — {report.kode}")
    print(f"- Report: {'baru' if created else 'existing'}; item report setelah import: {stats.report_item_count}")
    print(f"- Source III.A: {stats.source_iiia}/{EXPECTED_SOURCE_ITEMS}")
    print(f"- Source III.B: {stats.source_iiib}/{EXPECTED_SOURCE_ITEMS}")
    print(f"- Item MonthlyRiskReport dibuat: {stats.created_items}")
    print(f"- Item tersentuh/update: {stats.updated_items}; perubahan field: {stats.field_changes}")
    print(f"- III.D perubahan profil/strategi: {stats.changes_count}")
    print(f"- III.E loss event: {stats.loss_events_count}")
    print(
        f"- KRI bulan berjalan terisi: {stats.kri_rows_with_current_month_data}; "
        f"kosong: {len(stats.kri_rows_blank_current_month)}"
    )
    if stats.kri_rows_blank_current_month:
        print("  KRI kosong (tidak fallback bulan lain): " + ", ".join(stats.kri_rows_blank_current_month))
    if stats.empty_quarter_fields:
        print("- Field quarter kosong (dipertahankan kosong; tanpa fallback):")
        for msg in stats.empty_quarter_fields:
            print(f"  * {msg}")
    if stats.skipped_rows:
        print("- Rows dilewati:")
        for msg in stats.skipped_rows:
            print(f"  * {msg}")


def main():
    parser = argparse.ArgumentParser(description="Import HCGA Juni & Juli 2026")
    parser.add_argument("--june", required=True, type=Path, help="Path XLSX laporan Juni HCGA")
    parser.add_argument("--july", required=True, type=Path, help="Path XLSX laporan Juli HCGA")
    parser.add_argument("--apply", action="store_true", help="Commit perubahan. Default: DRY RUN/rollback.")
    args = parser.parse_args()

    specs = [ImportSpec(6, args.june.expanduser().resolve()), ImportSpec(7, args.july.expanduser().resolve())]
    for spec in specs:
        if not spec.path.exists():
            raise FileNotFoundError(spec.path)

    mode = "APPLY" if args.apply else "DRY RUN"
    print(f"HCGA JUNI–JULI {YEAR} | MODE={mode}")
    for spec in specs:
        print(
            f"SOURCE {MONTH_NAMES[spec.month]}: {spec.path} | "
            f"size={spec.path.stat().st_size} | sha256={file_sha256(spec.path)}"
        )

    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user aktif untuk prepared_by.")
    profile = get_hcga_profile()
    master_map = build_master_map(profile)
    print(f"PROFILE: id={profile.id} | {profile.judul} | unit={profile.unit_bisnis}")
    print_master_map(master_map)

    if args.apply:
        create_sqlite_backup_if_possible()

    results = []
    with transaction.atomic():
        tahun_buku, _ = TahunBuku.objects.get_or_create(
            tahun=YEAR, defaults={"aktif": True}
        )
        for spec in specs:
            results.append(process_month(spec, profile, master_map, tahun_buku, prepared_by))
        if not args.apply:
            transaction.set_rollback(True)

    for report, created, stats in results:
        print_stats(report, created, stats)

    if args.apply:
        print("\nRESULT: APPLY BERHASIL — transaksi Juni dan Juli sudah di-commit.")
    else:
        print("\nRESULT: DRY RUN BERHASIL — database TIDAK berubah (rollback).")
        print("Jika seluruh mapping/count sudah benar, ulangi perintah yang sama dengan --apply.")


if __name__ == "__main__":
    try:
        main()
    except (RuntimeError, ValidationError, FileNotFoundError) as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        sys.exit(2)
