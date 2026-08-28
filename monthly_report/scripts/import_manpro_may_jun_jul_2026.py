#!/usr/bin/env python3
"""Safe importer MANPRO May, June, July 2026.

Design decisions:
- Never mutates legacy MANPRO profile id=8, so Feb-Apr reports remain historical.
- Creates source-driven KM revisions plus separate May-Jun and July profile revisions.
- May/Jun: III.B is canonical for event/cause/treatment. III.A is used for risk
  assessment only when its event is known to still align with III.B.
- July: reads the 2026 MANPRO block from sheet "Juli dilapor Agustus".
- Default is DRY RUN; --apply is required to commit.
- Never trusts inconsistent legacy ReAssessmentItem->KM links for fallback.
- Missing May/Jun objectives are matched directly to legacy KM with MW guards or derived deterministically from III.B.
- Entire May+Jun+Jul operation is one database transaction.
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import os
import re
import shutil
import sys
import warnings
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django  # noqa: E402

django.setup()

from django.contrib.auth import get_user_model  # noqa: E402
from django.db import connection, transaction  # noqa: E402

from masterdata.models import PeriodeLaporan, TahunBuku  # noqa: E402
from monthly_report.models import (  # noqa: E402
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from monthly_report.services import refresh_monthly_report_summary  # noqa: E402
from risk.models import (  # noqa: E402
    BagianKontrakManajemen,
    ItemKontrakManajemen,
    KontrakManajemen,
    KategoriRisiko,
    MasterEfektivitasKontrol,
    MasterJenisExistingControl,
    MasterKategoriDampak,
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
    TaksonomiT3,
)

YEAR = 2026
BASE_PROFILE_ID = 8
MAY_JUNE_PROFILE_TITLE = "Profil Risiko MANPRO - Mei-Juni 2026"
JULY_PROFILE_TITLE = "Profil Risiko MANPRO - Juli 2026"
MAY_JUNE_KM_TITLE = "VPMANPRO - Mei-Juni 2026"
JULY_KM_TITLE = "VPMANPRO - Juli 2026"
MONTH_NAMES = {5: "Mei", 6: "Juni", 7: "Juli"}
CODE_RE = re.compile(r"^\s*BID\s+MANPRO\s*-\s*(\d+)\s*-\s*([a-z])\s*$", re.I)

# In the May/June source, these code groups have a materially different event
# between III.A and III.B. III.B is therefore authoritative and III.A risk
# assessment values are deliberately not copied for these groups.
MAY_JUNE_IIIA_MISMATCH_RISKS = {1, 2, 3, 8}


@dataclass
class SourceRow:
    code: str
    risk_no: int
    cause: str
    event: str
    description: str | None = None
    cause_text: str | None = None
    objective: str | None = None
    taxonomy: str | None = None
    risk_category: str | None = None
    kri: str | None = None
    kri_unit: str | None = None
    threshold_safe: str | None = None
    threshold_caution: str | None = None
    threshold_danger: str | None = None
    existing_control_type: str | None = None
    existing_control: str | None = None
    control_effectiveness: str | None = None
    impact_category: str | None = None
    impact_description: str | None = None
    exposure_period: str | None = None
    qualitative_explanation: str | None = None
    quantitative_assumption: str | None = None
    base_impact_value: Decimal | None = None
    base_impact_scale: int | None = None
    base_probability: Decimal | None = None
    base_probability_scale: int | None = None
    q_impact: dict[int, Decimal | None] = field(default_factory=dict)
    q_impact_scale: dict[int, int | None] = field(default_factory=dict)
    q_probability: dict[int, Decimal | None] = field(default_factory=dict)
    q_probability_scale: dict[int, int | None] = field(default_factory=dict)
    treatment_plan: str | None = None
    treatment_output: str | None = None
    treatment_budget: Decimal | None = None
    treatment_pic: str | None = None
    actual_treatment: str | None = None
    actual_output: str | None = None
    actual_cost: Decimal | None = None
    absorption_percent: Decimal | None = None
    actual_pic: str | None = None
    status: str | None = None
    status_note: str | None = None
    progress_percent: Decimal | None = None
    current_kri_status: str | None = None
    current_kri_score: str | None = None
    current_kri_value: Decimal | None = None
    actual_assumption: str | None = None
    actual_impact: Decimal | None = None
    actual_impact_scale: int | None = None
    actual_probability: Decimal | None = None
    actual_probability_scale: int | None = None
    actual_exposure: Decimal | None = None
    actual_score: int | None = None
    actual_level: str | None = None
    source_note: str | None = None


@dataclass
class MonthStats:
    month: int
    source_count: int = 0
    report_created: bool = False
    report_items: int = 0
    item_updates: int = 0
    field_changes: int = 0
    risk_assessment_skipped: list[str] = field(default_factory=list)
    blank_current_kri: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def normalize(value) -> str:
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def text_or_none(value):
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text or normalize(text) in {"n a", "na", "none", "nan"}:
        return None
    return text


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        raw = value.strip()
        if normalize(raw) in {"n a", "na", "no data", "div 0", "value", "ref"} or raw.startswith("#"):
            return None
        multiplier = Decimal("1")
        low = raw.lower()
        if "miliar" in low or "milyar" in low:
            multiplier = Decimal("1000000000")
        elif "juta" in low:
            multiplier = Decimal("1000000")
        cleaned = re.sub(r"[^0-9,\.\-]", "", raw)
        if not cleaned:
            return None
        if "," in cleaned and "." in cleaned:
            # Indonesian format 1.234,56
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        value = cleaned
    else:
        multiplier = Decimal("1")
    try:
        return Decimal(str(value).strip()) * multiplier
    except (InvalidOperation, ValueError):
        return None


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    if Decimal("0") <= number <= Decimal("1"):
        number *= Decimal("100")
    if number < 0 or number > 100:
        return None
    return number


def int_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    try:
        return int(number)
    except (TypeError, ValueError):
        return None


def parse_code(value):
    match = CODE_RE.match(str(value or ""))
    if not match:
        return None
    return int(match.group(1)), match.group(2).lower()


def canonical_code(risk_no: int, cause: str) -> str:
    return f"BID MANPRO-{risk_no}-{cause.lower()}"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def qround_for_field(instance, field_name, value):
    if value is None or not isinstance(value, Decimal):
        return value
    fld = instance._meta.get_field(field_name)
    places = getattr(fld, "decimal_places", None)
    if places is None:
        return value
    quantum = Decimal("1").scaleb(-places)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def scale_obj(model, level):
    level = int_or_none(level)
    if level is None:
        return None
    return model.objects.filter(urutan=level).first()


def is_qualitative_label(label) -> bool:
    n = normalize(label)
    return "kualitatif" in n or "kualilatif" in n


def category_obj(label):
    n = normalize(label)
    if not n:
        return None
    objs = list(MasterKategoriDampak.objects.all())
    for obj in objs:
        on = normalize(obj.nama)
        if on == n:
            return obj
    if "kualitatif" in n or "kualilatif" in n:
        for obj in objs:
            on = normalize(obj.nama)
            if "kualitatif" in on or "kualilatif" in on:
                return obj
    if "kuantitatif" in n:
        for obj in objs:
            if "kuantitatif" in normalize(obj.nama):
                return obj
    return None


def master_text_obj(model, label):
    n = normalize(label)
    if not n:
        return None
    objs = list(model.objects.all())
    for obj in objs:
        if normalize(str(obj)) == n or normalize(getattr(obj, "nama", "")) == n:
            return obj
    # Conservative containment fallback for simple masters.
    hits = [obj for obj in objs if n in normalize(str(obj)) or normalize(str(obj)) in n]
    return hits[0] if len(hits) == 1 else None


def taxonomy_obj(label):
    text = text_or_none(label)
    if not text:
        return None
    m = re.match(r"\s*([0-9]+(?:\.[0-9]+)*)", text)
    if m:
        obj = TaksonomiT3.objects.filter(kode=m.group(1)).first()
        if obj:
            return obj
    n = normalize(text)
    hits = [o for o in TaksonomiT3.objects.all() if normalize(o.nama) in n or n in normalize(o.nama)]
    return hits[0] if len(hits) == 1 else None


def risk_category_obj(label):
    text = text_or_none(label)
    if not text:
        return None
    m = re.match(r"\s*([0-9]+)", text)
    if m:
        obj = KategoriRisiko.objects.filter(kode=m.group(1)).first()
        if obj:
            return obj
    n = normalize(text)
    hits = [o for o in KategoriRisiko.objects.all() if normalize(o.nama) in n or n in normalize(o.nama)]
    return hits[0] if len(hits) == 1 else None


def treatment_status(value):
    n = normalize(value)
    if not n:
        return None
    if "discontinue" in n:
        return "discontinue"
    if "continue" in n:
        return "continue"
    return None


def treatment_effectiveness(value):
    n = normalize(value)
    if not n:
        return None
    if "tidak efektif" in n:
        return "tidak_efektif"
    if "efektif" in n:
        return "efektif"
    return None


def tokenize_for_match(value):
    stop = {
        "keterlambatan", "cod", "proyek", "pembangunan", "target", "selesai",
        "sesuai", "s", "curve", "pembangkit", "sewa", "aktual", "mw",
    }
    return {t for t in normalize(value).split() if t not in stop and len(t) > 1}


def text_match_score(query, candidate):
    qn, cn = normalize(query), normalize(candidate)
    if not qn or not cn:
        return 0.0
    if qn == cn:
        return 1.0
    qt, ct = tokenize_for_match(qn), tokenize_for_match(cn)
    union = qt | ct
    jaccard = (len(qt & ct) / len(union)) if union else 0.0
    ratio = SequenceMatcher(None, qn, cn).ratio()
    return 0.7 * jaccard + 0.3 * ratio


def resolve_km_item(profile, source: SourceRow):
    """Resolve only by the source's explicit KM objective.

    The revision KM is built from source objectives before profile items are synced,
    therefore fuzzy/event fallback is intentionally forbidden.  This prevents a
    new MANPRO project from being linked to an unrelated legacy KPI.
    """
    objective = text_or_none(source.objective)
    if not objective:
        raise RuntimeError(
            f"Sasaran KM kosong untuk {source.code} | {source.event}; "
            "mapping tidak boleh ditebak."
        )
    on = normalize(objective)
    candidates = list(
        ItemKontrakManajemen.objects.filter(kontrak=profile.kontrak_manajemen)
        .select_related("master_bagian")
        .order_by("master_bagian__urutan", "no_urut", "id")
    )
    exact = [c for c in candidates if normalize(c.indikator_kinerja_kunci) == on]
    if len(exact) == 1:
        return exact[0], "objective-exact", 1.0
    if len(exact) > 1:
        raise RuntimeError(
            f"Sasaran KM duplikat pada KM revisi {profile.kontrak_manajemen_id}: "
            f"{objective!r} -> {[x.id for x in exact]}"
        )
    raise RuntimeError(
        f"Sasaran KM source tidak ditemukan pada KM revisi untuk {source.code}: {objective!r}."
    )


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.order_by("id").first()
    )


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(YEAR, month)
    return PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{YEAR}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {YEAR}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{YEAR}-{month:02d}-01",
            "tanggal_selesai": f"{YEAR}-{month:02d}-{last_day:02d}",
        },
    )[0]


def load_book(path):
    from openpyxl import load_workbook

    warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
    warnings.filterwarnings("ignore", message="wmf image format is not supported.*")
    return load_workbook(path, data_only=True, read_only=True, keep_links=False)


def val(row, col1):
    idx = col1 - 1
    return row[idx] if idx < len(row) else None


def find_header_col(ws, needle):
    target = normalize(needle)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True):
        for idx, cell in enumerate(row, start=1):
            if normalize(cell) == target:
                return idx
    return None


def parse_standard_month(path: Path, month: int):
    wb = load_book(path)
    required = {"III.A", "III.B"}
    if not required.issubset(set(wb.sheetnames)):
        raise RuntimeError(f"{path.name}: sheet III.A/III.B tidak lengkap.")

    # III.B is canonical.
    ws_b = wb["III.B"]
    event_by_risk = {}
    desc_by_risk = {}
    rows_b = {}
    q = 2  # May and June are both Q2.
    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[q]
    kri_status_col = find_header_col(ws_b, f"Realisasi Threshold KRI {MONTH_NAMES[month]}")
    kri_score_col = kri_status_col + 1 if kri_status_col else None

    for row_no, cells in enumerate(ws_b.iter_rows(min_row=10, values_only=True), start=10):
        row = list(cells)
        parsed = parse_code(val(row, 6))  # F
        if not parsed:
            continue
        risk_no, cause = parsed
        if text_or_none(val(row, 3)):
            event_by_risk[risk_no] = text_or_none(val(row, 3))
        if text_or_none(val(row, 4)):
            desc_by_risk[risk_no] = text_or_none(val(row, 4))
        event = event_by_risk.get(risk_no)
        if not event:
            raise RuntimeError(f"{path.name} III.B row {row_no}: event risk {risk_no} tidak ditemukan.")
        code = canonical_code(risk_no, cause)
        status_value = text_or_none(val(row, 28))
        kri_status = text_or_none(val(row, kri_status_col)) if kri_status_col else None
        kri_score_raw = val(row, kri_score_col) if kri_score_col else None
        rows_b[code.lower()] = SourceRow(
            code=code,
            risk_no=risk_no,
            cause=cause,
            event=event,
            description=desc_by_risk.get(risk_no),
            cause_text=text_or_none(val(row, 7)),
            treatment_plan=text_or_none(val(row, 8)),
            treatment_output=text_or_none(val(row, 9)),
            treatment_budget=decimal_or_none(val(row, 10)),
            actual_treatment=text_or_none(val(row, 11)),
            actual_output=text_or_none(val(row, 12)),
            actual_cost=decimal_or_none(val(row, 13)),
            absorption_percent=percent_or_none(val(row, 14)),
            actual_pic=text_or_none(val(row, 15)),
            status=treatment_status(status_value),
            status_note=text_or_none(val(row, 29)),
            progress_percent=percent_or_none(val(row, progress_col)),
            current_kri_status=kri_status,
            current_kri_score=text_or_none(kri_score_raw),
            current_kri_value=decimal_or_none(kri_score_raw),
            source_note=f"III.B row {row_no}",
        )

    # III.A profile/risk assessment data by code. Use forward-fill only within same risk number.
    ws_a = wb["III.A"]
    rows_a = {}
    last_risk = None
    shared = {}
    for row_no, cells in enumerate(ws_a.iter_rows(min_row=9, values_only=True), start=9):
        row = list(cells)
        parsed = parse_code(val(row, 15))  # O
        if not parsed:
            continue
        risk_no, cause = parsed
        if risk_no != last_risk:
            shared = {}
            last_risk = risk_no
        for key, col in {
            "objective": 7, "taxonomy": 8, "risk_category": 9,
            "event": 12, "description": 13,
        }.items():
            if text_or_none(val(row, col)):
                shared[key] = text_or_none(val(row, col))
        code = canonical_code(risk_no, cause)
        impact_label = text_or_none(val(row, 25))  # Y
        qualitative = is_qualitative_label(impact_label)
        impact = None if qualitative else decimal_or_none(val(row, 31))  # AE
        q_impact = {1: decimal_or_none(val(row, 41)), 2: decimal_or_none(val(row, 42)), 3: decimal_or_none(val(row, 43)), 4: decimal_or_none(val(row, 44))}
        if qualitative:
            q_impact = {k: None for k in q_impact}
        rows_a[code.lower()] = SourceRow(
            code=code,
            risk_no=risk_no,
            cause=cause,
            event=shared.get("event") or "",
            description=shared.get("description"),
            objective=shared.get("objective"),
            taxonomy=shared.get("taxonomy"),
            risk_category=shared.get("risk_category"),
            cause_text=text_or_none(val(row, 16)),
            kri=text_or_none(val(row, 17)),
            kri_unit=text_or_none(val(row, 18)),
            threshold_safe=text_or_none(val(row, 19)),
            threshold_caution=text_or_none(val(row, 20)),
            threshold_danger=text_or_none(val(row, 21)),
            existing_control_type=text_or_none(val(row, 22)),
            existing_control=text_or_none(val(row, 23)),
            control_effectiveness=text_or_none(val(row, 24)),
            impact_category=impact_label,
            impact_description=text_or_none(val(row, 26)),
            exposure_period=text_or_none(val(row, 27)),
            qualitative_explanation=text_or_none(val(row, 28)),
            quantitative_assumption=text_or_none(val(row, 29)),
            base_impact_value=impact,
            base_impact_scale=int_or_none(val(row, 32)),
            base_probability=percent_or_none(val(row, 33)),
            base_probability_scale=int_or_none(val(row, 34)),
            q_impact=q_impact,
            q_impact_scale={1: int_or_none(val(row, 46)), 2: int_or_none(val(row, 47)), 3: int_or_none(val(row, 48)), 4: int_or_none(val(row, 49))},
            q_probability={1: percent_or_none(val(row, 50)), 2: percent_or_none(val(row, 51)), 3: percent_or_none(val(row, 52)), 4: percent_or_none(val(row, 53))},
            q_probability_scale={1: int_or_none(val(row, 54)), 2: int_or_none(val(row, 55)), 3: int_or_none(val(row, 56)), 4: int_or_none(val(row, 57))},
            actual_assumption=text_or_none(val(row, 40)) or text_or_none(val(row, 28)) or text_or_none(val(row, 29)),
            actual_impact=None if qualitative else decimal_or_none(val(row, 42)),  # AP Q2
            actual_impact_scale=int_or_none(val(row, 47)),  # AU Q2
            actual_probability=percent_or_none(val(row, 51)),  # AY Q2
            actual_probability_scale=int_or_none(val(row, 55)),  # BC Q2
            actual_exposure=decimal_or_none(val(row, 59)),  # BG Q2
            actual_score=int_or_none(val(row, 63)),  # BK Q2
            actual_level=text_or_none(val(row, 67)),  # BO Q2
            source_note=f"III.A row {row_no}",
        )

    if len(rows_b) != 26:
        raise RuntimeError(f"{path.name}: III.B MANPRO harus 26 kode, ditemukan {len(rows_b)}.")
    if len(rows_a) != 26:
        raise RuntimeError(f"{path.name}: III.A MANPRO valid harus 26 kode, ditemukan {len(rows_a)}.")
    if set(rows_b) != set(rows_a):
        raise RuntimeError(
            f"{path.name}: kode III.A dan III.B berbeda. "
            f"Only III.B={sorted(set(rows_b)-set(rows_a))}; Only III.A={sorted(set(rows_a)-set(rows_b))}"
        )

    # Merge: III.B canonical. III.A contributes profile/assessment where safe.
    merged = {}
    for key, b in rows_b.items():
        a = rows_a[key]
        trust_iiia = b.risk_no not in MAY_JUNE_IIIA_MISMATCH_RISKS
        for field_name in (
            "objective", "taxonomy", "risk_category", "kri", "kri_unit",
            "threshold_safe", "threshold_caution", "threshold_danger",
            "existing_control_type", "existing_control", "control_effectiveness",
            "impact_category", "impact_description", "exposure_period",
            "qualitative_explanation", "quantitative_assumption", "base_impact_value",
            "base_impact_scale", "base_probability", "base_probability_scale",
            "q_impact", "q_impact_scale", "q_probability", "q_probability_scale",
        ):
            if trust_iiia:
                setattr(b, field_name, getattr(a, field_name))
        if trust_iiia:
            b.actual_assumption = a.actual_assumption
            b.actual_impact = a.actual_impact
            b.actual_impact_scale = a.actual_impact_scale
            b.actual_probability = a.actual_probability
            b.actual_probability_scale = a.actual_probability_scale
            b.actual_exposure = a.actual_exposure
            b.actual_score = a.actual_score
            b.actual_level = a.actual_level
        else:
            b.source_note = f"{b.source_note}; III.A risk assessment deliberately ignored (source event mismatch)"
        merged[key] = b

    return wb, merged


def parse_july(path: Path):
    wb = load_book(path)
    sheet_name = "Juli dilapor Agustus"
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"{path.name}: sheet '{sheet_name}' tidak ditemukan.")
    ws = wb[sheet_name]
    rows = {}
    last_risk = None
    shared = {}
    for row_no, cells in enumerate(ws.iter_rows(min_row=50, max_row=88, values_only=True), start=50):
        row = list(cells)
        if normalize(val(row, 5)) != "bid manpro":
            continue
        parsed = parse_code(val(row, 16))  # P
        if not parsed:
            continue
        risk_no, cause = parsed
        if risk_no != last_risk:
            shared = {}
            last_risk = risk_no
        for key, col in {
            "objective": 8, "taxonomy": 10, "risk_category": 11,
            "event": 13, "description": 14,
        }.items():
            if text_or_none(val(row, col)):
                shared[key] = text_or_none(val(row, col))
        event = text_or_none(val(row, 13)) or shared.get("event")
        if not event:
            raise RuntimeError(f"Juli row {row_no}: event kosong untuk {val(row, 16)}")
        code = canonical_code(risk_no, cause)
        impact_label = text_or_none(val(row, 26))  # Z
        qualitative = is_qualitative_label(impact_label)
        # July source has explicit target Q1/Q2/Q3/Q4 and explicit realization only for Q2/Q4.
        # For July (Q3), no explicit realization-Q3 exists; monthly risk realization is kept blank.
        q_impact = {1: decimal_or_none(val(row, 42)), 2: decimal_or_none(val(row, 43)), 3: decimal_or_none(val(row, 45)), 4: decimal_or_none(val(row, 46))}
        if qualitative:
            q_impact = {k: None for k in q_impact}
        cm = text_or_none(val(row, 91))  # CM: Realisasi Output
        cn = text_or_none(val(row, 92))  # CN: header says % progress, content is narrative in this file
        progress = percent_or_none(val(row, 92))
        if cn and progress is None:
            progress = None
        rows[code.lower()] = SourceRow(
            code=code,
            risk_no=risk_no,
            cause=cause,
            event=event,
            description=text_or_none(val(row, 14)) or shared.get("description"),
            objective=text_or_none(val(row, 8)) or shared.get("objective"),
            taxonomy=text_or_none(val(row, 10)) or shared.get("taxonomy"),
            risk_category=text_or_none(val(row, 11)) or shared.get("risk_category"),
            cause_text=text_or_none(val(row, 17)),
            kri=text_or_none(val(row, 18)),
            kri_unit=text_or_none(val(row, 19)),
            threshold_safe=text_or_none(val(row, 20)),
            threshold_caution=text_or_none(val(row, 21)),
            threshold_danger=text_or_none(val(row, 22)),
            existing_control_type=text_or_none(val(row, 23)),
            existing_control=text_or_none(val(row, 24)),
            control_effectiveness=text_or_none(val(row, 25)),
            impact_category=impact_label,
            impact_description=text_or_none(val(row, 27)),
            exposure_period=text_or_none(val(row, 28)),
            qualitative_explanation=text_or_none(val(row, 29)),
            quantitative_assumption=text_or_none(val(row, 30)),
            base_impact_value=None if qualitative else decimal_or_none(val(row, 32)),  # AF
            base_impact_scale=int_or_none(val(row, 33)),  # AG
            base_probability=percent_or_none(val(row, 34)),  # AH
            base_probability_scale=int_or_none(val(row, 35)),  # AI
            q_impact=q_impact,
            q_impact_scale={1: int_or_none(val(row, 48)), 2: int_or_none(val(row, 49)), 3: int_or_none(val(row, 50)), 4: int_or_none(val(row, 51))},
            q_probability={1: percent_or_none(val(row, 52)), 2: percent_or_none(val(row, 53)), 3: percent_or_none(val(row, 55)), 4: percent_or_none(val(row, 56))},
            q_probability_scale={1: int_or_none(val(row, 58)), 2: int_or_none(val(row, 59)), 3: int_or_none(val(row, 60)), 4: int_or_none(val(row, 61))},
            treatment_plan=text_or_none(val(row, 89)),  # CK
            treatment_output=text_or_none(val(row, 90)),  # CL
            treatment_budget=decimal_or_none(val(row, 93)),  # CO
            treatment_pic=text_or_none(val(row, 96)),  # CR
            actual_output=cm,
            status_note=cn,
            progress_percent=progress,
            actual_cost=decimal_or_none(val(row, 94)),  # CP
            actual_pic=text_or_none(val(row, 96)),
            source_note=f"{sheet_name} row {row_no}; no explicit Q3 realization columns in source",
        )
    if len(rows) != 37:
        raise RuntimeError(f"{path.name}: blok MANPRO 2026 Juli harus 37 kode; ditemukan {len(rows)}.")
    return wb, rows




def _mw_values(value):
    """Return MW capacities explicitly written in text, for safe project matching."""
    raw = str(value or "").lower().replace(",", ".")
    out = set()
    for token in re.findall(r"(?<![\d.])(\d+(?:\.\d+)?)\s*mw\b", raw):
        try:
            out.add(Decimal(token))
        except Exception:
            pass
    return out


def _objective_from_source_event(source: SourceRow):
    """Build a deterministic revision-KM objective from the canonical III.B event.

    This is only used when the May/Jun III.A row is known to describe a different
    project than III.B (risk groups 1,2,3,8) and no safe legacy KM KPI matches.
    The text is derived from the source event itself; it is not a fuzzy guess.
    """
    event = text_or_none(source.event)
    if not event:
        return None
    project = re.sub(
        r"^\s*Keterlambatan\s+COD(?:\s*/\s*Serah\s+Terima)?\s+(?:Proyek\s+)?",
        "",
        event,
        flags=re.I,
    ).strip(" -")
    if not project or normalize(project) == normalize(event):
        return None
    return f"Pembangunan {project} Target Selesai sesuai S-Curve"


def _legacy_objective_for_missing_source(base, source: SourceRow):
    """Resolve a missing source objective without trusting legacy profile links.

    The legacy MANPRO profile contains historical ReAssessmentItem->KM links that
    are themselves inconsistent (for example identical 120 MW Kabil events can
    point to both PLTGU Batam#1 and Compliance).  Therefore this function never
    uses ReAssessmentItem.km_item as evidence.

    Resolution order:
    1. Match the canonical III.B event directly to the legacy KM KPI, with a hard
       MW-capacity conflict guard and a conservative confidence/margin rule.
    2. If no safe KPI exists, create a deterministic objective from the III.B
       event text. This is permitted only for known May/Jun mismatch risk groups.
    """
    km_candidates = []
    src_mw = _mw_values(source.event)
    for km_item in (
        ItemKontrakManajemen.objects.filter(kontrak=base.kontrak_manajemen)
        .select_related("master_bagian")
        .order_by("master_bagian__urutan", "no_urut", "id")
    ):
        kpi = text_or_none(km_item.indikator_kinerja_kunci)
        if not kpi:
            continue
        kpi_mw = _mw_values(kpi)
        # Never map a source project to a legacy KPI with a conflicting explicit
        # MW capacity.  This deliberately prevents Sagulung 60 MW -> 90 MW.
        if src_mw and kpi_mw and src_mw.isdisjoint(kpi_mw):
            continue
        score = text_match_score(source.event, kpi)
        km_candidates.append((score, km_item, kpi))
    km_candidates.sort(key=lambda x: (-x[0], x[1].id))

    if km_candidates:
        best_score, best_item, best_kpi = km_candidates[0]
        runner = km_candidates[1][0] if len(km_candidates) > 1 else 0.0
        # The event and KPI often differ only by boilerplate such as
        # "Keterlambatan COD Proyek" vs "Pembangunan ... 120 MW - Kabil".
        # A 0.45 score is accepted only with a wide margin, making PLTGU
        # Batam#1 -> KM239 safe while unrelated candidates remain rejected.
        if best_score >= 0.45 and (best_score - runner) >= 0.15:
            return best_kpi, "legacy-km-project", best_score, best_item.id

    # Only the explicitly known source-conflict groups may derive a new KM
    # objective from III.B.  This avoids silently fabricating objectives elsewhere.
    if source.risk_no in MAY_JUNE_IIIA_MISMATCH_RISKS:
        derived = _objective_from_source_event(source)
        if derived:
            return derived, "source-event-objective", 1.0, None

    top_km = "; ".join(
        f"{score:.3f}:KM{item.id}:{kpi}" for score, item, kpi in km_candidates[:3]
    ) or "-"
    raise RuntimeError(
        f"Sasaran KM kosong dan tidak dapat ditentukan aman untuk {source.code} | {source.event}. "
        f"KM candidates: {top_km}"
    )

def fill_missing_objectives_from_legacy(base, rows: dict[str, SourceRow]):
    resolutions = []
    for source in sorted(rows.values(), key=lambda x: (x.risk_no, x.cause)):
        if text_or_none(source.objective):
            continue
        objective, method, score, legacy_id = _legacy_objective_for_missing_source(base, source)
        source.objective = objective
        source.source_note = (source.source_note or "") + f"; objective={method}:{legacy_id}"
        resolutions.append((source.code, source.event, objective, method, score, legacy_id))
    return resolutions


def fill_missing_july_objectives_from_prior(prior_rows: dict[str, SourceRow], july_rows: dict[str, SourceRow]):
    """Fill blank July KM objectives by exact event identity from the prior month.

    The July source contains two carried-forward project risks (Sekupang 40 MW and
    Sagulung 60 MW) whose objective cells are blank although the risk/treatment
    rows are otherwise complete.  The same canonical events exist in June.
    Reuse the already-resolved June objective only when the normalized event is
    an exact match and maps to exactly one objective.  No fuzzy matching is used.
    """
    by_event = {}
    for source in prior_rows.values():
        event_key = normalize(source.event)
        objective = text_or_none(source.objective)
        if not event_key or not objective:
            continue
        by_event.setdefault(event_key, set()).add(objective)

    resolutions = []
    for source in sorted(july_rows.values(), key=lambda x: (x.risk_no, x.cause)):
        if text_or_none(source.objective):
            continue
        event_key = normalize(source.event)
        candidates = by_event.get(event_key, set())
        if len(candidates) == 1:
            objective = next(iter(candidates))
            source.objective = objective
            source.source_note = (source.source_note or "") + "; objective=prior-month-exact-event"
            resolutions.append((source.code, source.event, objective, "prior-month-exact-event", 1.0, None))
        elif len(candidates) > 1:
            raise RuntimeError(
                f"Sasaran Juli ambigu untuk {source.code} | {source.event}; "
                f"event yang sama pada bulan sebelumnya memiliki beberapa sasaran: {sorted(candidates)}"
            )
    return resolutions

def _base_km_exact_index(base):
    out = {}
    for item in (
        ItemKontrakManajemen.objects.filter(kontrak=base.kontrak_manajemen)
        .select_related("master_bagian")
        .order_by("master_bagian__urutan", "no_urut", "id")
    ):
        key = normalize(item.indikator_kinerja_kunci)
        if key:
            out.setdefault(key, []).append(item)
    return out


def get_or_create_revision_km(base, title, rows: dict[str, SourceRow]):
    """Create a source-driven KM revision without mutating legacy KM id=12.

    Only unique non-empty objectives present in the source are represented.  If
    an objective exactly exists in the legacy KM, its section/metadata are
    copied.  New objectives are placed in the first configured KM section with
    neutral metadata (bobot 0, target blank) rather than guessed values.
    """
    base_km = base.kontrak_manajemen
    defaults = {"status": "Draft"}
    for field_name in ("template", "pihak_pertama", "pihak_kedua", "tanggal_kontrak"):
        if hasattr(base_km, field_name):
            defaults[field_name] = getattr(base_km, field_name)
    km, created = KontrakManajemen.objects.get_or_create(
        judul=title,
        tahun=YEAR,
        unit_bisnis=base.unit_bisnis,
        defaults=defaults,
    )

    # Build canonical unique source objectives in risk/cause order.
    objectives = []
    seen = set()
    for source in sorted(rows.values(), key=lambda x: (x.risk_no, x.cause)):
        obj = text_or_none(source.objective)
        if not obj:
            raise RuntimeError(f"Sasaran KM kosong untuk {source.code} | {source.event}")
        key = normalize(obj)
        if key not in seen:
            seen.add(key)
            objectives.append((key, obj))

    existing_items = list(
        ItemKontrakManajemen.objects.filter(kontrak=km)
        .select_related("master_bagian")
        .order_by("master_bagian__urutan", "no_urut", "id")
    )
    if existing_items:
        existing_keys = {normalize(x.indikator_kinerja_kunci) for x in existing_items if normalize(x.indikator_kinerja_kunci)}
        source_keys = {key for key, _ in objectives}
        if existing_keys != source_keys:
            missing = [obj for key, obj in objectives if key not in existing_keys]
            extra = [x.indikator_kinerja_kunci for x in existing_items if normalize(x.indikator_kinerja_kunci) not in source_keys]
            raise RuntimeError(
                f"KM revisi existing {km.id} tidak sama dengan source. "
                f"Missing={missing[:10]}; Extra={extra[:10]}"
            )
        return km, created, [(x.id, x.master_bagian.kode_bagian, x.no_urut, x.indikator_kinerja_kunci, "existing") for x in existing_items]

    base_items = list(
        ItemKontrakManajemen.objects.filter(kontrak=base_km)
        .select_related("master_bagian")
        .order_by("master_bagian__urutan", "no_urut", "id")
    )
    if not base_items:
        raise RuntimeError("KM legacy MANPRO tidak memiliki ItemKontrakManajemen.")
    base_index = _base_km_exact_index(base)
    default_master = base_items[0].master_bagian
    counters = {}
    created_rows = []

    for key, objective in objectives:
        matches = base_index.get(key, [])
        if len(matches) > 1:
            raise RuntimeError(f"Objective legacy ambigu: {objective!r} -> {[x.id for x in matches]}")
        legacy = matches[0] if matches else None
        master = legacy.master_bagian if legacy else default_master
        bagian, _ = BagianKontrakManajemen.objects.get_or_create(
            kontrak=km,
            kode_bagian=master.kode_bagian,
            defaults={"nama_bagian": master.nama_bagian},
        )
        counters.setdefault(master.id, 0)
        counters[master.id] += 1
        no_urut = counters[master.id]
        item = ItemKontrakManajemen.objects.create(
            kontrak=km,
            bagian=bagian,
            master_bagian=master,
            no_urut=no_urut,
            indikator_kinerja_kunci=objective,
            formula=(legacy.formula if legacy else None),
            satuan=(legacy.satuan if legacy else None),
            bobot=(legacy.bobot if legacy else 0),
            target=(legacy.target if legacy else None),
            polaritas=(legacy.polaritas if legacy else "positif"),
        )
        created_rows.append(
            (item.id, master.kode_bagian, no_urut, objective, "legacy-exact" if legacy else "source-new")
        )
    return km, created, created_rows


def clone_or_get_profile(base, title, revision_km):
    profile = ReAssessmentSummary.objects.filter(
        judul=title,
        tahun=YEAR,
        unit_bisnis=base.unit_bisnis,
    ).order_by("id").first()
    created = False
    if profile is None:
        profile = ReAssessmentSummary.objects.create(
            judul=title,
            tahun=YEAR,
            unit_bisnis=base.unit_bisnis,
            kontrak_manajemen=revision_km,
            risk_matrix=base.risk_matrix,
            rkm=base.rkm,
            status=ReAssessmentSummary.STATUS_DRAFT,
        )
        created = True
    elif profile.kontrak_manajemen_id != revision_km.id:
        if profile.item.exists() or MonthlyRiskReport.objects.filter(reassessment=profile).exists():
            raise RuntimeError(
                f"Profil revisi existing {profile.id} memakai KM {profile.kontrak_manajemen_id}, "
                f"bukan KM revisi {revision_km.id}; tidak aman diganti."
            )
        profile.kontrak_manajemen = revision_km
        profile.save(update_fields=["kontrak_manajemen"])
    return profile, created


def apply_master_item(item: ReAssessmentItem, source: SourceRow, km_item, *, seq: int):
    item.no_item = seq
    item.no_risiko = source.risk_no
    item.no_penyebab_risiko = source.cause
    item.km_item = km_item
    item.peristiwa_risiko = source.event
    item.deskripsi_peristiwa_risiko = source.description or "-"
    item.penyebab_risiko = source.cause_text
    item.key_risk_indicators = source.kri
    item.unit_satuan_kri = source.kri_unit
    item.threshold_aman = source.threshold_safe
    item.threshold_hati_hati = source.threshold_caution
    item.threshold_bahaya = source.threshold_danger
    item.taksonomi_t3 = taxonomy_obj(source.taxonomy)
    item.kategori_risiko = risk_category_obj(source.risk_category)
    item.jenis_existing_control = master_text_obj(MasterJenisExistingControl, source.existing_control_type)
    item.existing_control = source.existing_control
    item.penilaian_efektivitas_kontrol = master_text_obj(MasterEfektivitasKontrol, source.control_effectiveness)
    item.kategori_dampak = category_obj(source.impact_category)
    item.deskripsi_dampak = source.impact_description
    item.perkiraan_waktu_terpapar_risiko = source.exposure_period
    item.asumsi_perhitungan_dampak = source.qualitative_explanation or source.quantitative_assumption
    item.nilai_dampak = qround_for_field(item, "nilai_dampak", source.base_impact_value)
    item.skala_probabilitas = scale_obj(MasterSkalaProbabilitas, source.base_probability_scale)
    item.nilai_probabilitas = qround_for_field(item, "nilai_probabilitas", source.base_probability)
    for q in range(1, 5):
        setattr(item, f"nilai_dampak_q{q}", qround_for_field(item, f"nilai_dampak_q{q}", source.q_impact.get(q)))
        setattr(item, f"skala_dampak_q{q}", scale_obj(MasterSkalaDampak, source.q_impact_scale.get(q)))
        setattr(item, f"nilai_probabilitas_q{q}", qround_for_field(item, f"nilai_probabilitas_q{q}", source.q_probability.get(q)))
        setattr(item, f"skala_probabilitas_q{q}", scale_obj(MasterSkalaProbabilitas, source.q_probability_scale.get(q)))
    item.rencana_perlakuan_risiko = source.treatment_plan
    item.output_perlakuan_risiko = source.treatment_output
    item.biaya_perlakuan_risiko = qround_for_field(item, "biaya_perlakuan_risiko", source.treatment_budget)
    item.pic = source.treatment_pic
    item.save()
    return item


def sync_profile_items(profile, rows: dict[str, SourceRow]):
    # Rerun-safe: move existing rows out of 1..N before restoring source order.
    existing = list(ReAssessmentItem.objects.filter(summary=profile).order_by("id"))
    for offset, item in enumerate(existing, start=9000):
        if item.no_item != offset:
            item.no_item = offset
            item.save(update_fields=["no_item"])

    by_key = {
        (int(item.no_risiko), normalize(item.no_penyebab_risiko)): item
        for item in ReAssessmentItem.objects.filter(summary=profile)
    }
    out = {}
    mappings = []
    source_keys = set()
    ordered = sorted(rows.values(), key=lambda x: (x.risk_no, x.cause))
    for seq, source in enumerate(ordered, start=1):
        source_keys.add((source.risk_no, normalize(source.cause)))
        km_item, method, score = resolve_km_item(profile, source)
        key = (source.risk_no, normalize(source.cause))
        item = by_key.get(key)
        if item is None:
            item = ReAssessmentItem(
                summary=profile,
                km_item=km_item,
                no_item=seq,
                no_risiko=source.risk_no,
                no_penyebab_risiko=source.cause,
                peristiwa_risiko=source.event,
                deskripsi_peristiwa_risiko=source.description or "-",
            )
        item = apply_master_item(item, source, km_item, seq=seq)
        out[source.code.lower()] = item
        mappings.append((source.code, item.id, method, score, km_item.no_urut, km_item.indikator_kinerja_kunci))

    extras = [
        (item.id, item.no_risiko, item.no_penyebab_risiko, item.peristiwa_risiko)
        for item in ReAssessmentItem.objects.filter(summary=profile)
        if (int(item.no_risiko), normalize(item.no_penyebab_risiko)) not in source_keys
    ]
    if extras:
        raise RuntimeError(f"Profil revisi {profile.id} memiliki master item ekstra yang tidak ada di source: {extras[:10]}")
    return out, mappings


def get_or_create_report(profile, month, tahun_buku, prepared_by):
    period = get_period(tahun_buku, month)
    code = f"MRR-MANPRO-{YEAR}-{month:02d}"
    code_conflict = MonthlyRiskReport.objects.filter(kode=code).exclude(reassessment=profile).first()
    if code_conflict:
        raise RuntimeError(f"Kode {code} sudah dipakai report ID={code_conflict.id} profile={code_conflict.reassessment_id}.")
    report = MonthlyRiskReport.objects.filter(reassessment=profile, periode=period, versi=1).first()
    created = False
    if report is None:
        report = MonthlyRiskReport.objects.create(
            reassessment=profile,
            periode=period,
            tahun_buku=tahun_buku,
            versi=1,
            kode=code,
            status="draft",
            prepared_by=prepared_by,
        )
        created = True
    else:
        if report.status not in {"draft", "revision"} or getattr(report, "is_locked", False):
            raise RuntimeError(f"{report.kode} tidak aman diimpor: status={report.status}, locked={report.is_locked}")
    return report, created


def apply_item_fields(item, values, stats: MonthStats, *, overwrite_none=()):
    changed = []
    for name, value in values.items():
        value = qround_for_field(item, name, value)
        if value is None and name not in overwrite_none:
            continue
        old = getattr(item, name)
        old_cmp = old.pk if hasattr(old, "pk") else old
        new_cmp = value.pk if hasattr(value, "pk") else value
        if old_cmp != new_cmp:
            setattr(item, name, value)
            changed.append(name)
            stats.field_changes += 1
    if changed:
        item.full_clean()
        item.save(update_fields=changed + ["updated_at"])
        stats.item_updates += 1


def apply_month_report(profile, master_map, rows, month, tahun_buku, prepared_by, *, july=False):
    stats = MonthStats(month=month, source_count=len(rows))
    report, created = get_or_create_report(profile, month, tahun_buku, prepared_by)
    stats.report_created = created
    touched_ids = []
    ordered = sorted(rows.values(), key=lambda x: (x.risk_no, x.cause))
    for source in ordered:
        master = master_map[source.code.lower()]
        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=master)
        touched_ids.append(item.id)
        qualitative = is_qualitative_label(source.impact_category) or (
            master.kategori_dampak_id and is_qualitative_label(master.kategori_dampak.nama)
        )
        values = {
            "jenis_risiko": "kualitatif" if qualitative else "kuantitatif" if source.impact_category else None,
            "realisasi_asumsi_dampak": source.actual_assumption,
            "realisasi_nilai_dampak": None if qualitative else source.actual_impact,
            "realisasi_skala_dampak": scale_obj(MasterSkalaDampak, source.actual_impact_scale),
            "realisasi_nilai_probabilitas": source.actual_probability,
            "realisasi_skala_probabilitas": scale_obj(MasterSkalaProbabilitas, source.actual_probability_scale),
            "realisasi_eksposur": source.actual_exposure,
            "realisasi_skor_risiko": source.actual_score,
            "realisasi_level_risiko": source.actual_level,
            "realisasi_rencana_perlakuan": source.actual_treatment,
            "realisasi_output_perlakuan": source.actual_output,
            "realisasi_biaya_perlakuan": source.actual_cost,
            "persentase_serapan_biaya": source.absorption_percent,
            "realisasi_pic": source.actual_pic,
            "status_rencana_perlakuan": source.status,
            "penjelasan_status_rencana": source.status_note,
            "progress_pelaksanaan_percent": source.progress_percent,
        }
        # For July, explicit Q3 risk realization does not exist in the source; keep those fields empty.
        overwrite = {
            "realisasi_nilai_dampak", "realisasi_skala_dampak",
            "realisasi_nilai_probabilitas", "realisasi_skala_probabilitas",
            "realisasi_eksposur", "realisasi_skor_risiko", "realisasi_level_risiko",
        }
        apply_item_fields(item, values, stats, overwrite_none=overwrite)

        kri_values = {}
        if source.current_kri_status is not None:
            kri_values["realisasi_threshold_kri"] = source.current_kri_status
        if source.current_kri_score is not None:
            kri_values["realisasi_threshold_kri_skor"] = source.current_kri_score
        if source.current_kri_value is not None:
            kri_values["realisasi_nilai_kri"] = source.current_kri_value
        if kri_values:
            apply_item_fields(item, kri_values, stats)
        elif not july:
            stats.blank_current_kri.append(source.code)

        if source.source_note and "III.A risk assessment deliberately ignored" in source.source_note:
            stats.risk_assessment_skipped.append(source.code)
        if july:
            stats.risk_assessment_skipped.append(source.code)

    # Reports are new; if rerun, remove only report items no longer in source.
    report.items.exclude(id__in=touched_ids).delete()
    refresh_monthly_report_summary(report)
    stats.report_items = report.items.count()
    return report, stats


def import_changes(workbook, report):
    if "III.D" not in workbook.sheetnames:
        return 0
    ws = workbook["III.D"]
    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }
    MonthlyRiskReportChange.objects.filter(report=report).delete()
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        change_type = mapping.get(normalize(row[1] if len(row) > 1 else None))
        if not change_type:
            continue
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=row[2] if len(row) > 2 else None,
            penjelasan=row[3] if len(row) > 3 else None,
        )
        count += 1
    return count


def import_loss_events(workbook, report):
    if "III.E" not in workbook.sheetnames:
        return 0
    ws = workbook["III.E"]
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not text_or_none(name):
            continue
        MonthlyRiskReportLossEvent.objects.create(
            report=report,
            nama_kejadian=name,
            identifikasi_kejadian=row[2] if len(row) > 2 else None,
            kategori_kejadian=row[3] if len(row) > 3 else None,
            sumber_penyebab_kejadian=(
                "external" if "eksternal" in normalize(row[4] if len(row) > 4 else None)
                else "internal" if "internal" in normalize(row[4] if len(row) > 4 else None)
                else None
            ),
            penyebab_kejadian=row[5] if len(row) > 5 else None,
            penanganan_saat_kejadian=row[6] if len(row) > 6 else None,
            deskripsi_kejadian_risk_event=row[7] if len(row) > 7 else None,
            kategori_risiko_bumn=row[8] if len(row) > 8 else None,
            kategori_risiko_t2_t3_kbumn=row[9] if len(row) > 9 else None,
            penjelasan_kerugian=row[10] if len(row) > 10 else None,
            nilai_kerugian=decimal_or_none(row[11] if len(row) > 11 else None),
            kejadian_berulang=(
                "ya" if "ya" in normalize(row[12] if len(row) > 12 else None)
                else "tidak" if "tidak" in normalize(row[12] if len(row) > 12 else None)
                else None
            ),
            frekuensi_kejadian=row[13] if len(row) > 13 else None,
            mitigasi_direncanakan=row[14] if len(row) > 14 else None,
            realisasi_mitigasi=row[15] if len(row) > 15 else None,
            perbaikan_mendatang=row[16] if len(row) > 16 else None,
            pihak_terkait=row[17] if len(row) > 17 else None,
            status_asuransi=(
                "ya" if "ya" in normalize(row[18] if len(row) > 18 else None)
                else "tidak" if "tidak" in normalize(row[18] if len(row) > 18 else None)
                else None
            ),
            nilai_premi=decimal_or_none(row[19] if len(row) > 19 else None),
            nilai_klaim=decimal_or_none(row[20] if len(row) > 20 else None),
        )
        count += 1
    return count


def backup_sqlite_if_needed():
    if connection.vendor != "sqlite":
        print(f"BACKUP: database vendor={connection.vendor}; backup file SQLite tidak diperlukan oleh script.")
        return None
    name = connection.settings_dict.get("NAME")
    if not name:
        return None
    src = Path(name)
    if not src.exists():
        return None
    backup_dir = PROJECT_DIR / "backups"
    backup_dir.mkdir(exist_ok=True)
    from datetime import datetime
    dst = backup_dir / f"db_before_manpro_may_jun_jul_2026_{datetime.now():%Y%m%d_%H%M%S}.sqlite3"
    shutil.copy2(src, dst)
    print(f"BACKUP: {dst}")
    return dst


def print_profile_mapping(title, profile, created, mappings):
    print(f"\nPROFILE: {profile.id} | {title} | {'baru' if created else 'existing'} | master={len(mappings)}")
    for code, re_id, method, score, km_no, km_text in mappings:
        print(f"- {code:<17} -> RE={re_id:>4} | KM#{km_no} | {method} {score:.3f} | {km_text[:100]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--may", required=True, type=Path)
    parser.add_argument("--june", required=True, type=Path)
    parser.add_argument("--july", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    for p in (args.may, args.june, args.july):
        if not p.exists():
            raise FileNotFoundError(p)

    mode = "APPLY" if args.apply else "DRY RUN"
    print(f"MANPRO MEI-JUNI-JULI 2026 | MODE={mode}")
    print(f"SOURCE Mei : {args.may} | size={args.may.stat().st_size} | sha256={sha256(args.may)}")
    print(f"SOURCE Juni: {args.june} | size={args.june.stat().st_size} | sha256={sha256(args.june)}")
    print(f"SOURCE Juli: {args.july} | size={args.july.stat().st_size} | sha256={sha256(args.july)}")

    base = ReAssessmentSummary.objects.select_related("unit_bisnis", "kontrak_manajemen").get(pk=BASE_PROFILE_ID)
    if base.tahun != YEAR or normalize(base.unit_bisnis.name) != "bid manpro":
        raise RuntimeError(f"Base profile ID={BASE_PROFILE_ID} bukan MANPRO 2026: {base} / {base.unit_bisnis}")
    existing_codes = list(
        MonthlyRiskReport.objects.filter(kode__in=["MRR-MANPRO-2026-05", "MRR-MANPRO-2026-06", "MRR-MANPRO-2026-07"])
        .values_list("id", "kode", "reassessment_id", "status")
    )
    if existing_codes:
        print(f"EXISTING target reports: {existing_codes}")

    may_wb, may_rows = parse_standard_month(args.may, 5)
    june_wb, june_rows = parse_standard_month(args.june, 6)
    july_wb, july_rows = parse_july(args.july)
    if set(may_rows) != set(june_rows):
        raise RuntimeError("Struktur kode MANPRO Mei dan Juni berbeda.")
    structural_diffs = []
    for key in may_rows:
        a, b = may_rows[key], june_rows[key]
        if normalize(a.event) != normalize(b.event) or normalize(a.cause_text) != normalize(b.cause_text):
            structural_diffs.append((a.code, a.event, b.event, a.cause_text, b.cause_text))
    if structural_diffs:
        raise RuntimeError(f"Master canonical Mei/Juni berbeda: {structural_diffs[:8]}")

    print(f"SOURCE COUNT: Mei={len(may_rows)} | Juni={len(june_rows)} | Juli={len(july_rows)}")
    print("May/Jun III.A risk assessment deliberately ignored for risk groups 1,2,3,8 because event text conflicts with III.B.")
    print("July has no explicit Q3 realization-risk columns; July monthly risk-realization fields will remain empty.")

    # For May/June groups whose III.A event is stale, recover the KM objective
    # only from legacy MANPRO relationships/project identity. This is done
    # before creating the revision KM so the resulting mapping remains exact.
    mj_objective_fallbacks = fill_missing_objectives_from_legacy(base, june_rows)

    # July rows 13/14 are carried-forward Sekupang/Sagulung risks with blank
    # objective cells in the July sheet. Resolve them first by exact event identity
    # against the already-resolved June source; only then use the conservative
    # legacy resolver for any objective that is still blank.
    july_prior_objectives = fill_missing_july_objectives_from_prior(june_rows, july_rows)
    july_objective_fallbacks = fill_missing_objectives_from_legacy(base, july_rows)
    if mj_objective_fallbacks:
        print("\nMAY/JUN OBJECTIVE FALLBACKS (legacy-grounded; no invented sasaran):")
        for code, event, objective, method, score, legacy_id in mj_objective_fallbacks:
            print(f"- {code} | {method} {score:.3f} id={legacy_id} | {event} -> {objective}")
    if july_prior_objectives:
        print("\nJULY OBJECTIVES FROM JUNE (exact event; no fuzzy matching):")
        for code, event, objective, method, score, legacy_id in july_prior_objectives:
            print(f"- {code} | {method} | {event} -> {objective}")
    if july_objective_fallbacks:
        print("\nJULY OBJECTIVE FALLBACKS (legacy-grounded; no invented sasaran):")
        for code, event, objective, method, score, legacy_id in july_objective_fallbacks:
            print(f"- {code} | {method} {score:.3f} id={legacy_id} | {event} -> {objective}")

    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user prepared_by.")

    # Backup the committed SQLite database BEFORE opening the write transaction.
    # This guarantees the backup is a clean pre-import snapshot.
    if args.apply:
        backup_sqlite_if_needed()

    with transaction.atomic():
        tahun_buku, _ = TahunBuku.objects.get_or_create(tahun=YEAR, defaults={"aktif": True})
        kmrev_mj, kmrev_mj_created, kmrev_mj_items = get_or_create_revision_km(base, MAY_JUNE_KM_TITLE, june_rows)
        kmrev_jul, kmrev_jul_created, kmrev_jul_items = get_or_create_revision_km(base, JULY_KM_TITLE, july_rows)
        p_mj, p_mj_created = clone_or_get_profile(base, MAY_JUNE_PROFILE_TITLE, kmrev_mj)
        p_jul, p_jul_created = clone_or_get_profile(base, JULY_PROFILE_TITLE, kmrev_jul)

        print(f"\nKM REVISION {MAY_JUNE_KM_TITLE}: id={kmrev_mj.id} | {'baru' if kmrev_mj_created else 'existing'} | source objectives={len(kmrev_mj_items)}")
        for row in kmrev_mj_items:
            print(f"- KM-ID={row[0]} | {row[1]}.{row[2]} | {row[4]} | {row[3]}")
        print(f"\nKM REVISION {JULY_KM_TITLE}: id={kmrev_jul.id} | {'baru' if kmrev_jul_created else 'existing'} | source objectives={len(kmrev_jul_items)}")
        for row in kmrev_jul_items:
            print(f"- KM-ID={row[0]} | {row[1]}.{row[2]} | {row[4]} | {row[3]}")

        # Use June source for the May-Jun revision master (same canonical structure, latest file).
        map_mj, km_mj = sync_profile_items(p_mj, june_rows)
        map_jul, km_jul = sync_profile_items(p_jul, july_rows)
        print_profile_mapping(MAY_JUNE_PROFILE_TITLE, p_mj, p_mj_created, km_mj)
        print_profile_mapping(JULY_PROFILE_TITLE, p_jul, p_jul_created, km_jul)

        may_report, may_stats = apply_month_report(p_mj, map_mj, may_rows, 5, tahun_buku, prepared_by)
        june_report, june_stats = apply_month_report(p_mj, map_mj, june_rows, 6, tahun_buku, prepared_by)
        july_report, july_stats = apply_month_report(p_jul, map_jul, july_rows, 7, tahun_buku, prepared_by, july=True)

        may_changes = import_changes(may_wb, may_report)
        may_losses = import_loss_events(may_wb, may_report)
        june_changes = import_changes(june_wb, june_report)
        june_losses = import_loss_events(june_wb, june_report)

        print("\nIMPORT SUMMARY")
        for report, stats, chg, loss in (
            (may_report, may_stats, may_changes, may_losses),
            (june_report, june_stats, june_changes, june_losses),
            (july_report, july_stats, 0, 0),
        ):
            print(f"\n{MONTH_NAMES[stats.month].upper()} 2026 — {report.kode}")
            print(f"- Report: {'baru' if stats.report_created else 'existing'} | items={stats.report_items}")
            print(f"- Source canonical: {stats.source_count}/{stats.source_count}")
            print(f"- Item tersentuh/update: {stats.item_updates}; perubahan field: {stats.field_changes}")
            print(f"- III.D changes: {chg}; III.E loss events: {loss}")
            if stats.risk_assessment_skipped:
                print(f"- Risk realization sengaja tidak diisi dari source: {len(stats.risk_assessment_skipped)}")
                print("  " + ", ".join(stats.risk_assessment_skipped))
            if stats.blank_current_kri:
                print(f"- KRI bulan berjalan kosong: {len(stats.blank_current_kri)}")
                print("  " + ", ".join(stats.blank_current_kri))

        if args.apply:
            print("\nRESULT: APPLY BERHASIL — transaksi Mei, Juni, Juli akan di-commit.")
        else:
            transaction.set_rollback(True)
            print("\nRESULT: DRY RUN BERHASIL — database TIDAK berubah (rollback).")
            print("Jika mapping KM dan count sudah benar, ulangi command yang sama dengan --apply.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
