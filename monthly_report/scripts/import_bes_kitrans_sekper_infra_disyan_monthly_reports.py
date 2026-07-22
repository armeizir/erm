from dataclasses import dataclass
from pathlib import Path

from django.db import transaction

from masterdata.models import TahunBuku
from monthly_report.models import MonthlyRiskReport, MonthlyRiskReportItem
from monthly_report.scripts.import_ops_manpro_keu_hcga_monthly_reports import (
    MONTH_NAMES,
    YEAR,
    build_maps,
    decimal_or_none,
    get_period,
    get_prepared_by,
    import_changes,
    import_loss_events,
    import_standard_iiia,
    import_standard_iiib,
    int_or_none,
    percent_or_none,
    read_rows,
    row_value,
    scale_by_level,
    selected_quarter,
    treatment_effectiveness,
    treatment_status,
    update_item_common,
)
from risk.models import MasterSkalaDampak, MasterSkalaProbabilitas, ReAssessmentSummary


@dataclass(frozen=True)
class ImportFile:
    code: str
    reassessment_id: int
    month: int
    path: Path
    template: str


FILES = [
    ImportFile(
        "BES",
        7,
        2,
        Path("/Users/armeizir/Downloads/2026/15. BES/02 Laporan Realisasi ManRisk UB BES Februari 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "BES",
        7,
        3,
        Path("/Users/armeizir/Downloads/2026/15. BES/03 Laporan Realisasi ManRisk UB BES Maret 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "KITRANS",
        14,
        2,
        Path("/Users/armeizir/Downloads/2026/14. KITRANS/Laporan Manajemen Risiko UBKITRANS Februari 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "KITRANS",
        14,
        3,
        Path("/Users/armeizir/Downloads/2026/14. KITRANS/Laporan Manajemen Risiko UBKITRANS Maret 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "KITRANS",
        14,
        4,
        Path("/Users/armeizir/Downloads/2026/14. KITRANS/Laporan Manajemen Risiko UBKITRANS April 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "SETPER",
        13,
        2,
        Path("/Users/armeizir/Downloads/2026/13. SEKPER/Laporan Mitigasi .xlsx"),
        "old_compact",
    ),
    ImportFile(
        "SETPER",
        13,
        3,
        Path("/Users/armeizir/Downloads/2026/13. SEKPER/Laporan MR Setper sd Maret 2026.xlsx"),
        "old_compact",
    ),
    ImportFile(
        "SETPER",
        13,
        4,
        Path("/Users/armeizir/Downloads/2026/13. SEKPER/Laporan Rev-1 SEKPER.xlsx"),
        "old_compact",
    ),
    ImportFile(
        "INFRA",
        3,
        3,
        Path("/Users/armeizir/Downloads/2026/12. INFRA/Laporan Realisasi 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "INFRA",
        3,
        4,
        Path("/Users/armeizir/Downloads/2026/12. INFRA/Laporan Realisasi Evaluasi Mitigasi Risiko Bulan April 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "DISYAN",
        12,
        2,
        Path("/Users/armeizir/Downloads/2026/11. DISYAN/Laporan Risiko Bulan Februari 2026.xlsx"),
        "old_disyan",
    ),
    ImportFile(
        "DISYAN",
        12,
        3,
        Path("/Users/armeizir/Downloads/2026/11. DISYAN/Laporan Risiko Bulan Maret 2026.xlsx"),
        "old_disyan",
    ),
    ImportFile(
        "DISYAN",
        12,
        4,
        Path("/Users/armeizir/Downloads/2026/11. DISYAN/Laporan Risiko UB Disyan Bulan April 2026.xlsx"),
        "old_disyan",
    ),
]


def risk_by_no_item(maps, value):
    _, by_no_item, _ = maps
    number = int_or_none(value)
    if number is None:
        return None
    return by_no_item.get(number)


def import_old_residual(workbook, report, maps, month, template, skipped):
    ws = workbook["LAP REAL III.A"]
    q = selected_quarter(month)
    if template == "old_disyan":
        no_col = 2
        assumption_col = 12
        cols = {
            "nilai_dampak": {1: 13, 2: 14, 3: 15, 4: 16}[q],
            "skala_dampak": {1: 17, 2: 18, 3: 19, 4: 20}[q],
            "nilai_prob": {1: 21, 2: 22, 3: 23, 4: 24}[q],
            "skala_prob": {1: 25, 2: 26, 3: 27, 4: 28}[q],
        }
        effect_col = 41
    else:
        no_col = 1
        assumption_col = 2
        cols = {
            "nilai_dampak": {1: 3, 2: 4, 3: 5, 4: 6}[q],
            "skala_dampak": {1: 7, 2: 8, 3: 9, 4: 10}[q],
            "nilai_prob": {1: 11, 2: 12, 3: 13, 4: 14}[q],
            "skala_prob": {1: 15, 2: 16, 3: 17, 4: 18}[q],
        }
        effect_col = 31

    imported = 0
    seen = set()
    for row_idx, row in read_rows(ws, 12):
        risk_event = risk_by_no_item(maps, row_value(row, no_col))
        if not risk_event:
            continue
        if risk_event.id in seen:
            continue
        seen.add(risk_event.id)
        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        item.realisasi_asumsi_dampak = row_value(row, assumption_col)
        item.realisasi_nilai_dampak = decimal_or_none(row_value(row, cols["nilai_dampak"]))
        item.realisasi_skala_dampak = scale_by_level(MasterSkalaDampak, row_value(row, cols["skala_dampak"]))
        item.realisasi_nilai_probabilitas = percent_or_none(row_value(row, cols["nilai_prob"]))
        item.realisasi_skala_probabilitas = scale_by_level(
            MasterSkalaProbabilitas,
            row_value(row, cols["skala_prob"]),
        )
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(row_value(row, effect_col))
        item.save()
        imported += 1
    return imported


def import_old_treatment(workbook, report, maps, month, template, skipped):
    ws = workbook["LAP REAL III.B"]
    q = selected_quarter(month)

    if template == "old_compact":
        no_col = 1
        cols = {
            "rencana": 2,
            "output": 3,
            "biaya": 4,
            "serapan": 5,
            "pic": 6,
            "status": 26,
            "explanation": 27,
            "progress": {1: 28, 2: 29, 3: 30, 4: 31}[q],
            "threshold": {2: 20, 3: 22, 4: 24}.get(month, 20),
        }
    else:
        no_col = 2
        full_template = row_value(next(ws.iter_rows(min_row=4, max_row=4, values_only=True)), 3) == "Peristiwa Risiko"
        if full_template:
            cols = {
                "rencana": 12,
                "output": 13,
                "biaya": 14,
                "serapan": 17,
                "pic": 18,
                "status": 37,
                "explanation": 38,
                "progress": 39,
                "threshold": 35,
            }
        else:
            cols = {
                "rencana": 3,
                "output": 4,
                "biaya": 8,
                "serapan": 9,
                "pic": 10,
                "status": 26,
                "explanation": 27,
                "progress": {1: 28, 2: 29, 3: 30, 4: 31}[q],
                "threshold": 24,
            }

    imported = 0
    current_risk_event = None
    for row_idx, row in read_rows(ws, 10):
        risk_event = risk_by_no_item(maps, row_value(row, no_col))
        if risk_event:
            current_risk_event = risk_event
        if not current_risk_event:
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=current_risk_event,
        )
        update_item_common(
            item,
            rencana=row_value(row, cols["rencana"]),
            output=row_value(row, cols["output"]),
            biaya=row_value(row, cols["biaya"]),
            serapan=row_value(row, cols["serapan"]),
            pic=row_value(row, cols["pic"]),
        )
        item.status_rencana_perlakuan = treatment_status(row_value(row, cols["status"]))
        item.penjelasan_status_rencana = row_value(row, cols["explanation"])
        item.progress_pelaksanaan_percent = percent_or_none(row_value(row, cols["progress"]))
        item.realisasi_threshold_kri = row_value(row, cols["threshold"])
        score = row_value(row, cols["threshold"] + 1)
        item.realisasi_threshold_kri_skor = str(score) if score not in (None, "") else None
        item.save()
        imported += 1
    return imported


def import_one(workbook, report, maps, import_file):
    skipped = []
    if import_file.template == "standard":
        iiia = import_standard_iiia(workbook, report, maps, import_file.month, skipped)
        iiib = import_standard_iiib(workbook, report, maps, import_file.month, skipped)
        iiid = import_changes(workbook, report)
        iiie = import_loss_events(workbook, report)
    else:
        iiia = import_old_residual(
            workbook,
            report,
            maps,
            import_file.month,
            import_file.template,
            skipped,
        )
        iiib = import_old_treatment(
            workbook,
            report,
            maps,
            import_file.month,
            import_file.template,
            skipped,
        )
        iiid = 0
        iiie = 0
    return iiia, iiib, iiid, iiie, skipped


def run():
    from openpyxl import load_workbook

    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user untuk prepared_by.")

    tahun_buku, _ = TahunBuku.objects.get_or_create(tahun=YEAR, defaults={"aktif": True})
    cache = {}
    summaries = []
    all_skipped = []

    with transaction.atomic():
        for import_file in FILES:
            if not import_file.path.exists():
                raise FileNotFoundError(import_file.path)

            reassessment = ReAssessmentSummary.objects.get(pk=import_file.reassessment_id)
            maps = cache.setdefault(reassessment.id, build_maps(reassessment))
            workbook = load_workbook(
                import_file.path,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
            period = get_period(tahun_buku, import_file.month)
            report, _ = MonthlyRiskReport.objects.update_or_create(
                reassessment=reassessment,
                periode=period,
                versi=1,
                defaults={
                    "kode": f"MRR-{import_file.code}-{YEAR}-{import_file.month:02d}",
                    "tahun_buku": tahun_buku,
                    "status": "draft",
                    "prepared_by": prepared_by,
                },
            )

            iiia, iiib, iiid, iiie, skipped = import_one(workbook, report, maps, import_file)

            report.total_risiko = report.items.count()
            report.total_high = report.items.filter(realisasi_level_risiko__icontains="tinggi").count()
            report.total_mitigasi_terlambat = report.items.filter(mitigation_status="delayed").count()
            report.total_selesai = report.items.filter(status_rencana_perlakuan="discontinue").count()
            report.save()

            summaries.append(
                {
                    "code": report.kode,
                    "items": report.items.count(),
                    "iiia": iiia,
                    "iiib": iiib,
                    "iiid": iiid,
                    "iiie": iiie,
                    "skipped": len(skipped),
                }
            )
            all_skipped.extend((report.kode, *item) for item in skipped)

    print("IMPORT SUMMARY")
    for item in summaries:
        print(
            f"- {item['code']}: items {item['items']}, "
            f"III.A {item['iiia']}, III.B {item['iiib']}, "
            f"III.D {item['iiid']}, III.E {item['iiie']}, skipped {item['skipped']}"
        )

    print("\nSKIPPED")
    if not all_skipped:
        print("- Tidak ada item utama yang dilewati.")
    for item in all_skipped[:80]:
        print(f"- {item}")
    if len(all_skipped) > 80:
        print(f"- ... {len(all_skipped) - 80} skipped lainnya")


if __name__ == "__main__":
    run()
