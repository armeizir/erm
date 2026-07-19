"""
One-time repair for duplicate BID AGA / Profil Risiko NIAGA June 2026 reports.

Goals:
1. Back up the active SQLite database before changing data.
2. Use the supplied June BID AGA workbook as the source of truth.
3. Synchronize the NIAGA risk-event names/details to the 14 numbered risks in the workbook.
4. Keep exactly one June 2026 MonthlyRiskReport (canonical code MRR-BIDAGA-2026-06).
5. Rebuild its monthly items from III.A and III.B so realization data is attached to the
   correct numbered risk instead of being shifted by row order.
6. Delete duplicate June reports only after the canonical report has been rebuilt.

Run from project root:
    python monthly_report/scripts/repair_bidaga_june_2026.py \
      "/Users/armeizir/Downloads/Laporan Realisasi Juni 2026 BID AGA.xlsx" --apply

Without --apply the script only audits and prints what would change.
"""

from __future__ import annotations

import argparse
import calendar
import os
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django

django.setup()

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook

from masterdata.models import PeriodeLaporan, TahunBuku
from monthly_report.models import MonthlyRiskReport, MonthlyRiskReportItem
from monthly_report.scripts.import_aga_monthly_reports import (
    decimal_or_none,
    import_sheet_iiia,
    import_sheet_iiib,
    import_sheet_iiid,
    import_sheet_iiie,
    normalize,
)
from risk.models import ReAssessmentItem, ReAssessmentSummary


YEAR = 2026
MONTH = 6
UNIT_NAME = "BID AGA"
PROFILE_TITLE = "Profil Risiko NIAGA"
REPORT_CODE = "MRR-BIDAGA-2026-06"
DEFAULT_SOURCE_FILE = "/Users/armeizir/Downloads/Laporan Realisasi Juni 2026 BID AGA.xlsx"
EXPECTED_RISK_COUNT = 14


@dataclass
class ExcelRisk:
    number: int
    event: str
    description: str | None = None
    cause_no: str | None = None
    cause: str | None = None
    treatment: str | None = None
    treatment_output: str | None = None
    treatment_budget: Any = None
    pic: str | None = None
    kri: str | None = None
    kri_unit: str | None = None
    threshold_safe: str | None = None
    threshold_caution: str | None = None
    threshold_danger: str | None = None
    impact_assumption: str | None = None


def as_int(value: Any) -> int | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return None
    return number


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def workbook_risks(workbook) -> list[ExcelRisk]:
    """Read the numbered NIAGA risk block (1..14) from III.B and enrich from III.A."""
    iiib = workbook["III.B"]
    risks: dict[int, ExcelRisk] = {}

    for row in iiib.iter_rows(min_row=10, values_only=True):
        number = as_int(row[1] if len(row) > 1 else None)
        event = clean_text(row[2] if len(row) > 2 else None)
        if number is None or not (1 <= number <= EXPECTED_RISK_COUNT) or not event:
            continue
        # The workbook contains exactly one numbered BID AGA block 1..14. If an older
        # numbered block appears first in a future template, the last occurrence wins.
        risks[number] = ExcelRisk(
            number=number,
            event=event,
            description=clean_text(row[3] if len(row) > 3 else None),
            cause_no=clean_text(row[4] if len(row) > 4 else None),
            cause=clean_text(row[6] if len(row) > 6 else None),
            treatment=clean_text(row[7] if len(row) > 7 else None),
            treatment_output=clean_text(row[8] if len(row) > 8 else None),
            treatment_budget=row[9] if len(row) > 9 else None,
            pic=clean_text(row[14] if len(row) > 14 else None),
            kri=clean_text(row[33] if len(row) > 33 else None),
            kri_unit=clean_text(row[34] if len(row) > 34 else None),
            threshold_safe=clean_text(row[35] if len(row) > 35 else None),
            threshold_caution=clean_text(row[36] if len(row) > 36 else None),
            threshold_danger=clean_text(row[37] if len(row) > 37 else None),
        )

    iiia = workbook["III.A"]
    for row in iiia.iter_rows(min_row=10, values_only=True):
        number = as_int(row[1] if len(row) > 1 else None)
        event = clean_text(row[2] if len(row) > 2 else None)
        if number not in risks or not event:
            continue
        if normalize(event) != normalize(risks[number].event):
            continue
        risks[number].impact_assumption = clean_text(row[4] if len(row) > 4 else None)

    ordered = [risks[number] for number in sorted(risks)]
    if [risk.number for risk in ordered] != list(range(1, EXPECTED_RISK_COUNT + 1)):
        found = [risk.number for risk in ordered]
        raise RuntimeError(
            f"Blok risiko NIAGA pada Excel tidak lengkap. Ditemukan nomor {found}; "
            f"diharapkan 1..{EXPECTED_RISK_COUNT}."
        )
    return ordered


def get_profile() -> ReAssessmentSummary:
    profile = (
        ReAssessmentSummary.objects.filter(
            tahun=YEAR,
            judul__iexact=PROFILE_TITLE,
            unit_bisnis__name=UNIT_NAME,
        )
        .select_related("unit_bisnis", "kontrak_manajemen")
        .order_by("id")
        .first()
    )
    if not profile:
        raise RuntimeError(f"{PROFILE_TITLE} / {UNIT_NAME} tahun {YEAR} tidak ditemukan.")
    return profile


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.order_by("id").first()
    )


def get_period(tahun_buku: TahunBuku) -> PeriodeLaporan:
    _, last_day = calendar.monthrange(YEAR, MONTH)
    return PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{YEAR}-{MONTH:02d}",
        defaults={
            "nama_periode": f"Juni {YEAR}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{YEAR}-{MONTH:02d}-01",
            "tanggal_selesai": f"{YEAR}-{MONTH:02d}-{last_day:02d}",
        },
    )[0]


def backup_sqlite_database() -> Path | None:
    db_name = settings.DATABASES["default"].get("NAME")
    if not db_name:
        return None
    db_path = Path(str(db_name)).resolve()
    if db_path.suffix.lower() not in {".sqlite", ".sqlite3", ".db"} or not db_path.exists():
        return None
    backup_dir = PROJECT_DIR / "backups"
    backup_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"db_before_repair_niaga_june_2026_{timestamp}.sqlite3"
    shutil.copy2(db_path, backup_path)
    return backup_path


def sync_profile(profile: ReAssessmentSummary, excel_risks: list[ExcelRisk]) -> list[ReAssessmentItem]:
    """
    Synchronize the 14 numbered risk rows without deleting the profile.

    Existing items 1..13 retain their structural FK configuration (KM/taxonomy/category)
    but their event/detail fields are corrected to the numbered Excel rows. Risk 14 is
    created from the structural configuration of the last existing item so the monthly
    report can represent all 14 Excel risks. This avoids row-shifted realization data.
    """
    existing = {
        int(item.no_risiko): item
        for item in ReAssessmentItem.objects.filter(summary=profile).order_by("no_risiko", "id")
        if item.no_risiko
    }
    if not existing:
        raise RuntimeError("Profil Risiko NIAGA tidak memiliki item risiko untuk dijadikan dasar sinkronisasi.")

    template = existing.get(max(existing)) or next(iter(existing.values()))
    result: list[ReAssessmentItem] = []

    for risk in excel_risks:
        item = existing.get(risk.number)
        if item is None:
            item = ReAssessmentItem(
                summary=profile,
                no_item=risk.number,
                no_risiko=risk.number,
                unit_bisnis=profile.unit_bisnis,
                km_item=template.km_item,
                sasaran_kbumn=template.sasaran_kbumn,
                taksonomi_t3=template.taksonomi_t3,
                kategori_risiko=template.kategori_risiko,
                jenis_existing_control=template.jenis_existing_control,
                penilaian_efektivitas_kontrol=template.penilaian_efektivitas_kontrol,
                kategori_dampak=template.kategori_dampak,
                jenis_program_dalam_rkap=template.jenis_program_dalam_rkap,
                opsi_perlakuan_risiko=template.opsi_perlakuan_risiko,
                pos_anggaran=template.pos_anggaran,
            )

        is_new = item.pk is None
        item.no_item = risk.number
        item.no_risiko = risk.number
        # Existing items keep their structural profile/KM configuration. Only the numbered
        # risk identity is aligned so June realization rows can be mapped correctly by name.
        item.peristiwa_risiko = risk.event
        item.deskripsi_peristiwa_risiko = risk.description or risk.event
        if is_new:
            # Risk 14 does not yet exist in the current profile, so seed its descriptive
            # fields from the authoritative workbook while reusing the template's structural FKs.
            item.no_penyebab_risiko = risk.cause_no
            item.penyebab_risiko = risk.cause
            item.rencana_perlakuan_risiko = risk.treatment
            item.output_perlakuan_risiko = risk.treatment_output
            budget = decimal_or_none(risk.treatment_budget)
            if budget is not None:
                item.biaya_perlakuan_risiko = budget
            item.pic = risk.pic
            item.key_risk_indicators = risk.kri
            item.unit_satuan_kri = risk.kri_unit
            item.threshold_aman = risk.threshold_safe
            item.threshold_hati_hati = risk.threshold_caution
            item.threshold_bahaya = risk.threshold_danger
            if risk.impact_assumption:
                item.asumsi_perhitungan_dampak = risk.impact_assumption
        item.save()
        result.append(item)

    # Do not silently keep obsolete numbered items beyond the authoritative Excel block.
    extras = ReAssessmentItem.objects.filter(summary=profile).exclude(
        no_risiko__in=[risk.number for risk in excel_risks]
    )
    if extras.exists():
        extra_labels = list(extras.values_list("no_risiko", "peristiwa_risiko"))
        raise RuntimeError(
            "Profil memiliki item di luar nomor 1..14. Repair dihentikan agar tidak menghapus data "
            f"tanpa review: {extra_labels}"
        )

    return result


def choose_canonical_report(profile, period) -> tuple[MonthlyRiskReport | None, list[MonthlyRiskReport]]:
    reports = list(
        MonthlyRiskReport.objects.filter(
            reassessment=profile,
            periode=period,
            versi=1,
        ).order_by("id")
    )
    canonical = next((report for report in reports if report.kode == REPORT_CODE), None)
    if canonical is None and reports:
        canonical = reports[0]
    duplicates = [report for report in reports if canonical is not None and report.pk != canonical.pk]
    return canonical, duplicates


def print_audit(profile, excel_risks, period) -> None:
    profile_items = list(
        ReAssessmentItem.objects.filter(summary=profile)
        .order_by("no_risiko", "id")
        .values_list("no_risiko", "peristiwa_risiko")
    )
    reports = []
    if period is not None:
        reports = list(
            MonthlyRiskReport.objects.filter(reassessment=profile, periode=period, versi=1)
            .order_by("id")
        )

    print("AUDIT NIAGA JUNI 2026")
    print(f"- Profil: {profile.id} {profile.judul}")
    print(f"- Risiko pada Profil saat ini: {len(profile_items)}")
    print(f"- Risiko bernomor pada Excel: {len(excel_risks)}")
    print(f"- Laporan Juni ditemukan: {len(reports)}")
    for report in reports:
        print(
            f"  - report_id={report.id}, kode={report.kode!r}, status={report.status}, "
            f"items={report.items.count()}"
        )

    print("\nPERBANDINGAN NAMA RISIKO")
    current_by_no = {int(no): name for no, name in profile_items if no}
    for risk in excel_risks:
        current = current_by_no.get(risk.number)
        marker = "OK" if current and normalize(current) == normalize(risk.event) else "BEDA"
        print(f"- {risk.number:02d} [{marker}] Excel: {risk.event}")
        if marker == "BEDA":
            print(f"     Profil: {current or '(belum ada)'}")


def apply_repair(path: Path, workbook, profile, excel_risks) -> None:
    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user yang dapat digunakan sebagai prepared_by.")

    backup_path = backup_sqlite_database()
    if backup_path:
        print(f"Backup database dibuat: {backup_path}")
    else:
        print("Catatan: database bukan SQLite lokal atau file DB tidak ditemukan; backup otomatis dilewati.")

    with transaction.atomic():
        tahun_buku, _ = TahunBuku.objects.get_or_create(tahun=YEAR, defaults={"aktif": True})
        period = get_period(tahun_buku)

        sync_profile(profile, excel_risks)

        canonical, duplicates = choose_canonical_report(profile, period)
        if canonical is None:
            canonical = MonthlyRiskReport.objects.create(
                reassessment=profile,
                periode=period,
                versi=1,
                kode=REPORT_CODE,
                tahun_buku=tahun_buku,
                status="draft",
                prepared_by=prepared_by,
            )
        else:
            canonical.kode = REPORT_CODE
            canonical.tahun_buku = tahun_buku
            canonical.prepared_by = canonical.prepared_by or prepared_by
            canonical.save(update_fields=["kode", "tahun_buku", "prepared_by", "updated_at"])

        # Rebuild from source workbook so no shifted/misaligned realization survives.
        MonthlyRiskReportItem.objects.filter(report=canonical).delete()
        event_map = {
            normalize(item.peristiwa_risiko): item
            for item in ReAssessmentItem.objects.filter(summary=profile)
        }
        skipped: list[dict[str, Any]] = []
        iiia = import_sheet_iiia(workbook, canonical, event_map, MONTH, skipped)
        iiib = import_sheet_iiib(workbook, canonical, event_map, MONTH, skipped)
        iiid = import_sheet_iiid(workbook, canonical)
        iiie = import_sheet_iiie(workbook, canonical, skipped)

        canonical.total_risiko = canonical.items.count()
        canonical.total_high = canonical.items.filter(
            realisasi_level_risiko__icontains="tinggi"
        ).count()
        canonical.total_mitigasi_terlambat = canonical.items.filter(
            mitigation_status="delayed"
        ).count()
        canonical.total_selesai = canonical.items.filter(
            status_rencana_perlakuan="discontinue"
        ).count()
        canonical.save(
            update_fields=[
                "total_risiko",
                "total_high",
                "total_mitigasi_terlambat",
                "total_selesai",
                "updated_at",
            ]
        )

        if canonical.items.count() != EXPECTED_RISK_COUNT:
            raise RuntimeError(
                f"Hasil rebuild hanya {canonical.items.count()} item, bukan {EXPECTED_RISK_COUNT}. "
                f"Transaksi dibatalkan. Skipped={skipped}"
            )

        # Delete duplicates only after a complete canonical report exists.
        duplicate_ids = [report.id for report in duplicates]
        for duplicate in duplicates:
            duplicate.delete()

        remaining = MonthlyRiskReport.objects.filter(
            reassessment=profile,
            periode=period,
            versi=1,
        )
        if remaining.count() != 1:
            raise RuntimeError(
                f"Validasi gagal: masih ada {remaining.count()} laporan Juni NIAGA. Transaksi dibatalkan."
            )

    print("\nREPAIR BERHASIL")
    print(f"- Sumber Excel: {path}")
    print(f"- Report canonical: id={canonical.id}, kode={canonical.kode}")
    print(f"- Jumlah item: {canonical.items.count()} (diharapkan {EXPECTED_RISK_COUNT})")
    print(f"- III.A imported: {iiia}")
    print(f"- III.B imported: {iiib}")
    print(f"- III.D imported: {iiid}")
    print(f"- III.E imported: {iiie}")
    print(f"- Duplicate report yang dihapus: {duplicate_ids or '-'}")
    if skipped:
        print("- Catatan baris yang dilewati:")
        for row in skipped:
            print(f"  {row}")
    else:
        print("- Tidak ada baris risiko bernomor yang terlewat.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Repair duplicate Monthly Risk Report NIAGA Juni 2026")
    parser.add_argument("source", nargs="?", default=DEFAULT_SOURCE_FILE, help="Path file Excel Juni BID AGA")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Terapkan perubahan. Tanpa flag ini hanya audit/read-only.",
    )
    args = parser.parse_args()

    path = Path(args.source).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    excel_risks = workbook_risks(workbook)
    profile = get_profile()
    tahun_buku = TahunBuku.objects.filter(tahun=YEAR).first()
    period = None
    if tahun_buku is not None:
        period = PeriodeLaporan.objects.filter(
            tahun_buku=tahun_buku,
            kode_periode=f"{YEAR}-{MONTH:02d}",
        ).first()

    print_audit(profile, excel_risks, period)
    if not args.apply:
        print("\nDRY RUN: tidak ada data yang diubah.")
        print("Jalankan ulang dengan --apply setelah hasil audit diperiksa.")
        return

    apply_repair(path, workbook, profile, excel_risks)


if __name__ == "__main__":
    main()
