#!/usr/bin/env python3
"""
Import laporan bulanan BID STRADA Juni dan Juli 2026 ke monthly_report.

Safety:
- standalone;
- default DRY RUN / rollback;
- --apply untuk commit;
- hanya menerima tepat 11 risiko STRADA;
- mapping source -> master memakai nomor item + validasi event ter-normalisasi;
- Juni memakai residual Q2, Juli memakai residual Q3;
- field quarter kosong dipertahankan kosong, tanpa fallback quarter sebelumnya;
- risiko kualitatif tidak dipaksa memiliki nilai dampak numerik;
- KRI aktual tidak diimpor karena workbook sumber tidak memiliki kolom realisasi KRI bulanan;
- III.D kosong; III.E berisi contoh legacy 2024 dan sengaja tidak diimpor sebagai loss event STRADA 2026.

Run:
  export DJANGO_SETTINGS_MODULE=riskproject.settings.prod
  python monthly_report/scripts/import_strada_jun_jul_2026.py --june ... --july ...

Apply setelah dry-run bersih:
  python monthly_report/scripts/import_strada_jun_jul_2026.py --june ... --july ... --apply
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

if not os.environ.get("DJANGO_SETTINGS_MODULE"):
    raise RuntimeError(
        "DJANGO_SETTINGS_MODULE belum diset. Untuk PROD jalankan: "
        "export DJANGO_SETTINGS_MODULE=riskproject.settings.prod"
    )

import django  # noqa: E402

django.setup()

from django.conf import settings  # noqa: E402
from django.contrib.auth import get_user_model  # noqa: E402
from django.db import transaction  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from masterdata.models import PeriodeLaporan, TahunBuku  # noqa: E402
from monthly_report.models import (  # noqa: E402
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from monthly_report.services import refresh_monthly_report_summary  # noqa: E402
from risk.models import (  # noqa: E402
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)

YEAR = 2026
EXPECTED_SOURCE_ITEMS = 11
MONTH_NAMES = {6: "Juni", 7: "Juli"}
SOURCE_PREFIX = "bid strada"


def col0(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


# STRADA workbook residual layout (III.A) — Q1..Q4.
IIIA_QUARTER_COLUMNS = {
    "realisasi_nilai_dampak": {1: col0("AO"), 2: col0("AP"), 3: col0("AQ"), 4: col0("AR")},
    "realisasi_skala_dampak": {1: col0("AS"), 2: col0("AT"), 3: col0("AU"), 4: col0("AV")},
    "realisasi_nilai_probabilitas": {1: col0("AW"), 2: col0("AX"), 3: col0("AY"), 4: col0("AZ")},
    "realisasi_skala_probabilitas": {1: col0("BA"), 2: col0("BB"), 3: col0("BC"), 4: col0("BD")},
    "realisasi_eksposur": {1: col0("BE"), 2: col0("BF"), 3: col0("BG"), 4: col0("BH")},
    "realisasi_skor_risiko": {1: col0("BI"), 2: col0("BJ"), 3: col0("BK"), 4: col0("BL")},
    "realisasi_level_risiko_bumn": {1: col0("BM"), 2: col0("BN"), 3: col0("BO"), 4: col0("BP")},
    "realisasi_level_risiko_kbumn": {1: col0("BQ"), 2: col0("BR"), 3: col0("BS"), 4: col0("BT")},
}
IIIA_ASSUMPTION_INDEX = col0("AC")
IIIA_EVENT_INDEX = col0("M")
IIIA_CODE_INDEX = col0("P")
IIIA_RISK_NO_INDEX = col0("L")
IIIA_CATEGORY_INDEX = col0("Z")

# III.B is used only for canonical row validation and to confirm treatment timeline.
IIIB_EVENT_INDEX = col0("M")
IIIB_CODE_INDEX = col0("P")
IIIB_RISK_NO_INDEX = col0("L")
IIIB_PLAN_INDEX = col0("CD")
IIIB_OUTPUT_INDEX = col0("CE")
IIIB_PIC_INDEX = col0("CJ")
IIIB_TIMELINE_FIRST = col0("CK")  # Jan; CL Feb ... CV Dec


@dataclass(frozen=True)
class ImportSpec:
    month: int
    path: Path


@dataclass
class MonthStats:
    month: int
    source_iiia: int = 0
    source_iiib: int = 0
    created_items: int = 0
    updated_items: int = 0
    field_changes: int = 0
    report_item_count: int = 0
    empty_quarter_fields: list[str] = field(default_factory=list)
    current_month_timeline_on: list[str] = field(default_factory=list)
    current_month_timeline_off: list[str] = field(default_factory=list)
    changes_count: int = 0
    loss_events_count: int = 0
    category_mismatches: list[str] = field(default_factory=list)


def normalize(value) -> str:
    text = str(value or "").casefold().replace("\xa0", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def text_or_none(value):
    if value is None:
        return None
    value = str(value).strip()
    if not value or normalize(value) in {"n a", "na", "none", "null"}:
        return None
    if str(value).strip().upper() in {"#NAME?", "#REF!", "#N/A", "#DIV/0!"}:
        return None
    return value


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        stripped = value.strip().casefold()
        if stripped in {"-", "n/a", "no data", "#div/0!", "#n/a", "#name?", "#ref!"}:
            return None
        multiplier = Decimal("1")
        if re.search(r"miliar|milyar", stripped):
            multiplier = Decimal("1000000000")
        elif "juta" in stripped:
            multiplier = Decimal("1000000")
        cleaned = re.sub(r"[^0-9,.\-]", "", stripped)
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        if not cleaned:
            return None
        value = cleaned
    else:
        multiplier = Decimal("1")
    try:
        return Decimal(str(value).strip()) * multiplier
    except (InvalidOperation, ValueError, TypeError):
        return None


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    if Decimal("0") <= number <= Decimal("100"):
        return number
    return None


def int_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    try:
        return int(number)
    except (ValueError, TypeError):
        return None


def value_at(row, zero_index):
    return row[zero_index] if len(row) > zero_index else None


def file_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def quarter_for_month(month):
    return ((month - 1) // 3) + 1


def source_risk_number(code):
    m = re.search(r"bid\s*strada\s*[- ]?(\d+)\s*[- ]?([a-z])?", normalize(code))
    return int(m.group(1)) if m else None


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.filter(is_active=True).order_by("id").first()
    )


def get_strada_profile():
    candidates = list(
        ReAssessmentSummary.objects.filter(tahun=YEAR)
        .select_related("unit_bisnis", "kontrak_manajemen")
        .order_by("id")
    )
    matched = [
        p for p in candidates
        if "strada" in normalize(p.judul)
        or (p.unit_bisnis_id and "strada" in normalize(p.unit_bisnis.name))
    ]
    if len(matched) != 1:
        raise RuntimeError(
            f"Profil STRADA {YEAR} harus tepat satu; ditemukan {len(matched)}: "
            + ", ".join(f"id={p.id} {p.judul}" for p in matched)
        )
    profile = matched[0]
    if profile.item.count() != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"Profil STRADA id={profile.id} harus memiliki {EXPECTED_SOURCE_ITEMS} item; "
            f"ditemukan {profile.item.count()}."
        )
    return profile


def build_master_map(profile):
    items = list(
        ReAssessmentItem.objects.filter(summary=profile)
        .select_related("kategori_dampak")
        .order_by("no_item", "no_risiko", "id")
    )
    mapping = {}
    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        found = [x for x in items if x.no_item == n]
        if len(found) != 1:
            raise RuntimeError(
                f"Master STRADA item {n} harus tepat satu; ditemukan {len(found)}: "
                + ", ".join(f"RE={x.id}/R{x.no_risiko} {x.peristiwa_risiko}" for x in found)
            )
        mapping[n] = found[0]
    if len({x.pk for x in mapping.values()}) != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError("Master mapping STRADA tidak menghasilkan 11 target unik.")
    return mapping


def is_qualitative_master(item):
    category = normalize(getattr(getattr(item, "kategori_dampak", None), "nama", ""))
    return "kual" in category


def find_start_row(ws):
    for row in range(1, min(ws.max_row, 80) + 1):
        if "start pengisian" in normalize(ws.cell(row, 1).value):
            return row
    raise RuntimeError(f"Anchor 'Start pengisian' tidak ditemukan pada {ws.title}.")


def parse_source_rows(workbook, sheet_name, stats, attr_name):
    ws = workbook[sheet_name]
    start = find_start_row(ws)
    parsed = {}
    code_idx = IIIA_CODE_INDEX if sheet_name == "III.A" else IIIB_CODE_INDEX
    event_idx = IIIA_EVENT_INDEX if sheet_name == "III.A" else IIIB_EVENT_INDEX
    risk_idx = IIIA_RISK_NO_INDEX if sheet_name == "III.A" else IIIB_RISK_NO_INDEX

    for row_no, cells in enumerate(
        ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True), start=start
    ):
        row = list(cells)
        code = value_at(row, code_idx)
        if SOURCE_PREFIX not in normalize(code):
            continue
        n = source_risk_number(code)
        if n is None or not 1 <= n <= EXPECTED_SOURCE_ITEMS:
            raise RuntimeError(f"{sheet_name} row {row_no}: kode STRADA tidak valid: {code!r}")
        row_risk = int_or_none(value_at(row, risk_idx))
        if row_risk is not None and row_risk != n:
            raise RuntimeError(
                f"{sheet_name} row {row_no}: nomor pada kolom L={row_risk} tidak sama dengan kode {code}."
            )
        if n in parsed:
            raise RuntimeError(f"{sheet_name}: risiko {n} duplikat (row {row_no}).")
        event = text_or_none(value_at(row, event_idx))
        if not event:
            raise RuntimeError(f"{sheet_name} row {row_no}: event kosong untuk {code}.")
        parsed[n] = (row_no, row)

    setattr(stats, attr_name, len(parsed))
    if len(parsed) != EXPECTED_SOURCE_ITEMS:
        missing = sorted(set(range(1, EXPECTED_SOURCE_ITEMS + 1)) - set(parsed))
        raise RuntimeError(
            f"{sheet_name} harus menghasilkan {EXPECTED_SOURCE_ITEMS} risiko; "
            f"ditemukan {len(parsed)}. Missing={missing}."
        )
    return parsed


def validate_source_against_master(master_map, parsed_iiia, parsed_iiib, month):
    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        master = master_map[n]
        row_a, vals_a = parsed_iiia[n]
        row_b, vals_b = parsed_iiib[n]
        event_a = normalize(value_at(vals_a, IIIA_EVENT_INDEX))
        event_b = normalize(value_at(vals_b, IIIB_EVENT_INDEX))
        event_m = normalize(master.peristiwa_risiko)
        if event_a != event_b:
            raise RuntimeError(
                f"{MONTH_NAMES[month]} risiko {n}: event III.A row {row_a} berbeda dengan III.B row {row_b}."
            )
        if event_a != event_m:
            raise RuntimeError(
                f"{MONTH_NAMES[month]} risiko {n}: source tidak sama dengan master RE={master.id}.\n"
                f"MASTER={master.peristiwa_risiko}\nSOURCE={value_at(vals_a, IIIA_EVENT_INDEX)}"
            )


def scale_by_level(model, value):
    number = int_or_none(value)
    if number is None or not 1 <= number <= 5:
        return None
    return model.objects.filter(urutan=number).first()


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(YEAR, month)
    period, _ = PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{YEAR}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {YEAR}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{YEAR}-{month:02d}-01",
            "tanggal_selesai": f"{YEAR}-{month:02d}-{last_day:02d}",
        },
    )
    return period


def get_or_create_report(profile, period, tahun_buku, prepared_by, month):
    report = MonthlyRiskReport.objects.filter(
        reassessment=profile, periode=period, versi=1
    ).first()
    created = False
    if report is None:
        report = MonthlyRiskReport.objects.create(
            reassessment=profile,
            periode=period,
            tahun_buku=tahun_buku,
            versi=1,
            kode=f"MRR-STRADA-{YEAR}-{month:02d}",
            status="draft",
            prepared_by=prepared_by,
        )
        created = True
    else:
        if report.status not in {"draft", "revision"}:
            raise RuntimeError(
                f"{report.kode or report.pk} berstatus {report.status}; hanya Draft/Revision yang boleh diimpor."
            )
        if getattr(report, "is_locked", False):
            raise RuntimeError(f"{report.kode or report.pk} sedang locked; import dibatalkan.")
        fields = []
        if not report.kode:
            report.kode = f"MRR-STRADA-{YEAR}-{month:02d}"
            fields.append("kode")
        if not report.prepared_by_id:
            report.prepared_by = prepared_by
            fields.append("prepared_by")
        if not report.tahun_buku_id:
            report.tahun_buku = tahun_buku
            fields.append("tahun_buku")
        if fields:
            report.save(update_fields=fields + ["updated_at"])
    return report, created


def _field_changed(old, new):
    if hasattr(old, "pk"):
        old = old.pk
    if hasattr(new, "pk"):
        new = new.pk
    return old != new


def _normalize_decimal_for_model_field(item, field_name, value):
    if value is None or not isinstance(value, Decimal):
        return value
    model_field = item._meta.get_field(field_name)
    decimal_places = getattr(model_field, "decimal_places", None)
    if decimal_places is None:
        return value
    quantum = Decimal("1").scaleb(-decimal_places)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def apply_fields(item, values, stats: MonthStats, *, overwrite_none_fields=()):
    changed_fields = []
    for field_name, new_value in values.items():
        new_value = _normalize_decimal_for_model_field(item, field_name, new_value)
        if new_value is None and field_name not in overwrite_none_fields:
            continue
        old_value = getattr(item, field_name)
        if _field_changed(old_value, new_value):
            setattr(item, field_name, new_value)
            changed_fields.append(field_name)
            stats.field_changes += 1
    if changed_fields:
        item.full_clean()
        item.save(update_fields=changed_fields + ["updated_at"])
        stats.updated_items += 1
    return changed_fields


def import_iiia(report, master_map, parsed, month, stats):
    quarter = quarter_for_month(month)
    overwrite_none = {
        "realisasi_nilai_dampak",
        "realisasi_skala_dampak",
        "realisasi_nilai_probabilitas",
        "realisasi_skala_probabilitas",
        "realisasi_eksposur",
        "realisasi_skor_risiko",
        "realisasi_level_risiko_bumn",
        "realisasi_level_risiko_kbumn",
    }

    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        master = master_map[n]
        row_no, row = parsed[n]
        item, created = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=master)
        if created:
            stats.created_items += 1

        source_category = normalize(value_at(row, IIIA_CATEGORY_INDEX))
        if "kuant" in source_category:
            qualitative = False
        elif "kual" in source_category:
            qualitative = True
        else:
            qualitative = is_qualitative_master(master)

        master_qualitative = is_qualitative_master(master)
        if qualitative != master_qualitative:
            stats.category_mismatches.append(
                f"R{n} RE={master.id}: master={getattr(getattr(master, 'kategori_dampak', None), 'nama', '-')} "
                f"| source={value_at(row, IIIA_CATEGORY_INDEX)} (monthly memakai kategori source; master tidak diubah)"
            )
        impact_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_nilai_dampak"][quarter])
        impact_scale_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_dampak"][quarter])
        prob_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_nilai_probabilitas"][quarter])
        prob_scale_raw = value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skala_probabilitas"][quarter])

        # For qualitative risks, 0 in the numeric impact columns is a template placeholder,
        # not an economic impact. Keep the numeric impact empty.
        impact_value = None if qualitative else decimal_or_none(impact_raw)
        impact_scale = scale_by_level(MasterSkalaDampak, impact_scale_raw)
        probability_value = percent_or_none(prob_raw)
        probability_scale = scale_by_level(MasterSkalaProbabilitas, prob_scale_raw)

        if impact_scale is None:
            stats.empty_quarter_fields.append(
                f"R{n} III.A row {row_no}: skala_dampak Q{quarter} kosong/tidak valid"
            )
        if probability_scale is None:
            stats.empty_quarter_fields.append(
                f"R{n} III.A row {row_no}: skala_probabilitas Q{quarter} kosong/tidak valid"
            )
        if not qualitative and impact_value is None:
            stats.empty_quarter_fields.append(
                f"R{n} III.A row {row_no}: nilai_dampak Q{quarter} kuantitatif kosong"
            )
        # Probability numeric is optional for qualitative according to current completeness rule.
        if not qualitative and probability_value is None:
            stats.empty_quarter_fields.append(
                f"R{n} III.A row {row_no}: nilai_probabilitas Q{quarter} kuantitatif kosong"
            )

        values = {
            "jenis_risiko": "kualitatif" if qualitative else "kuantitatif",
            "realisasi_asumsi_dampak": text_or_none(value_at(row, IIIA_ASSUMPTION_INDEX)),
            "realisasi_nilai_dampak": impact_value,
            "realisasi_skala_dampak": impact_scale,
            "realisasi_nilai_probabilitas": probability_value,
            "realisasi_skala_probabilitas": probability_scale,
            "realisasi_eksposur": decimal_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_eksposur"][quarter])
            ),
            "realisasi_skor_risiko": int_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_skor_risiko"][quarter])
            ),
            "realisasi_level_risiko_bumn": text_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_level_risiko_bumn"][quarter])
            ),
            "realisasi_level_risiko_kbumn": text_or_none(
                value_at(row, IIIA_QUARTER_COLUMNS["realisasi_level_risiko_kbumn"][quarter])
            ),
        }
        apply_fields(item, values, stats, overwrite_none_fields=overwrite_none)


def inspect_iiib_timeline(report, master_map, parsed, month, stats):
    # STRADA source has a planned implementation timeline (CK:CV), not an actual-progress field.
    # We do not convert that plan into actual progress. We only report it for audit visibility.
    timeline_idx = IIIB_TIMELINE_FIRST + (month - 1)
    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        master = master_map[n]
        row_no, row = parsed[n]
        MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=master)
        flag = decimal_or_none(value_at(row, timeline_idx))
        if flag is not None and flag != 0:
            stats.current_month_timeline_on.append(f"R{n} (III.B row {row_no})")
        else:
            stats.current_month_timeline_off.append(f"R{n} (III.B row {row_no})")


def import_changes(workbook, report):
    if "III.D" not in workbook.sheetnames:
        return 0
    ws = workbook["III.D"]
    start = find_start_row(ws)
    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }
    rows = []
    for cells in ws.iter_rows(min_row=start, values_only=True):
        row = list(cells)
        change_type = mapping.get(normalize(value_at(row, 1)))
        if change_type:
            rows.append((change_type, row))
    MonthlyRiskReportChange.objects.filter(report=report).delete()
    for change_type, row in rows:
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=value_at(row, 2),
            penjelasan=value_at(row, 3),
        )
    return len(rows)


def clear_loss_events_as_legacy_sample(workbook, report):
    # Both supplied STRADA files contain the same 2024 gas/currency example in III.E.
    # It is not a STRADA 2026 loss event; do not import it.
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()
    if "III.E" not in workbook.sheetnames:
        return 0
    ws = workbook["III.E"]
    try:
        start = find_start_row(ws)
    except RuntimeError:
        return 0
    sample_text = " ".join(
        str(ws.cell(start, c).value or "") for c in range(1, min(ws.max_column, 21) + 1)
    )
    if sample_text.strip():
        if "2024" not in sample_text and "harga gas" not in normalize(sample_text):
            raise RuntimeError(
                "III.E berisi data non-kosong yang tidak dikenali sebagai sample legacy 2024. "
                "Import dihentikan agar loss event tidak terlewat."
            )
    return 0


def create_sqlite_backup_if_possible():
    db = settings.DATABASES.get("default", {})
    if db.get("ENGINE") != "django.db.backends.sqlite3":
        print("BACKUP: database bukan SQLite; backup otomatis file tidak dibuat.")
        return None
    source = Path(db.get("NAME"))
    if not source.exists():
        print(f"BACKUP: file SQLite tidak ditemukan: {source}")
        return None
    backup_dir = PROJECT_ROOT / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"db_before_strada_jun_jul_2026_{stamp}.sqlite3"
    shutil.copy2(source, target)
    print(f"BACKUP: {target}")
    return target


def process_month(spec, profile, master_map, tahun_buku, prepared_by):
    stats = MonthStats(month=spec.month)
    workbook = load_workbook(spec.path, data_only=True, read_only=False, keep_links=False)
    required = {"III.A", "III.B"}
    missing = required - set(workbook.sheetnames)
    if missing:
        raise RuntimeError(f"{spec.path.name}: sheet wajib tidak ada: {sorted(missing)}")

    period = get_period(tahun_buku, spec.month)
    report, created = get_or_create_report(profile, period, tahun_buku, prepared_by, spec.month)
    parsed_a = parse_source_rows(workbook, "III.A", stats, "source_iiia")
    parsed_b = parse_source_rows(workbook, "III.B", stats, "source_iiib")
    validate_source_against_master(master_map, parsed_a, parsed_b, spec.month)

    import_iiia(report, master_map, parsed_a, spec.month, stats)
    inspect_iiib_timeline(report, master_map, parsed_b, spec.month, stats)
    stats.changes_count = import_changes(workbook, report)
    stats.loss_events_count = clear_loss_events_as_legacy_sample(workbook, report)

    refresh_monthly_report_summary(report)
    report.refresh_from_db()
    stats.report_item_count = report.items.count()
    if stats.report_item_count != EXPECTED_SOURCE_ITEMS:
        raise RuntimeError(
            f"{report.kode}: item report harus {EXPECTED_SOURCE_ITEMS}, ditemukan {stats.report_item_count}."
        )
    return report, created, stats


def print_master_map(master_map):
    print("MASTER MAPPING STRADA:")
    for n in range(1, EXPECTED_SOURCE_ITEMS + 1):
        item = master_map[n]
        cat = getattr(getattr(item, "kategori_dampak", None), "nama", "-")
        print(
            f"- R{n:02d} -> RE={item.id:>4} | item={item.no_item} | risk={item.no_risiko} | "
            f"cause={item.no_penyebab_risiko or '-'} | {cat} | {item.peristiwa_risiko}"
        )


def print_stats(report, created, stats):
    q = quarter_for_month(stats.month)
    print(f"\n{MONTH_NAMES[stats.month].upper()} {YEAR} — {report.kode}")
    print(f"- Report: {'baru' if created else 'existing'} | items={stats.report_item_count}")
    print(f"- Source III.A: {stats.source_iiia}/{EXPECTED_SOURCE_ITEMS}")
    print(f"- Source III.B: {stats.source_iiib}/{EXPECTED_SOURCE_ITEMS}")
    print(f"- Residual source: Q{q}")
    print(f"- Item dibuat: {stats.created_items}")
    print(f"- Item tersentuh/update: {stats.updated_items}; perubahan field: {stats.field_changes}")
    print(f"- III.D changes: {stats.changes_count}")
    print("- III.E loss events: 0 (sample legacy 2024 sengaja tidak diimpor)")
    print("- KRI aktual bulanan: tidak tersedia eksplisit di workbook; tidak diisi/fallback")
    print("- Realisasi perlakuan/progress aktual: tidak tersedia eksplisit; timeline III.B hanya dipakai sebagai audit plan")
    print(
        f"- Timeline rencana bulan berjalan = 1: {len(stats.current_month_timeline_on)} | "
        f"kosong/0: {len(stats.current_month_timeline_off)}"
    )
    if stats.current_month_timeline_off:
        print("  Timeline kosong/0: " + ", ".join(stats.current_month_timeline_off))
    if stats.category_mismatches:
        print("- Perbedaan kategori dampak master vs source (master TIDAK diubah):")
        for msg in stats.category_mismatches:
            print(f"  * {msg}")
    if stats.empty_quarter_fields:
        print("- Field residual Q{} kosong/tidak valid (dipertahankan kosong; tanpa fallback):".format(q))
        for msg in stats.empty_quarter_fields:
            print(f"  * {msg}")


def main():
    parser = argparse.ArgumentParser(description="Import STRADA Juni & Juli 2026")
    parser.add_argument("--june", required=True, type=Path, help="Path XLSX laporan Juni STRADA")
    parser.add_argument("--july", required=True, type=Path, help="Path XLSX laporan Juli STRADA")
    parser.add_argument("--apply", action="store_true", help="Commit perubahan. Default DRY RUN/rollback.")
    args = parser.parse_args()

    specs = [
        ImportSpec(6, args.june.expanduser().resolve()),
        ImportSpec(7, args.july.expanduser().resolve()),
    ]
    for spec in specs:
        if not spec.path.exists():
            raise FileNotFoundError(spec.path)

    mode = "APPLY" if args.apply else "DRY RUN"
    print(f"STRADA JUNI–JULI {YEAR} | MODE={mode}")
    for spec in specs:
        print(
            f"SOURCE {MONTH_NAMES[spec.month]}: {spec.path} | size={spec.path.stat().st_size} | "
            f"sha256={file_sha256(spec.path)}"
        )

    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user aktif untuk prepared_by.")
    profile = get_strada_profile()
    master_map = build_master_map(profile)
    print(f"PROFILE: id={profile.id} | {profile.judul} | unit={profile.unit_bisnis} | master={profile.item.count()}")
    print_master_map(master_map)

    if args.apply:
        create_sqlite_backup_if_possible()

    results = []
    with transaction.atomic():
        tahun_buku, _ = TahunBuku.objects.get_or_create(tahun=YEAR, defaults={"aktif": True})
        for spec in specs:
            results.append(process_month(spec, profile, master_map, tahun_buku, prepared_by))
        if not args.apply:
            transaction.set_rollback(True)

    for result in results:
        print_stats(*result)

    if args.apply:
        print("\nRESULT: APPLY BERHASIL — transaksi Juni dan Juli sudah di-commit.")
    else:
        print("\nRESULT: DRY RUN BERHASIL — database TIDAK berubah (rollback).")
        print("Jika mapping/count sudah benar, ulangi command yang sama dengan --apply.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
