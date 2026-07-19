import calendar
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.db import transaction

from masterdata.models import PeriodeLaporan, TahunBuku
from monthly_report.models import (
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from risk.models import (
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)


BASE_DIR = Path("/Users/armeizir/Downloads/2026/7. Bid. Bis")
REASSESSMENT_ID = 2


@dataclass(frozen=True)
class ImportFile:
    month: int
    path: Path


FILES = [
    ImportFile(3, BASE_DIR / "Mitigasi Risiko dan KRI BIS Maret 2026.xlsx"),
    ImportFile(4, BASE_DIR / "Mitigasi Risiko dan KRI BIS April 2026.xlsx"),
]

MONTH_NAMES = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}

CODE_PREFIX_EVENT = {
    "bid bis 1": "Klausul HoA yang multitafsir",
    "bid bis 2": "Kelayakan teknis proyek belum tervalidasi",
    "bid bis 3": "Parameter komersial proyek belum matang",
    "bid bis 4": "Kapabilitas mitra tidak memadai",
    "bid bis 1 e": "Tidak tercapainya penyerapan anggaran investasi dan anggaran kas investasi sesuai peruntukannya",
    "bid bis 1 f": "Tidak terpenuhinya Nilai Hasil Assesment Manajemen SDM (HCR-OCR & Produktivitas Pegawai)",
    "bid bis 1 g": "Tidak terpenuhinya aspek-aspek compliance",
}


def normalize(value):
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.strip().lower() in {
        "-",
        "n/a",
        "no data",
        "#div/0!",
    }:
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    return number


def scale_by_level(model, value):
    number = decimal_or_none(value)
    if number is None:
        return None
    return model.objects.filter(urutan=int(number)).first()


def treatment_effectiveness(value):
    text = normalize(value)
    if not text:
        return None
    if "tidak" in text:
        return "tidak_efektif"
    if "cukup" in text:
        return "cukup_efektif"
    if "efektif" in text:
        return "efektif"
    return None


def treatment_status(value):
    text = normalize(value)
    if not text:
        return None
    if "discontinue" in text:
        return "discontinue"
    if "continue" in text:
        return "continue"
    return None


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.order_by("id").first()
    )


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(tahun_buku.tahun, month)
    return PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{tahun_buku.tahun}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {tahun_buku.tahun}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{tahun_buku.tahun}-{month:02d}-01",
            "tanggal_selesai": f"{tahun_buku.tahun}-{month:02d}-{last_day:02d}",
        },
    )[0]


def build_event_map(reassessment):
    result = {}
    for item in ReAssessmentItem.objects.filter(summary=reassessment).order_by(
        "no_item",
        "no_risiko",
        "id",
    ):
        result.setdefault(normalize(item.peristiwa_risiko), item)
    return result


def code_based_event(row):
    code_values = [
        row[1] if len(row) > 1 else None,
        row[5] if len(row) > 5 else None,
    ]
    for value in code_values:
        code = normalize(value)
        if not code:
            continue
        if code in {"bid bis 1 e", "bid bis 1 f", "bid bis 1 g"}:
            return CODE_PREFIX_EVENT[code]
        if code.startswith("bid bis 1"):
            return CODE_PREFIX_EVENT["bid bis 1"]
        if code.startswith("bid bis 2"):
            return CODE_PREFIX_EVENT["bid bis 2"]
        if code.startswith("bid bis 3"):
            return CODE_PREFIX_EVENT["bid bis 3"]
        if code.startswith("bid bis 4"):
            return CODE_PREFIX_EVENT["bid bis 4"]
    return None


def strip_project_suffix(event_name):
    text = str(event_name or "").strip()
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    return text


def resolve_risk_event(row, event_map):
    candidates = []
    code_event = code_based_event(row)
    if code_event:
        candidates.append(code_event)
    event_name = row[2] if len(row) > 2 else None
    candidates.append(strip_project_suffix(event_name))
    candidates.append(event_name)

    for candidate in candidates:
        normalized = normalize(candidate)
        if normalized in event_map:
            return event_map[normalized]
    return None


def read_rows(ws, start_row=1):
    for offset, row in enumerate(
        ws.iter_rows(min_row=start_row, values_only=True),
        start=start_row,
    ):
        yield offset, list(row)


def selected_quarter(month):
    return ((month - 1) // 3) + 1


def import_sheet_iiia(workbook, report, event_map, month, skipped):
    ws = workbook["III.A"]
    q = selected_quarter(month)
    cols = {
        "nilai_dampak": {1: 13, 2: 14, 3: 15, 4: 16}[q],
        "skala_dampak": {1: 17, 2: 18, 3: 19, 4: 20}[q],
        "nilai_prob": {1: 25, 2: 26, 3: 27, 4: 28}[q],
        "skala_prob": {1: 29, 2: 30, 3: 31, 4: 32}[q],
    }
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        code = row[1] if len(row) > 1 else None
        event_name = row[2] if len(row) > 2 else None
        if not event_name or not code:
            continue

        risk_event = resolve_risk_event(row, event_map)
        if not risk_event:
            if "bid bis" in normalize(code):
                skipped.append(
                    {
                        "sheet": "III.A",
                        "row": row_idx,
                        "no": code,
                        "event": event_name,
                        "reason": "Peristiwa risiko tidak ditemukan di Profil Risiko BIS.",
                    }
                )
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=risk_event,
        )
        item.realisasi_asumsi_dampak = row[11] if len(row) > 11 else None
        item.realisasi_nilai_dampak = decimal_or_none(row[cols["nilai_dampak"] - 1])
        item.realisasi_skala_dampak = scale_by_level(
            MasterSkalaDampak,
            row[cols["skala_dampak"] - 1],
        )
        item.realisasi_nilai_probabilitas = percent_or_none(row[cols["nilai_prob"] - 1])
        item.realisasi_skala_probabilitas = scale_by_level(
            MasterSkalaProbabilitas,
            row[cols["skala_prob"] - 1],
        )
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(
            row[56] if len(row) > 56 else None
        )
        item.save()
        imported += 1
    return imported


def import_sheet_iiib(workbook, report, event_map, month, skipped):
    ws = workbook["III.B"]
    q = selected_quarter(month)
    threshold_col = 39 + ((month - 1) * 2)
    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[q]
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        if not any(row[idx] if len(row) > idx else None for idx in (1, 2, 5)):
            continue

        risk_event = resolve_risk_event(row, event_map)
        code = row[1] if len(row) > 1 else row[5] if len(row) > 5 else None
        event_name = row[2] if len(row) > 2 else None
        if not risk_event:
            if "bid bis" in normalize(code) or "bis" in normalize(row[5] if len(row) > 5 else None):
                skipped.append(
                    {
                        "sheet": "III.B",
                        "row": row_idx,
                        "no": code,
                        "event": event_name,
                        "reason": "Peristiwa risiko tidak ditemukan di Profil Risiko BIS.",
                    }
                )
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=risk_event,
        )
        item.realisasi_rencana_perlakuan = row[10] if len(row) > 10 else None
        item.realisasi_output_perlakuan = row[11] if len(row) > 11 else None
        budget = decimal_or_none(row[9] if len(row) > 9 else None)
        if budget is not None:
            risk_event.biaya_perlakuan_risiko = budget
            risk_event.save(update_fields=["biaya_perlakuan_risiko"])
        item.realisasi_biaya_perlakuan = decimal_or_none(row[12] if len(row) > 12 else None)
        item.realisasi_pic = row[14] if len(row) > 14 else None
        item.status_rencana_perlakuan = treatment_status(row[27] if len(row) > 27 else None)
        item.penjelasan_status_rencana = row[28] if len(row) > 28 else None
        item.progress_pelaksanaan_percent = percent_or_none(
            row[progress_col - 1] if len(row) >= progress_col else None
        )
        item.realisasi_threshold_kri = (
            row[threshold_col - 1] if len(row) >= threshold_col else None
        )
        threshold_score = row[threshold_col] if len(row) > threshold_col else None
        item.realisasi_threshold_kri_skor = (
            str(threshold_score) if threshold_score not in (None, "") else None
        )
        item.save()
        imported += 1
    return imported


def import_sheet_iiid(workbook, report):
    ws = workbook["III.D"]
    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }
    MonthlyRiskReportChange.objects.filter(report=report).delete()
    imported = 0
    for _, row in read_rows(ws, 8):
        change_type_text = row[1] if len(row) > 1 else None
        impacted = row[2] if len(row) > 2 else None
        explanation = row[3] if len(row) > 3 else None
        if not any([change_type_text, impacted, explanation]):
            continue
        change_type = mapping.get(normalize(change_type_text))
        if not change_type:
            continue
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=impacted,
            penjelasan=explanation,
        )
        imported += 1
    return imported


def import_sheet_iiie(workbook, report, event_map, skipped):
    ws = workbook["III.E"]
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()
    imported = 0
    for row_idx, row in read_rows(ws, 8):
        name = row[1] if len(row) > 1 else None
        normalized_name = normalize(name)
        if not name or "belum ada laporan loss event" in normalized_name:
            continue
        if normalized_name not in event_map:
            skipped.append(
                {
                    "sheet": "III.E",
                    "row": row_idx,
                    "event": name,
                    "reason": "Tidak diimpor otomatis karena tidak terkait item risiko BIS.",
                }
            )
            continue
        MonthlyRiskReportLossEvent.objects.create(
            report=report,
            nama_kejadian=name,
            identifikasi_kejadian=row[2] if len(row) > 2 else None,
            kategori_kejadian=row[3] if len(row) > 3 else None,
            sumber_penyebab_kejadian="external"
            if "eksternal" in normalize(row[4] if len(row) > 4 else None)
            else "internal"
            if "internal" in normalize(row[4] if len(row) > 4 else None)
            else None,
            penyebab_kejadian=row[5] if len(row) > 5 else None,
            penanganan_saat_kejadian=row[6] if len(row) > 6 else None,
            deskripsi_kejadian_risk_event=row[7] if len(row) > 7 else None,
            kategori_risiko_bumn=row[8] if len(row) > 8 else None,
            kategori_risiko_t2_t3_kbumn=row[9] if len(row) > 9 else None,
            penjelasan_kerugian=row[10] if len(row) > 10 else None,
            nilai_kerugian=decimal_or_none(row[11] if len(row) > 11 else None),
            kejadian_berulang="ya"
            if "ya" in normalize(row[12] if len(row) > 12 else None)
            else "tidak"
            if "tidak" in normalize(row[12] if len(row) > 12 else None)
            else None,
            frekuensi_kejadian=row[13] if len(row) > 13 else None,
            mitigasi_direncanakan=row[14] if len(row) > 14 else None,
            realisasi_mitigasi=row[15] if len(row) > 15 else None,
            perbaikan_mendatang=row[16] if len(row) > 16 else None,
            pihak_terkait=row[17] if len(row) > 17 else None,
            status_asuransi="ya"
            if "ya" in normalize(row[18] if len(row) > 18 else None)
            else "tidak"
            if "tidak" in normalize(row[18] if len(row) > 18 else None)
            else None,
            nilai_premi=decimal_or_none(row[19] if len(row) > 19 else None),
            nilai_klaim=decimal_or_none(row[20] if len(row) > 20 else None),
        )
        imported += 1
    return imported


def run():
    from openpyxl import load_workbook

    reassessment = ReAssessmentSummary.objects.get(pk=REASSESSMENT_ID)
    tahun_buku, _ = TahunBuku.objects.get_or_create(
        tahun=reassessment.tahun,
        defaults={"aktif": True},
    )
    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user untuk prepared_by.")

    event_map = build_event_map(reassessment)
    summary = []
    all_skipped = []

    with transaction.atomic():
        for import_file in FILES:
            if not import_file.path.exists():
                raise FileNotFoundError(import_file.path)

            workbook = load_workbook(
                import_file.path,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
            period = get_period(tahun_buku, import_file.month)
            report, _ = MonthlyRiskReport.objects.update_or_create(
                reassessment=reassessment,
                periode=period,
                versi=1,
                defaults={
                    "kode": f"MRR-BIS-{tahun_buku.tahun}-{import_file.month:02d}",
                    "tahun_buku": tahun_buku,
                    "status": "draft",
                    "prepared_by": prepared_by,
                },
            )

            skipped = []
            iiia_count = import_sheet_iiia(
                workbook,
                report,
                event_map,
                import_file.month,
                skipped,
            )
            iiib_count = import_sheet_iiib(
                workbook,
                report,
                event_map,
                import_file.month,
                skipped,
            )
            iiid_count = import_sheet_iiid(workbook, report)
            iiie_count = import_sheet_iiie(workbook, report, event_map, skipped)

            report.total_risiko = report.items.count()
            report.total_high = report.items.filter(
                realisasi_level_risiko__icontains="tinggi",
            ).count()
            report.total_mitigasi_terlambat = report.items.filter(
                mitigation_status="delayed",
            ).count()
            report.total_selesai = report.items.filter(
                status_rencana_perlakuan="discontinue",
            ).count()
            report.save()

            summary.append(
                {
                    "month": MONTH_NAMES[import_file.month],
                    "report_id": report.id,
                    "items": report.items.count(),
                    "iiia": iiia_count,
                    "iiib": iiib_count,
                    "iiid": iiid_count,
                    "iiie": iiie_count,
                    "skipped": len(skipped),
                }
            )
            all_skipped.extend(
                {"month": MONTH_NAMES[import_file.month], **item}
                for item in skipped
            )

    print("IMPORT SUMMARY")
    for item in summary:
        print(
            f"- {item['month']}: report #{item['report_id']}, "
            f"items {item['items']}, III.A {item['iiia']}, "
            f"III.B {item['iiib']}, III.D {item['iiid']}, "
            f"III.E {item['iiie']}, skipped {item['skipped']}"
        )

    print("\nSKIPPED")
    for item in all_skipped:
        label = f"{item['month']} {item['sheet']} row {item['row']}"
        event = item.get("event")
        no = item.get("no")
        if no is not None:
            label += f" no {no}"
        print(f"- {label}: {event} -> {item['reason']}")


if __name__ == "__main__":
    run()
