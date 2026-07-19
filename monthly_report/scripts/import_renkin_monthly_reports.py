import calendar
import re
import tempfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from django.contrib.auth import get_user_model
from django.db import transaction

from masterdata.models import PeriodeLaporan, TahunBuku
from monthly_report.models import (
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from risk.models import (
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)


BASE_DIR = Path("/Users/armeizir/Downloads/2026/8. Bid. Renkin")
REASSESSMENT_ID = 10


@dataclass(frozen=True)
class ImportFile:
    month: int
    path: Path


FILES = [
    ImportFile(3, BASE_DIR / "Laporan MR BID RENKIN Maret 2026.xlsx"),
    ImportFile(4, BASE_DIR / "Laporan MR BID RENKIN April 2026.xlsx"),
]

MONTH_NAMES = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


def sanitize_workbook(path):
    src = Path(path)
    dst = Path(tempfile.gettempdir()) / f"{src.stem}.sanitized.xlsx"
    with ZipFile(src) as zin, ZipFile(dst, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/workbook.xml":
                text = data.decode("utf-8", errors="replace")
                text = re.sub(r"<definedNames>.*?</definedNames>", "", text, flags=re.S)
                data = text.encode("utf-8")
            zout.writestr(info, data)
    return dst


def normalize(value):
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.strip().lower() in {"-", "n/a", "no data"}:
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    return number


def scale_by_level(model, value):
    number = decimal_or_none(value)
    if number is None:
        return None
    return model.objects.filter(urutan=int(number)).first()


def treatment_effectiveness(value):
    text = normalize(value)
    if not text:
        return None
    if "tidak" in text:
        return "tidak_efektif"
    if "cukup" in text:
        return "cukup_efektif"
    if "efektif" in text:
        return "efektif"
    return None


def treatment_status(value):
    text = normalize(value)
    if not text:
        return None
    if "discontinue" in text:
        return "discontinue"
    if "continue" in text:
        return "continue"
    return None


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.order_by("id").first()
    )


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(tahun_buku.tahun, month)
    return PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{tahun_buku.tahun}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {tahun_buku.tahun}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{tahun_buku.tahun}-{month:02d}-01",
            "tanggal_selesai": f"{tahun_buku.tahun}-{month:02d}-{last_day:02d}",
        },
    )[0]


def number_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    return int(number)


def build_event_maps(reassessment):
    by_name = {}
    by_risk_number = {}
    for item in ReAssessmentItem.objects.filter(summary=reassessment).order_by("no_risiko", "id"):
        by_name.setdefault(normalize(item.peristiwa_risiko), []).append(item)
        if item.no_risiko is not None:
            by_risk_number[int(item.no_risiko)] = item
    return by_name, by_risk_number


def resolve_risk_event(row, by_name, by_risk_number):
    number = number_or_none(row[1] if len(row) > 1 else None)
    event_name = normalize(row[2] if len(row) > 2 else None)
    candidates = by_name.get(event_name, [])
    if number is not None and candidates:
        same_item = [item for item in candidates if item.no_item == number]
        if same_item:
            return same_item[0]
    if len(candidates) == 1:
        return candidates[0]
    if number in by_risk_number and normalize(by_risk_number[number].peristiwa_risiko) == event_name:
        return by_risk_number[number]
    return candidates[0] if candidates else None


def read_rows(ws, start_row=1):
    for offset, row in enumerate(
        ws.iter_rows(min_row=start_row, values_only=True),
        start=start_row,
    ):
        yield offset, list(row)


def selected_quarter(month):
    return ((month - 1) // 3) + 1


def import_sheet_iiia(workbook, report, by_name, by_risk_number, month, skipped):
    ws = workbook["III.A"]
    q = selected_quarter(month)
    cols = {
        "nilai_dampak": {1: 15, 2: 16, 3: 17, 4: 18}[q],
        "skala_dampak": {1: 19, 2: 20, 3: 21, 4: 22}[q],
        "nilai_prob": {1: 27, 2: 28, 3: 29, 4: 30}[q],
        "skala_prob": {1: 31, 2: 32, 3: 33, 4: 34}[q],
    }
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        no_risiko = row[1] if len(row) > 1 else None
        event_name = row[2] if len(row) > 2 else None
        if not isinstance(no_risiko, (int, float)) or not event_name:
            continue

        risk_event = resolve_risk_event(row, by_name, by_risk_number)
        if not risk_event:
            skipped.append(
                {
                    "sheet": "III.A",
                    "row": row_idx,
                    "no": no_risiko,
                    "event": event_name,
                    "reason": "Peristiwa risiko tidak ditemukan di Profil Risiko RENKIN.",
                }
            )
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=risk_event,
        )
        item.realisasi_asumsi_dampak = row[13] if len(row) > 13 else None
        item.realisasi_nilai_dampak = decimal_or_none(row[cols["nilai_dampak"] - 1])
        item.realisasi_skala_dampak = scale_by_level(
            MasterSkalaDampak,
            row[cols["skala_dampak"] - 1],
        )
        item.realisasi_nilai_probabilitas = percent_or_none(row[cols["nilai_prob"] - 1])
        item.realisasi_skala_probabilitas = scale_by_level(
            MasterSkalaProbabilitas,
            row[cols["skala_prob"] - 1],
        )
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(
            row[58] if len(row) > 58 else None
        )
        item.save()
        imported += 1
    return imported


def import_sheet_iiib(workbook, report, by_name, by_risk_number, month, skipped):
    ws = workbook["III.B"]
    q = selected_quarter(month)
    threshold_col = 39 + ((month - 1) * 2)
    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[q]
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        no_risiko = row[1] if len(row) > 1 else None
        event_name = row[2] if len(row) > 2 else None
        if not isinstance(no_risiko, (int, float)) or not event_name:
            continue

        risk_event = resolve_risk_event(row, by_name, by_risk_number)
        if not risk_event:
            skipped.append(
                {
                    "sheet": "III.B",
                    "row": row_idx,
                    "no": no_risiko,
                    "event": event_name,
                    "reason": "Peristiwa risiko tidak ditemukan di Profil Risiko RENKIN.",
                }
            )
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(
            report=report,
            risk_event=risk_event,
        )
        item.realisasi_rencana_perlakuan = row[10] if len(row) > 10 else None
        item.realisasi_output_perlakuan = row[11] if len(row) > 11 else None
        budget = decimal_or_none(row[9] if len(row) > 9 else None)
        if budget is not None:
            risk_event.biaya_perlakuan_risiko = budget
            risk_event.save(update_fields=["biaya_perlakuan_risiko"])
        item.realisasi_biaya_perlakuan = decimal_or_none(row[12] if len(row) > 12 else None)
        item.realisasi_pic = row[14] if len(row) > 14 else None
        item.status_rencana_perlakuan = treatment_status(row[27] if len(row) > 27 else None)
        item.penjelasan_status_rencana = row[28] if len(row) > 28 else None
        item.progress_pelaksanaan_percent = percent_or_none(
            row[progress_col - 1] if len(row) >= progress_col else None
        )
        item.realisasi_threshold_kri = (
            row[threshold_col - 1] if len(row) >= threshold_col else None
        )
        threshold_score = row[threshold_col] if len(row) > threshold_col else None
        item.realisasi_threshold_kri_skor = (
            str(threshold_score) if threshold_score not in (None, "") else None
        )
        item.save()
        imported += 1
    return imported


def import_sheet_iiid(workbook, report):
    ws = workbook["III.D"]
    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }
    MonthlyRiskReportChange.objects.filter(report=report).delete()
    imported = 0
    for _, row in read_rows(ws, 8):
        change_type_text = row[1] if len(row) > 1 else None
        impacted = row[2] if len(row) > 2 else None
        explanation = row[3] if len(row) > 3 else None
        if not any([change_type_text, impacted, explanation]):
            continue
        change_type = mapping.get(normalize(change_type_text))
        if not change_type:
            continue
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=impacted,
            penjelasan=explanation,
        )
        imported += 1
    return imported


def import_sheet_iiie(workbook, report, by_name, skipped):
    ws = workbook["III.E"]
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()
    imported = 0
    for row_idx, row in read_rows(ws, 8):
        name = row[1] if len(row) > 1 else None
        normalized_name = normalize(name)
        if not name:
            continue
        if normalized_name not in by_name:
            skipped.append(
                {
                    "sheet": "III.E",
                    "row": row_idx,
                    "event": name,
                    "reason": "Tidak diimpor otomatis karena tidak terkait item risiko RENKIN.",
                }
            )
            continue
        MonthlyRiskReportLossEvent.objects.create(
            report=report,
            nama_kejadian=name,
            identifikasi_kejadian=row[2] if len(row) > 2 else None,
            kategori_kejadian=row[3] if len(row) > 3 else None,
            sumber_penyebab_kejadian="external"
            if "eksternal" in normalize(row[4] if len(row) > 4 else None)
            else "internal"
            if "internal" in normalize(row[4] if len(row) > 4 else None)
            else None,
            penyebab_kejadian=row[5] if len(row) > 5 else None,
            penanganan_saat_kejadian=row[6] if len(row) > 6 else None,
            deskripsi_kejadian_risk_event=row[7] if len(row) > 7 else None,
            kategori_risiko_bumn=row[8] if len(row) > 8 else None,
            kategori_risiko_t2_t3_kbumn=row[9] if len(row) > 9 else None,
            penjelasan_kerugian=row[10] if len(row) > 10 else None,
            nilai_kerugian=decimal_or_none(row[11] if len(row) > 11 else None),
            kejadian_berulang="ya"
            if "ya" in normalize(row[12] if len(row) > 12 else None)
            else "tidak"
            if "tidak" in normalize(row[12] if len(row) > 12 else None)
            else None,
            frekuensi_kejadian=row[13] if len(row) > 13 else None,
            mitigasi_direncanakan=row[14] if len(row) > 14 else None,
            realisasi_mitigasi=row[15] if len(row) > 15 else None,
            perbaikan_mendatang=row[16] if len(row) > 16 else None,
            pihak_terkait=row[17] if len(row) > 17 else None,
            status_asuransi="ya"
            if "ya" in normalize(row[18] if len(row) > 18 else None)
            else "tidak"
            if "tidak" in normalize(row[18] if len(row) > 18 else None)
            else None,
            nilai_premi=decimal_or_none(row[19] if len(row) > 19 else None),
            nilai_klaim=decimal_or_none(row[20] if len(row) > 20 else None),
        )
        imported += 1
    return imported


def run():
    from openpyxl import load_workbook

    reassessment = ReAssessmentSummary.objects.get(pk=REASSESSMENT_ID)
    tahun_buku, _ = TahunBuku.objects.get_or_create(
        tahun=reassessment.tahun,
        defaults={"aktif": True},
    )
    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user untuk prepared_by.")

    by_name, by_risk_number = build_event_maps(reassessment)
    summary = []
    all_skipped = []

    with transaction.atomic():
        for import_file in FILES:
            if not import_file.path.exists():
                raise FileNotFoundError(import_file.path)

            workbook = load_workbook(
                sanitize_workbook(import_file.path),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
            period = get_period(tahun_buku, import_file.month)
            report, _ = MonthlyRiskReport.objects.update_or_create(
                reassessment=reassessment,
                periode=period,
                versi=1,
                defaults={
                    "kode": f"MRR-RENKIN-{tahun_buku.tahun}-{import_file.month:02d}",
                    "tahun_buku": tahun_buku,
                    "status": "draft",
                    "prepared_by": prepared_by,
                },
            )

            skipped = []
            iiia_count = import_sheet_iiia(
                workbook,
                report,
                by_name,
                by_risk_number,
                import_file.month,
                skipped,
            )
            iiib_count = import_sheet_iiib(
                workbook,
                report,
                by_name,
                by_risk_number,
                import_file.month,
                skipped,
            )
            iiid_count = import_sheet_iiid(workbook, report)
            iiie_count = import_sheet_iiie(workbook, report, by_name, skipped)

            report.total_risiko = report.items.count()
            report.total_high = report.items.filter(
                realisasi_level_risiko__icontains="tinggi",
            ).count()
            report.total_mitigasi_terlambat = report.items.filter(
                mitigation_status="delayed",
            ).count()
            report.total_selesai = report.items.filter(
                status_rencana_perlakuan="discontinue",
            ).count()
            report.save()

            summary.append(
                {
                    "month": MONTH_NAMES[import_file.month],
                    "report_id": report.id,
                    "items": report.items.count(),
                    "iiia": iiia_count,
                    "iiib": iiib_count,
                    "iiid": iiid_count,
                    "iiie": iiie_count,
                    "skipped": len(skipped),
                }
            )
            all_skipped.extend(
                {"month": MONTH_NAMES[import_file.month], **item}
                for item in skipped
            )

    print("IMPORT SUMMARY")
    for item in summary:
        print(
            f"- {item['month']}: report #{item['report_id']}, "
            f"items {item['items']}, III.A {item['iiia']}, "
            f"III.B {item['iiib']}, III.D {item['iiid']}, "
            f"III.E {item['iiie']}, skipped {item['skipped']}"
        )

    print("\nSKIPPED")
    for item in all_skipped:
        label = f"{item['month']} {item['sheet']} row {item['row']}"
        event = item.get("event")
        no = item.get("no")
        if no is not None:
            label += f" no {no}"
        print(f"- {label}: {event} -> {item['reason']}")


if __name__ == "__main__":
    run()
