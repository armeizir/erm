import calendar
import re
import tempfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from django.contrib.auth import get_user_model
from django.db import transaction

from masterdata.models import PeriodeLaporan, TahunBuku
from monthly_report.models import (
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from risk.models import (
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)


YEAR = 2026


@dataclass(frozen=True)
class ImportFile:
    code: str
    reassessment_id: int
    month: int
    path: Path
    template: str


FILES = [
    ImportFile(
        "OPS",
        9,
        2,
        Path("/Users/armeizir/Downloads/2026/6. Bid. Ops/Profil Risk BidOPS 2026 - MONEV FEB.xlsx"),
        "ops_old",
    ),
    ImportFile(
        "OPS",
        9,
        3,
        Path("/Users/armeizir/Downloads/2026/6. Bid. Ops/Monitoring Mitigasi Risiko MARET 2026 (1710).xlsx"),
        "ops_old",
    ),
    ImportFile(
        "OPS",
        9,
        4,
        Path("/Users/armeizir/Downloads/2026/6. Bid. Ops/Laporan Realisasi BIDOPS APRIL 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "MANPRO",
        8,
        2,
        Path("/Users/armeizir/Downloads/2026/5. Bid. Manpro/Laporan Bulanan ManRisk .xlsx"),
        "manpro",
    ),
    ImportFile(
        "MANPRO",
        8,
        3,
        Path("/Users/armeizir/Downloads/2026/5. Bid. Manpro/Laporan  MR MANPRO MARET 25 .xlsx"),
        "manpro",
    ),
    ImportFile(
        "MANPRO",
        8,
        4,
        Path("/Users/armeizir/Downloads/2026/5. Bid. Manpro/Laporan MR MANPRO APRIL 25 MEI 25.xlsx"),
        "manpro",
    ),
    ImportFile(
        "KEU",
        6,
        3,
        Path("/Users/armeizir/Downloads/2026/3. Bid. Keu/Laporan Mitigasi Risiko Keuangan Maret 2026( dr Bid. KEU).xlsx"),
        "keu",
    ),
    ImportFile(
        "KEU",
        6,
        4,
        Path("/Users/armeizir/Downloads/2026/3. Bid. Keu/Laporan Mitigasi Risiko Keuangan April 2026.xlsx"),
        "keu",
    ),
    ImportFile(
        "HCGA",
        4,
        2,
        Path("/Users/armeizir/Downloads/2026/2. Bid. HCGA/Laporan Realisasi MR BID HCGA Feb 2026.xlsx"),
        "standard",
    ),
    ImportFile(
        "HCGA",
        4,
        3,
        Path("/Users/armeizir/Downloads/2026/2. Bid. HCGA/Laporan Realisasi MR BID HCGA Mar (1).xlsx"),
        "standard",
    ),
    ImportFile(
        "HCGA",
        4,
        4,
        Path("/Users/armeizir/Downloads/2026/2. Bid. HCGA/Laporan Realisasi MR BID HCGA April Final.xlsx"),
        "standard",
    ),
]


MONTH_NAMES = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


def sanitize_workbook(path):
    src = Path(path)
    dst = Path(tempfile.gettempdir()) / f"{src.stem}.sanitized.xlsx"
    with ZipFile(src) as zin, ZipFile(dst, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/workbook.xml":
                text = data.decode("utf-8", errors="replace")
                text = re.sub(r"<definedNames>.*?</definedNames>", "", text, flags=re.S)
                data = text.encode("utf-8")
            zout.writestr(info, data)
    return dst


def normalize(value):
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        stripped = value.strip().lower()
        if stripped in {"-", "n/a", "no data", "#div/0!", "#n/a"}:
            return None
        multiplier = Decimal("1")
        if re.search(r"miliar|milyar", stripped):
            multiplier = Decimal("1000000000")
        elif re.search(r"\bm\b", stripped):
            multiplier = Decimal("1000000")
        elif "juta" in stripped:
            multiplier = Decimal("1000000")
        cleaned = re.sub(r"[^0-9,.\-]", "", stripped)
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        if not cleaned:
            return None
        value = cleaned
    else:
        multiplier = Decimal("1")
    try:
        return Decimal(str(value).strip()) * multiplier
    except (InvalidOperation, ValueError):
        return None


def percent_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    if number > Decimal("100") or number < Decimal("0"):
        return None
    return number


def scale_by_level(model, value):
    number = decimal_or_none(value)
    if number is None:
        return None
    return model.objects.filter(urutan=int(number)).first()


def treatment_effectiveness(value):
    text = normalize(value)
    if not text:
        return None
    if "tidak" in text:
        return "tidak_efektif"
    if "sebagian" in text or "cukup" in text:
        return "cukup_efektif"
    if "efektif" in text:
        return "efektif"
    return None


def treatment_status(value):
    text = normalize(value)
    if not text:
        return None
    if "discontinue" in text:
        return "discontinue"
    if "continue" in text:
        return "continue"
    return None


def row_value(row, col):
    idx = col - 1
    return row[idx] if len(row) > idx else None


def int_or_none(value):
    number = decimal_or_none(value)
    return int(number) if number is not None else None


def selected_quarter(month):
    return ((month - 1) // 3) + 1


def get_prepared_by():
    User = get_user_model()
    return (
        User.objects.filter(username="armeizir").first()
        or User.objects.filter(is_superuser=True).order_by("id").first()
        or User.objects.order_by("id").first()
    )


def get_period(tahun_buku, month):
    _, last_day = calendar.monthrange(tahun_buku.tahun, month)
    return PeriodeLaporan.objects.get_or_create(
        tahun_buku=tahun_buku,
        kode_periode=f"{tahun_buku.tahun}-{month:02d}",
        defaults={
            "nama_periode": f"{MONTH_NAMES[month]} {tahun_buku.tahun}",
            "jenis_periode": "bulanan",
            "tanggal_mulai": f"{tahun_buku.tahun}-{month:02d}-01",
            "tanggal_selesai": f"{tahun_buku.tahun}-{month:02d}-{last_day:02d}",
        },
    )[0]


def build_maps(reassessment):
    by_name = {}
    by_no_item = {}
    by_no_risiko = {}
    for item in ReAssessmentItem.objects.filter(summary=reassessment).order_by(
        "no_item",
        "no_risiko",
        "id",
    ):
        by_name.setdefault(normalize(item.peristiwa_risiko), item)
        by_no_item.setdefault(int(item.no_item), item)
        by_no_risiko.setdefault(int(item.no_risiko), item)
    return by_name, by_no_item, by_no_risiko


def resolve_default(row, by_name, by_no_item, by_no_risiko):
    event_name = normalize(row_value(row, 3))
    if event_name in by_name:
        return by_name[event_name]

    number = int_or_none(row_value(row, 2))
    if number in by_no_risiko and normalize(by_no_risiko[number].peristiwa_risiko) == event_name:
        return by_no_risiko[number]
    if number in by_no_item and normalize(by_no_item[number].peristiwa_risiko) == event_name:
        return by_no_item[number]
    return None


def resolve_by_number_and_name(row, by_name, by_no_item, by_no_risiko):
    return resolve_default(row, by_name, by_no_item, by_no_risiko)


def resolve_manpro(row, by_name, by_no_item, by_no_risiko):
    combined = normalize(
        " ".join(
            str(value or "")
            for value in [
                row_value(row, 3),
                row_value(row, 6),
                row_value(row, 7),
                row_value(row, 12),
                row_value(row, 15),
            ]
        )
    )

    rules = [
        (("sengkuang 80",), 1),
        (("tanjung sengkuang", "mpi"), 2),
        (("sekupang 40",), 3),
        (("sagulung", "60"), 4),
        (("pltmg 50", "sekupang"), 5),
        (("epc pltmg sekupang 50",), 5),
        (("reroute",), 6),
        (("t 26",), 6),
        (("nongsa",), 7),
        (("batu besar",), 8),
        (("gis kabil",), 10),
        (("kabil 1 dan kabil 2",), 10),
        (("120 mw",), 11),
        (("kabil 2",), 11),
    ]
    for keywords, no_item in rules:
        if all(keyword in combined for keyword in keywords):
            return by_no_item.get(no_item)
    return resolve_default(row, by_name, by_no_item, by_no_risiko)


def read_rows(ws, start_row=1):
    for offset, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        yield offset, list(row)


def update_item_common(item, *, rencana=None, output=None, biaya=None, serapan=None, pic=None):
    if rencana not in (None, ""):
        item.realisasi_rencana_perlakuan = rencana
    if output not in (None, ""):
        item.realisasi_output_perlakuan = output
    if biaya not in (None, ""):
        item.realisasi_biaya_perlakuan = decimal_or_none(biaya)
    if serapan not in (None, ""):
        item.persentase_serapan_biaya = percent_or_none(serapan)
    if pic not in (None, ""):
        item.realisasi_pic = str(pic)[:255]


def import_standard_iiia(workbook, report, maps, month, skipped, resolver=resolve_default):
    ws = workbook["III.A"]
    by_name, by_no_item, by_no_risiko = maps
    q = selected_quarter(month)
    cols = {
        "nilai_dampak": {1: 15, 2: 16, 3: 17, 4: 18}[q],
        "skala_dampak": {1: 19, 2: 20, 3: 21, 4: 22}[q],
        "nilai_prob": {1: 23, 2: 24, 3: 25, 4: 26}[q],
        "skala_prob": {1: 27, 2: 28, 3: 29, 4: 30}[q],
    }
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        if not row_value(row, 3):
            continue
        risk_event = resolver(row, by_name, by_no_item, by_no_risiko)
        if not risk_event:
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        item.realisasi_asumsi_dampak = row_value(row, 14) or row_value(row, 5)
        item.realisasi_nilai_dampak = decimal_or_none(row_value(row, cols["nilai_dampak"]))
        item.realisasi_skala_dampak = scale_by_level(MasterSkalaDampak, row_value(row, cols["skala_dampak"]))
        item.realisasi_nilai_probabilitas = percent_or_none(row_value(row, cols["nilai_prob"]))
        item.realisasi_skala_probabilitas = scale_by_level(
            MasterSkalaProbabilitas,
            row_value(row, cols["skala_prob"]),
        )
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(row_value(row, 47))
        item.save()
        imported += 1
    return imported


def import_standard_iiib(workbook, report, maps, month, skipped, resolver=resolve_default):
    ws = workbook["III.B"]
    by_name, by_no_item, by_no_risiko = maps
    q = selected_quarter(month)
    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[q]
    threshold_col = 39 + ((month - 1) * 2)
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        if not row_value(row, 3):
            continue
        risk_event = resolver(row, by_name, by_no_item, by_no_risiko)
        if not risk_event:
            continue

        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        update_item_common(
            item,
            rencana=row_value(row, 11),
            output=row_value(row, 12),
            biaya=row_value(row, 13),
            serapan=row_value(row, 14),
            pic=row_value(row, 15),
        )
        item.status_rencana_perlakuan = treatment_status(row_value(row, 28))
        item.penjelasan_status_rencana = row_value(row, 29)
        item.progress_pelaksanaan_percent = percent_or_none(row_value(row, progress_col))
        item.realisasi_threshold_kri = row_value(row, threshold_col)
        score = row_value(row, threshold_col + 1)
        item.realisasi_threshold_kri_skor = str(score) if score not in (None, "") else None
        item.save()
        imported += 1
    return imported


def import_manpro_iiia(workbook, report, maps, month, skipped):
    ws = workbook["III.A"]
    by_name, by_no_item, by_no_risiko = maps
    imported = 0
    seen = set()
    for row_idx, row in read_rows(ws, 20):
        if "bid manpro" not in normalize(row_value(row, 5)):
            continue
        synthetic = [None, row_value(row, 11), row_value(row, 12), None, None, row_value(row, 15)]
        risk_event = resolve_manpro(synthetic, by_name, by_no_item, by_no_risiko)
        if not risk_event:
            event_text = row_value(row, 12) or row_value(row, 7)
            skipped.append(("III.A", row_idx, event_text))
            continue
        if risk_event.id in seen:
            continue
        seen.add(risk_event.id)

        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        item.realisasi_asumsi_dampak = row_value(row, 28) or row_value(row, 13)
        item.realisasi_nilai_dampak = decimal_or_none(row_value(row, 31))
        item.realisasi_skala_dampak = scale_by_level(MasterSkalaDampak, row_value(row, 32))
        item.realisasi_nilai_probabilitas = percent_or_none(row_value(row, 33))
        item.realisasi_skala_probabilitas = scale_by_level(MasterSkalaProbabilitas, row_value(row, 34))
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(row_value(row, 47))
        item.save()
        imported += 1
    return imported


def import_keu_iiia(workbook, report, maps, month, skipped):
    ws = workbook["III.A-Korporat-Bidang"]
    by_name, by_no_item, by_no_risiko = maps
    q = selected_quarter(month)
    cols = {
        "nilai_dampak": {1: 15, 2: 16, 3: 17, 4: 18}[q],
        "skala_dampak": {1: 19, 2: 20, 3: 21, 4: 22}[q],
        "nilai_prob": {1: 23, 2: 24, 3: 25, 4: 26}[q],
        "skala_prob": {1: 27, 2: 28, 3: 29, 4: 30}[q],
    }
    imported = 0
    for row_idx, row in read_rows(ws, 11):
        if not row_value(row, 3):
            continue
        risk_event = resolve_by_number_and_name(row, by_name, by_no_item, by_no_risiko)
        if not risk_event:
            skipped.append(("III.A-Korporat-Bidang", row_idx, row_value(row, 3)))
            continue
        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        item.realisasi_asumsi_dampak = row_value(row, 14) or row_value(row, 5)
        item.realisasi_nilai_dampak = decimal_or_none(row_value(row, cols["nilai_dampak"]))
        item.realisasi_skala_dampak = scale_by_level(MasterSkalaDampak, row_value(row, cols["skala_dampak"]))
        item.realisasi_nilai_probabilitas = percent_or_none(row_value(row, cols["nilai_prob"]))
        item.realisasi_skala_probabilitas = scale_by_level(
            MasterSkalaProbabilitas,
            row_value(row, cols["skala_prob"]),
        )
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(row_value(row, 47))
        item.save()
        imported += 1
    return imported


def import_keu_iiib(workbook, report, maps, month, skipped):
    ws = workbook["III.B-Korporat-Bidang"]
    by_name, by_no_item, by_no_risiko = maps
    q = selected_quarter(month)
    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[q]
    threshold_col = 39 + ((month - 1) * 2)
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        if not row_value(row, 3):
            continue
        risk_event = resolve_by_number_and_name(row, by_name, by_no_item, by_no_risiko)
        if not risk_event:
            continue
        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        update_item_common(
            item,
            rencana=row_value(row, 11),
            output=row_value(row, 12),
            biaya=row_value(row, 13),
            serapan=row_value(row, 14),
            pic=row_value(row, 15),
        )
        item.status_rencana_perlakuan = treatment_status(row_value(row, 28))
        item.penjelasan_status_rencana = row_value(row, 29)
        item.progress_pelaksanaan_percent = percent_or_none(row_value(row, progress_col))
        item.realisasi_threshold_kri = row_value(row, threshold_col)
        score = row_value(row, threshold_col + 1)
        item.realisasi_threshold_kri_skor = str(score) if score not in (None, "") else None
        item.save()
        imported += 1
    return imported


def import_ops_old_iiia(workbook, report, maps, month, skipped):
    ws = workbook["LAP REAL III.A"]
    by_name, by_no_item, by_no_risiko = maps
    q = selected_quarter(month)
    cols = {
        "nilai_dampak": {1: 4, 2: 5, 3: 6, 4: 7}[q],
        "skala_dampak": {1: 8, 2: 9, 3: 10, 4: 11}[q],
        "nilai_prob": {1: 12, 2: 13, 3: 14, 4: 15}[q],
        "skala_prob": {1: 16, 2: 17, 3: 18, 4: 19}[q],
    }
    imported = 0
    for row_idx, row in read_rows(ws, 12):
        event_name = normalize(row_value(row, 2))
        if not event_name:
            continue
        risk_event = by_name.get(event_name) or by_no_item.get(int_or_none(row_value(row, 1)))
        if not risk_event:
            skipped.append(("LAP REAL III.A", row_idx, row_value(row, 2)))
            continue
        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        item.realisasi_asumsi_dampak = row_value(row, 3)
        item.realisasi_nilai_dampak = decimal_or_none(row_value(row, cols["nilai_dampak"]))
        item.realisasi_skala_dampak = scale_by_level(MasterSkalaDampak, row_value(row, cols["skala_dampak"]))
        item.realisasi_nilai_probabilitas = percent_or_none(row_value(row, cols["nilai_prob"]))
        item.realisasi_skala_probabilitas = scale_by_level(
            MasterSkalaProbabilitas,
            row_value(row, cols["skala_prob"]),
        )
        item.efektivitas_perlakuan_risiko = treatment_effectiveness(row_value(row, 60))
        item.save()
        imported += 1
    return imported


def import_ops_old_iiib(workbook, report, maps, month, skipped):
    ws = workbook["LAP REAL III.B"]
    by_name, by_no_item, by_no_risiko = maps
    q = selected_quarter(month)
    threshold_col = 21 + ((month - 2) * 2)
    progress_col = {1: 26, 2: 27, 3: 28, 4: 29}[q]
    status_col = 24 if month == 2 else 26
    explanation_col = status_col + 1
    imported = 0
    for row_idx, row in read_rows(ws, 10):
        event_name = normalize(row_value(row, 2))
        if not event_name:
            continue
        risk_event = by_name.get(event_name) or by_no_item.get(int_or_none(row_value(row, 1)))
        if not risk_event:
            skipped.append(("LAP REAL III.B", row_idx, row_value(row, 2)))
            continue
        item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
        update_item_common(
            item,
            rencana=row_value(row, 4),
            output=row_value(row, 5),
            biaya=row_value(row, 6),
            serapan=row_value(row, 7),
            pic=row_value(row, 8),
        )
        item.status_rencana_perlakuan = treatment_status(row_value(row, status_col))
        item.penjelasan_status_rencana = row_value(row, explanation_col)
        item.progress_pelaksanaan_percent = percent_or_none(row_value(row, progress_col))
        item.realisasi_threshold_kri = row_value(row, threshold_col)
        score = row_value(row, threshold_col + 1)
        item.realisasi_threshold_kri_skor = str(score) if score not in (None, "") else None
        item.save()
        imported += 1
    return imported


def import_changes(workbook, report):
    sheet_name = "III.D" if "III.D" in workbook.sheetnames else None
    if not sheet_name:
        return 0
    ws = workbook[sheet_name]
    mapping = {
        "perubahan profil risiko": MonthlyRiskReportChange.CHANGE_TYPE_PROFILE,
        "penambahan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_ADD_ITEM,
        "pengurangan item risiko": MonthlyRiskReportChange.CHANGE_TYPE_REMOVE_ITEM,
        "perubahan strategi risiko": MonthlyRiskReportChange.CHANGE_TYPE_STRATEGY,
    }
    MonthlyRiskReportChange.objects.filter(report=report).delete()
    imported = 0
    for _, row in read_rows(ws, 8):
        change_type = mapping.get(normalize(row_value(row, 2)))
        if not change_type:
            continue
        MonthlyRiskReportChange.objects.create(
            report=report,
            jenis_perubahan=change_type,
            peristiwa_risiko_terdampak=row_value(row, 3),
            penjelasan=row_value(row, 4),
        )
        imported += 1
    return imported


def import_loss_events(workbook, report):
    sheet_name = "III.E" if "III.E" in workbook.sheetnames else None
    if not sheet_name:
        return 0
    ws = workbook[sheet_name]
    MonthlyRiskReportLossEvent.objects.filter(report=report).delete()
    imported = 0
    for _, row in read_rows(ws, 8):
        name = row_value(row, 2)
        if not name:
            continue
        MonthlyRiskReportLossEvent.objects.create(
            report=report,
            nama_kejadian=name,
            identifikasi_kejadian=row_value(row, 3),
            kategori_kejadian=row_value(row, 4),
            sumber_penyebab_kejadian="external"
            if "eksternal" in normalize(row_value(row, 5))
            else "internal"
            if "internal" in normalize(row_value(row, 5))
            else None,
            penyebab_kejadian=row_value(row, 6),
            penanganan_saat_kejadian=row_value(row, 7),
            deskripsi_kejadian_risk_event=row_value(row, 8),
            kategori_risiko_bumn=row_value(row, 9),
            kategori_risiko_t2_t3_kbumn=row_value(row, 10),
            penjelasan_kerugian=row_value(row, 11),
            nilai_kerugian=decimal_or_none(row_value(row, 12)),
            kejadian_berulang="ya"
            if "ya" in normalize(row_value(row, 13))
            else "tidak"
            if "tidak" in normalize(row_value(row, 13))
            else None,
            frekuensi_kejadian=row_value(row, 14),
            mitigasi_direncanakan=row_value(row, 15),
            realisasi_mitigasi=row_value(row, 16),
            perbaikan_mendatang=row_value(row, 17),
            pihak_terkait=row_value(row, 18),
            status_asuransi="ya"
            if "ya" in normalize(row_value(row, 19))
            else "tidak"
            if "tidak" in normalize(row_value(row, 19))
            else None,
            nilai_premi=decimal_or_none(row_value(row, 20)),
            nilai_klaim=decimal_or_none(row_value(row, 21)),
        )
        imported += 1
    return imported


def import_one(workbook, report, maps, import_file):
    skipped = []
    if import_file.template == "ops_old":
        iiia = import_ops_old_iiia(workbook, report, maps, import_file.month, skipped)
        iiib = import_ops_old_iiib(workbook, report, maps, import_file.month, skipped)
        iiid = 0
        iiie = 0
    elif import_file.template == "keu":
        iiia = import_keu_iiia(workbook, report, maps, import_file.month, skipped)
        iiib = import_keu_iiib(workbook, report, maps, import_file.month, skipped)
        iiid = 0
        iiie = 0
    elif import_file.template == "manpro":
        iiia = import_manpro_iiia(workbook, report, maps, import_file.month, skipped)
        iiib = import_standard_iiib(
            workbook,
            report,
            maps,
            import_file.month,
            skipped,
            resolver=resolve_manpro,
        )
        iiid = import_changes(workbook, report)
        iiie = import_loss_events(workbook, report)
    else:
        iiia = import_standard_iiia(workbook, report, maps, import_file.month, skipped)
        iiib = import_standard_iiib(workbook, report, maps, import_file.month, skipped)
        iiid = import_changes(workbook, report)
        iiie = import_loss_events(workbook, report)
    return iiia, iiib, iiid, iiie, skipped


def run():
    from openpyxl import load_workbook

    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user untuk prepared_by.")

    tahun_buku, _ = TahunBuku.objects.get_or_create(tahun=YEAR, defaults={"aktif": True})
    cache = {}
    summaries = []
    all_skipped = []

    with transaction.atomic():
        for import_file in FILES:
            if not import_file.path.exists():
                raise FileNotFoundError(import_file.path)

            reassessment = ReAssessmentSummary.objects.get(pk=import_file.reassessment_id)
            maps = cache.setdefault(reassessment.id, build_maps(reassessment))
            workbook_path = (
                sanitize_workbook(import_file.path)
                if import_file.template == "keu"
                else import_file.path
            )
            workbook = load_workbook(
                workbook_path,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
            period = get_period(tahun_buku, import_file.month)
            report, _ = MonthlyRiskReport.objects.update_or_create(
                reassessment=reassessment,
                periode=period,
                versi=1,
                defaults={
                    "kode": f"MRR-{import_file.code}-{YEAR}-{import_file.month:02d}",
                    "tahun_buku": tahun_buku,
                    "status": "draft",
                    "prepared_by": prepared_by,
                },
            )

            iiia, iiib, iiid, iiie, skipped = import_one(workbook, report, maps, import_file)

            report.total_risiko = report.items.count()
            report.total_high = report.items.filter(realisasi_level_risiko__icontains="tinggi").count()
            report.total_mitigasi_terlambat = report.items.filter(mitigation_status="delayed").count()
            report.total_selesai = report.items.filter(status_rencana_perlakuan="discontinue").count()
            report.save()

            summaries.append(
                {
                    "code": report.kode,
                    "items": report.items.count(),
                    "iiia": iiia,
                    "iiib": iiib,
                    "iiid": iiid,
                    "iiie": iiie,
                    "skipped": len(skipped),
                }
            )
            for sheet, row_idx, event in skipped:
                all_skipped.append((report.kode, sheet, row_idx, event))

    print("IMPORT SUMMARY")
    for item in summaries:
        print(
            f"- {item['code']}: items {item['items']}, "
            f"III.A {item['iiia']}, III.B {item['iiib']}, "
            f"III.D {item['iiid']}, III.E {item['iiie']}, skipped {item['skipped']}"
        )

    print("\nSKIPPED")
    if not all_skipped:
        print("- Tidak ada item utama yang dilewati.")
    for code, sheet, row_idx, event in all_skipped[:80]:
        print(f"- {code} {sheet} row {row_idx}: {event}")
    if len(all_skipped) > 80:
        print(f"- ... {len(all_skipped) - 80} skipped lainnya")


if __name__ == "__main__":
    run()
