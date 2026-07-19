from pathlib import Path
import os
import re
import sys

PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django

django.setup()

from django.db import transaction
from openpyxl import load_workbook

from masterdata.models import TahunBuku
from monthly_report.models import MonthlyRiskReport, MonthlyRiskReportItem
from monthly_report.scripts.import_bis_monthly_reports import (
    MONTH_NAMES,
    decimal_or_none,
    get_period,
    get_prepared_by,
    percent_or_none,
    scale_by_level,
    treatment_effectiveness,
    treatment_status,
)
from monthly_report.services import refresh_monthly_report_summary
from risk.models import (
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    ReAssessmentItem,
    ReAssessmentSummary,
)


DEFAULT_SOURCE_FILE = "/Users/armeizir/Downloads/Mitigasi Risiko dan KRI BIS Juni 2026.xlsx"
UNIT_NAME = "BID BIS"
YEAR = 2026
MONTH = int(os.environ.get("BIS_REPORT_MONTH") or (sys.argv[2] if len(sys.argv) > 2 else 6))
REPORT_CODE = f"MRR-BIS-{YEAR}-{MONTH:02d}"


CODE_RE = re.compile(r"bid\s*bis\s*-\s*(\d+)\s*-\s*([a-z]+)", re.I)


def source_file():
    return Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SOURCE_FILE)


def normalize(value):
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def base_event(value):
    text = str(value or "").strip()
    return text.split(" - ", 1)[0].strip() if " - " in text else text


def parse_bis_code(value):
    match = CODE_RE.search(str(value or ""))
    if not match:
        return None
    return int(match.group(1)), match.group(2).lower()


def selected_quarter(month):
    return ((month - 1) // 3) + 1


def read_rows(ws, start_row):
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
        yield row_idx, list(row)


def bis_reassessment():
    return (
        ReAssessmentSummary.objects.filter(
            tahun=YEAR,
            unit_bisnis__name=UNIT_NAME,
            judul__icontains="BIS",
        )
        .select_related("unit_bisnis")
        .order_by("id")
        .first()
    )


def iiia_rows(workbook):
    ws = workbook["III.A"]
    q = selected_quarter(MONTH)
    cols = {
        "nilai_dampak": {1: 13, 2: 14, 3: 15, 4: 16}[q],
        "skala_dampak": {1: 17, 2: 18, 3: 19, 4: 20}[q],
        "nilai_prob": {1: 25, 2: 26, 3: 27, 4: 28}[q],
        "skala_prob": {1: 29, 2: 30, 3: 31, 4: 32}[q],
        "efektivitas": 57,
    }
    rows = []
    skipped = []
    for row_idx, row in read_rows(ws, 10):
        code = row[1] if len(row) > 1 else None
        event = row[2] if len(row) > 2 else None
        parsed = parse_bis_code(code)
        if not any([code, event]):
            continue
        if not parsed:
            if "bis" in normalize(code):
                skipped.append(
                    f"III.A row {row_idx} no={code}: {event} -> kode tidak mengikuti pola BID BIS-x-y."
                )
            continue
        rows.append(
            {
                "row": row_idx,
                "code": str(code).strip(),
                "no_risiko": parsed[0],
                "no_penyebab": parsed[1],
                "event": str(event or "").strip(),
                "base_event": base_event(event),
                "deskripsi": row[4] if len(row) > 4 else None,
                "asumsi": row[11] if len(row) > 11 else None,
                "nilai_dampak": decimal_or_none(row[cols["nilai_dampak"] - 1]),
                "skala_dampak": scale_by_level(
                    MasterSkalaDampak,
                    row[cols["skala_dampak"] - 1],
                ),
                "nilai_probabilitas": percent_or_none(row[cols["nilai_prob"] - 1]),
                "skala_probabilitas": scale_by_level(
                    MasterSkalaProbabilitas,
                    row[cols["skala_prob"] - 1],
                ),
                "efektivitas": treatment_effectiveness(
                    row[cols["efektivitas"] - 1] if len(row) >= cols["efektivitas"] else None
                ),
            }
        )
    return rows, skipped


def iiib_rows(workbook):
    ws = workbook["III.B"]
    q = selected_quarter(MONTH)
    progress_col = {1: 30, 2: 31, 3: 32, 4: 33}[q]
    threshold_col = 39 + ((MONTH - 1) * 2)
    rows = {}
    skipped = []
    for row_idx, row in read_rows(ws, 10):
        code = row[5] if len(row) > 5 else None
        parsed = parse_bis_code(code)
        if not any(row[idx] if len(row) > idx else None for idx in (1, 2, 5)):
            continue
        if not parsed:
            raw_code = row[1] if len(row) > 1 else code
            event = row[2] if len(row) > 2 else None
            if "bis" in normalize(raw_code) or "bis" in normalize(code):
                skipped.append(
                    f"III.B row {row_idx} no={raw_code or code}: {event} -> kode tidak mengikuti pola BID BIS-x-y."
                )
            continue
        rows[str(code).strip().lower()] = {
            "row": row_idx,
            "code": str(code).strip(),
            "penyebab": row[6] if len(row) > 6 else None,
            "rencana": row[7] if len(row) > 7 else None,
            "output": row[8] if len(row) > 8 else None,
            "anggaran": decimal_or_none(row[9] if len(row) > 9 else None),
            "realisasi_rencana": row[10] if len(row) > 10 else None,
            "realisasi_output": row[11] if len(row) > 11 else None,
            "realisasi_biaya": decimal_or_none(row[12] if len(row) > 12 else None),
            "serapan": percent_or_none(row[13] if len(row) > 13 else None),
            "pic": row[14] if len(row) > 14 else None,
            "status": treatment_status(row[27] if len(row) > 27 else None),
            "penjelasan_status": row[28] if len(row) > 28 else None,
            "progress": percent_or_none(row[progress_col - 1] if len(row) >= progress_col else None),
            "threshold": row[threshold_col - 1] if len(row) >= threshold_col else None,
            "threshold_skor": row[threshold_col] if len(row) > threshold_col else None,
        }
    return rows, skipped


def find_existing_item(summary, data):
    target_base = normalize(data["base_event"])
    cause = data["no_penyebab"].lower()

    exact_code_match = (
        ReAssessmentItem.objects.filter(
            summary=summary,
            no_risiko=data["no_risiko"],
            no_penyebab_risiko__iexact=cause,
        )
        .order_by("id")
        .first()
    )
    if exact_code_match:
        return exact_code_match

    candidates = ReAssessmentItem.objects.filter(summary=summary).order_by("id")
    for item in candidates:
        if (
            normalize(base_event(item.peristiwa_risiko)) == target_base
            and normalize(item.no_penyebab_risiko) == cause
        ):
            return item

    if cause == "a":
        for item in candidates:
            if normalize(base_event(item.peristiwa_risiko)) == target_base:
                return item

    return None


def sync_master_item(summary, data, iiib_data, no_item):
    existing = find_existing_item(summary, data)
    created = False
    if not existing:
        template = ReAssessmentItem.objects.filter(summary=summary).order_by("no_item", "id").first()
        if not template:
            raise RuntimeError("Profil Risiko BIS belum memiliki template item untuk disalin.")
        existing = ReAssessmentItem(
            summary=summary,
            km_item=template.km_item,
            no_item=no_item,
            no_risiko=data["no_risiko"],
            no_penyebab_risiko=data["no_penyebab"],
            peristiwa_risiko=data["event"],
            deskripsi_peristiwa_risiko=data["deskripsi"] or "-",
        )
        created = True

    existing.no_item = no_item
    existing.no_risiko = data["no_risiko"]
    existing.no_penyebab_risiko = data["no_penyebab"]
    existing.peristiwa_risiko = data["event"]
    existing.deskripsi_peristiwa_risiko = data["deskripsi"] or existing.deskripsi_peristiwa_risiko or "-"
    if iiib_data:
        existing.penyebab_risiko = iiib_data.get("penyebab") or existing.penyebab_risiko
        existing.rencana_perlakuan_risiko = iiib_data.get("rencana") or existing.rencana_perlakuan_risiko
        existing.output_perlakuan_risiko = iiib_data.get("output") or existing.output_perlakuan_risiko
        if iiib_data.get("anggaran") is not None:
            existing.biaya_perlakuan_risiko = iiib_data["anggaran"]
    existing.save()
    return existing, created


def apply_report_item(report, risk_event, data, iiib_data):
    item, _ = MonthlyRiskReportItem.objects.get_or_create(report=report, risk_event=risk_event)
    item.realisasi_asumsi_dampak = data["asumsi"]
    item.realisasi_nilai_dampak = data["nilai_dampak"]
    item.realisasi_skala_dampak = data["skala_dampak"]
    item.realisasi_nilai_probabilitas = data["nilai_probabilitas"]
    item.realisasi_skala_probabilitas = data["skala_probabilitas"]
    item.efektivitas_perlakuan_risiko = data["efektivitas"]

    if iiib_data:
        item.realisasi_rencana_perlakuan = iiib_data.get("realisasi_rencana")
        item.realisasi_output_perlakuan = iiib_data.get("realisasi_output")
        item.realisasi_biaya_perlakuan = iiib_data.get("realisasi_biaya")
        item.realisasi_pic = iiib_data.get("pic")
        item.status_rencana_perlakuan = iiib_data.get("status")
        item.penjelasan_status_rencana = iiib_data.get("penjelasan_status")
        item.progress_pelaksanaan_percent = iiib_data.get("progress")
        item.realisasi_threshold_kri = iiib_data.get("threshold")
        threshold_score = iiib_data.get("threshold_skor")
        item.realisasi_threshold_kri_skor = (
            str(threshold_score) if threshold_score not in (None, "") else None
        )

    item.save()
    return item


def run():
    path = source_file()
    if not path.exists():
        raise FileNotFoundError(path)

    reassessment = bis_reassessment()
    if not reassessment:
        raise RuntimeError(f"Profil Risiko {UNIT_NAME} tahun {YEAR} tidak ditemukan.")

    prepared_by = get_prepared_by()
    if not prepared_by:
        raise RuntimeError("Tidak ada user untuk prepared_by.")

    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    rows_iiia, skipped = iiia_rows(workbook)
    rows_iiib, skipped_iiib = iiib_rows(workbook)
    skipped.extend(skipped_iiib)

    created_codes = []
    updated_codes = []
    missing_iiib = []
    imported_item_ids = []

    with transaction.atomic():
        tahun_buku, _ = TahunBuku.objects.get_or_create(tahun=YEAR, defaults={"aktif": True})
        period = get_period(tahun_buku, MONTH)
        report, _ = MonthlyRiskReport.objects.update_or_create(
            reassessment=reassessment,
            periode=period,
            versi=1,
            defaults={
                "kode": REPORT_CODE,
                "tahun_buku": tahun_buku,
                "status": "draft",
                "prepared_by": prepared_by,
            },
        )

        selected_items = []
        for idx, data in enumerate(rows_iiia, start=1):
            iiib_data = rows_iiib.get(data["code"].lower())
            if not iiib_data:
                missing_iiib.append(f"{data['code']} ({data['event']})")
            item = find_existing_item(reassessment, data)
            if item:
                selected_items.append(item)

        for offset, item in enumerate(selected_items, start=9000):
            item.no_item = offset
            item.save()

        for idx, data in enumerate(rows_iiia, start=1):
            iiib_data = rows_iiib.get(data["code"].lower())
            risk_event, created = sync_master_item(reassessment, data, iiib_data, idx)
            report_item = apply_report_item(report, risk_event, data, iiib_data)
            imported_item_ids.append(report_item.id)
            if created:
                created_codes.append(data["code"])
            else:
                updated_codes.append(data["code"])

        deleted_count, _ = report.items.exclude(id__in=imported_item_ids).delete()
        refresh_monthly_report_summary(report)

    print("IMPORT SUMMARY")
    print(f"- Report: {report.id} {report.kode} / {report}")
    print(f"- Profil Risiko: {reassessment.id} {reassessment.judul}")
    print(f"- Items report setelah sinkron: {report.items.count()}")
    print(f"- III.A diinput: {len(rows_iiia)}")
    print(f"- III.B diinput: {len(rows_iiib)}")
    print(f"- Master item dibuat baru: {len(created_codes)}")
    print(f"- Master item diperbarui: {len(updated_codes)}")
    print(f"- Item lama report yang dihapus karena tidak ada di Excel: {deleted_count}")
    print(f"- Total risiko: {report.total_risiko}")
    print(f"- Total high: {report.total_high}")

    print("\nCATATAN DATA")
    if created_codes:
        print("- Dibuat item Profil Risiko baru karena kode Excel belum ada di database:")
        for code in created_codes:
            print(f"  * {code}")
    if missing_iiib:
        print("- Ada baris III.A yang tidak punya pasangan III.B:")
        for note in missing_iiib:
            print(f"  * {note}")
    if skipped:
        print("- Baris tidak diinput karena kode/profil tidak cocok:")
        for note in skipped:
            print(f"  * {note}")
    if not any([created_codes, missing_iiib, skipped]):
        print("- Semua baris utama BIS cocok dengan pola Profil Risiko dan berhasil diinput.")


if __name__ == "__main__":
    run()
