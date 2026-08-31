#!/usr/bin/env python3
"""
Safe in-place import MRR UB KITRAN Maret 2026.

Target:
- ReAssessmentSummary/Profile : id=14 (Profil Risiko UBKITRANS)
- MonthlyRiskReport           : id=25 (MRR-KITRANS-2026-03)
- Canonical mapping reference : report id=84 (MRR-KITRANS-2026-06, 40 cause-level items)

Safety design:
- default DRY RUN; --apply is required to commit;
- does NOT create/modify ReAssessmentItem/profile rows;
- updates report id=25 in-place and preserves its identity/history;
- maps the 40 March source cause groups only to the 40 risk_event rows already
  proven by the June UBKITRANS report;
- source hash and structural counts are hard guards;
- Q1 residual is imported exactly from III.A (no fallback to another quarter);
- March timeline policy: Jan-Mar may be 1 from source; Apr-Dec forced to 0;
- KRI is preserved as source text/status because current UBKITRANS master has no
  configured KRI direction in the validated June import;
- III.D and III.E must contain no source data for this workbook;
- malformed Excel comment/person metadata is sanitized only in a temporary copy.

Usage (production):
  python monthly_report/scripts/import_mrr_ubkitrans_maret_2026_v1_safe.py \
    --source "/tmp/Laporan Manajemen Risiko UBKITRANS Maret 2026.xlsx"

Apply after clean dry-run:
  python monthly_report/scripts/import_mrr_ubkitrans_maret_2026_v1_safe.py \
    --source "/tmp/Laporan Manajemen Risiko UBKITRANS Maret 2026.xlsx" --apply
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import shutil
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from xml.sax.saxutils import escape

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django  # noqa: E402

django.setup()

from django.conf import settings  # noqa: E402
from django.core.exceptions import ValidationError  # noqa: E402
from django.db import transaction  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from monthly_report.models import (  # noqa: E402
    MonthlyRiskReport,
    MonthlyRiskReportChange,
    MonthlyRiskReportItem,
    MonthlyRiskReportLossEvent,
)
from monthly_report.services import refresh_monthly_report_summary  # noqa: E402
from risk.models import MasterSkalaDampak, MasterSkalaProbabilitas  # noqa: E402

YEAR = 2026
MONTH = 3
PROFILE_ID = 14
TARGET_REPORT_ID = 25
CANONICAL_REPORT_ID = 84
EXPECTED_SOURCE_SHA256 = "570a248c7df5ec0b550fbb69058cdf111e41ce8c439d33593accc65c48e717a6"
EXPECTED_EVENTS = 19
EXPECTED_CAUSE_GROUPS = 40
EXPECTED_FORMAL_ACTIVITIES = 106
EXPECTED_ACTUAL_ONLY_ROWS = 20
EXPECTED_PROGRESS_GROUPS = 33
EXPECTED_KRI_GROUPS = 24
EXPECTED_PLANNED_COST = Decimal("558910040366.41")
EXPECTED_ACTUAL_COST = Decimal("0.00")

# III.A UBKITRANS format, zero-based indexes.
IIIA_EVENT = 2               # C
IIIA_RISK_TYPE = 3           # D
IIIA_ASSUMPTION = 12         # M
IIIA_Q1_IMPACT = 13          # N
IIIA_Q1_IMPACT_SCALE = 17    # R
IIIA_Q1_IMPACT_KBUMN = 21    # V
IIIA_Q1_PROB = 25            # Z
IIIA_Q1_PROB_SCALE = 29      # AD
IIIA_Q1_PROB_KBUMN = 33      # AH
IIIA_Q1_EXPOSURE = 37        # AL
IIIA_Q1_SCORE = 41           # AP
IIIA_Q1_SCORE_KBUMN = 45     # AT
IIIA_Q1_LEVEL = 49           # AX
IIIA_Q1_LEVEL_KBUMN = 53     # BB
IIIA_EFFECTIVENESS = 57      # BF

# III.B fixed zero-based indexes.
B_RISK_NO = 1                # B
B_EVENT = 2                  # C
B_CAUSE_NO = 4               # E
B_CAUSE_CODE = 5             # F
B_CAUSE_TEXT = 6             # G
B_PLAN = 7                   # H
B_OUTPUT = 8                 # I
B_PLAN_COST = 9              # J
B_ACTUAL_PLAN = 10           # K
B_ACTUAL_OUTPUT = 11         # L
B_ACTUAL_COST = 12           # M
B_PIC = 14                   # O
B_TIMELINE_FIRST = 15        # P = Jan; P:AA = 12 months
B_STATUS = 27                # AB
B_STATUS_NOTE = 28           # AC
B_Q1_PROGRESS = 29           # AD
B_KRI = 33                   # AH
B_KRI_UNIT = 34              # AI
B_KRI_SAFE = 35              # AJ
B_KRI_CAUTION = 36           # AK
B_KRI_DANGER = 37            # AL
B_KRI_STATUS = 38            # AM
B_KRI_VALUE = 39             # AN


@dataclass
class SourceGroup:
    index: int
    start_row: int
    end_row: int
    risk_no: object
    event: str
    cause_no: str | None
    cause_code: str | None
    cause_text: str
    rows: list[list]
    master: object | None = None


def norm(value) -> str:
    text = str(value or "").casefold().replace("\xa0", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def txt(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def unique_texts(values):
    out = []
    seen = set()
    for value in values:
        t = txt(value)
        if not t:
            continue
        key = norm(t)
        if key not in seen:
            out.append(t)
            seen.add(key)
    return out


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    s = str(value).strip()
    if not s or s.casefold() in {"-", "n/a", "na", "none", "#n/a", "#div/0!", "#name?"}:
        return None
    s = s.replace("\xa0", " ")
    is_percent = "%" in s
    cleaned = re.sub(r"[^0-9,\.\-]", "", s)
    if not cleaned or cleaned in {"-", ".", ","}:
        return None

    # Indonesian thousands/decimal formatting.
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    elif cleaned.count(".") > 1:
        cleaned = cleaned.replace(".", "")

    try:
        number = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if is_percent:
        return number
    return number


def percent_or_none(value):
    if value in (None, ""):
        return None
    raw = str(value).strip() if isinstance(value, str) else ""
    number = decimal_or_none(value)
    if number is None:
        return None
    if "%" in raw:
        return number
    if Decimal("-1") <= number <= Decimal("1"):
        return number * Decimal("100")
    if Decimal("0") <= number <= Decimal("100"):
        return number
    return None


def integer_or_none(value):
    number = decimal_or_none(value)
    if number is None:
        return None
    try:
        return int(number)
    except (ValueError, TypeError, OverflowError):
        return None


def money_cent(value):
    number = decimal_or_none(value)
    return None if number is None else number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def file_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sanitize_xlsx_for_openpyxl(source: Path) -> Path:
    """Create a temporary copy with malformed comment/person metadata removed/fixed."""
    fd, temp_name = tempfile.mkstemp(prefix="ubkitrans_maret_", suffix=".xlsx")
    os.close(fd)
    target = Path(temp_name)

    person_rel_pattern = re.compile(
        rb'<Relationship\b[^>]*(?:Type="[^"]*/person"|Target="(?:\.\./)?persons/person\.xml")[^>]*/>',
        re.IGNORECASE,
    )
    person_ct_pattern = re.compile(
        rb'<Override\b[^>]*PartName="/xl/persons/person\.xml"[^>]*/>',
        re.IGNORECASE,
    )

    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            name = info.filename
            if name == "xl/persons/person.xml":
                continue
            data = zin.read(name)
            if name == "xl/_rels/workbook.xml.rels":
                data = person_rel_pattern.sub(b"", data)
            elif name == "[Content_Types].xml":
                data = person_ct_pattern.sub(b"", data)
            elif name.startswith("xl/comments") and name.endswith(".xml"):
                data = re.sub(rb"<author\s*/>", b"<author>PLN Batam</author>", data)
                data = re.sub(rb"<author>\s*</author>", b"<author>PLN Batam</author>", data)
            zout.writestr(info, data)
    return target


def find_start_row(ws):
    for row in range(1, min(ws.max_row, 80) + 1):
        if "start pengisian" in norm(ws.cell(row, 1).value):
            return row
    # UBKITRANS workbook has its source data starting at row 11 (III.A) / row 10 (III.B).
    return 1


def scale_by_level(model, value):
    level = integer_or_none(value)
    if level is None:
        return None
    obj = model.objects.filter(urutan=level).first()
    if obj is None:
        raise RuntimeError(f"Master scale {model.__name__} urutan={level} tidak ditemukan.")
    return obj


def treatment_effectiveness(raw):
    t = norm(raw)
    if not t:
        return None
    if "cukup efektif" in t:
        return "cukup_efektif"
    if "tidak efektif" in t:
        return "tidak_efektif"
    if t == "efektif" or t.startswith("efektif "):
        return "efektif"
    return None


def treatment_status(raw_values):
    values = [norm(v) for v in raw_values if txt(v)]
    if any("discontinue" in v for v in values):
        return "discontinue"
    if any("continue" in v or "rutin" in v for v in values):
        return "continue"
    return None


def parse_source(workbook):
    for required in ("III.A", "III.B", "III.D", "III.E"):
        if required not in workbook.sheetnames:
            raise RuntimeError(f"Sheet wajib {required!r} tidak ditemukan.")

    # III.A: source event rows carry an ID/no in column B and event in C.
    ws_a = workbook["III.A"]
    events = []
    for r in range(1, ws_a.max_row + 1):
        row = [ws_a.cell(r, c).value for c in range(1, ws_a.max_column + 1)]
        if r >= 11 and txt(row[1]) and txt(row[IIIA_EVENT]):
            events.append((r, row))
    if len(events) != EXPECTED_EVENTS:
        raise RuntimeError(f"III.A harus {EXPECTED_EVENTS} risk event; ditemukan {len(events)}.")
    event_map = {}
    for r, row in events:
        key = norm(row[IIIA_EVENT])
        if key in event_map:
            raise RuntimeError(f"III.A event duplikat setelah normalisasi pada row {r}: {row[IIIA_EVENT]!r}")
        event_map[key] = (r, row)

    # KRI header must be exactly for March.
    ws_b = workbook["III.B"]
    kri_header_found = False
    for r in range(1, min(ws_b.max_row, 12) + 1):
        for c in range(1, ws_b.max_column + 1):
            if norm(ws_b.cell(r, c).value) == norm("Realisasi Threshold KRI Maret"):
                kri_header_found = True
                break
        if kri_header_found:
            break
    if not kri_header_found:
        raise RuntimeError("Header 'Realisasi Threshold KRI Maret' tidak ditemukan pada III.B.")

    # A cause group starts when cause text exists and a cause marker exists.
    starts = []
    current_event = None
    for r in range(10, ws_b.max_row + 1):
        values = [ws_b.cell(r, c).value for c in range(1, ws_b.max_column + 1)]
        if txt(values[B_EVENT]):
            current_event = txt(values[B_EVENT])
        cause_text = txt(values[B_CAUSE_TEXT])
        marker = txt(values[B_CAUSE_NO]) or txt(values[B_CAUSE_CODE]) or txt(values[B_RISK_NO])
        if cause_text and marker:
            if not current_event:
                raise RuntimeError(f"III.B row {r}: cause group tanpa event induk.")
            starts.append((r, current_event, values))

    if len(starts) != EXPECTED_CAUSE_GROUPS:
        raise RuntimeError(f"III.B harus {EXPECTED_CAUSE_GROUPS} cause group; ditemukan {len(starts)}.")

    groups = []
    for idx, (start_row, event, start_values) in enumerate(starts, 1):
        end_row = starts[idx][0] - 1 if idx < len(starts) else ws_b.max_row
        rows = [
            [ws_b.cell(r, c).value for c in range(1, ws_b.max_column + 1)]
            for r in range(start_row, end_row + 1)
        ]
        groups.append(
            SourceGroup(
                index=idx,
                start_row=start_row,
                end_row=end_row,
                risk_no=start_values[B_RISK_NO],
                event=event,
                cause_no=txt(start_values[B_CAUSE_NO]),
                cause_code=txt(start_values[B_CAUSE_CODE]),
                cause_text=txt(start_values[B_CAUSE_TEXT]),
                rows=rows,
            )
        )

    # Structural hard guards.
    formal = 0
    actual_only = 0
    planned_total = Decimal("0")
    actual_total = Decimal("0")
    progress_groups = 0
    kri_groups = 0
    for g in groups:
        if g.rows[0][B_Q1_PROGRESS] not in (None, ""):
            progress_groups += 1
        if txt(g.rows[0][B_KRI_STATUS]) or txt(g.rows[0][B_KRI_VALUE]):
            kri_groups += 1
        for row in g.rows:
            plan = txt(row[B_PLAN])
            actual_plan = txt(row[B_ACTUAL_PLAN])
            actual_output = txt(row[B_ACTUAL_OUTPUT])
            actual_cost = money_cent(row[B_ACTUAL_COST])
            if plan:
                formal += 1
                p = money_cent(row[B_PLAN_COST])
                if p is not None:
                    planned_total += p
            elif actual_plan or actual_output or (actual_cost not in (None, Decimal("0.00"))):
                actual_only += 1
            if actual_cost is not None:
                actual_total += actual_cost

    planned_total = planned_total.quantize(Decimal("0.01"))
    actual_total = actual_total.quantize(Decimal("0.01"))
    guards = {
        "formal activities": (formal, EXPECTED_FORMAL_ACTIVITIES),
        "actual-only rows": (actual_only, EXPECTED_ACTUAL_ONLY_ROWS),
        "Q1 progress groups": (progress_groups, EXPECTED_PROGRESS_GROUPS),
        "KRI groups": (kri_groups, EXPECTED_KRI_GROUPS),
        "planned cost": (planned_total, EXPECTED_PLANNED_COST),
        "actual cost": (actual_total, EXPECTED_ACTUAL_COST),
    }
    for label, (actual, expected) in guards.items():
        if actual != expected:
            raise RuntimeError(f"Source guard gagal — {label}: actual={actual!r}, expected={expected!r}")

    # III.D / III.E must not carry real source rows.
    changes = 0
    ws_d = workbook["III.D"]
    for r in range(1, ws_d.max_row + 1):
        # Actual change rows use a meaningful change type in B and explanation in C/D.
        row_text = " ".join(norm(ws_d.cell(r, c).value) for c in range(1, ws_d.max_column + 1))
        if any(k in row_text for k in ("perubahan profil risiko", "penambahan item risiko", "pengurangan item risiko", "perubahan strategi risiko")):
            # Instruction row contains these labels; only count if row has substantive cells beyond the label.
            if r > 7 and any(txt(ws_d.cell(r, c).value) for c in range(2, ws_d.max_column + 1)):
                changes += 1
    losses = 0
    ws_e = workbook["III.E"]
    for r in range(1, ws_e.max_row + 1):
        # Data rows have event name in B after the template header area.
        if r > 8 and txt(ws_e.cell(r, 2).value):
            losses += 1
    if changes or losses:
        raise RuntimeError(f"III.D/III.E source expected kosong; changes={changes}, loss_events={losses}.")

    return event_map, groups, {
        "events": len(events),
        "groups": len(groups),
        "formal": formal,
        "actual_only": actual_only,
        "progress": progress_groups,
        "kri": kri_groups,
        "planned": planned_total,
        "actual": actual_total,
    }



# ---------------------------------------------------------------------------
# Historical event alias.
#
# Source Maret:
#   Penyerapan gas dalam BTU/kWh ...
#
# Current UBKITRANS profile:
#   Heatrate gas (BTU/kWh) ...
#
# Hanya event-name yang diberi alias.
# Cause dan cause_no tetap harus cocok dengan canonical report.
# ---------------------------------------------------------------------------

EVENT_ALIASES = {
    norm(
        "Penyerapan gas dalam BTU/kWh oleh pembangkit sendiri dan sewa "
        "melebihi asumsi yang ditetapkan"
    ): {
        norm(
            "Penyerapan gas dalam BTU/kWh oleh pembangkit sendiri dan sewa "
            "melebihi asumsi yang ditetapkan"
        ),
        norm(
            "Heatrate gas (BTU/kWh) pada pembangkit sendiri dan sewa "
            "melebihi asumsi yang ditetapkan"
        ),
    },
}


def event_matches(source_event, canonical_event):
    source_key = norm(source_event)
    canonical_key = norm(canonical_event)

    aliases = EVENT_ALIASES.get(source_key)
    if aliases is not None:
        return canonical_key in aliases

    return canonical_key == source_key


def build_canonical_mapping(groups):
    canonical_report = MonthlyRiskReport.objects.select_related("reassessment").get(pk=CANONICAL_REPORT_ID)
    if canonical_report.reassessment_id != PROFILE_ID:
        raise RuntimeError(
            f"Canonical MRR {CANONICAL_REPORT_ID} profile={canonical_report.reassessment_id}, expected={PROFILE_ID}."
        )
    canonical_items = list(
        MonthlyRiskReportItem.objects.filter(report=canonical_report)
        .select_related("risk_event", "risk_event__summary", "risk_event__km_item")
        .order_by("id")
    )
    if len(canonical_items) != EXPECTED_CAUSE_GROUPS:
        raise RuntimeError(
            f"Canonical MRR {CANONICAL_REPORT_ID} harus {EXPECTED_CAUSE_GROUPS} item; ditemukan {len(canonical_items)}."
        )

    unused = {x.risk_event_id: x.risk_event for x in canonical_items}
    mapped = []
    for group in groups:
        event_key = norm(group.event)
        cause_key = norm(group.cause_text)
        event_candidates = [
            revent for revent in unused.values()
            if event_matches(group.event, revent.peristiwa_risiko)
        ]
        cause_candidates = [
            revent for revent in event_candidates
            if norm(revent.penyebab_risiko) == cause_key
        ]
        chosen = None
        reason = None
        if len(cause_candidates) == 1:
            chosen = cause_candidates[0]
            reason = "event+cause exact"
        elif group.cause_no:
            by_letter = [
                revent for revent in event_candidates
                if norm(revent.no_penyebab_risiko) == norm(group.cause_no)
            ]
            if len(by_letter) == 1:
                chosen = by_letter[0]
                reason = "event+cause_no"
        if chosen is None and len(event_candidates) == 1:
            chosen = event_candidates[0]
            reason = "unique event"
        if chosen is None:
            details = ", ".join(
                f"RE={x.id} cause_no={x.no_penyebab_risiko!r} cause={(x.penyebab_risiko or '')[:80]!r}"
                for x in event_candidates
            ) or "(tidak ada kandidat event)"
            raise RuntimeError(
                f"Mapping SRC {group.index:02d} row {group.start_row} gagal. "
                f"event={group.event!r}; cause_no={group.cause_no!r}; cause={group.cause_text!r}; candidates={details}"
            )
        if chosen.summary_id != PROFILE_ID:
            raise RuntimeError(f"RE={chosen.id} bukan milik profile {PROFILE_ID}.")
        group.master = chosen
        unused.pop(chosen.id, None)
        mapped.append((group, chosen, reason))

    if unused:
        raise RuntimeError(
            "Canonical mapping tidak habis 40 item; sisa RE=" + ",".join(map(str, sorted(unused)))
        )
    return canonical_report, mapped


def aggregate_treatment(group: SourceGroup):
    activities = []
    current = None
    plan_cost_values = []
    actual_cost_values = []
    pics = []
    raw_statuses = []
    status_notes = []

    for row in group.rows:
        plan = txt(row[B_PLAN])
        planned_output = txt(row[B_OUTPUT])
        actual_plan = txt(row[B_ACTUAL_PLAN])
        actual_output = txt(row[B_ACTUAL_OUTPUT])
        plan_cost = money_cent(row[B_PLAN_COST])
        actual_cost = money_cent(row[B_ACTUAL_COST])
        pic = txt(row[B_PIC])
        raw_status = txt(row[B_STATUS])
        status_note = txt(row[B_STATUS_NOTE])

        if plan:
            current = {
                "plan": plan,
                "planned_output": planned_output,
                "actual_plan": [],
                "actual_output": [],
            }
            if actual_plan:
                current["actual_plan"].append(actual_plan)
            if actual_output:
                current["actual_output"].append(actual_output)
            activities.append(current)
            if plan_cost is not None:
                plan_cost_values.append(plan_cost)
        else:
            # Source continuation row: append actual realization to preceding formal activity;
            # never create a new activity from K/L-only continuation rows.
            if current is not None:
                if actual_plan:
                    current["actual_plan"].append(actual_plan)
                if actual_output:
                    current["actual_output"].append(actual_output)

        if actual_cost is not None:
            actual_cost_values.append(actual_cost)
        if pic:
            pics.append(pic)
        if raw_status:
            raw_statuses.append(raw_status)
        if status_note:
            status_notes.append(status_note)

    plan_parts = []
    output_parts = []
    for i, activity in enumerate(activities, 1):
        block = [f"{i:02d}. Rencana: {activity['plan']}"]
        if activity["actual_plan"]:
            block.append("    Realisasi source: " + "\n    ".join(unique_texts(activity["actual_plan"])))
        plan_parts.append("\n".join(block))

        if activity["planned_output"] or activity["actual_output"]:
            oblock = [f"{i:02d}. Output rencana: {activity['planned_output'] or '-'}"]
            if activity["actual_output"]:
                oblock.append("    Realisasi output: " + "\n    ".join(unique_texts(activity["actual_output"])))
            output_parts.append("\n".join(oblock))

    planned_cost = sum(plan_cost_values, Decimal("0.00")) if plan_cost_values else None
    actual_cost = sum(actual_cost_values, Decimal("0.00")) if actual_cost_values else None
    absorption = None
    if planned_cost is not None and actual_cost is not None:
        if planned_cost != 0:
            absorption = (actual_cost / planned_cost * Decimal("100")).quantize(Decimal("0.01"))
        elif actual_cost == 0:
            absorption = Decimal("0.00")

    # Item-level monthly timeline is a checkbox flag: any positive activity in that month => 1.
    timeline = {}
    for month in range(1, 13):
        if month > MONTH:
            timeline[month] = 0
            continue
        col = B_TIMELINE_FIRST + month - 1
        positive = False
        for row in group.rows:
            v = decimal_or_none(row[col])
            if v is not None and v > 0:
                positive = True
                break
        timeline[month] = 1 if positive else 0

    return {
        "realisasi_rencana_perlakuan": "\n\n".join(plan_parts) or None,
        "realisasi_output_perlakuan": "\n\n".join(output_parts) or None,
        "rencana_biaya_perlakuan": planned_cost,
        "realisasi_biaya_perlakuan": actual_cost,
        "persentase_serapan_biaya": absorption,
        "realisasi_pic": "\n".join(unique_texts(pics)) or None,
        "status_rencana_perlakuan": treatment_status(raw_statuses),
        "penjelasan_status_rencana": "\n\n".join(unique_texts(status_notes)) or None,
        "progress_pelaksanaan_percent": percent_or_none(group.rows[0][B_Q1_PROGRESS]),
        "timeline": timeline,
        "raw_statuses": raw_statuses,
    }


def kri_values(group: SourceGroup):
    row = group.rows[0]
    name = txt(row[B_KRI])
    unit = txt(row[B_KRI_UNIT])
    safe = txt(row[B_KRI_SAFE])
    caution = txt(row[B_KRI_CAUTION])
    danger = txt(row[B_KRI_DANGER])
    status = txt(row[B_KRI_STATUS])
    actual = txt(row[B_KRI_VALUE])
    if not any((name, unit, safe, caution, danger, status, actual)):
        return {
            "realisasi_threshold_kri": None,
            "realisasi_nilai_kri": None,
            "realisasi_kri_text": None,
            "realisasi_threshold_kri_skor": None,
        }
    text = (
        f"KRI sumber: {name or '-'}\n"
        f"Satuan: {unit or '-'}\n"
        f"Aman: {safe or '-'}\n"
        f"Hati-hati: {caution or '-'}\n"
        f"Bahaya: {danger or '-'}\n"
        f"Realisasi bulan source: kategori={status!r}; skor={actual!r}"
    )
    return {
        # Preserve source category/status literally; numeric evaluation is intentionally not inferred.
        "realisasi_threshold_kri": status,
        "realisasi_nilai_kri": None,
        "realisasi_kri_text": text,
        "realisasi_threshold_kri_skor": None,
    }


def residual_values(event_row, master):
    risk_type = norm(getattr(master, "jenis_risiko", None))
    if risk_type not in {"kuantitatif", "kualitatif"}:
        source_type = norm(event_row[IIIA_RISK_TYPE])
        risk_type = "kualitatif" if "kual" in source_type else "kuantitatif" if "kuant" in source_type else None

    level = txt(event_row[IIIA_Q1_LEVEL])
    eff_raw = txt(event_row[IIIA_EFFECTIVENESS])
    values = {
        "jenis_risiko": risk_type,
        "realisasi_asumsi_dampak": txt(event_row[IIIA_ASSUMPTION]),
        "realisasi_nilai_dampak": decimal_or_none(event_row[IIIA_Q1_IMPACT]),
        "realisasi_skala_dampak": scale_by_level(MasterSkalaDampak, event_row[IIIA_Q1_IMPACT_SCALE]),
        "realisasi_nilai_probabilitas": percent_or_none(event_row[IIIA_Q1_PROB]),
        "realisasi_skala_probabilitas": scale_by_level(MasterSkalaProbabilitas, event_row[IIIA_Q1_PROB_SCALE]),
        "realisasi_skala_dampak_kbumn": integer_or_none(event_row[IIIA_Q1_IMPACT_KBUMN]),
        "realisasi_skala_probabilitas_kbumn": integer_or_none(event_row[IIIA_Q1_PROB_KBUMN]),
        "realisasi_eksposur": decimal_or_none(event_row[IIIA_Q1_EXPOSURE]),
        "realisasi_skor_risiko": integer_or_none(event_row[IIIA_Q1_SCORE]),
        "realisasi_skala_nilai_risiko_kbumn": integer_or_none(event_row[IIIA_Q1_SCORE_KBUMN]),
        "realisasi_level_risiko": level,
        "realisasi_level_risiko_bumn": level,
        "realisasi_level_risiko_kbumn": txt(event_row[IIIA_Q1_LEVEL_KBUMN]),
        "efektivitas_perlakuan_risiko": treatment_effectiveness(eff_raw),
    }
    eff_note = None
    if eff_raw and values["efektivitas_perlakuan_risiko"] is None:
        eff_note = eff_raw
    return values, eff_note


def update_item(item, values):
    fields = {f.name for f in MonthlyRiskReportItem._meta.fields}
    changed = []
    for name, value in values.items():
        if name not in fields:
            continue
        old = getattr(item, name)
        old_cmp = old.pk if hasattr(old, "pk") else old
        new_cmp = value.pk if hasattr(value, "pk") else value
        if old_cmp != new_cmp:
            setattr(item, name, value)
            changed.append(name)
    if changed:
        item.save(update_fields=changed + (["updated_at"] if "updated_at" in fields else []))
    return changed


def create_sqlite_backup_if_possible():
    db = settings.DATABASES.get("default", {})
    if db.get("ENGINE") != "django.db.backends.sqlite3":
        print("BACKUP: DB bukan SQLite; file backup otomatis dilewati.")
        return None
    source = Path(db.get("NAME"))
    if not source.exists():
        raise RuntimeError(f"SQLite database tidak ditemukan: {source}")
    backup_dir = Path("/home/adminsvr/backup") if Path("/home/adminsvr").exists() else PROJECT_ROOT / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"db_before_import_mrr_ubkitrans_maret_2026_{stamp}.sqlite3"
    shutil.copy2(source, target)
    print(f"BACKUP: {target}")
    return target


def verify_target_report():
    report = (
        MonthlyRiskReport.objects.select_related("reassessment", "periode", "kontrak_manajemen")
        .get(pk=TARGET_REPORT_ID)
    )
    if report.reassessment_id != PROFILE_ID:
        raise RuntimeError(f"MRR {TARGET_REPORT_ID} profile={report.reassessment_id}, expected={PROFILE_ID}")
    period_code = getattr(report.periode, "kode_periode", "") or ""
    period_name = getattr(report.periode, "nama_periode", "") or ""
    if period_code != "2026-03" and "maret" not in norm(period_name):
        raise RuntimeError(f"MRR {TARGET_REPORT_ID} bukan periode Maret 2026: {period_code!r} / {period_name!r}")
    if report.status not in {"draft", "revision"}:
        raise RuntimeError(f"MRR {TARGET_REPORT_ID} status={report.status!r}; hanya draft/revision yang boleh diubah.")
    if getattr(report, "is_locked", False):
        raise RuntimeError(f"MRR {TARGET_REPORT_ID} sedang locked.")
    km_id = getattr(report, "kontrak_manajemen_id", None)
    if km_id not in (None, 10):
        raise RuntimeError(f"MRR {TARGET_REPORT_ID} kontrak_manajemen_id={km_id}, expected 10/None.")
    return report


def main():
    parser = argparse.ArgumentParser(description="Safe import MRR UB KITRAN Maret 2026")
    parser.add_argument("--source", required=True, type=Path, help="Path file XLSX sumber Maret 2026")
    parser.add_argument("--apply", action="store_true", help="Commit. Default: DRY RUN / rollback.")
    parser.add_argument(
        "--allow-different-hash",
        action="store_true",
        help="Bypass SHA256 guard only if intentionally using a revised source with identical structure.",
    )
    args = parser.parse_args()
    source = args.source.expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    source_hash = file_sha256(source)
    if source_hash != EXPECTED_SOURCE_SHA256 and not args.allow_different_hash:
        raise RuntimeError(
            "SHA256 source berbeda dari file yang diaudit. "
            f"actual={source_hash}; expected={EXPECTED_SOURCE_SHA256}. "
            "Jangan apply sebelum file revisi diaudit; jika memang disengaja gunakan --allow-different-hash untuk DRY RUN dulu."
        )

    print("=" * 150)
    print(f"IMPORT MRR UB KITRAN MARET {YEAR} V2 SAFE — {'APPLY' if args.apply else 'DRY RUN'}")
    print("=" * 150)
    print(f"SOURCE : {source}")
    print(f"SHA256 : {source_hash}")

    sanitized = sanitize_xlsx_for_openpyxl(source)
    try:
        workbook = load_workbook(sanitized, data_only=True, read_only=False, keep_links=False)
        event_map, groups, audit = parse_source(workbook)
    finally:
        try:
            sanitized.unlink(missing_ok=True)
        except Exception:
            pass

    report = verify_target_report()
    canonical_report, mapped = build_canonical_mapping(groups)

    print("\nSOURCE AUDIT")
    print(f"  III.A risk events            : {audit['events']}")
    print(f"  III.B cause groups           : {audit['groups']}")
    print(f"  Formal treatment rows (H)    : {audit['formal']}")
    print(f"  Actual-only continuation     : {audit['actual_only']} (tidak dibuat activity baru)")
    print(f"  Q1 progress populated        : {audit['progress']} / {audit['groups']}")
    print(f"  KRI Maret populated groups   : {audit['kri']} / {audit['groups']}")
    print(f"  Planned cost total           : {audit['planned']}")
    print(f"  Actual cost total            : {audit['actual']}")
    print("  III.D / III.E                : 0 / 0")
    print("  Timeline policy              : Jan-Mar source flags; Apr-Dec = 0")
    print("  KRI policy                   : preserve source status + text; no inferred numeric direction")

    print("\nTARGET / CANONICAL")
    print(f"  Target MRR                   : {report.id} {report.kode} | items_before={report.items.count()}")
    print(f"  Profile                      : {report.reassessment_id} {report.reassessment.judul}")
    print(f"  Canonical mapping report     : {canonical_report.id} {canonical_report.kode} | items=40")

    print("\nCANONICAL MAPPING")
    for group, master, reason in mapped:
        print(
            f"  SRC {group.index:02d} row={group.start_row:<3} -> RE={master.id:<4} "
            f"cause={str(master.no_penyebab_risiko or '-'):>3} | {reason:<18} | "
            f"event={group.event[:72]!r}"
        )

    if args.apply:
        create_sqlite_backup_if_possible()

    created = 0
    updated = 0
    field_changes = 0
    source_master_ids = {master.id for _, master, _ in mapped}

    with transaction.atomic():
        # Clear III.D/III.E child rows because audited source is empty.
        MonthlyRiskReportChange.objects.filter(report=report).delete()
        MonthlyRiskReportLossEvent.objects.filter(report=report).delete()

        for group, master, _reason in mapped:
            event_key = norm(group.event)
            if event_key not in event_map:
                raise RuntimeError(
                    f"SRC {group.index:02d}: event III.B tidak ditemukan di III.A: {group.event!r}"
                )
            iiia_row_no, iiia_row = event_map[event_key]

            item, was_created = MonthlyRiskReportItem.objects.get_or_create(
                report=report,
                risk_event=master,
                defaults={"km_item": getattr(master, "km_item", None)},
            )
            if was_created:
                created += 1

            values, eff_note = residual_values(iiia_row, master)
            treatment = aggregate_treatment(group)
            timeline = treatment.pop("timeline")
            raw_statuses = treatment.pop("raw_statuses")
            values.update(treatment)
            values.update(kri_values(group))
            values["km_item"] = getattr(master, "km_item", None)

            # Preserve source notes that cannot legally fit enum fields.
            extra_notes = []
            if values.get("penjelasan_status_rencana"):
                extra_notes.append(values["penjelasan_status_rencana"])
            if eff_note:
                extra_notes.append(f"[Catatan efektivitas source III.A row {iiia_row_no}: {eff_note}]")
            unrepresented_statuses = [
                s for s in unique_texts(raw_statuses)
                if "continue" not in norm(s) and "discontinue" not in norm(s) and "rutin" not in norm(s)
            ]
            if unrepresented_statuses:
                extra_notes.append("[Status source III.B: " + " | ".join(unrepresented_statuses) + "]")
            values["penjelasan_status_rencana"] = "\n\n".join(extra_notes) or None

            for month, flag in timeline.items():
                values[f"realisasi_timeline_{month}"] = flag

            changed = update_item(item, values)
            if changed:
                updated += 1
                field_changes += len(changed)

        # Hard guard: target report may contain only the canonical 40 risk_event rows.
        current_ids = set(report.items.values_list("risk_event_id", flat=True))
        extras = sorted(current_ids - source_master_ids)
        missing = sorted(source_master_ids - current_ids)
        if extras or missing:
            raise RuntimeError(f"Final item-set guard gagal: extras={extras}, missing={missing}")
        if report.items.count() != EXPECTED_CAUSE_GROUPS:
            raise RuntimeError(
                f"Final report item count={report.items.count()}, expected={EXPECTED_CAUSE_GROUPS}."
            )

        refresh_monthly_report_summary(report)
        report.refresh_from_db()

        if not args.apply:
            transaction.set_rollback(True)

    print("\n" + "=" * 150)
    print("RESULT")
    print("=" * 150)
    print(f"MRR ID                       : {report.id}")
    print(f"MRR code                     : {report.kode}")
    print(f"Existing items before        : 3 (historical baseline expected)")
    print(f"Items created                : {created}")
    print(f"Items updated/touched        : {updated}")
    print(f"Field changes                : {field_changes}")
    print(f"Final items                  : {EXPECTED_CAUSE_GROUPS}")
    print(f"Database                     : {'TERSIMPAN' if args.apply else 'BELUM DIUBAH — DRY RUN ROLLBACK'}")
    if args.apply:
        print("STATUS                       : APPLY BERHASIL")
    else:
        print("STATUS                       : DRY RUN BERHASIL")
        print("NEXT                         : review mapping/count; then rerun command with --apply")


if __name__ == "__main__":
    try:
        main()
    except (RuntimeError, ValidationError, FileNotFoundError) as exc:
        print(f"\nSTOP: {exc}", file=sys.stderr)
        sys.exit(2)
