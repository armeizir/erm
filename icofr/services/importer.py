from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from django.core.files.base import File
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from icofr.models import (
    RCMControl,
    RCMControlAttribute,
    RCMEntry,
    RCMEntryAssertion,
    RCMImportBatch,
    RCMRisk,
    RCMSet,
    RCMSupportingDocument,
    RCMType,
)


HEADER_ALIASES = {
    "entitas": "entity_name",
    "no sub proses": "subprocess_number",
    "nomor sub proses": "subprocess_number",
    "deskripsi sub proses": "subprocess_description",
    "deskripsi akun": "account_description",
    "asersi": "assertions_raw",
    "ref risiko": "risk_reference",
    "referensi risiko": "risk_reference",
    "deskripsi risiko": "risk_description",
    "tujuan coso": "coso_objective",
    "komponen coso": "coso_component",
    "elemen coso": "coso_element",
    "area kontrol": "control_area",
    "area control": "control_area",
    "sub area kontrol": "control_sub_area",
    "sub area control": "control_sub_area",
    "risiko kecurangan": "fraud_risk",
    "dampak": "impact",
    "kemungkinan terjadi": "likelihood",
    "tingkat risiko": "risk_level",
    "ref kontrol": "control_reference",
    "referensi kontrol": "control_reference",
    "tujuan kontrol": "control_objective",
    "deskripsi kontrol": "control_description",
    "jenis kontrol": "control_type",
    "kontrol utama": "is_key_control",
    "anti kecurangan": "anti_fraud",
    "aplikasi pendukung": "supporting_application",
    "lokasi": "location",
    "deskripsi lokasi": "location_description",
    "frekuensi": "frequency",
    "control preparer": "preparer_position",
    "control reviewer": "reviewer_position",
    "atribut kontrol": "attributes_raw",
    "dokumen pendukung": "documents_raw",
    "kontrol kompensasi": "compensating_control",
    "segmen": "segment",
}

REQUIRED_COMMON = {"entity_name", "risk_reference", "control_reference"}
REQUIRED_BY_TYPE = {
    RCMType.TLC: {"subprocess_number", "subprocess_description", "account_description", "assertions_raw"},
    RCMType.ELC: {"coso_element"},
    RCMType.ITGC: {"control_area", "control_sub_area"},
}


@dataclass
class ParsedRCM:
    rcm_type: str
    version: str
    rows: list[dict]
    header_row: int
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def summary(self):
        risks = {row.get("risk_reference") for row in self.rows if row.get("risk_reference")}
        controls = {row.get("control_reference") for row in self.rows if row.get("control_reference")}
        pairs = {
            (row.get("risk_reference"), row.get("control_reference"))
            for row in self.rows
            if row.get("risk_reference") and row.get("control_reference")
        }
        return {
            "rows": len(self.rows),
            "risks": len(risks),
            "controls": len(controls),
            "risk_control_pairs": len(pairs),
            "segments": len({row.get("segment") for row in self.rows if row.get("segment")}),
            "preparer_positions": len({row.get("preparer_position") for row in self.rows if row.get("preparer_position")}),
            "reviewer_positions": len({row.get("reviewer_position") for row in self.rows if row.get("reviewer_position")}),
        }


def normalize_header(value):
    text = str(value or "").strip().lower()
    text = text.replace(".", " ").replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_yes_no(value):
    normalized = text(value).lower()
    if normalized in {"ya", "yes", "y", "1", "true"}:
        return True
    if normalized in {"tidak", "no", "n", "0", "false"}:
        return False
    return None


def split_items(value):
    raw = text(value)
    if not raw or raw.upper() in {"N/A", "NA", "-"}:
        return []
    parts = [part.strip(" \t\r\n-•") for part in re.split(r";|\n(?=\s*(?:\d+[.)]|[-•]))", raw)]
    return [part for part in parts if part]


def file_sha256(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def detect_metadata(ws):
    rcm_type = ""
    version = ""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        row_values = [text(v) for v in row]
        joined = " | ".join(row_values).lower()
        if "transaction level control" in joined:
            rcm_type = RCMType.TLC
        elif "entity level control" in joined:
            rcm_type = RCMType.ELC
        elif "it general control" in joined:
            rcm_type = RCMType.ITGC

        if any(value.lower() == "versi" for value in row_values if value):
            candidates = [
                value.lstrip(": ")
                for value in row_values
                if value and value.lower() != "versi" and value not in {":"}
            ]
            if candidates:
                version = candidates[-1]
    return rcm_type, version


def find_header(ws):
    max_scan = min(ws.max_row, 40)
    matrix = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for row_no, row in enumerate(matrix, start=1):
        values = [normalize_header(value) for value in row]
        normalized = {HEADER_ALIASES.get(value) for value in values if value}
        if "risk_reference" in normalized and "control_reference" in normalized:
            mapping = {}
            max_cols = max(len(row), ws.max_column)
            for col in range(1, max_cols + 1):
                # Merged headers such as Entitas can live 1-2 rows above the detailed header row.
                field_name = None
                for lookback in range(0, 3):
                    source_row = row_no - 1 - lookback
                    if source_row < 0 or source_row >= len(matrix):
                        continue
                    source = matrix[source_row]
                    value = normalize_header(source[col - 1] if col - 1 < len(source) else None)
                    candidate = HEADER_ALIASES.get(value)
                    if candidate:
                        field_name = candidate
                        break
                if field_name and field_name not in mapping:
                    mapping[field_name] = col
            return row_no, mapping
    return None, {}


def parse_workbook(path) -> ParsedRCM:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rcm_type, version = detect_metadata(ws)
    header_row, column_map = find_header(ws)
    errors = []
    warnings = []

    if not rcm_type:
        errors.append("Jenis RCM tidak dapat dideteksi dari metadata workbook.")
    if not version:
        errors.append("Versi RCM tidak dapat dideteksi dari metadata workbook.")
    if not header_row:
        errors.append("Header RCM tidak ditemukan. Pastikan ada kolom Ref. Risiko dan Ref. Kontrol.")
        return ParsedRCM(rcm_type, version, [], 0, errors, warnings)

    required = set(REQUIRED_COMMON)
    required |= set(REQUIRED_BY_TYPE.get(rcm_type, set()))
    missing = sorted(required - set(column_map))
    if missing:
        errors.append("Kolom wajib belum tersedia: " + ", ".join(missing))

    rows = []
    if not errors:
        for row_no, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            payload = {
                field: text(values[col - 1] if col - 1 < len(values) else None)
                for field, col in column_map.items()
            }
            if not any(payload.values()):
                continue
            if not payload.get("risk_reference") and not payload.get("control_reference"):
                continue
            if not payload.get("risk_reference"):
                errors.append(f"Baris {row_no}: Referensi Risiko kosong.")
                continue
            if not payload.get("control_reference"):
                errors.append(f"Baris {row_no}: Referensi Kontrol kosong.")
                continue
            payload["source_row_number"] = row_no
            rows.append(payload)

    if not rows and not errors:
        errors.append("Tidak ada baris data RCM yang dapat diimpor.")

    # Canonical-definition variation is preserved in raw_payload, but warned for review.
    for ref_field, compare_fields, label in [
        ("risk_reference", ("risk_description", "coso_objective", "coso_component"), "risiko"),
        ("control_reference", ("control_objective", "control_description", "control_type"), "kontrol"),
    ]:
        seen = {}
        varied = set()
        for row in rows:
            ref = row.get(ref_field)
            signature = tuple(row.get(field, "") for field in compare_fields)
            if ref in seen and seen[ref] != signature:
                varied.add(ref)
            seen.setdefault(ref, signature)
        if varied:
            warnings.append(
                f"{len(varied)} referensi {label} memiliki variasi definisi antarbaris. "
                "Definisi pertama menjadi canonical; seluruh nilai asli tetap disimpan pada raw payload."
            )

    return ParsedRCM(rcm_type, version, rows, header_row, errors, warnings)


def validate_batch(batch: RCMImportBatch):
    parsed = parse_workbook(batch.upload.path)
    errors = list(parsed.errors)
    warnings = list(parsed.warnings)
    if parsed.rcm_type and parsed.version:
        existing = RCMSet.objects.filter(rcm_type=parsed.rcm_type, version=parsed.version).first()
        if existing:
            errors.append(
                f"RCM {parsed.rcm_type} versi {parsed.version} sudah ada (ID {existing.pk}). "
                "Gunakan versi baru; RCM existing tidak ditimpa."
            )
    batch.detected_type = parsed.rcm_type
    batch.detected_version = parsed.version
    batch.row_count = len(parsed.rows)
    batch.summary = parsed.summary
    batch.validation_errors = errors
    batch.validation_warnings = warnings
    batch.source_sha256 = file_sha256(batch.upload.path)
    batch.status = RCMImportBatch.Status.FAILED if errors else RCMImportBatch.Status.VALIDATED
    batch.save()
    return parsed


def _fingerprint(row):
    canonical = json.dumps(row, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _first_nonempty(existing, value):
    return existing or value or ""


@transaction.atomic
def import_batch(batch: RCMImportBatch, *, user=None):
    batch = RCMImportBatch.objects.select_for_update().get(pk=batch.pk)
    if not batch.can_import:
        raise ValueError("Batch belum tervalidasi, memiliki error, atau sudah pernah diimpor.")
    parsed = parse_workbook(batch.upload.path)
    if parsed.errors:
        raise ValueError("Workbook tidak lagi valid: " + "; ".join(parsed.errors))
    if RCMSet.objects.filter(rcm_type=parsed.rcm_type, version=parsed.version).exists():
        raise ValueError(f"RCM {parsed.rcm_type} versi {parsed.version} sudah ada.")

    first_entity = next((row.get("entity_name", "") for row in parsed.rows if row.get("entity_name")), "")
    rcm_set = RCMSet.objects.create(
        rcm_type=parsed.rcm_type,
        version=parsed.version,
        entity_name=first_entity,
        source_filename=batch.original_filename,
        source_sha256=batch.source_sha256 or file_sha256(batch.upload.path),
        source_row_count=len(parsed.rows),
        imported_by=user,
        imported_at=timezone.now(),
    )

    # Canonical risk/control definitions use the first occurrence; raw_payload on each
    # RCMEntry preserves the exact source row when definitions vary across occurrences.
    risk_seed = {}
    control_seed = {}
    for row in parsed.rows:
        risk_seed.setdefault(row["risk_reference"], row)
        control_seed.setdefault(row["control_reference"], row)

    RCMRisk.objects.bulk_create(
        [
            RCMRisk(
                rcm_set=rcm_set,
                reference=ref,
                description=row.get("risk_description", ""),
                coso_objective=row.get("coso_objective", ""),
                coso_component=row.get("coso_component", ""),
                fraud_risk=row.get("fraud_risk", ""),
                impact=row.get("impact", ""),
                likelihood=row.get("likelihood", ""),
                risk_level=row.get("risk_level", ""),
                coso_element=row.get("coso_element", ""),
                control_area=row.get("control_area", ""),
                control_sub_area=row.get("control_sub_area", ""),
            )
            for ref, row in risk_seed.items()
        ],
        batch_size=1000,
    )
    RCMControl.objects.bulk_create(
        [
            RCMControl(
                rcm_set=rcm_set,
                reference=ref,
                objective=row.get("control_objective", ""),
                description=row.get("control_description", ""),
                control_type=row.get("control_type", ""),
                is_key_control=parse_yes_no(row.get("is_key_control")),
                anti_fraud=parse_yes_no(row.get("anti_fraud")),
                supporting_application=row.get("supporting_application", ""),
            )
            for ref, row in control_seed.items()
        ],
        batch_size=1000,
    )

    risks = {obj.reference: obj for obj in RCMRisk.objects.filter(rcm_set=rcm_set)}
    controls = {obj.reference: obj for obj in RCMControl.objects.filter(rcm_set=rcm_set)}

    entry_rows = []
    entries = []
    for row in parsed.rows:
        entry_rows.append(row)
        entries.append(
            RCMEntry(
                rcm_set=rcm_set,
                risk=risks[row["risk_reference"]],
                control=controls[row["control_reference"]],
                entity_name=row.get("entity_name", ""),
                subprocess_number=row.get("subprocess_number", ""),
                subprocess_description=row.get("subprocess_description", ""),
                account_description=row.get("account_description", ""),
                assertions_raw=row.get("assertions_raw", ""),
                location=row.get("location", ""),
                location_description=row.get("location_description", ""),
                frequency=row.get("frequency", ""),
                preparer_position=row.get("preparer_position", ""),
                reviewer_position=row.get("reviewer_position", ""),
                compensating_control=row.get("compensating_control", ""),
                segment=row.get("segment", ""),
                source_row_number=row["source_row_number"],
                source_fingerprint=_fingerprint(row),
                raw_payload=row,
            )
        )
    RCMEntry.objects.bulk_create(entries, batch_size=500)

    # Re-read in source order so child rows always have stable database PKs on all supported DBs.
    persisted_entries = {
        entry.source_row_number: entry
        for entry in RCMEntry.objects.filter(rcm_set=rcm_set).order_by("source_row_number")
    }
    assertion_items = []
    attribute_items = []
    document_items = []
    for row in entry_rows:
        entry = persisted_entries[row["source_row_number"]]
        for sequence, item in enumerate(split_items(row.get("assertions_raw")), start=1):
            assertion_items.append(RCMEntryAssertion(entry=entry, sequence=sequence, text=item))
        for sequence, item in enumerate(split_items(row.get("attributes_raw")), start=1):
            attribute_items.append(RCMControlAttribute(entry=entry, sequence=sequence, text=item))
        for sequence, item in enumerate(split_items(row.get("documents_raw")), start=1):
            document_items.append(RCMSupportingDocument(entry=entry, sequence=sequence, text=item))

    RCMEntryAssertion.objects.bulk_create(assertion_items, batch_size=1000)
    RCMControlAttribute.objects.bulk_create(attribute_items, batch_size=1000)
    RCMSupportingDocument.objects.bulk_create(document_items, batch_size=1000)

    batch.status = RCMImportBatch.Status.IMPORTED
    batch.imported_rcm = rcm_set
    batch.save(update_fields=("status", "imported_rcm", "updated_at"))
    return rcm_set
