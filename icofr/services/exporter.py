from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from icofr.models import RCMType


COMMON_RISK = [
    ("risk_reference", "Ref. Risiko"),
    ("risk_description", "Deskripsi Risiko"),
]


def _bool_label(value):
    if value is True:
        return "Ya"
    if value is False:
        return "Tidak"
    return ""


def _headers(rcm_type):
    if rcm_type == RCMType.TLC:
        prefix = [
            ("entity_name", "Entitas"),
            ("subprocess_number", "No Sub Proses"),
            ("subprocess_description", "Deskripsi Sub Proses"),
            ("account_description", "Deskripsi Akun"),
            ("assertions_raw", "Asersi"),
        ]
    else:
        prefix = [("entity_name", "Entitas")]

    risk_extra = []
    if rcm_type == RCMType.ELC:
        risk_extra.append(("coso_element", "Elemen COSO"))
    elif rcm_type == RCMType.ITGC:
        risk_extra.extend([
            ("control_area", "Area Kontrol"),
            ("control_sub_area", "Sub Area Kontrol"),
        ])

    return prefix + COMMON_RISK + risk_extra + [
        ("coso_objective", "Tujuan COSO"),
        ("coso_component", "Komponen COSO"),
        ("fraud_risk", "Risiko Kecurangan"),
        ("impact", "Dampak"),
        ("likelihood", "Kemungkinan Terjadi"),
        ("risk_level", "Tingkat Risiko"),
        ("control_reference", "Ref. Kontrol"),
        ("control_objective", "Tujuan Kontrol"),
        ("control_description", "Deskripsi Kontrol"),
        ("control_type", "Jenis Kontrol"),
        ("is_key_control", "Kontrol Utama"),
        ("anti_fraud", "Anti Kecurangan"),
        ("supporting_application", "Aplikasi Pendukung"),
        ("location", "Lokasi"),
        ("location_description", "Deskripsi Lokasi"),
        ("frequency", "Frekuensi"),
        ("preparer_position", "Control Preparer"),
        ("reviewer_position", "Control Reviewer"),
        ("attributes", "Atribut Kontrol"),
        ("documents", "Dokumen Pendukung"),
        ("compensating_control", "Kontrol Kompensasi"),
        ("segment", "Segmen"),
    ]


def build_rcm_export(rcm_set):
    wb = Workbook()
    ws = wb.active
    ws.title = "RCM"
    ws["A1"] = "RISK CONTROL MATRIX (RCM)"
    ws["A2"] = "Jenis"
    ws["B2"] = rcm_set.get_rcm_type_display()
    ws["A3"] = "Versi"
    ws["B3"] = rcm_set.version
    headers = _headers(rcm_set.rcm_type)
    header_row = 5
    for col, (_, label) in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    entries = (
        rcm_set.entries
        .select_related("risk", "control")
        .prefetch_related("control_attributes", "supporting_documents")
        .order_by("source_row_number")
    )
    for row_no, entry in enumerate(entries, start=header_row + 1):
        risk = entry.risk
        control = entry.control
        payload = {
            "entity_name": entry.entity_name,
            "subprocess_number": entry.subprocess_number,
            "subprocess_description": entry.subprocess_description,
            "account_description": entry.account_description,
            "assertions_raw": entry.assertions_raw,
            "risk_reference": risk.reference,
            "risk_description": risk.description,
            "coso_element": risk.coso_element,
            "control_area": risk.control_area,
            "control_sub_area": risk.control_sub_area,
            "coso_objective": risk.coso_objective,
            "coso_component": risk.coso_component,
            "fraud_risk": risk.fraud_risk,
            "impact": risk.impact,
            "likelihood": risk.likelihood,
            "risk_level": risk.risk_level,
            "control_reference": control.reference,
            "control_objective": control.objective,
            "control_description": control.description,
            "control_type": control.control_type,
            "is_key_control": _bool_label(control.is_key_control),
            "anti_fraud": _bool_label(control.anti_fraud),
            "supporting_application": control.supporting_application,
            "location": entry.location,
            "location_description": entry.location_description,
            "frequency": entry.frequency,
            "preparer_position": entry.preparer_position,
            "reviewer_position": entry.reviewer_position,
            "attributes": "; ".join(entry.control_attributes.values_list("text", flat=True)),
            "documents": "; ".join(entry.supporting_documents.values_list("text", flat=True)),
            "compensating_control": entry.compensating_control,
            "segment": entry.segment,
        }
        for col, (key, _) in enumerate(headers, start=1):
            ws.cell(row_no, col, payload.get(key, ""))

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(12, max(len(str(cell.value or "")) for cell in col[:50]) + 2), 40)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
