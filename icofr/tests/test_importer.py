from pathlib import Path
from tempfile import NamedTemporaryFile

from django.contrib.auth import get_user_model
from django.core.files import File
from django.test import TestCase
from openpyxl import Workbook

from icofr.models import RCMEntry, RCMImportBatch, RCMMapping, RCMSet
from icofr.services.importer import import_batch, parse_workbook, validate_batch
from icofr.services.mapping import auto_map_rcm
from risk.models import RiwayatJabatanUser


User = get_user_model()


class RCMWorkbookMixin:
    def make_workbook(self, rcm_label, version, headers, rows):
        tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb = Workbook()
        ws = wb.active
        ws["A9"] = "RISK CONTROL MATRIX (RCM)"
        ws["A10"] = "Jenis"
        ws["E10"] = rcm_label
        ws["A11"] = "Versi"
        ws["E11"] = version
        for col, header in enumerate(headers, start=1):
            ws.cell(17, col, header)
        for row_no, values in enumerate(rows, start=18):
            for col, value in enumerate(values, start=1):
                ws.cell(row_no, col, value)
        wb.save(tmp.name)
        return Path(tmp.name)


class RCMImporterTests(RCMWorkbookMixin, TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser("admin", "admin@example.com", "x")

    def _create_batch(self, path):
        with path.open("rb") as handle:
            batch = RCMImportBatch.objects.create(
                upload=File(handle, name=path.name),
                original_filename=path.name,
                uploaded_by=self.user,
            )
        return batch

    def test_tlc_import_preserves_occurrences_and_many_to_many_risk_control_usage(self):
        headers = [
            "Entitas", "No Sub Proses", "Deskripsi Sub Proses", "Deskripsi Akun", "Asersi",
            "Ref. Risiko", "Deskripsi Risiko", "Tujuan COSO", "Komponen COSO", "Risiko Kecurangan",
            "Dampak", "Kemungkinan Terjadi", "Tingkat Risiko", "Ref. Kontrol", "Tujuan Kontrol",
            "Deskripsi Kontrol", "Jenis Kontrol", "Kontrol Utama", "Anti Kecurangan", "Aplikasi Pendukung",
            "Lokasi", "Deskripsi Lokasi", "Frekuensi", "Control Preparer", "Control Reviewer",
            "Atribut Kontrol", "Dokumen Pendukung", "Kontrol Kompensasi", "Segmen",
        ]
        base = [
            "PT PLN Batam", "1A.01.01", "Pendaftaran", "Pendapatan", "Completeness;Accuracy",
            None, None, "Pelaporan Keuangan", "Control Activities", "Tidak", "Signifikan", "Sedang", "Tinggi",
            "C.001", "Memastikan transaksi akurat", "Reviewer memeriksa transaksi", "Manual - Preventif", "Ya", "Tidak", "SAP",
            "KANTOR PUSAT", "BIDANG KEUANGAN", "Sesuai Transaksi", "MAN AKUNTANSI", "VP KEUANGAN",
            "Ada paraf;Nilai sesuai", "BAST;BAPP", "N/A", "Biaya Operasi",
        ]
        row1 = list(base); row1[5] = "R.001"; row1[6] = "Risiko pertama"
        row2 = list(base); row2[5] = "R.002"; row2[6] = "Risiko kedua"
        path = self.make_workbook("Transaction Level Controls", "2026/TW1-0", headers, [row1, row2])
        batch = self._create_batch(path)
        validate_batch(batch)
        batch.refresh_from_db()
        self.assertEqual(batch.status, RCMImportBatch.Status.VALIDATED)
        self.assertEqual(batch.summary["rows"], 2)
        self.assertEqual(batch.summary["risks"], 2)
        self.assertEqual(batch.summary["controls"], 1)
        self.assertEqual(batch.summary["risk_control_pairs"], 2)

        rcm = import_batch(batch, user=self.user)
        self.assertEqual(rcm.risks.count(), 2)
        self.assertEqual(rcm.controls.count(), 1)
        self.assertEqual(rcm.entries.count(), 2)
        self.assertEqual(RCMEntry.objects.filter(control__reference="C.001").count(), 2)
        entry = rcm.entries.order_by("source_row_number").first()
        self.assertEqual(entry.assertions.count(), 2)
        self.assertEqual(entry.control_attributes.count(), 2)
        self.assertEqual(entry.supporting_documents.count(), 2)

    def test_detects_elc_and_itgc_specific_headers(self):
        common = [
            "Entitas", "Ref. Risiko", "Deskripsi Risiko", "Tujuan COSO", "Komponen COSO",
            "Risiko Kecurangan", "Dampak", "Kemungkinan Terjadi", "Tingkat Risiko", "Ref. Kontrol",
            "Tujuan Kontrol", "Deskripsi Kontrol", "Jenis Kontrol", "Kontrol Utama", "Anti Kecurangan",
            "Lokasi", "Deskripsi Lokasi", "Frekuensi", "Control Preparer", "Control Reviewer",
            "Atribut Kontrol", "Dokumen Pendukung", "Kontrol Kompensasi", "Segmen",
        ]
        elc_headers = common[:3] + ["Elemen COSO"] + common[3:]
        elc_row = ["PT PLN Batam", "R.001", "Risk", "Commitment", "Pelaporan Keuangan", "Control Environment", "Tidak", "Signifikan", "Sedang", "Tinggi", "C.ELC1001", "Objective", "Control", "Manual - Preventif", "Ya", "Tidak", "KANTOR PUSAT", "HCGA", "Tahunan", "MAN HC", "VP HC", "Attribute", "Policy", "N/A", "N/A"]
        elc = parse_workbook(self.make_workbook("Entity Level Controls", "2026/ELC", elc_headers, [elc_row]))
        self.assertFalse(elc.errors)
        self.assertEqual(elc.rcm_type, "ELC")
        self.assertEqual(elc.rows[0]["coso_element"], "Commitment")

        itgc_headers = common[:3] + ["Area Kontrol", "Sub Area Kontrol"] + common[3:]
        itgc_row = ["PT PLN Batam", "R.001", "Risk", "Pengembangan Program", "Pengelolaan Kebutuhan TI", "Pelaporan Keuangan", "Control Activities", "Tidak", "Signifikan", "Sedang", "Tinggi", "C.ITGC1001", "Objective", "Control", "Manual - Preventif", "Ya", "Tidak", "UNIT BISNIS", "UB INFRA TI", "Tahunan", "MAN TI", "SM TI", "Attribute", "BRD", "N/A", "N/A"]
        itgc = parse_workbook(self.make_workbook("IT General Controls", "12/2026", itgc_headers, [itgc_row]))
        self.assertFalse(itgc.errors)
        self.assertEqual(itgc.rcm_type, "ITGC")
        self.assertEqual(itgc.rows[0]["control_area"], "Pengembangan Program")
        self.assertEqual(itgc.rows[0]["control_sub_area"], "Pengelolaan Kebutuhan TI")

    def test_final_rcm_is_locked_flag(self):
        rcm = RCMSet.objects.create(rcm_type="TLC", version="v1")
        self.assertFalse(rcm.is_locked)
        rcm.status = RCMSet.Status.FINAL
        self.assertTrue(rcm.is_locked)


class RCMMappingTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin", "admin@example.com", "x")
        self.preparer = User.objects.create_user("prep", is_active=True, is_staff=True)
        self.reviewer = User.objects.create_user("review", is_active=True, is_staff=True)
        RiwayatJabatanUser.objects.create(user=self.preparer, jabatan="MAN AKUNTANSI", tanggal_mulai="2026-01-01")
        RiwayatJabatanUser.objects.create(user=self.reviewer, jabatan="VP KEUANGAN", tanggal_mulai="2026-01-01")

    def test_strict_position_mapping_maps_unique_active_users(self):
        from icofr.models import RCMControl, RCMRisk
        rcm = RCMSet.objects.create(rcm_type="TLC", version="v-map")
        risk = RCMRisk.objects.create(rcm_set=rcm, reference="R.1")
        control = RCMControl.objects.create(rcm_set=rcm, reference="C.1")
        RCMEntry.objects.create(
            rcm_set=rcm,
            risk=risk,
            control=control,
            source_row_number=18,
            preparer_position="MAN AKUNTANSI",
            reviewer_position="VP KEUANGAN",
        )
        result = auto_map_rcm(rcm, user=self.admin)
        self.assertEqual(result["mapped"], 1)
        mapping = RCMMapping.objects.get(entry__rcm_set=rcm)
        self.assertEqual(mapping.preparer_user, self.preparer)
        self.assertEqual(mapping.reviewer_user, self.reviewer)
        self.assertEqual(mapping.status, RCMMapping.Status.MAPPED)
