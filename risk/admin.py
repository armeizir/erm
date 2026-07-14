from django import forms
from django.contrib import admin, messages
from django.contrib.auth.admin import GroupAdmin as BaseGroupAdmin, UserAdmin as BaseUserAdmin
from django.contrib.auth.models import Group, User
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect
from django.urls import path, reverse
from django.utils.html import format_html, format_html_join

from django.http import HttpResponse, HttpResponseNotAllowed
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import Image, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.styles import ParagraphStyle
from masterdata.models import MasterBUMN, TahunBuku, PeriodeLaporan

from django.contrib.auth import get_user_model
from django.db import models, transaction
from calendar import monthrange
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook, load_workbook

from corporate_risk.models import (
    MonteCarloMetricHistory,
    MultiMetricAIInsightKorporat,
    MultiMetricMonteCarloResult,
    RiskMetric,
)
from corporate_risk.pdf_reports import render_quarterly_lmr_pdf

from .models import (
    AppSetting,
    KnowledgeBaseArticle,
    KnowledgeBaseCategory,
    KontrakManajemen,
    BagianKontrakManajemen,
    ItemKontrakManajemen,
    RKMSummary,
    RKMItem,
    ReAssessmentSummary,
    ReAssessmentItem,
    KPMRSummary,
    KPMRItem,
    ProfilRisikoKorporatSummary,
    ProfilRisikoKorporatItem,
    ProfilRisikoKorporatPenyebab,  # 🔥 WAJIB TAMBAH
    ProfilRisikoKorporatSumber,
    SasaranKBUMN,
    TaksonomiT3,
    KategoriRisiko,
    MasterJenisExistingControl,
    MasterEfektivitasKontrol,
    MasterKategoriDampak,
    MasterSkalaDampak,
    MasterSkalaProbabilitas,
    MasterOpsiPerlakuanRisiko,
    MasterJenisRencanaPerlakuanRisiko,
    MasterPosAnggaran,
    MasterJenisProgramRKAP,
    MasterLevelRisiko,
    RiskMatrix,
    RiskMatrixCell,
    PenugasanUnitBisnis,
    RisikoInherenKuantitatif,
    RencanaPerlakuanRisikoKorporat,
    KPMRPeriode,
    KPMRIndikatorResmi,
    KPMRSubIndikatorResmi,
    RiskManagementReview,
    KinerjaPeriode,
    KinerjaIndikator,
    KompositRisikoTriwulan,
    RoadmapProgram,
    RoadmapPenilaianSemester,
    RKAPItem,
    MasterTemplateKM,
    MasterBagianKM,
    RiwayatJabatanUser,
    BagianKontrakManajemen,
)
from riskproject.admin_site import risk_admin_site



admin.site.site_header = "Manajemen Risiko PLN Batam"
admin.site.site_title = "Manajemen Risiko PLN Batam"
admin.site.index_title = "Dashboard Manajemen Risiko"


# =========================================================
# AUTH / GROUP
# =========================================================

def assigned_unit_businesses_for_user(user):
    if not user.is_authenticated:
        return Group.objects.none()
    if user.is_superuser:
        return Group.objects.all()
    return Group.objects.filter(
        penugasan_pengguna__user=user,
        penugasan_pengguna__aktif=True,
    ).distinct()


def user_can_access_unit(request, unit_id):
    if request.user.is_superuser:
        return True
    if not unit_id:
        return False
    return assigned_unit_businesses_for_user(request.user).filter(pk=unit_id).exists()

try:
    admin.site.unregister(Group)
except admin.sites.NotRegistered:
    pass

try:
    admin.site.unregister(User)
except admin.sites.NotRegistered:
    pass

Group._meta.verbose_name = "Bidang / Unit Bisnis"
Group._meta.verbose_name_plural = "Bidang / Unit Bisnis"


class PenugasanUnitBisnisUserInline(admin.TabularInline):
    model = PenugasanUnitBisnis
    fk_name = "user"
    extra = 0
    autocomplete_fields = ("unit_bisnis",)
    fields = ("unit_bisnis", "peran", "aktif", "catatan")
    ordering = ("unit_bisnis", "peran")


class PenugasanUnitBisnisGroupInline(admin.TabularInline):
    model = PenugasanUnitBisnis
    fk_name = "unit_bisnis"
    extra = 0
    autocomplete_fields = ("user",)
    fields = ("user", "peran", "aktif", "catatan")
    ordering = ("peran", "user")


@admin.register(Group)
class CustomGroupAdmin(BaseGroupAdmin):
    inlines = [PenugasanUnitBisnisGroupInline]
    list_display = ("name", "jenis_group", "jumlah_permission")
    list_filter = ("permissions__content_type__app_label",)
    search_fields = ("name",)
    ordering = ("name",)

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related("permissions").order_by("name")

    def get_inline_instances(self, request, obj=None):
        if obj and self._is_permission_role(obj):
            return []
        return super().get_inline_instances(request, obj)

    def _is_permission_role(self, obj):
        return str(obj.name or "").startswith("ROLE - ")

    @admin.display(description="Jenis")
    def jenis_group(self, obj):
        return "Role Permission" if self._is_permission_role(obj) else "Bidang / Unit Bisnis"

    @admin.display(description="Permissions")
    def jumlah_permission(self, obj):
        return obj.permissions.count()


class AppSettingForm(forms.ModelForm):
    ai_api_key = forms.CharField(
        required=False,
        widget=forms.PasswordInput(render_value=True),
        label="API Key AI",
        help_text="Kosongkan jika belum menggunakan AI. Nilai disimpan di database lokal aplikasi.",
    )
    email_host_password = forms.CharField(
        required=False,
        widget=forms.PasswordInput(render_value=True),
        label="SMTP Password",
        help_text="Kosongkan jika SMTP belum digunakan. Nilai disimpan di database lokal aplikasi.",
    )

    class Meta:
        model = AppSetting
        fields = "__all__"


@admin.register(AppSetting)
class AppSettingAdmin(admin.ModelAdmin):
    form = AppSettingForm
    list_display = (
        "nama_aplikasi",
        "logo_preview",
        "ldap_status",
        "ai_status",
        "smtp_status",
        "masked_ai_key",
        "diperbarui_pada",
    )
    readonly_fields = ("logo_preview", "masked_ai_key", "masked_email_password", "diperbarui_pada")
    fieldsets = (
        ("Identitas Aplikasi & Logo", {
            "fields": (
                "nama_aplikasi",
                "subtitle_aplikasi",
                "logo",
                "logo_preview",
                "tampilkan_logo",
                "warna_header",
                "warna_teks_header",
            )
        }),
        ("LDAP / Active Directory", {
            "fields": (
                "ldap_aktif",
                "ldap_server",
                "ldap_base_dn",
                "ldap_domain",
                "ldap_user_filter",
                "ldap_email_domain",
                "ldap_debug",
            ),
            "description": "Pengaturan login LDAP. Jika LDAP dimatikan, login lokal superuser tetap bisa digunakan.",
        }),
        ("API AI / Bahasa Manajemen", {
            "fields": (
                "ai_aktif",
                "ai_provider",
                "ai_model",
                "ai_base_url",
                "ai_api_key",
                "masked_ai_key",
                "ai_temperature",
            ),
            "description": (
                "Digunakan untuk memoles AI Insight menjadi bahasa manajemen. "
                "Untuk free tier, pilih Google Gemini API dan isi API key dari Google AI Studio. "
                "Jika AI tidak aktif atau API gagal, sistem tetap memakai insight rule-based."
            ),
        }),
        ("SMTP Email", {
            "fields": (
                "email_smtp_aktif",
                "email_host",
                "email_port",
                "email_host_user",
                "email_host_password",
                "masked_email_password",
                "email_use_tls",
                "email_use_ssl",
                "default_from_email",
            ),
            "description": (
                "Dipakai untuk mengirim notifikasi email dari aplikasi, termasuk Risk Awareness. "
                "Jika tidak aktif, sistem memakai konfigurasi EMAIL_* dari environment."
            ),
        }),
        ("Lain-lain", {
            "fields": (
                "support_email",
                "footer_laporan",
                "diperbarui_pada",
            )
        }),
    )

    def has_module_permission(self, request):
        return super().has_module_permission(request)

    def has_view_permission(self, request, obj=None):
        return super().has_view_permission(request, obj)

    def has_change_permission(self, request, obj=None):
        return super().has_change_permission(request, obj)

    def has_add_permission(self, request):
        return (
            super().has_add_permission(request)
            and not AppSetting.objects.exists()
        )

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.display(description="SMTP Password Tersimpan")
    def masked_email_password(self, obj):
        return obj.masked_email_host_password if obj else "-"

    def logo_preview(self, obj):
        if obj and obj.logo:
            return format_html(
                '<img src="{}" style="max-width: 260px; max-height: 90px; object-fit: contain;" />',
                obj.logo.url,
            )
        return "-"

    logo_preview.short_description = "Preview Logo"

    def ldap_status(self, obj):
        return "Aktif" if obj.ldap_aktif else "Nonaktif"

    ldap_status.short_description = "LDAP"

    def ai_status(self, obj):
        return "Aktif" if obj.ai_aktif else "Nonaktif"

    ai_status.short_description = "AI"

    def smtp_status(self, obj):
        return "Aktif" if obj.email_smtp_aktif else "Nonaktif"

    smtp_status.short_description = "SMTP Email"

    def masked_ai_key(self, obj):
        return obj.masked_ai_api_key if obj else "-"

    masked_ai_key.short_description = "API Key Tersimpan"


@admin.register(User)
class CustomUserAdmin(BaseUserAdmin):
    inlines = [PenugasanUnitBisnisUserInline]
    list_display = (
        "username",
        "email",
        "first_name",
        "last_name",
        "groups_display",
        "is_staff",
    )
    list_select_related = BaseUserAdmin.list_select_related

    @admin.display(description="Groups", ordering="groups_order")
    def groups_display(self, obj):
        group_names = [group.name for group in obj.groups.all()]
        return ", ".join(group_names) if group_names else "-"

    def get_queryset(self, request):
        return (
            super().get_queryset(request)
            .prefetch_related("groups")
            .annotate(groups_order=models.Min("groups__name"))
        )


@admin.register(PenugasanUnitBisnis)
class PenugasanUnitBisnisAdmin(admin.ModelAdmin):
    list_display = ("user", "unit_bisnis", "peran", "aktif", "dibuat_pada")
    list_filter = ("peran", "aktif", "unit_bisnis")
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__email",
        "unit_bisnis__name",
    )
    autocomplete_fields = ("user", "unit_bisnis")
    ordering = ("unit_bisnis__name", "peran", "user__username")


@admin.register(RiwayatJabatanUser)
class RiwayatJabatanUserAdmin(admin.ModelAdmin):
    list_display = (
        "user",
        "jabatan",
        "tanggal_mulai",
        "tanggal_selesai",
    )
    list_filter = ("user",)
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__email",
        "jabatan",
    )
    autocomplete_fields = ("user",)
    ordering = ("user", "-tanggal_mulai")



@admin.register(MasterKategoriDampak)
class MasterKategoriDampakAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterSkalaDampak)
class MasterSkalaDampakAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterSkalaProbabilitas)
class MasterSkalaProbabilitasAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterOpsiPerlakuanRisiko)
class MasterOpsiPerlakuanRisikoAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterJenisRencanaPerlakuanRisiko)
class MasterJenisRencanaPerlakuanRisikoAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterPosAnggaran)
class MasterPosAnggaranAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterJenisProgramRKAP)
class MasterJenisProgramRKAPAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterJenisExistingControl)
class MasterJenisExistingControlAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterEfektivitasKontrol)
class MasterEfektivitasKontrolAdmin(admin.ModelAdmin):
    list_display = ("nama", "aktif", "urutan")
    list_editable = ("aktif", "urutan")
    search_fields = ("nama",)
    ordering = ("urutan", "nama")


@admin.register(MasterLevelRisiko)
class MasterLevelRisikoAdmin(admin.ModelAdmin):
    list_display = ("kode", "nama", "urutan", "aktif")
    search_fields = ("kode", "nama")
    list_filter = ("aktif",)
    ordering = ("urutan", "nama")


class RiskMatrixCellInline(admin.TabularInline):
    model = RiskMatrixCell
    extra = 0


@admin.register(RiskMatrix)
class RiskMatrixAdmin(admin.ModelAdmin):
    list_display = ("kode", "nama", "ukuran", "is_default", "aktif")
    list_filter = ("aktif", "is_default")
    search_fields = ("kode", "nama")
    ordering = ("kode", "nama")
    inlines = [RiskMatrixCellInline]


# =========================================================
# MASTER DATA
# =========================================================

@admin.register(TaksonomiT3)
class TaksonomiT3Admin(admin.ModelAdmin):
    list_display = ("kode", "nama", "aktif")
    search_fields = ("kode", "nama")
    list_filter = ("aktif",)
    ordering = ("kode", "nama")


@admin.register(KategoriRisiko)
class KategoriRisikoAdmin(admin.ModelAdmin):
    list_display = ("kode", "nama", "aktif")
    search_fields = ("kode", "nama")
    list_filter = ("aktif",)
    ordering = ("kode", "nama")


@admin.register(SasaranKBUMN)
class SasaranKBUMNAdmin(admin.ModelAdmin):
    list_display = ("kode", "nama", "aktif")
    search_fields = ("kode", "nama")
    list_filter = ("aktif",)
    ordering = ("kode", "nama")


@admin.register(MasterBUMN)
class MasterBUMNAdmin(admin.ModelAdmin):
    list_display = ("nama", "kode")
    search_fields = ("nama", "kode")
    ordering = ("nama",)


# =========================================================
# RKAP (SIMPLIFIED)
# =========================================================

@admin.register(RKAPItem)
class RKAPItemAdmin(admin.ModelAdmin):
    list_display = (
        "tahun",
        "kode",
        "sasaran",
        "indikator",
        "target",
        "satuan",
        "unit_penanggung_jawab",
        "aktif",
    )
    list_filter = ("tahun", "aktif", "unit_penanggung_jawab")
    search_fields = ("kode", "sasaran", "indikator", "asumsi")
    ordering = ("-tahun", "kode", "sasaran")
    autocomplete_fields = ("unit_penanggung_jawab",)

    fieldsets = (
        ("Informasi Utama", {
            "fields": (
                "tahun",
                "kode",
                "sasaran",
                "indikator",
            )
        }),
        ("Target", {
            "fields": (
                "target",
                "satuan",
            )
        }),
        ("Konteks", {
            "fields": (
                "asumsi",
                "unit_penanggung_jawab",
                "aktif",
            )
        }),
    )

class MasterBagianKMInline(admin.TabularInline):
    model = MasterBagianKM
    extra = 0
    fields = ("urutan", "kode_bagian", "nama_bagian")
    ordering = ("urutan", "kode_bagian")


@admin.register(MasterTemplateKM)
class MasterTemplateKMAdmin(admin.ModelAdmin):
    list_display = ("tahun", "nama")
    search_fields = ("nama",)
    ordering = ("-tahun",)
    inlines = [MasterBagianKMInline]

@admin.register(MasterBagianKM)
class MasterBagianKMAdmin(admin.ModelAdmin):
    list_display = (
        "template",
        "urutan",
        "kode_bagian",
        "nama_bagian",
    )

    search_fields = (
        "template__nama",
        "kode_bagian",
        "nama_bagian",
    )

    list_filter = (
        "template__tahun",
        "template",
    )

    ordering = (
        "template__tahun",
        "urutan",
        "kode_bagian",
    )

    def save_model(self, request, obj, form, change):
        if obj.kontrak_id and obj.master_bagian_id and not obj.bagian_id:
            bagian, created = BagianKontrakManajemen.objects.get_or_create(
                kontrak=obj.kontrak,
                kode_bagian=obj.master_bagian.kode_bagian,
                defaults={
                    "nama_bagian": obj.master_bagian.nama_bagian,
                }
            )
            obj.bagian = bagian

        super().save_model(request, obj, form, change)

    def has_module_permission(self, request):
        return False


# =========================================================
# KONTRAK MANAJEMEN
# =========================================================

class BagianKontrakInline(admin.TabularInline):
    model = BagianKontrakManajemen
    extra = 0
    fields = ("kode_bagian", "nama_bagian")
    show_change_link = False
    verbose_name = ""
    verbose_name_plural = "Bagian Kontrak Manajemen"

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        formset.form.__str__ = lambda self: ""
        return formset


@admin.register(KontrakManajemen)
class KontrakManajemenAdmin(admin.ModelAdmin):
    list_display = (
        "judul",
        "tahun",
        "tanggal_kontrak",
        "template",
        "unit_bisnis",
        "pihak_pertama",
        "pihak_kedua",
        "status",
        "dibuat_pada",
        "pdf_button",
    )

    autocomplete_fields = (
        "template",
        "unit_bisnis",
        "pihak_pertama",
        "pihak_kedua",
    )

    fields = (
        "judul",
        "tahun",
        "tanggal_kontrak",
        "template",
        "unit_bisnis",
        "status",
        "pihak_pertama",
        "pihak_kedua",
    )

    list_filter = ("tahun", "status", "unit_bisnis", "template")
    search_fields = ("judul", "unit_bisnis__name", "template__nama")
    ordering = ("-tahun", "judul")

    def _has_km_permission(self, request, action):
        return (
            request.user.has_perm(f"risk.{action}_kontrakmanajemen")
            or request.user.has_perm(f"km.{action}_kontrakmanajemen")
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(unit_bisnis__in=assigned_unit_businesses_for_user(request.user))

    def has_module_permission(self, request):
        return self._has_km_permission(request, "view")

    def has_view_permission(self, request, obj=None):
        allowed = self._has_km_permission(request, "view") or self._has_km_permission(request, "change")
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = self._has_km_permission(request, "change")
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_add_permission(self, request):
        return self._has_km_permission(request, "add")

    def has_delete_permission(self, request, obj=None):
        allowed = self._has_km_permission(request, "delete")
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser and db_field.name == "unit_bisnis":
            kwargs["queryset"] = assigned_unit_businesses_for_user(request.user)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        is_new = obj.pk is None
        super().save_model(request, obj, form, change)

        if is_new and obj.template:
            for bagian in obj.template.bagian_list.all():
                ItemKontrakManajemen.objects.get_or_create(
                    kontrak=obj,
                    master_bagian=bagian,
                    no_urut=bagian.urutan,
                    defaults={
                        "indikator_kinerja_kunci": "",
                        "formula": "",
                        "satuan": "",
                        "target": "",
                        "bobot": 0,
                    },
                )

    def pdf_button(self, obj):
        url = reverse("admin:risk_kontrakmanajemen_pdf", args=[obj.pk])
        max_month = (
            RKMSummary.objects
            .filter(kontrak_manajemen=obj, tahun=obj.tahun)
            .order_by("-bulan")
            .values_list("bulan", flat=True)
            .first()
        )
        if max_month:
            month_links = [
                (url, obj.tahun, month, month)
                for month in range(1, min(max_month, 12) + 1)
            ]
            return format_html(
                "PDF KM: {}",
                format_html_join(
                    " ",
                    '<a class="button" href="{}?tahun={}&bulan={}" target="_blank">{}</a>',
                    month_links,
                ),
            )
        return format_html(
            '<a class="button" href="{}" target="_blank">PDF</a>',
            url,
        )

    pdf_button.short_description = "Laporan KM"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:kontrak_id>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="risk_kontrakmanajemen_pdf",
            ),
        ]
        return custom_urls + urls

    def nama_user(self, user):
        if not user:
            return "-"
        return user.get_full_name() or user.username

    def jabatan_user(self, user, tanggal):
        if not user or not tanggal:
            return "-"

        riwayat = (
            user.riwayat_jabatan
            .filter(tanggal_mulai__lte=tanggal)
            .filter(
                models.Q(tanggal_selesai__isnull=True)
                | models.Q(tanggal_selesai__gte=tanggal)
            )
            .order_by("-tanggal_mulai")
            .first()
        )

        return riwayat.jabatan if riwayat else "-"

    def bulan_indonesia(self, nomor_bulan):
        bulan = [
            "",
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "November",
            "Desember",
        ]
        if 1 <= nomor_bulan <= 12:
            return bulan[nomor_bulan]
        return "-"

    def format_angka_km(self, nilai):
        if nilai in (None, ""):
            return ""
        try:
            angka = Decimal(str(nilai))
        except Exception:
            return str(nilai)

        if angka == angka.to_integral_value():
            return str(int(angka))
        return f"{angka:.2f}".rstrip("0").rstrip(".")

    def paragraph_km(self, text, style, max_chars=None):
        text = " ".join(str(text or "").split())
        if max_chars and len(text) > max_chars:
            text = f"{text[:max_chars].rstrip()}..."
        return Paragraph(escape(text), style)

    def decimal_km(self, nilai):
        if nilai in (None, ""):
            return None
        try:
            cleaned = str(nilai).strip().replace("%", "").replace(",", ".")
            if "-" in cleaned and not cleaned.startswith("-"):
                cleaned = cleaned.split("-", 1)[0].strip()
            return Decimal(cleaned)
        except Exception:
            return None

    def calculate_km_score(self, item, target, realisasi):
        if target is None or realisasi is None or target == 0:
            return None, None
        if item.polaritas == "negatif":
            pencapaian = (target / realisasi * Decimal("100")) if realisasi else None
        else:
            pencapaian = realisasi / target * Decimal("100")
        if pencapaian is None:
            return None, None
        bobot = Decimal(str(item.bobot or 0))
        nilai = bobot * pencapaian / Decimal("100")
        return pencapaian, nilai

    def indikator_km(self, pencapaian):
        if pencapaian is None:
            return "-"
        if pencapaian >= Decimal("100"):
            return "Tercapai"
        if pencapaian >= Decimal("95"):
            return "Hampir Tercapai"
        return "Perlu Peningkatan"

    def realisasi_field_for_bulan(self, bulan):
        fields = {
            1: "realisasi_januari",
            2: "realisasi_februari",
            3: "realisasi_maret",
            4: "realisasi_april",
            5: "realisasi_mei",
            6: "realisasi_juni",
            7: "realisasi_juli",
            8: "realisasi_agustus",
            9: "realisasi_september",
            10: "realisasi_oktober",
            11: "realisasi_november",
            12: "realisasi_desember",
        }
        return fields.get(bulan)

    def target_field_for_bulan(self, bulan):
        fields = {
            1: "target_januari",
            2: "target_februari",
            3: "target_maret",
            4: "target_april",
            5: "target_mei",
            6: "target_juni",
            7: "target_juli",
            8: "target_agustus",
            9: "target_september",
            10: "target_oktober",
            11: "target_november",
            12: "target_desember",
        }
        return fields.get(bulan)

    def rkm_laporan_km(self, kontrak, request):
        rkm_qs = RKMSummary.objects.filter(kontrak_manajemen=kontrak).order_by("-tahun", "-bulan")
        tahun = request.GET.get("tahun")
        bulan = request.GET.get("bulan")
        if tahun:
            rkm_qs = rkm_qs.filter(tahun=tahun)
        if bulan:
            exact_rkm = (
                rkm_qs
                .filter(bulan=bulan)
                .select_related("penandatangan_laporan_km", "unit_bisnis")
                .first()
            )
            if exact_rkm:
                return exact_rkm
            rkm_qs = rkm_qs.filter(bulan__gte=bulan).order_by("bulan")
        return rkm_qs.select_related("penandatangan_laporan_km", "unit_bisnis").first()

    def km_logo_flowable(self):
        logo_path = None
        app_setting = AppSetting.objects.first()
        if app_setting and app_setting.logo:
            candidate = Path(app_setting.logo.path)
            if candidate.exists():
                logo_path = candidate

        if not logo_path:
            candidate = Path(settings.MEDIA_ROOT) / "system/logo/pln_batam_logo.png"
            if candidate.exists():
                logo_path = candidate

        if logo_path:
            return Image(str(logo_path), width=100, height=50)

        return Table(
            [[
                Paragraph("<b>PLN</b>", ParagraphStyle(
                    "KMLogoPLN",
                    parent=getSampleStyleSheet()["Normal"],
                    fontName="Helvetica-Bold",
                    fontSize=11,
                    leading=12,
                    alignment=TA_CENTER,
                )),
                Paragraph("Batam", ParagraphStyle(
                    "KMLogoBatam",
                    parent=getSampleStyleSheet()["Normal"],
                    fontName="Helvetica-Bold",
                    fontSize=10,
                    leading=11,
                )),
            ]],
            colWidths=[34, 58],
        )

    def pdf_view(self, request, kontrak_id):
        kontrak = get_object_or_404(KontrakManajemen, pk=kontrak_id)
        rkm = self.rkm_laporan_km(kontrak, request)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="KM_{kontrak.judul}_{kontrak.tahun}.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=14,
            leftMargin=14,
            topMargin=14,
            bottomMargin=14,
        )

        styles = getSampleStyleSheet()
        normal = ParagraphStyle(
            "KMNormal",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=5.8,
            leading=6.8,
            alignment=TA_LEFT,
            spaceAfter=0,
        )
        normal_center = ParagraphStyle(
            "KMNormalCenter",
            parent=normal,
            alignment=TA_CENTER,
        )
        header_style = ParagraphStyle(
            "KMHeader",
            parent=normal_center,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        section_style = ParagraphStyle(
            "KMSection",
            parent=normal,
            fontName="Helvetica-Bold",
        )
        title_style = ParagraphStyle(
            "KMTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            alignment=TA_CENTER,
            spaceAfter=2,
        )

        elements = []

        tanggal = kontrak.tanggal_kontrak
        try:
            requested_month = int(request.GET.get("bulan") or 0)
        except (TypeError, ValueError):
            requested_month = 0
        bulan_laporan = requested_month or (rkm.bulan if rkm else (tanggal.month if tanggal else 1))
        bulan_laporan = max(1, min(bulan_laporan, 12))
        tahun_label = rkm.tahun if rkm else (tanggal.year if tanggal else kontrak.tahun)
        bulan_label = self.bulan_indonesia(bulan_laporan)
        periode_label = f"S.D {bulan_label.upper()} {tahun_label}"
        unit_label = str(kontrak.unit_bisnis or kontrak.judul or "").upper()
        target_field = self.target_field_for_bulan(bulan_laporan)
        realisasi_field = self.realisasi_field_for_bulan(bulan_laporan)
        rkm_items = {
            item.km_item_id: item
            for item in rkm.item.select_related("km_item")
        } if rkm else {}

        header_table = Table(
            [[
                self.km_logo_flowable(),
                [
                    Paragraph(
                        f"PENCAPAIAN KONTRAK MANAJEMEN TAHUN {kontrak.tahun}",
                        title_style,
                    ),
                    Paragraph(unit_label, title_style),
                    Paragraph(periode_label, title_style),
                ],
                "",
            ]],
            colWidths=[110, 590, 110],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(header_table)

        data = [
            [
                self.paragraph_km("NO", header_style),
                self.paragraph_km("INDIKATOR KINERJA KUNCI", header_style),
                self.paragraph_km("FORMULA", header_style),
                self.paragraph_km("SATUAN", header_style),
                self.paragraph_km("BOBOT", header_style),
                self.paragraph_km(f"TARGET {kontrak.tahun}", header_style),
                self.paragraph_km(periode_label, header_style),
                "",
                "",
                "",
                "",
                "",
            ],
            [
                self.paragraph_km("1", header_style),
                self.paragraph_km("2", header_style),
                self.paragraph_km("3", header_style),
                self.paragraph_km("4", header_style),
                self.paragraph_km("5", header_style),
                self.paragraph_km("6", header_style),
                self.paragraph_km("TARGET", header_style),
                self.paragraph_km("REALISASI", header_style),
                self.paragraph_km("PENCAPAIAN", header_style),
                self.paragraph_km("NILAI", header_style),
                self.paragraph_km("INDIKATOR", header_style),
                self.paragraph_km("KETERANGAN", header_style),
            ],
        ]

        items = (
            ItemKontrakManajemen.objects
            .filter(kontrak=kontrak)
            .select_related("master_bagian")
            .order_by("master_bagian__urutan", "no_urut")
        )

        current_bagian = None
        section_rows = []
        total_bobot = Decimal("0")
        total_nilai = Decimal("0")
        has_nilai = False
        items_list = list(items)

        for item in items_list:
            if not (
                item.indikator_kinerja_kunci
                or item.formula
                or item.target
                or item.bobot
            ):
                continue

            if item.master_bagian_id and item.master_bagian != current_bagian:
                current_bagian = item.master_bagian
                subtotal_bobot = sum(
                    Decimal(str(i.bobot or 0))
                    for i in items_list
                    if i.master_bagian_id == current_bagian.id
                    and (
                        i.indikator_kinerja_kunci
                        or i.formula
                        or i.target
                        or i.bobot
                    )
                )
                section_rows.append(len(data))
                data.append([
                    self.paragraph_km(current_bagian.kode_bagian, section_style),
                    self.paragraph_km(current_bagian.nama_bagian, section_style),
                    "",
                    "",
                    self.paragraph_km(self.format_angka_km(subtotal_bobot), section_style),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ])

            total_bobot += Decimal(str(item.bobot or 0))
            rkm_item = rkm_items.get(item.id)
            target_bulanan = ""
            if rkm_item:
                target_bulanan = (
                    (getattr(rkm_item, target_field, None) if target_field else None)
                    or rkm_item.target_akumulasi
                    or rkm_item.target_bulanan
                    or ""
                )
            realisasi = ""
            if rkm_item:
                realisasi = (
                    (getattr(rkm_item, realisasi_field, None) if realisasi_field else None)
                    or rkm_item.jumlah_realisasi
                    or rkm_item.realisasi
                    or ""
                )
            target_value = self.decimal_km(target_bulanan) or self.decimal_km(item.target)
            realisasi_value = self.decimal_km(realisasi)
            pencapaian, nilai = self.calculate_km_score(item, target_value, realisasi_value)
            if nilai is not None:
                has_nilai = True
                total_nilai += nilai
            polaritas = (
                item.get_polaritas_display()
                if hasattr(item, "get_polaritas_display")
                else item.polaritas
            )
            indikator = item.indikator_kinerja_kunci or ""
            if polaritas:
                indikator = f"{indikator}\n({polaritas})"

            keterangan = ""
            if rkm_item:
                keterangan = (
                    rkm_item.hasil_analisa_program_kerja
                    or rkm_item.keterangan
                    or ""
                )

            data.append([
                self.paragraph_km(item.no_urut, normal_center),
                self.paragraph_km(indikator, normal, max_chars=220),
                self.paragraph_km(item.formula or "", normal, max_chars=260),
                self.paragraph_km(item.satuan or "", normal_center),
                self.paragraph_km(self.format_angka_km(item.bobot), normal_center),
                self.paragraph_km(item.target or "", normal_center),
                self.paragraph_km(target_bulanan or item.target or "", normal_center),
                self.paragraph_km(realisasi, normal_center),
                self.paragraph_km(f"{self.format_angka_km(pencapaian)}%" if pencapaian is not None else "", normal_center),
                self.paragraph_km(self.format_angka_km(nilai), normal_center),
                self.paragraph_km(self.indikator_km(pencapaian), normal_center),
                self.paragraph_km(keterangan, normal, max_chars=180),
            ])

        total_row_index = len(data)
        data.append([
            "",
            self.paragraph_km("TOTAL", section_style),
            "",
            "",
            self.paragraph_km(self.format_angka_km(total_bobot), section_style),
            "",
            "",
            "",
            "",
            self.paragraph_km(self.format_angka_km(total_nilai) if has_nilai else "", section_style),
            "",
            "",
        ])

        table = Table(
            data,
            colWidths=[28, 150, 190, 42, 36, 52, 52, 52, 52, 42, 44, 70],
            repeatRows=2,
        )

        table_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("SPAN", (6, 0), (11, 0)),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#0070C0")),
            ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 5.8),
            ("LEADING", (0, 0), (-1, -1), 6.8),
            ("ALIGN", (0, 0), (-1, 1), "CENTER"),
            ("ALIGN", (0, 2), (0, -1), "CENTER"),
            ("ALIGN", (3, 2), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 2), (-1, -1), colors.white),
            ("BACKGROUND", (0, total_row_index), (-1, total_row_index), colors.HexColor("#D9EAF7")),
            ("FONTNAME", (0, total_row_index), (-1, total_row_index), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 1.6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.6),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ])
        for row_index in section_rows:
            table_style.add(
                "BACKGROUND",
                (0, row_index),
                (-1, row_index),
                colors.HexColor("#FFC000"),
            )
            table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")

        table.setStyle(table_style)

        elements.append(table)
        elements.append(Spacer(1, 8))

        legend_data = [
            ["-", "Belum dilakukan proses pengukuran"],
            ["", "Tercapai (NKO >= 100)"],
            ["", "Hampir Tercapai (95 <= NKO < 100)"],
            ["", "Perlu Peningkatan (NKO < 95)"],
        ]
        legend_table = Table(legend_data, colWidths=[16, 170])
        legend_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEADING", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#00B050")),
            ("BACKGROUND", (0, 2), (0, 2), colors.HexColor("#FFFF00")),
            ("BACKGROUND", (0, 3), (0, 3), colors.HexColor("#FF0000")),
            ("BOX", (0, 1), (0, 3), 0.25, colors.black),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))

        penandatangan = (
            rkm.penandatangan_laporan_km
            if rkm and rkm.penandatangan_laporan_km_id
            else kontrak.pihak_kedua
        )
        nama_p2 = self.nama_user(penandatangan) if penandatangan else ""

        jabatan_p2 = self.jabatan_user(
            penandatangan,
            rkm.tanggal_selesai if rkm and rkm.tanggal_selesai else kontrak.tanggal_kontrak,
        )
        if jabatan_p2 == "-":
            jabatan_p2 = ""

        tanggal_text = (
            f"{tanggal.day:02d} {self.bulan_indonesia(tanggal.month)} {tanggal.year}"
            if tanggal
            else f"{self.bulan_indonesia(2)} {kontrak.tahun}"
        )
        sign_style = ParagraphStyle(
            "KMSign",
            parent=normal_center,
            fontSize=7,
            leading=8,
        )
        sign_bold = ParagraphStyle(
            "KMSignBold",
            parent=sign_style,
            fontName="Helvetica-Bold",
        )
        signature_table = Table(
            [
                [Paragraph(f"Batam, {tanggal_text}", sign_style)],
                [Paragraph(escape(jabatan_p2), sign_style)],
                [""],
                [Paragraph(f"<u>{escape(nama_p2)}</u>", sign_bold) if nama_p2 else ""],
            ],
            colWidths=[220],
            rowHeights=[12, 18, 30, 14],
        )
        signature_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))

        footer_table = Table(
            [[legend_table, "", signature_table]],
            colWidths=[230, 360, 220],
        )
        footer_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        elements.append(footer_table)

        doc.build(elements)
        return response


class ItemKontrakInline(admin.TabularInline):
    model = ItemKontrakManajemen
    extra = 0
    fields = (
        "no_urut",
        "indikator_kinerja_kunci",
        "formula",
        "satuan",
        "target",
        "bobot",
    )
    ordering = ("no_urut",)


class KnowledgeBaseArticleForm(forms.ModelForm):
    class Meta:
        model = KnowledgeBaseArticle
        fields = "__all__"
        widgets = {
            "ringkasan": forms.Textarea(attrs={"rows": 3}),
            "konten": forms.Textarea(attrs={"class": "kb-rich-editor", "rows": 24}),
        }


@admin.register(KnowledgeBaseCategory)
class KnowledgeBaseCategoryAdmin(admin.ModelAdmin):
    list_display = ("nama", "urutan", "aktif", "jumlah_artikel")
    list_filter = ("aktif",)
    search_fields = ("nama", "deskripsi")
    prepopulated_fields = {"slug": ("nama",)}
    ordering = ("urutan", "nama")

    @admin.display(description="Artikel")
    def jumlah_artikel(self, obj):
        return obj.artikel.count()

    def has_module_permission(self, request):
        return bool(request.user and request.user.is_active and request.user.is_staff)

    def has_view_permission(self, request, obj=None):
        return bool(request.user and request.user.is_active and request.user.is_staff)


@admin.register(KnowledgeBaseArticle)
class KnowledgeBaseArticleAdmin(admin.ModelAdmin):
    form = KnowledgeBaseArticleForm
    list_display = (
        "judul",
        "kategori",
        "audience",
        "status",
        "dibuat_oleh",
        "dipublikasikan_pada",
        "diperbarui_pada",
    )
    list_filter = ("status", "audience", "kategori", "dipublikasikan_pada")
    search_fields = ("judul", "ringkasan", "konten", "tags")
    prepopulated_fields = {"slug": ("judul",)}
    autocomplete_fields = ("kategori", "dibuat_oleh", "diperbarui_oleh")
    readonly_fields = ("dibuat_pada", "diperbarui_pada")
    ordering = ("kategori__urutan", "judul")
    date_hierarchy = "diperbarui_pada"

    fieldsets = (
        ("Informasi Artikel", {
            "fields": (
                "kategori",
                "judul",
                "slug",
                "ringkasan",
                "tags",
                "audience",
                "status",
            )
        }),
        ("Konten", {
            "fields": (
                "konten",
                "lampiran",
            )
        }),
        ("Publikasi", {
            "fields": (
                "dipublikasikan_pada",
                "dibuat_oleh",
                "diperbarui_oleh",
                "dibuat_pada",
                "diperbarui_pada",
            )
        }),
    )

    class Media:
        js = (
            "https://cdn.ckeditor.com/4.22.1/full/ckeditor.js",
            "risk/admin/knowledge_base_editor.js",
        )
        css = {
            "all": ("risk/admin/knowledge_base_editor.css",)
        }

    def save_model(self, request, obj, form, change):
        if not obj.dibuat_oleh_id:
            obj.dibuat_oleh = request.user
        obj.diperbarui_oleh = request.user
        super().save_model(request, obj, form, change)

    def has_module_permission(self, request):
        return bool(request.user and request.user.is_active and request.user.is_staff)

    def has_view_permission(self, request, obj=None):
        return bool(request.user and request.user.is_active and request.user.is_staff)


@admin.register(BagianKontrakManajemen)
class BagianKontrakManajemenAdmin(admin.ModelAdmin):
    list_display = ("kontrak", "kode_bagian", "nama_bagian")
    search_fields = ("kode_bagian", "nama_bagian", "kontrak__judul")
    ordering = ("kontrak", "kode_bagian", "nama_bagian")

    def has_module_permission(self, request):
        return False


@admin.register(ItemKontrakManajemen)
class ItemKontrakManajemenAdmin(admin.ModelAdmin):
    list_display = (
        "kontrak",
        "master_bagian",
        "no_urut",
        "indikator_kinerja_kunci",
        "satuan",
        "bobot",
        "target",
        "polaritas",
    )

    list_filter = (
        "kontrak__tahun",
        "kontrak__unit_bisnis",
        "master_bagian",
        "polaritas",
    )

    search_fields = (
        "kontrak__judul",
        "kontrak__unit_bisnis__name",
        "master_bagian__kode_bagian",
        "master_bagian__nama_bagian",
        "indikator_kinerja_kunci",
        "formula",
        "satuan",
        "target",
    )

    autocomplete_fields = (
        "kontrak",
        "master_bagian",
    )

    ordering = (
        "kontrak",
        "master_bagian__urutan",
        "no_urut",
    )

    fields = (
        "kontrak",
        "master_bagian",
        "no_urut",
        "indikator_kinerja_kunci",
        "formula",
        "satuan",
        "bobot",
        "target",
        "polaritas",
    )

    def _has_km_permission(self, request, action):
        return (
            request.user.has_perm(f"risk.{action}_itemkontrakmanajemen")
            or request.user.has_perm(f"km.{action}_kontrakmanajemenitem")
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(
            kontrak__unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
        )

    def has_module_permission(self, request):
        return self._has_km_permission(request, "view")

    def has_view_permission(self, request, obj=None):
        allowed = self._has_km_permission(request, "view") or self._has_km_permission(request, "change")
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.kontrak.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = self._has_km_permission(request, "change")
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.kontrak.unit_bisnis_id)

    def has_add_permission(self, request):
        return self._has_km_permission(request, "add")

    def has_delete_permission(self, request, obj=None):
        allowed = self._has_km_permission(request, "delete")
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.kontrak.unit_bisnis_id)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser and db_field.name == "kontrak":
            kwargs["queryset"] = KontrakManajemen.objects.filter(
                unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):

        if obj.master_bagian and obj.kontrak:

            bagian = BagianKontrakManajemen.objects.filter(
                kontrak=obj.kontrak,
                kode_bagian=obj.master_bagian.kode_bagian
            ).first()

            if not bagian:
                bagian = BagianKontrakManajemen.objects.create(
                    kontrak=obj.kontrak,
                    kode_bagian=obj.master_bagian.kode_bagian,
                    nama_bagian=obj.master_bagian.nama_bagian,
                )

            obj.bagian = bagian

        super().save_model(request, obj, form, change)


# =========================================================
# RKM
# =========================================================

RKM_MONTH_FIELD_PAIRS = {
    1: ("target_januari", "realisasi_januari"),
    2: ("target_februari", "realisasi_februari"),
    3: ("target_maret", "realisasi_maret"),
    4: ("target_april", "realisasi_april"),
    5: ("target_mei", "realisasi_mei"),
    6: ("target_juni", "realisasi_juni"),
    7: ("target_juli", "realisasi_juli"),
    8: ("target_agustus", "realisasi_agustus"),
    9: ("target_september", "realisasi_september"),
    10: ("target_oktober", "realisasi_oktober"),
    11: ("target_november", "realisasi_november"),
    12: ("target_desember", "realisasi_desember"),
}
RKM_MONTH_LABELS = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


class RKMItemInline(admin.TabularInline):
    model = RKMItem
    extra = 0
    can_delete = True
    base_fields = (
        "no_item",
        "kategori_rkm",
        "km_item",
        "kpi_indikator",
        "kpi_satuan",
        "kpi_target",
        "inisiatif_strategis",
        "program_kerja_utama",
        "anggaran_rp_ribu",
        "target_akumulasi",
        "target_akumulasi_satuan",
    )
    result_fields = (
    )
    fields = base_fields + RKM_MONTH_FIELD_PAIRS[1] + result_fields
    ordering = ("no_item",)

    def get_fields(self, request, obj=None):
        month_fields = RKM_MONTH_FIELD_PAIRS.get(getattr(obj, "bulan", None), ())
        return self.base_fields + month_fields + self.result_fields

    def get_readonly_fields(self, request, obj=None):
        if obj and obj.is_approved:
            return self.get_fields(request, obj)
        month_fields = RKM_MONTH_FIELD_PAIRS.get(getattr(obj, "bulan", None), ())
        return (month_fields[0],) if month_fields else ()

    def has_add_permission(self, request, obj=None):
        allowed = super().has_add_permission(request, obj)
        if not allowed:
            return False
        return not (obj and obj.is_approved)

    def has_delete_permission(self, request, obj=None):
        allowed = super().has_delete_permission(request, obj)
        if not allowed:
            return False
        return not (obj and obj.is_approved)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(
            summary__unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser and db_field.name == "km_item":
            kwargs["queryset"] = ItemKontrakManajemen.objects.filter(
                kontrak__unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(RKMSummary)
class RKMSummaryAdmin(admin.ModelAdmin):
    list_display = (
        "judul",
        "tahun",
        "bulan",
        "unit_bisnis",
        "kontrak_manajemen",
        "penandatangan_laporan_km",
        "penandatangan_laporan_rkm",
        "status",
        "status_pengajuan",
        "rkm_pdf_button",
    )
    list_filter = ("tahun", "bulan", "status", "status_pengajuan", "unit_bisnis")
    search_fields = ("judul", "unit_bisnis__name", "kontrak_manajemen__judul")
    ordering = ("-tahun", "bulan", "judul")
    inlines = [RKMItemInline]
    actions = ("finalize_sign_selected", "cancel_sign_selected")
    readonly_fields = ("rkm_lock_info",)
    autocomplete_fields = (
        "unit_bisnis",
        "kontrak_manajemen",
        "penandatangan_laporan_km",
        "penandatangan_laporan_rkm",
    )

    fieldsets = (
        ("Informasi Utama", {
            "fields": (
                "judul",
                "tahun",
                "bulan",
                "unit_bisnis",
                "kontrak_manajemen",
                "status",
                "rkm_lock_info",
            )
        }),
        ("Penandatangan Laporan", {
            "fields": (
                "penandatangan_laporan_km",
                "penandatangan_laporan_rkm",
            )
        }),
        ("Periode", {
            "fields": (
                "tanggal_mulai",
                "tanggal_selesai",
            )
        }),
        ("Pengajuan", {
            "fields": (
                "pic",
                "deadline_pengajuan",
                "tanggal_pengajuan",
                "status_pengajuan",
            )
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:rkm_id>/generate-items/",
                self.admin_site.admin_view(self.generate_items_view),
                name="risk_rkmsummary_generate_items",
            ),
            path(
                "<int:rkm_id>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="risk_rkmsummary_pdf",
            ),
        ]
        return custom_urls + urls

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(unit_bisnis__in=assigned_unit_businesses_for_user(request.user))

    def get_readonly_fields(self, request, obj=None):
        if obj and obj.is_approved:
            return (
                "judul",
                "tahun",
                "bulan",
                "unit_bisnis",
                "kontrak_manajemen",
                "status",
                "tanggal_mulai",
                "tanggal_selesai",
                "penandatangan_laporan_km",
                "penandatangan_laporan_rkm",
                "pic",
                "deadline_pengajuan",
                "tanggal_pengajuan",
                "status_pengajuan",
                "rkm_lock_info",
            )
        return self.readonly_fields

    def get_list_filter(self, request):
        if request.user.is_superuser:
            return self.list_filter
        return tuple(
            item for item in self.list_filter if item != "unit_bisnis"
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser:
            if db_field.name == "unit_bisnis":
                kwargs["queryset"] = assigned_unit_businesses_for_user(request.user)
            elif db_field.name == "kontrak_manajemen":
                kwargs["queryset"] = KontrakManajemen.objects.filter(
                    unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
                )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_view_permission(self, request, obj=None):
        allowed = super().has_view_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = super().has_change_permission(request, obj)
        if not allowed or obj is None:
            return allowed
        if obj.is_approved:
            return request.user.is_superuser
        if request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_delete_permission(self, request, obj=None):
        allowed = super().has_delete_permission(request, obj)
        if not allowed or obj is None:
            return allowed
        if obj.is_approved:
            return False
        if request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    @admin.action(description="Final/Sign RKM terpilih")
    def finalize_sign_selected(self, request, queryset):
        accessible = queryset
        if not request.user.is_superuser:
            accessible = accessible.filter(unit_bisnis__in=assigned_unit_businesses_for_user(request.user))
        updated = accessible.exclude(status="Final", status_pengajuan="Disetujui").update(
            status="Final",
            status_pengajuan="Disetujui",
            tanggal_pengajuan=date.today(),
        )
        self.message_user(request, f"{updated} RKM berhasil difinalkan/sign.")

    @admin.action(description="Batalkan Final/Sign RKM terpilih")
    def cancel_sign_selected(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(
                request,
                "Hanya admin yang dapat membatalkan Final/Sign RKM.",
                level=messages.ERROR,
            )
            return
        updated = queryset.filter(
            models.Q(status="Final") | models.Q(status_pengajuan="Disetujui")
        ).update(
            status="Draft",
            status_pengajuan="Belum",
        )
        self.message_user(request, f"{updated} RKM berhasil dibatalkan Final/Sign-nya.")

    def generate_button(self, obj):
        if obj.is_approved:
            return "RKM sudah disetujui/final"
        url = reverse("admin:risk_rkmsummary_generate_items", args=[obj.pk])
        return format_html('<a class="button" href="{}">Generate RKM dari KM</a>', url)
    generate_button.short_description = "Generate"

    def km_pdf_button(self, obj):
        url = reverse("admin:risk_kontrakmanajemen_pdf", args=[obj.kontrak_manajemen_id])
        month_links = [
            (url, obj.tahun, month, month)
            for month in range(1, min(obj.bulan or 1, 12) + 1)
        ]
        return format_html(
            'PDF KM: {}',
            format_html_join(
                " ",
                '<a class="button" href="{}?tahun={}&bulan={}" target="_blank">{}</a>',
                month_links,
            ),
        )
    km_pdf_button.short_description = "Laporan KM"

    def rkm_pdf_button(self, obj):
        url = reverse("admin:risk_rkmsummary_pdf", args=[obj.pk])
        month_links = [
            (url, month, month)
            for month in range(1, min(obj.bulan or 1, 12) + 1)
        ]
        return format_html(
            'PDF RKM: {}',
            format_html_join(
                " ",
                '<a class="button" href="{}?bulan={}" target="_blank">{}</a>',
                month_links,
            ),
        )

    rkm_pdf_button.short_description = "Laporan RKM"

    @admin.display(description="Status Kunci")
    def rkm_lock_info(self, obj):
        if obj and obj.is_approved:
            return "RKM sudah disetujui/final dan tidak dapat diedit. Realisasi dicatat melalui Laporan Risiko Bulanan."
        return "RKM masih dapat diedit sebelum status Final/Disetujui."

    def generate_items_view(self, request, rkm_id, *args, **kwargs):
        rkm = get_object_or_404(RKMSummary, pk=rkm_id)
        if rkm.is_approved:
            self.message_user(
                request,
                "RKM sudah disetujui/final sehingga item tidak dapat digenerate ulang.",
                level=messages.ERROR,
            )
            change_url = reverse("admin:risk_rkmsummary_change", args=[rkm.pk])
            return redirect(change_url)
        if not user_can_access_unit(request, rkm.unit_bisnis_id):
            raise PermissionDenied
        created_count = rkm.generate_items_from_km()

        self.message_user(
            request,
            f"Berhasil generate {created_count} item RKM dari KM.",
            level=messages.SUCCESS,
        )

        change_url = reverse("admin:risk_rkmsummary_change", args=[rkm.pk])
        return redirect(change_url)

    def bulan_indonesia(self, nomor_bulan):
        bulan = [
            "",
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "November",
            "Desember",
        ]
        if 1 <= nomor_bulan <= 12:
            return bulan[nomor_bulan]
        return "-"

    def format_angka_pdf(self, nilai):
        if nilai in (None, ""):
            return ""
        try:
            angka = Decimal(str(nilai))
        except Exception:
            return str(nilai)
        if angka == angka.to_integral_value():
            return str(int(angka))
        return f"{angka:.2f}".rstrip("0").rstrip(".")

    def paragraph_pdf(self, text, style):
        return Paragraph(escape(str(text or "")), style)

    def logo_flowable_pdf(self):
        logo_path = None
        app_setting = AppSetting.objects.first()
        if app_setting and app_setting.logo:
            candidate = Path(app_setting.logo.path)
            if candidate.exists():
                logo_path = candidate

        if not logo_path:
            candidate = Path(settings.MEDIA_ROOT) / "system/logo/pln_batam_logo.png"
            if candidate.exists():
                logo_path = candidate

        if logo_path:
            return Image(str(logo_path), width=100, height=50)

        return Table(
            [[
                Paragraph("<b>PLN</b>", ParagraphStyle(
                    "RKMLogoPLN",
                    parent=getSampleStyleSheet()["Normal"],
                    fontName="Helvetica-Bold",
                    fontSize=11,
                    leading=12,
                    alignment=TA_CENTER,
                )),
                Paragraph("Batam", ParagraphStyle(
                    "RKMLogoBatam",
                    parent=getSampleStyleSheet()["Normal"],
                    fontName="Helvetica-Bold",
                    fontSize=10,
                    leading=11,
                )),
            ]],
            colWidths=[34, 58],
        )

    def nama_user_pdf(self, user):
        if not user:
            return ""
        return user.get_full_name() or user.username

    def jabatan_user_pdf(self, user, tanggal):
        if not user or not tanggal:
            return ""
        riwayat = (
            user.riwayat_jabatan
            .filter(tanggal_mulai__lte=tanggal)
            .filter(
                models.Q(tanggal_selesai__isnull=True)
                | models.Q(tanggal_selesai__gte=tanggal)
            )
            .order_by("-tanggal_mulai")
            .first()
        )
        return riwayat.jabatan if riwayat else ""

    def pdf_view(self, request, rkm_id):
        rkm = get_object_or_404(
            RKMSummary.objects.select_related(
                "unit_bisnis",
                "kontrak_manajemen",
                "penandatangan_laporan_rkm",
            ),
            pk=rkm_id,
        )
        if not user_can_access_unit(request, rkm.unit_bisnis_id):
            raise PermissionDenied

        try:
            report_month = int(request.GET.get("bulan") or rkm.bulan or 1)
        except (TypeError, ValueError):
            report_month = rkm.bulan or 1
        report_month = max(1, min(report_month, rkm.bulan or 12, 12))
        report_month_label = self.bulan_indonesia(report_month)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="RKM_{rkm.unit_bisnis.name}_{report_month}_{rkm.tahun}.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=14,
            leftMargin=14,
            topMargin=14,
            bottomMargin=14,
        )

        styles = getSampleStyleSheet()
        normal = ParagraphStyle(
            "RKMNormal",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=5.6,
            leading=6.6,
            alignment=TA_LEFT,
            spaceAfter=0,
        )
        normal_center = ParagraphStyle("RKMNormalCenter", parent=normal, alignment=TA_CENTER)
        header_style = ParagraphStyle(
            "RKMHeader",
            parent=normal_center,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        matrix_title_style = ParagraphStyle(
            "RKMMatrixTitle",
            parent=normal_center,
            fontName="Helvetica-Bold",
            fontSize=6.5,
            leading=7.5,
            textColor=colors.black,
        )
        matrix_header_style = ParagraphStyle(
            "RKMMatrixHeader",
            parent=normal_center,
            fontName="Helvetica-Bold",
            fontSize=5.8,
            leading=6.5,
            textColor=colors.HexColor("#17365D"),
        )
        section_style = ParagraphStyle("RKMSection", parent=normal, fontName="Helvetica-Bold")
        title_style = ParagraphStyle(
            "RKMTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            alignment=TA_CENTER,
            spaceAfter=2,
        )

        periode_label = f"{report_month_label.upper()} {rkm.tahun}"
        unit_label = str(rkm.unit_bisnis or "").upper()
        elements = []

        header_table = Table(
            [[
                self.logo_flowable_pdf(),
                [
                    Paragraph(f"RENCANA KERJA MANAJEMEN TAHUN {rkm.tahun}", title_style),
                    Paragraph(unit_label, title_style),
                    Paragraph(periode_label, title_style),
                ],
                "",
            ]],
            colWidths=[110, 590, 110],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(header_table)

        info_table = Table(
            [[
                self.paragraph_pdf("Judul", section_style),
                self.paragraph_pdf(rkm.judul, normal),
                self.paragraph_pdf("Kontrak Manajemen", section_style),
                self.paragraph_pdf(rkm.kontrak_manajemen, normal),
                self.paragraph_pdf("Status", section_style),
                self.paragraph_pdf(f"{rkm.status} / {rkm.status_pengajuan}", normal),
            ]],
            colWidths=[45, 260, 80, 220, 40, 165],
        )
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C9D6")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF3F8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 6))

        items = list(
            rkm.item
            .select_related("km_item", "km_item__master_bagian")
            .order_by("no_item", "km_item__master_bagian__urutan")
        )
        month_fields = [
            (1, "januari"),
            (2, "februari"),
            (3, "maret"),
            (4, "april"),
            (5, "mei"),
            (6, "juni"),
            (7, "juli"),
            (8, "agustus"),
            (9, "september"),
            (10, "oktober"),
            (11, "november"),
            (12, "desember"),
        ]
        matrix_header = [
            self.paragraph_pdf("MONTH", matrix_header_style),
        ]
        matrix_subheader = [""]
        for item in items:
            code = ""
            if item.km_item_id and item.km_item.master_bagian_id:
                code = f"{item.km_item.master_bagian.kode_bagian}{item.km_item.no_urut}"
            else:
                code = str(item.no_item)
            matrix_header.extend([
                self.paragraph_pdf(code, matrix_header_style),
                "",
            ])
            matrix_subheader.extend([
                self.paragraph_pdf("TGT", matrix_header_style),
                self.paragraph_pdf("REAL", matrix_header_style),
            ])

        matrix_data = [
            [self.paragraph_pdf("MONTHLY KEY PERFORMANCE INDICATOR RESULT", matrix_title_style)]
            + ["" for _ in range(len(items) * 2)],
            matrix_header,
            matrix_subheader,
        ]
        for month_number, field_suffix in month_fields:
            row = [self.paragraph_pdf(month_number, normal_center)]
            for item in items:
                if month_number > report_month:
                    row.extend(["", ""])
                    continue
                row.extend([
                    self.paragraph_pdf(getattr(item, f"target_{field_suffix}", "") or "", normal_center),
                    self.paragraph_pdf(getattr(item, f"realisasi_{field_suffix}", "") or "", normal_center),
                ])
            matrix_data.append(row)

        table_width = doc.width
        month_col_width = 34
        value_col_count = max(len(items) * 2, 1)
        value_col_width = (table_width - month_col_width) / value_col_count

        matrix_table = Table(
            matrix_data,
            colWidths=[month_col_width] + [value_col_width for _ in range(len(items) * 2)],
            repeatRows=3,
            hAlign="CENTER",
        )
        matrix_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
            ("SPAN", (0, 0), (-1, 0)),
            ("SPAN", (0, 1), (0, 2)),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFE699")),
            ("BACKGROUND", (0, 1), (-1, 2), colors.HexColor("#D9EAF7")),
            ("BACKGROUND", (0, 3), (0, -1), colors.HexColor("#FFC000")),
            ("TEXTCOLOR", (0, 0), (-1, 2), colors.black),
            ("FONTNAME", (0, 0), (-1, 2), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 6.5),
            ("LEADING", (0, 0), (-1, 0), 7.5),
            ("FONTSIZE", (0, 1), (-1, 2), 5.8),
            ("LEADING", (0, 1), (-1, 2), 6.5),
            ("FONTSIZE", (0, 3), (-1, -1), 5.6),
            ("LEADING", (0, 3), (-1, -1), 6.4),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ])
        for index in range(len(items)):
            start_col = 1 + (index * 2)
            matrix_style.add("SPAN", (start_col, 1), (start_col + 1, 1))
        current_month_row = 2 + report_month
        if 3 <= current_month_row < len(matrix_data):
            matrix_style.add("TEXTCOLOR", (1, current_month_row), (-1, current_month_row), colors.red)
            matrix_style.add("FONTNAME", (1, current_month_row), (-1, current_month_row), "Helvetica-Bold")
        matrix_table.setStyle(matrix_style)
        elements.append(matrix_table)
        elements.append(Spacer(1, 8))

        penandatangan = rkm.penandatangan_laporan_rkm
        nama = self.nama_user_pdf(penandatangan)
        jabatan = self.jabatan_user_pdf(penandatangan, rkm.tanggal_selesai) if penandatangan else ""
        tanggal = rkm.tanggal_selesai
        if not tanggal or tanggal.month != report_month or tanggal.year != rkm.tahun:
            tanggal = date(
                rkm.tahun,
                report_month,
                monthrange(rkm.tahun, report_month)[1],
            )
        tanggal_text = f"{tanggal.day:02d} {self.bulan_indonesia(tanggal.month)} {tanggal.year}"
        sign_style = ParagraphStyle("RKMSign", parent=normal_center, fontSize=7, leading=8)
        sign_bold = ParagraphStyle("RKMSignBold", parent=sign_style, fontName="Helvetica-Bold")
        signature_table = Table(
            [
                [Paragraph(f"Batam, {tanggal_text}", sign_style)],
                [Paragraph(escape(jabatan), sign_style)],
                [""],
                [Paragraph(f"<u>{escape(nama)}</u>", sign_bold) if nama else ""],
            ],
            colWidths=[220],
            rowHeights=[12, 18, 30, 14],
        )
        signature_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        footer_table = Table([["", signature_table]], colWidths=[590, 220])
        footer_table.setStyle(TableStyle([
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(footer_table)

        doc.build(elements)
        return response


@admin.register(RKMItem)
class RKMItemAdmin(admin.ModelAdmin):
    list_display = (
        "summary",
        "no_item",
        "kategori_rkm",
        "km_item",
        "kpi_indikator",
        "program_kerja_utama",
        "target_akumulasi",
        "jumlah_realisasi",
        "persen_capaian",
        "pic_rkm",
    )
    list_filter = ("summary__tahun", "summary__bulan", "summary__unit_bisnis", "kategori_rkm")
    search_fields = (
        "summary__judul",
        "km_item__indikator_kinerja_kunci",
        "sasaran",
        "kpi_indikator",
        "program_kerja_utama",
        "pic_rkm",
    )
    ordering = ("summary", "no_item")
    base_fieldsets = (
        ("Acuan", {
            "fields": (
                "summary",
                "no_item",
                "kategori_rkm",
                "km_item",
            )
        }),
        ("KPI", {
            "fields": (
                "kpi_indikator",
                "kpi_satuan",
                "kpi_target",
            )
        }),
        ("Program Kerja", {
            "fields": (
                "inisiatif_strategis",
                "program_kerja_utama",
                "anggaran_rp_ribu",
                "target_akumulasi",
                "target_akumulasi_satuan",
            )
        }),
    )
    legacy_fieldset = (
        "Data Lama",
        {
            "classes": ("collapse",),
            "fields": (
                "sasaran",
                "target_bulanan",
                "realisasi",
                "deviasi",
                "keterangan",
            )
        },
    )
    fieldsets = base_fieldsets + (
        ("Realisasi Bulanan", {
            "fields": (
                "target_januari",
                "realisasi_januari",
                "target_februari",
                "realisasi_februari",
                "target_maret",
                "realisasi_maret",
                "target_april",
                "realisasi_april",
                "target_mei",
                "realisasi_mei",
                "target_juni",
                "realisasi_juni",
                "target_juli",
                "realisasi_juli",
                "target_agustus",
                "realisasi_agustus",
                "target_september",
                "realisasi_september",
                "target_oktober",
                "realisasi_oktober",
                "target_november",
                "realisasi_november",
                "target_desember",
                "realisasi_desember",
            )
        }),
    ) + (legacy_fieldset,)

    def _month_fields_for_item_form(self, request, obj=None):
        month = getattr(getattr(obj, "summary", None), "bulan", None)
        if not month:
            summary_id = request.POST.get("summary") or request.GET.get("summary")
            if summary_id:
                month = RKMSummary.objects.filter(pk=summary_id).values_list("bulan", flat=True).first()
        return RKM_MONTH_FIELD_PAIRS.get(month)

    def get_fieldsets(self, request, obj=None):
        month_fields = self._month_fields_for_item_form(request, obj)
        if not month_fields:
            return super().get_fieldsets(request, obj)
        month_label = RKM_MONTH_LABELS.get(obj.summary.bulan, "Bulan Laporan") if obj and obj.summary_id else "Bulan Laporan"
        return self.base_fieldsets + (
            (f"Realisasi {month_label}", {"fields": month_fields}),
            self.legacy_fieldset,
        )

    def get_readonly_fields(self, request, obj=None):
        if obj and obj.summary.is_approved:
            readonly = []
            for _, options in self.get_fieldsets(request, obj):
                readonly.extend(options.get("fields", ()))
            return tuple(readonly)
        month_fields = self._month_fields_for_item_form(request, obj)
        if month_fields:
            return (month_fields[0],)
        return super().get_readonly_fields(request, obj)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(summary__unit_bisnis__in=assigned_unit_businesses_for_user(request.user))

    def get_list_filter(self, request):
        if request.user.is_superuser:
            return self.list_filter
        return tuple(
            item for item in self.list_filter if item != "summary__unit_bisnis"
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "summary":
            summary_qs = RKMSummary.objects.exclude(
                status="Final",
            ).exclude(
                status_pengajuan="Disetujui",
            )
            if not request.user.is_superuser:
                summary_qs = summary_qs.filter(
                    unit_bisnis__in=assigned_unit_businesses_for_user(request.user),
                )
            kwargs["queryset"] = summary_qs
        elif not request.user.is_superuser:
            if db_field.name == "km_item":
                kwargs["queryset"] = ItemKontrakManajemen.objects.filter(
                    kontrak__unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
                )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_view_permission(self, request, obj=None):
        allowed = super().has_view_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.summary.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = super().has_change_permission(request, obj)
        if not allowed or obj is None:
            return allowed
        if obj.summary.is_approved:
            return False
        if request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.summary.unit_bisnis_id)

    def has_delete_permission(self, request, obj=None):
        allowed = super().has_delete_permission(request, obj)
        if not allowed or obj is None:
            return allowed
        if obj.summary.is_approved:
            return False
        if request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.summary.unit_bisnis_id)


# =========================================================
# RE-ASSESSMENT
# =========================================================

class ReAssessmentItemInline(admin.TabularInline):
    model = ReAssessmentItem
    extra = 0
    ordering = ("no_item",)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "km_item":
            summary_id = request.POST.get("summary") or request.GET.get("summary")

            object_id = request.resolver_match.kwargs.get("object_id")
            if not summary_id and object_id:
                obj = ReAssessmentItem.objects.filter(pk=object_id).first()
                if obj:
                    summary_id = obj.summary_id

            if summary_id:
                summary = ReAssessmentSummary.objects.filter(pk=summary_id).first()

                if summary and summary.kontrak_manajemen_id:
                    kwargs["queryset"] = ItemKontrakManajemen.objects.filter(
                        kontrak_id=summary.kontrak_manajemen_id
                    ).order_by(
                        "master_bagian__urutan",
                        "no_urut",
                    )
                else:
                    kwargs["queryset"] = ItemKontrakManajemen.objects.none()
            else:
                kwargs["queryset"] = ItemKontrakManajemen.objects.all().order_by(
                    "kontrak__judul",
                    "master_bagian__urutan",
                    "no_urut",
                )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

@admin.register(ReAssessmentSummary)
class ReAssessmentSummaryAdmin(admin.ModelAdmin):
    list_display = (
        "judul",
        "tahun",
        "unit_bisnis",
        "kontrak_manajemen",
        "rkm",
        "dibuat_pada",
        "pdf_button",
    )

    list_filter = ("tahun", "unit_bisnis")
    search_fields = ("judul", "unit_bisnis__name", "kontrak_manajemen__judul")
    ordering = ("-tahun", "judul")

    fields = (
        "judul",
        "tahun",
        "unit_bisnis",
        "kontrak_manajemen",
        "rkm",
        "risk_matrix",
    )

    autocomplete_fields = (
        "unit_bisnis",
        "kontrak_manajemen",
        "rkm",
        "risk_matrix",
    )

    inlines = [ReAssessmentItemInline]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:summary_id>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="risk_reassessmentsummary_pdf",
            ),
        ]
        return custom_urls + urls

    def pdf_button(self, obj):
        url = reverse("admin:risk_reassessmentsummary_pdf", args=[obj.pk])
        return format_html('<a class="button" href="{}" target="_blank">PDF Profil Risiko</a>', url)

    pdf_button.short_description = "Laporan"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(unit_bisnis__in=assigned_unit_businesses_for_user(request.user))

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser:
            if db_field.name == "unit_bisnis":
                kwargs["queryset"] = assigned_unit_businesses_for_user(request.user)
            elif db_field.name == "kontrak_manajemen":
                kwargs["queryset"] = KontrakManajemen.objects.filter(
                    unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
                )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_view_permission(self, request, obj=None):
        allowed = super().has_view_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = super().has_change_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_delete_permission(self, request, obj=None):
        allowed = super().has_delete_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def logo_flowable_pdf(self):
        logo_path = None
        app_setting = AppSetting.objects.first()
        if app_setting and app_setting.logo:
            candidate = Path(app_setting.logo.path)
            if candidate.exists():
                logo_path = candidate

        if not logo_path:
            candidate = Path(settings.MEDIA_ROOT) / "system/logo/pln_batam_logo.png"
            if candidate.exists():
                logo_path = candidate

        if logo_path:
            return Image(str(logo_path), width=100, height=50)

        return Table(
            [[
                Paragraph("<b>PLN</b>", ParagraphStyle(
                    "ProfilLogoPLN",
                    parent=getSampleStyleSheet()["Normal"],
                    fontName="Helvetica-Bold",
                    fontSize=11,
                    leading=12,
                    alignment=TA_CENTER,
                )),
                Paragraph("Batam", ParagraphStyle(
                    "ProfilLogoBatam",
                    parent=getSampleStyleSheet()["Normal"],
                    fontName="Helvetica-Bold",
                    fontSize=10,
                    leading=11,
                )),
            ]],
            colWidths=[34, 58],
        )

    def paragraph_pdf(self, text, style):
        return Paragraph(escape(str(text or "")), style)

    def pdf_view(self, request, summary_id):
        summary = get_object_or_404(
            ReAssessmentSummary.objects.select_related(
                "unit_bisnis",
                "kontrak_manajemen",
                "rkm",
                "risk_matrix",
            ),
            pk=summary_id,
        )
        if not user_can_access_unit(request, summary.unit_bisnis_id):
            raise PermissionDenied

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="Profil_Risiko_{summary.unit_bisnis.name}_{summary.tahun}.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=12,
            leftMargin=12,
            topMargin=12,
            bottomMargin=12,
        )
        styles = getSampleStyleSheet()
        normal = ParagraphStyle(
            "ProfilNormal",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=5.2,
            leading=6.2,
            alignment=TA_LEFT,
            spaceAfter=0,
        )
        normal_center = ParagraphStyle("ProfilNormalCenter", parent=normal, alignment=TA_CENTER)
        header_style = ParagraphStyle(
            "ProfilHeader",
            parent=normal_center,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        title_style = ParagraphStyle(
            "ProfilTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            alignment=TA_CENTER,
            spaceAfter=2,
        )
        section_style = ParagraphStyle("ProfilSection", parent=normal, fontName="Helvetica-Bold")

        elements = []
        header_table = Table(
            [[
                self.logo_flowable_pdf(),
                [
                    Paragraph(f"PROFIL RISIKO BIDANG/UNIT BISNIS TAHUN {summary.tahun}", title_style),
                    Paragraph(str(summary.unit_bisnis or "").upper(), title_style),
                    Paragraph(str(summary.judul or ""), title_style),
                ],
                "",
            ]],
            colWidths=[110, 590, 110],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(header_table)

        info_table = Table(
            [[
                self.paragraph_pdf("Kontrak Manajemen", section_style),
                self.paragraph_pdf(summary.kontrak_manajemen, normal),
                self.paragraph_pdf("RKM", section_style),
                self.paragraph_pdf(summary.rkm or "-", normal),
                self.paragraph_pdf("Matriks Risiko", section_style),
                self.paragraph_pdf(summary.risk_matrix or "-", normal),
            ]],
            colWidths=[82, 235, 32, 245, 58, 158],
        )
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C9D6")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF3F8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 6))

        data = [[
            self.paragraph_pdf("NO", header_style),
            self.paragraph_pdf("KM/KPI", header_style),
            self.paragraph_pdf("PERISTIWA RISIKO", header_style),
            self.paragraph_pdf("PENYEBAB / KRI", header_style),
            self.paragraph_pdf("EXISTING CONTROL", header_style),
            self.paragraph_pdf("DAMPAK", header_style),
            self.paragraph_pdf("Q1", header_style),
            self.paragraph_pdf("Q2", header_style),
            self.paragraph_pdf("Q3", header_style),
            self.paragraph_pdf("Q4", header_style),
            self.paragraph_pdf("RENCANA / OUTPUT PERLAKUAN", header_style),
            self.paragraph_pdf("PIC", header_style),
        ]]

        items = (
            summary.item
            .select_related("km_item", "km_item__master_bagian")
            .order_by("no_item", "no_risiko", "no_penyebab_risiko", "id")
        )
        for item in items:
            km_label = ""
            if item.km_item_id:
                prefix = item.km_item.master_bagian.kode_bagian if item.km_item.master_bagian_id else ""
                km_label = f"{prefix}{item.km_item.no_urut} - {item.km_item.indikator_kinerja_kunci}"
            q_values = [
                f"{getattr(item, f'level_nilai_risiko_q{q}') or ''}\n{getattr(item, f'eksposur_risiko_q{q}') or ''}"
                for q in range(1, 5)
            ]
            treatment = "\n".join(filter(None, [item.rencana_perlakuan_risiko, item.output_perlakuan_risiko]))
            cause_kri = "\n".join(filter(None, [item.penyebab_risiko, item.key_risk_indicators]))
            data.append([
                self.paragraph_pdf(f"{item.no_item}.{item.no_risiko}{item.no_penyebab_risiko or ''}", normal_center),
                self.paragraph_pdf(km_label, normal),
                self.paragraph_pdf(item.peristiwa_risiko, normal),
                self.paragraph_pdf(cause_kri, normal),
                self.paragraph_pdf(item.existing_control, normal),
                self.paragraph_pdf(item.deskripsi_dampak or item.asumsi_perhitungan_dampak, normal),
                self.paragraph_pdf(q_values[0], normal_center),
                self.paragraph_pdf(q_values[1], normal_center),
                self.paragraph_pdf(q_values[2], normal_center),
                self.paragraph_pdf(q_values[3], normal_center),
                self.paragraph_pdf(treatment, normal),
                self.paragraph_pdf(item.pic, normal),
            ])

        table = Table(
            data,
            colWidths=[30, 105, 105, 110, 95, 95, 40, 40, 40, 40, 95, 55],
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0070C0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 5.2),
            ("LEADING", (0, 0), (-1, -1), 6.2),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("ALIGN", (6, 1), (9, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(table)
        doc.build(elements)
        return response

class ProfilRisikoKorporatSumberByReassessmentInline(admin.TabularInline):
    model = ProfilRisikoKorporatSumber
    fk_name = "reassessment_item"
    extra = 0
    autocomplete_fields = ("risiko_korporat",)
    fields = (
        "risiko_korporat",
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
        "penyebab_risiko",
        "keterangan",
    )
    readonly_fields = (
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
        "penyebab_risiko",
    )
    verbose_name = "Relasi ke Profil Risiko Korporat"
    verbose_name_plural = "Relasi ke Profil Risiko Korporat"


@admin.register(ReAssessmentItem)
class ReAssessmentItemAdmin(admin.ModelAdmin):
    fields = (
        "summary",
        "km_item",
        "no_item",
        "taksonomi_t3",
        "sasaran_kbumn",
        "kategori_risiko",
        "no_risiko",
        "peristiwa_risiko",
        "deskripsi_peristiwa_risiko",
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
        "penyebab_risiko",
        "key_risk_indicators",
        "unit_satuan_kri",
        "threshold_aman",
        "threshold_hati_hati",
        "threshold_bahaya",
        "jenis_existing_control",
        "existing_control",
        "penilaian_efektivitas_kontrol",
        "kategori_dampak",
        "deskripsi_dampak",
        "perkiraan_waktu_terpapar_risiko",
        "asumsi_perhitungan_dampak",
        "nilai_dampak",
        "nilai_dampak_q1",
        "nilai_dampak_q2",
        "nilai_dampak_q3",
        "nilai_dampak_q4",
        "skala_dampak_q1",
        "skala_dampak_q2",
        "skala_dampak_q3",
        "skala_dampak_q4",
        "nilai_probabilitas",
        "nilai_probabilitas_q1",
        "nilai_probabilitas_q2",
        "nilai_probabilitas_q3",
        "nilai_probabilitas_q4",
        "skala_probabilitas",
        "skala_probabilitas_q1",
        "skala_probabilitas_q2",
        "skala_probabilitas_q3",
        "skala_probabilitas_q4",
        "eksposur_risiko_q1",
        "eksposur_risiko_q2",
        "eksposur_risiko_q3",
        "eksposur_risiko_q4",
        "skala_risiko_q1",
        "skala_risiko_q2",
        "skala_risiko_q3",
        "skala_risiko_q4",
        "level_nilai_risiko_q1",
        "level_nilai_risiko_q2",
        "level_nilai_risiko_q3",
        "level_nilai_risiko_q4",
        "opsi_perlakuan_risiko",
        "jenis_rencana_perlakuan_risiko",
        "rencana_perlakuan_risiko",
        "output_perlakuan_risiko",
        "biaya_perlakuan_risiko",
        "pos_anggaran",
        "prk",
        "jenis_program_dalam_rkap",
        "pic",
        "timeline_1",
        "timeline_2",
        "timeline_3",
        "timeline_4",
        "timeline_5",
        "timeline_6",
        "timeline_7",
        "timeline_8",
        "timeline_9",
        "timeline_10",
        "timeline_11",
        "timeline_12",
    )
    readonly_fields = (
        "kode_penyebab_risiko",
        "nilai_dampak_q1",
        "nilai_probabilitas_q1",
        "nilai_probabilitas_q2",
        "nilai_probabilitas_q3",
        "nilai_probabilitas_q4",
        "eksposur_risiko_q1",
        "eksposur_risiko_q2",
        "eksposur_risiko_q3",
        "eksposur_risiko_q4",
        "skala_risiko_q1",
        "skala_risiko_q2",
        "skala_risiko_q3",
        "skala_risiko_q4",
        "level_nilai_risiko_q1",
        "level_nilai_risiko_q2",
        "level_nilai_risiko_q3",
        "level_nilai_risiko_q4",
    )
    list_display = (
        "summary",
        "no_item",
        "unit_bisnis_summary",
        "km_item",
        "sasaran_kbumn",
        "taksonomi_t3",
        "kategori_risiko",
        "no_risiko",
        "jumlah_relasi_korporat",
    )
    list_filter = (
        "summary__tahun",
        "summary__unit_bisnis",
        "sasaran_kbumn",
        "taksonomi_t3",
        "kategori_risiko",
    )
    search_fields = (
        "peristiwa_risiko",
        "deskripsi_peristiwa_risiko",
        "km_item__indikator_kinerja_kunci",
        "sasaran_kbumn__kode",
        "sasaran_kbumn__nama",
        "taksonomi_t3__kode",
        "taksonomi_t3__nama",
        "kategori_risiko__kode",
        "kategori_risiko__nama",
        "summary__judul",
        "summary__unit_bisnis__name",
        "mendukung_risiko_korporat__risiko_korporat__peristiwa_risiko",
        "mendukung_risiko_korporat__risiko_korporat__summary__judul",
    )
    ordering = ("summary", "no_item")
    autocomplete_fields = (
        "summary",
        "taksonomi_t3",
        "sasaran_kbumn",
        "kategori_risiko",
        "jenis_existing_control",
        "penilaian_efektivitas_kontrol",
        "kategori_dampak",
        "opsi_perlakuan_risiko",
        "pos_anggaran",
        "jenis_program_dalam_rkap",
    )
    inlines = [ProfilRisikoKorporatSumberByReassessmentInline]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(
            summary__unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
        )

    def has_view_permission(self, request, obj=None):
        allowed = super().has_view_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.summary.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = super().has_change_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.summary.unit_bisnis_id)

    def has_delete_permission(self, request, obj=None):
        allowed = super().has_delete_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.summary.unit_bisnis_id)

    def unit_bisnis_summary(self, obj):
        return obj.summary.unit_bisnis if obj.summary_id else "-"
    unit_bisnis_summary.short_description = "Bidang / Unit Bisnis"

    def jumlah_relasi_korporat(self, obj):
        return obj.mendukung_risiko_korporat.count()
    jumlah_relasi_korporat.short_description = "Relasi Korporat"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser and db_field.name == "summary":
            kwargs["queryset"] = ReAssessmentSummary.objects.filter(
                unit_bisnis__in=assigned_unit_businesses_for_user(request.user)
            )

        if db_field.name == "km_item":
            object_id = request.resolver_match.kwargs.get("object_id")

            if object_id:
                obj = ReAssessmentItem.objects.filter(pk=object_id).first()
                if obj and obj.summary_id:
                    kwargs["queryset"] = ItemKontrakManajemen.objects.filter(
                        kontrak_id=obj.summary.kontrak_manajemen_id
                    ).order_by("master_bagian__urutan", "no_urut")
                else:
                    kwargs["queryset"] = ItemKontrakManajemen.objects.none()
            else:
                kwargs["queryset"] = ItemKontrakManajemen.objects.all().order_by(
                    "kontrak__judul",
                    "master_bagian__urutan",
                    "no_urut",
                )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

# =========================================================
# KPMR
# =========================================================

class KPMRItemInline(admin.TabularInline):
    model = KPMRItem
    extra = 0
    can_delete = False

    fields = (
        "no_item",
        "reassessment_item",
        "perlakuan_risiko",
        "bukti",
        "nilai_kpmr",
        "status_kpmr",
        "catatan",
    )

    readonly_fields = (
        "no_item",
        "reassessment_item",
        "perlakuan_risiko",
        "bukti",
        "nilai_kpmr",
        "status_kpmr",
        "catatan",
    )

    ordering = ("no_item",)

    def has_add_permission(self, request, obj=None):
        return False

@admin.register(KPMRSummary)
class KPMRSummaryAdmin(admin.ModelAdmin):
    list_display = (
        "judul",
        "tahun",
        "unit_bisnis",
        "reassessment",
        "dibuat_pada",
    )
    list_filter = ("tahun", "unit_bisnis")
    search_fields = ("judul", "unit_bisnis__name", "reassessment__judul")
    ordering = ("-tahun", "judul")
    readonly_fields = ("dibuat_pada",)

    inlines = [KPMRItemInline]

    fieldsets = (
        ("Informasi Utama", {
            "fields": (
                "judul",
                "tahun",
                "unit_bisnis",
                "reassessment",
                "dibuat_pada",
            )
        }),
    )

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if obj and obj.reassessment_id:
            readonly.extend(["unit_bisnis", "tahun"])
        return readonly

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        self.message_user(
            request,
            "KPMR berhasil disimpan dan item dihitung otomatis dari Profil Risiko Bidang/Unit Bisnis.",
            level=messages.SUCCESS,
        )

@admin.register(KPMRItem)
class KPMRItemAdmin(admin.ModelAdmin):
    list_display = (
        "summary",
        "no_item",
        "reassessment_item",
        "status_kpmr",
        "nilai_kpmr",
    )
    list_filter = (
        "summary__tahun",
        "summary__unit_bisnis",
        "status_kpmr",
    )
    search_fields = (
        "reassessment_item__peristiwa_risiko",
        "perlakuan_risiko",
        "catatan",
    )
    ordering = ("summary", "no_item")
    readonly_fields = (
        "summary",
        "no_item",
        "reassessment_item",
        "perlakuan_risiko",
        "bukti",
        "nilai_kpmr",
        "status_kpmr",
        "catatan",
    )

    def has_add_permission(self, request):
        return False


@admin.register(RiskManagementReview)
class RiskManagementReviewAdmin(admin.ModelAdmin):
    list_display = (
        "title",
        "tahun",
        "unit_bisnis",
        "profil_risiko",
        "rkm",
        "kpmr",
        "status",
        "review_date",
        "pdf_button",
    )
    list_filter = ("tahun", "status", "unit_bisnis", "review_date")
    search_fields = (
        "title",
        "unit_bisnis__name",
        "profil_risiko__judul",
        "rkm__judul",
        "kpmr__judul",
        "review_summary",
        "recommendation",
    )
    ordering = ("-tahun", "-review_date", "unit_bisnis__name")
    autocomplete_fields = (
        "unit_bisnis",
        "kontrak_manajemen",
        "rkm",
        "profil_risiko",
        "kpmr",
        "pairing_officer",
        "man_risk",
        "vp_mrk",
    )
    readonly_fields = ("created_at", "updated_at")
    fieldsets = (
        ("Dokumen Review", {
            "fields": (
                "title",
                "tahun",
                "unit_bisnis",
                "kontrak_manajemen",
                "rkm",
                "profil_risiko",
                "kpmr",
                "review_date",
                "status",
            )
        }),
        ("Hasil Review Sub Bidang Manajemen Risiko", {
            "fields": (
                "review_summary",
                "km_notes",
                "rkm_notes",
                "profil_risiko_notes",
                "kpmr_notes",
                "recommendation",
            )
        }),
        ("Penandatangan", {
            "fields": (
                "pairing_officer",
                "man_risk",
                "vp_mrk",
            )
        }),
        ("Audit", {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:review_id>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="risk_riskmanagementreview_pdf",
            ),
        ]
        return custom_urls + urls

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(unit_bisnis__in=assigned_unit_businesses_for_user(request.user))

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser:
            units = assigned_unit_businesses_for_user(request.user)
            if db_field.name == "unit_bisnis":
                kwargs["queryset"] = units
            elif db_field.name == "kontrak_manajemen":
                kwargs["queryset"] = KontrakManajemen.objects.filter(unit_bisnis__in=units)
            elif db_field.name == "rkm":
                kwargs["queryset"] = RKMSummary.objects.filter(unit_bisnis__in=units)
            elif db_field.name == "profil_risiko":
                kwargs["queryset"] = ReAssessmentSummary.objects.filter(unit_bisnis__in=units)
            elif db_field.name == "kpmr":
                kwargs["queryset"] = KPMRSummary.objects.filter(unit_bisnis__in=units)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_view_permission(self, request, obj=None):
        allowed = super().has_view_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_change_permission(self, request, obj=None):
        allowed = super().has_change_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def has_delete_permission(self, request, obj=None):
        allowed = super().has_delete_permission(request, obj)
        if not allowed or obj is None or request.user.is_superuser:
            return allowed
        return user_can_access_unit(request, obj.unit_bisnis_id)

    def pdf_button(self, obj):
        url = reverse("admin:risk_riskmanagementreview_pdf", args=[obj.pk])
        return format_html('<a class="button" href="{}" target="_blank">PDF Review</a>', url)

    pdf_button.short_description = "PDF"

    def _logo_flowable(self):
        logo_path = None
        app_setting = AppSetting.objects.first()
        if app_setting and app_setting.logo:
            candidate = Path(app_setting.logo.path)
            if candidate.exists():
                logo_path = candidate
        if not logo_path:
            candidate = Path(settings.MEDIA_ROOT) / "system/logo/pln_batam_logo.png"
            if candidate.exists():
                logo_path = candidate
        if logo_path:
            return Image(str(logo_path), width=100, height=50)
        return Paragraph("<b>PLN Batam</b>", getSampleStyleSheet()["Normal"])

    def _p(self, text, style):
        return Paragraph(escape(str(text or "")), style)

    def _user_name(self, user):
        if not user:
            return ""
        return user.get_full_name() or user.username

    def _user_position(self, user, tanggal):
        if not user or not tanggal:
            return ""
        riwayat = (
            user.riwayat_jabatan
            .filter(tanggal_mulai__lte=tanggal)
            .filter(
                models.Q(tanggal_selesai__isnull=True)
                | models.Q(tanggal_selesai__gte=tanggal)
            )
            .order_by("-tanggal_mulai")
            .first()
        )
        return riwayat.jabatan if riwayat else ""

    def _signature_cell(self, label, user, tanggal, normal_center, sign_bold):
        name = self._user_name(user)
        position = self._user_position(user, tanggal) or label
        return [
            Paragraph(escape(label), normal_center),
            Paragraph(escape(position), normal_center),
            Spacer(1, 34),
            Paragraph(f"<u>{escape(name)}</u>", sign_bold) if name else Paragraph("", sign_bold),
        ]

    def pdf_view(self, request, review_id):
        review = get_object_or_404(
            RiskManagementReview.objects.select_related(
                "unit_bisnis",
                "kontrak_manajemen",
                "rkm",
                "profil_risiko",
                "kpmr",
                "pairing_officer",
                "man_risk",
                "vp_mrk",
            ),
            pk=review_id,
        )
        if not user_can_access_unit(request, review.unit_bisnis_id):
            raise PermissionDenied

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="Review_MR_{review.unit_bisnis.name}_{review.tahun}.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=28,
            leftMargin=28,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        normal = ParagraphStyle(
            "MRReviewNormal",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
        )
        normal_center = ParagraphStyle("MRReviewCenter", parent=normal, alignment=TA_CENTER)
        title_style = ParagraphStyle(
            "MRReviewTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            alignment=TA_CENTER,
        )
        header_style = ParagraphStyle(
            "MRReviewHeader",
            parent=normal_center,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        label_style = ParagraphStyle("MRReviewLabel", parent=normal, fontName="Helvetica-Bold")

        elements = []
        header_table = Table(
            [[
                self._logo_flowable(),
                [
                    Paragraph("HASIL REVIEW SUB BIDANG MANAJEMEN RISIKO", title_style),
                    Paragraph(str(review.unit_bisnis or "").upper(), title_style),
                    Paragraph(f"TAHUN {review.tahun}", title_style),
                ],
            ]],
            colWidths=[105, 435],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(header_table)

        info_rows = [
            ["Judul Review", review.title, "Tanggal Review", review.review_date.strftime("%d-%m-%Y")],
            ["Kontrak Manajemen", review.kontrak_manajemen, "Status", review.get_status_display()],
            ["RKM", review.rkm, "KPMR", review.kpmr or "-"],
            ["Profil Risiko", review.profil_risiko, "Unit", review.unit_bisnis],
        ]
        info_table = Table(
            [[self._p(a, label_style), self._p(b, normal), self._p(c, label_style), self._p(d, normal)] for a, b, c, d in info_rows],
            colWidths=[95, 190, 80, 175],
        )
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C9D6")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF3F8")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        review_rows = [
            ["Ringkasan Hasil Review", review.review_summary],
            ["Catatan KM", review.km_notes],
            ["Catatan RKM", review.rkm_notes],
            ["Catatan Profil Risiko", review.profil_risiko_notes],
            ["Catatan KPMR", review.kpmr_notes],
            ["Rekomendasi / Tindak Lanjut", review.recommendation],
        ]
        review_table = Table(
            [[self._p(label, label_style), self._p(value or "-", normal)] for label, value in review_rows],
            colWidths=[145, 395],
        )
        review_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9EAF7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(review_table)
        elements.append(Spacer(1, 16))

        tanggal_text = review.review_date.strftime("%d-%m-%Y")
        elements.append(Paragraph(f"Batam, {tanggal_text}", normal))
        elements.append(Spacer(1, 8))

        sign_bold = ParagraphStyle("MRReviewSignBold", parent=normal_center, fontName="Helvetica-Bold")
        signature_rows = [
            [
                self._signature_cell("Pairing Officer", review.pairing_officer, review.review_date, normal_center, sign_bold),
                self._signature_cell("MAN RISK", review.man_risk, review.review_date, normal_center, sign_bold),
                self._signature_cell("VP MRK", review.vp_mrk, review.review_date, normal_center, sign_bold),
            ]
        ]
        signature_table = Table(signature_rows, colWidths=[180, 180, 180], rowHeights=[88])
        signature_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C9D6")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(signature_table)

        doc.build(elements)
        return response

# =========================================================
# KPMR PLN 2026 (RESMI)
# =========================================================

class KPMRSubIndikatorResmiInline(admin.TabularInline):
    model = KPMRSubIndikatorResmi
    extra = 0
    fields = (
        "kode",
        "nama",
        "bobot",
        "jawaban",
        "hasil",
        "skor",
        "keterangan",
    )
    ordering = ("kode",)


class KPMRIndikatorResmiInline(admin.TabularInline):
    model = KPMRIndikatorResmi
    extra = 0
    fields = (
        "kode",
        "nama",
        "bobot",
        "hasil",
        "skor",
        "dokumen_referensi",
        "keterangan",
    )
    ordering = ("kode",)
    show_change_link = True


@admin.register(KPMRPeriode)
class KPMRPeriodeAdmin(admin.ModelAdmin):
    list_display = (
        "tahun",
        "triwulan",
        "unit_bisnis",
        "status",
        "skor_total",
        "rating",
        "dibuat_pada",
    )
    list_filter = ("tahun", "triwulan", "status", "rating", "unit_bisnis")
    search_fields = ("unit_bisnis__name", "catatan")
    ordering = ("-tahun", "triwulan", "unit_bisnis__name")
    inlines = [KPMRIndikatorResmiInline]


@admin.register(KPMRIndikatorResmi)
class KPMRIndikatorResmiAdmin(admin.ModelAdmin):
    list_display = (
        "periode",
        "kode",
        "nama",
        "bobot",
        "hasil",
        "skor",
    )
    list_filter = ("kode", "periode__tahun", "periode__triwulan", "periode__unit_bisnis")
    search_fields = ("nama", "periode__unit_bisnis__name")
    ordering = ("periode", "kode")
    inlines = [KPMRSubIndikatorResmiInline]


@admin.register(KPMRSubIndikatorResmi)
class KPMRSubIndikatorResmiAdmin(admin.ModelAdmin):
    list_display = (
        "indikator",
        "kode",
        "nama",
        "bobot",
        "hasil",
        "skor",
    )
    list_filter = (
        "kode",
        "indikator__periode__tahun",
        "indikator__periode__triwulan",
        "indikator__periode__unit_bisnis",
    )
    search_fields = ("nama", "indikator__periode__unit_bisnis__name")
    ordering = ("indikator", "kode")


class KinerjaIndikatorInline(admin.TabularInline):
    model = KinerjaIndikator
    extra = 0
    fields = ("nama", "bobot", "hasil", "skor", "keterangan")
    ordering = ("nama",)


@admin.register(KinerjaPeriode)
class KinerjaPeriodeAdmin(admin.ModelAdmin):
    list_display = (
        "tahun",
        "triwulan",
        "unit_bisnis",
        "skor_total",
        "rating",
    )
    list_filter = ("tahun", "triwulan", "rating", "unit_bisnis")
    search_fields = ("unit_bisnis__name", "catatan")
    ordering = ("-tahun", "triwulan", "unit_bisnis__name")
    inlines = [KinerjaIndikatorInline]


@admin.register(KinerjaIndikator)
class KinerjaIndikatorAdmin(admin.ModelAdmin):
    list_display = ("periode", "nama", "bobot", "hasil", "skor")
    list_filter = ("nama", "periode__tahun", "periode__triwulan", "periode__unit_bisnis")
    search_fields = ("periode__unit_bisnis__name",)
    ordering = ("periode", "nama")


@admin.register(KompositRisikoTriwulan)
class KompositRisikoTriwulanAdmin(admin.ModelAdmin):
    list_display = (
        "periode_kpmr",
        "periode_kinerja",
        "skor_kpmr",
        "skor_kinerja",
        "peringkat_komposit",
    )
    search_fields = (
        "periode_kpmr__unit_bisnis__name",
        "periode_kinerja__unit_bisnis__name",
        "catatan_review_spi",
    )
    ordering = ("-periode_kpmr__tahun", "periode_kpmr__triwulan")


@admin.register(RoadmapProgram)
class RoadmapProgramAdmin(admin.ModelAdmin):
    list_display = ("tahun", "nomor_urut", "nama_program", "aktif")
    list_filter = ("tahun", "aktif")
    search_fields = ("nama_program",)
    ordering = ("tahun", "nomor_urut")


@admin.register(RoadmapPenilaianSemester)
class RoadmapPenilaianSemesterAdmin(admin.ModelAdmin):
    list_display = (
        "tahun",
        "semester",
        "unit_bisnis",
        "program",
        "nilai_kuantitas",
        "nilai_kualitas",
        "nilai_waktu",
        "nilai_program",
    )
    list_filter = ("tahun", "semester", "unit_bisnis")
    search_fields = ("unit_bisnis__name", "program__nama_program", "catatan")
    ordering = ("-tahun", "semester", "unit_bisnis__name", "program__nomor_urut")


# =========================================================
# PROFIL RISIKO KORPORAT
# =========================================================

class ProfilRisikoKorporatItemInline(admin.TabularInline):
    model = ProfilRisikoKorporatItem
    extra = 0
    ordering = ("no_item",)
    fields = (
        "no_item",
        "no_risiko",
        "bumn",
        "rkap_item",
        "sasaran_korporat",
        "sasaran_kbumn",
        "kategori_risiko",
        "taksonomi_t3",
        "peristiwa_risiko",
        "deskripsi_peristiwa_risiko",
        "dampak",
        "kemungkinan",
        "level_risiko",
        "matrix_cell_inheren",
        "residual_dampak",
        "residual_kemungkinan",
        "residual_level_risiko",
        "matrix_cell_residual",
        "status",
    )
    readonly_fields = (
        "no_risiko",
        "level_risiko",
        "matrix_cell_inheren",
        "residual_level_risiko",
        "matrix_cell_residual",
    )
    autocomplete_fields = ("bumn", "rkap_item", "sasaran_kbumn", "kategori_risiko", "taksonomi_t3")

@admin.register(ProfilRisikoKorporatSummary)
class ProfilRisikoKorporatSummaryAdmin(admin.ModelAdmin):
    change_form_template = "admin/risk/profilrisikokorporatsummary/change_form.html"
    list_display = (
        "judul",
        "tahun",
        "nama_perusahaan",
        "kode_perusahaan",
        "status",
        "dibuat_pada",
        "metric_history_shortcut",
        "lmr_button",
        "pdf_button",
    )
    list_filter = ("tahun", "status")
    search_fields = ("judul", "nama_perusahaan", "kode_perusahaan")
    ordering = ("-tahun", "judul")
    inlines = [ProfilRisikoKorporatItemInline]

    fieldsets = (
        ("Informasi Utama", {
            "fields": (
                "judul",
                "tahun",
                "nama_perusahaan",
                "kode_perusahaan",
                "status",
            )
        }),
        ("Catatan", {
            "fields": ("catatan",)
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:summary_id>/metric-history/save/",
                self.admin_site.admin_view(self.metric_history_save_view),
                name="risk_profilrisikokorporatsummary_metric_history_save",
            ),
            path(
                "<int:summary_id>/metric-history/upload/",
                self.admin_site.admin_view(self.metric_history_upload_view),
                name="risk_profilrisikokorporatsummary_metric_history_upload",
            ),
            path(
                "<int:summary_id>/metric-history/template/",
                self.admin_site.admin_view(self.metric_history_template_view),
                name="risk_profilrisikokorporatsummary_metric_history_template",
            ),
            path(
                "<int:summary_id>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="risk_profilrisikokorporatsummary_pdf",
            ),
            path(
                "<int:summary_id>/lmr-triwulan/<int:periode_id>/pdf/",
                self.admin_site.admin_view(self.lmr_quarterly_pdf_view),
                name="risk_profilrisikokorporatsummary_lmr_quarterly_pdf",
            ),
            path(
                "<int:summary_id>/lmr-triwulan/pdf/",
                self.admin_site.admin_view(self.lmr_quarterly_pdf_view),
                name="risk_profilrisikokorporatsummary_lmr_pdf",
            ),
        ]
        return custom_urls + urls

    def _can_view_metric_panel(self, request, summary):
        if request.user.is_superuser:
            return True
        return (
            self.has_view_permission(request, summary)
            and request.user.has_perm("corporate_risk.view_riskmetric")
        )

    def _can_manage_metric_history(self, request, summary):
        if request.user.is_superuser:
            return True
        return (
            self.has_change_permission(request, summary)
            and (
                request.user.has_perm("corporate_risk.add_montecarlometrichistory")
                or request.user.has_perm("corporate_risk.change_montecarlometrichistory")
            )
        )

    def _metric_queryset(self, summary):
        return (
            RiskMetric.objects.filter(corporate_risk_item__summary=summary)
            .select_related("corporate_risk_item", "rkap_item")
            .prefetch_related("metric_histories__periode")
            .order_by("corporate_risk_item__no_item", "name")
        )

    def _metric_history_redirect(self, summary_id):
        return redirect(
            f"{reverse(f'{self.admin_site.name}:risk_profilrisikokorporatsummary_change', args=[summary_id])}"
            "#monte-carlo-korporat"
        )

    def _get_scoped_metric(self, request, summary, metric_id, require_manage=False):
        if require_manage and not self._can_manage_metric_history(request, summary):
            raise PermissionDenied
        if not require_manage and not self._can_view_metric_panel(request, summary):
            raise PermissionDenied
        return get_object_or_404(self._metric_queryset(summary), pk=metric_id)

    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
        if obj and self._can_view_metric_panel(request, obj):
            metrics = list(self._metric_queryset(obj))
            metric_ids = [metric.pk for metric in metrics]
            context["monte_carlo_panel"] = {
                "summary": obj,
                "metrics": metrics,
                "periods": PeriodeLaporan.objects.all().order_by("tanggal_mulai", "kode_periode"),
                "quarter_periods": PeriodeLaporan.objects.filter(
                    tahun_buku__tahun=obj.tahun,
                    jenis_periode="triwulan",
                ).order_by("tanggal_mulai", "kode_periode"),
                "can_manage_history": self._can_manage_metric_history(request, obj),
                "save_url": reverse(
                    f"{self.admin_site.name}:risk_profilrisikokorporatsummary_metric_history_save",
                    args=[obj.pk],
                ),
                "upload_url": reverse(
                    f"{self.admin_site.name}:risk_profilrisikokorporatsummary_metric_history_upload",
                    args=[obj.pk],
                ),
                "template_url": reverse(
                    f"{self.admin_site.name}:risk_profilrisikokorporatsummary_metric_history_template",
                    args=[obj.pk],
                ),
                "lmr_fallback_url": reverse(
                    f"{self.admin_site.name}:risk_profilrisikokorporatsummary_lmr_pdf",
                    args=[obj.pk],
                ),
                "results": MultiMetricMonteCarloResult.objects.filter(
                    corporate_risk_item__summary=obj
                ).select_related("corporate_risk_item", "forecast_periode").order_by(
                    "corporate_risk_item__no_item", "-created_at"
                ),
                "insights": MultiMetricAIInsightKorporat.objects.filter(
                    multi_metric_result__corporate_risk_item__summary=obj
                ).select_related("multi_metric_result", "multi_metric_result__corporate_risk_item"),
                "history_count": MonteCarloMetricHistory.objects.filter(metric_id__in=metric_ids).count(),
            }
        return super().render_change_form(request, context, add, change, form_url, obj)

    def metric_history_shortcut(self, obj):
        url = reverse(
            f"{self.admin_site.name}:risk_profilrisikokorporatsummary_change",
            args=[obj.pk],
        )
        return format_html('<a class="button" href="{}#monte-carlo-korporat">Input Histori</a>', url)
    metric_history_shortcut.short_description = "Monte Carlo"

    def lmr_button(self, obj):
        period = (
            PeriodeLaporan.objects.filter(tahun_buku__tahun=obj.tahun, jenis_periode="triwulan")
            .order_by("tanggal_mulai", "kode_periode")
            .first()
        )
        if period:
            url = reverse(
                f"{self.admin_site.name}:risk_profilrisikokorporatsummary_lmr_quarterly_pdf",
                args=[obj.pk, period.pk],
            )
            label = "LMR TW"
        else:
            url = reverse(
                f"{self.admin_site.name}:risk_profilrisikokorporatsummary_lmr_pdf",
                args=[obj.pk],
            )
            label = "LMR PDF"
        return format_html('<a class="button" href="{}" target="_blank">{}</a>', url, label)
    lmr_button.short_description = "LMR Triwulan"

    def pdf_button(self, obj):
        url = reverse(f"{self.admin_site.name}:risk_profilrisikokorporatsummary_pdf", args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank">PDF</a>',
            url,
        )
    pdf_button.short_description = "PDF"

    def lmr_quarterly_pdf_view(self, request, summary_id, periode_id=None):
        summary = get_object_or_404(ProfilRisikoKorporatSummary, pk=summary_id)
        if not self.has_view_permission(request, summary):
            raise PermissionDenied
        period = None
        if periode_id is not None:
            period = get_object_or_404(
                PeriodeLaporan,
                pk=periode_id,
                tahun_buku__tahun=summary.tahun,
                jenis_periode="triwulan",
            )

        pdf_bytes = render_quarterly_lmr_pdf(summary, period)
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        period_code = period.kode_periode if period else "SEMUA_PERIODE"
        response["Content-Disposition"] = (
            f'inline; filename="LMR_Profil_Risiko_Monte_Carlo_{period_code}_{summary.tahun}.pdf"'
        )
        return response

    def metric_history_save_view(self, request, summary_id):
        if request.method != "POST":
            return HttpResponseNotAllowed(["POST"])
        summary = get_object_or_404(ProfilRisikoKorporatSummary, pk=summary_id)
        metric = self._get_scoped_metric(
            request,
            summary,
            request.POST.get("metric_id"),
            require_manage=True,
        )

        history_id = request.POST.get("history_id")
        periode_id = request.POST.get("periode")
        tanggal_data = request.POST.get("tanggal_data")
        metric_value = request.POST.get("metric_value")
        target_value = request.POST.get("target_value")
        keterangan = request.POST.get("keterangan", "")

        if not periode_id or not tanggal_data or metric_value in (None, ""):
            self.message_user(
                request,
                "Periode, tanggal data, dan nilai aktual wajib diisi.",
                level=messages.ERROR,
            )
            return self._metric_history_redirect(summary.pk)

        try:
            defaults = {
                "tanggal_data": tanggal_data,
                "metric_value": Decimal(str(metric_value)),
                "target_value": Decimal(str(target_value)) if target_value not in (None, "") else None,
                "keterangan": keterangan,
            }
        except Exception:
            self.message_user(request, "Nilai aktual/target harus berupa angka.", level=messages.ERROR)
            return self._metric_history_redirect(summary.pk)

        with transaction.atomic():
            if request.POST.get("delete") == "1" and history_id:
                deleted, _ = MonteCarloMetricHistory.objects.filter(metric=metric, pk=history_id).delete()
                self.message_user(request, f"{deleted} histori metric dihapus.")
                return self._metric_history_redirect(summary.pk)

            if history_id:
                history = get_object_or_404(MonteCarloMetricHistory, metric=metric, pk=history_id)
                for field, value in defaults.items():
                    setattr(history, field, value)
                history.periode_id = periode_id
                history.save()
                action = "diperbarui"
            else:
                MonteCarloMetricHistory.objects.update_or_create(
                    metric=metric,
                    periode_id=periode_id,
                    defaults=defaults,
                )
                action = "disimpan"

        self.message_user(request, f"Histori metric {metric.name} berhasil {action}.")
        return self._metric_history_redirect(summary.pk)

    def metric_history_upload_view(self, request, summary_id):
        if request.method != "POST":
            return HttpResponseNotAllowed(["POST"])
        summary = get_object_or_404(ProfilRisikoKorporatSummary, pk=summary_id)
        metric = self._get_scoped_metric(
            request,
            summary,
            request.POST.get("metric_id"),
            require_manage=True,
        )
        uploaded_file = request.FILES.get("excel_file")

        if not uploaded_file:
            self.message_user(request, "Pilih file Excel terlebih dahulu.", level=messages.ERROR)
            return self._metric_history_redirect(summary.pk)
        if not uploaded_file.name.lower().endswith(".xlsx"):
            self.message_user(request, "Upload hanya menerima file .xlsx.", level=messages.ERROR)
            return self._metric_history_redirect(summary.pk)
        if uploaded_file.size > 5 * 1024 * 1024:
            self.message_user(request, "Ukuran file maksimum 5 MB.", level=messages.ERROR)
            return self._metric_history_redirect(summary.pk)

        try:
            with transaction.atomic():
                imported, skipped = self._import_metric_history_excel(metric, uploaded_file)
            self.message_user(
                request,
                f"Upload Excel berhasil. {imported} baris tersimpan, {skipped} baris dilewati.",
            )
        except Exception as exc:
            self.message_user(request, f"Gagal upload Excel: {exc}", level=messages.ERROR)
        return self._metric_history_redirect(summary.pk)

    def metric_history_template_view(self, request, summary_id):
        summary = get_object_or_404(ProfilRisikoKorporatSummary, pk=summary_id)
        if not self._can_view_metric_panel(request, summary):
            raise PermissionDenied

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Metric History"
        ws.append([
            "risk_metric",
            "nomor_item_risiko",
            "peristiwa_risiko",
            "profil_risiko_korporat",
            "periode",
            "tanggal_data",
            "nilai_aktual",
            "target",
            "keterangan",
        ])
        first_metric = self._metric_queryset(summary).first()
        first_item = first_metric.corporate_risk_item if first_metric else None
        ws.append([
            first_metric.name if first_metric else "Nama metric",
            first_item.no_item if first_item else "11",
            first_item.get_peristiwa_risiko_text() if first_item else "Contoh peristiwa risiko",
            str(summary),
            "2026-01",
            "2026-01-31",
            1000,
            1200,
            "Contoh data",
        ])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="template_metric_history_{summary.pk}.xlsx"'
        )
        return response

    def _normalize_header(self, value):
        return str(value or "").strip().lower().replace(" ", "_")

    def _parse_decimal(self, value):
        if value in (None, ""):
            return None
        return Decimal(str(value).strip().replace(",", ""))

    def _parse_excel_date(self, value):
        if value in (None, ""):
            return None
        if hasattr(value, "date"):
            return value.date()
        return value

    def _resolve_period(self, value, tanggal_data=None):
        if value not in (None, ""):
            text = str(value).strip()
            periode = (
                PeriodeLaporan.objects.filter(kode_periode__iexact=text).first()
                or PeriodeLaporan.objects.filter(nama_periode__iexact=text).first()
            )
            if periode:
                return periode
            if text.isdigit():
                return PeriodeLaporan.objects.filter(pk=int(text)).first()
        if tanggal_data:
            return PeriodeLaporan.objects.filter(
                tanggal_mulai__lte=tanggal_data,
                tanggal_selesai__gte=tanggal_data,
            ).first()
        return None

    def _import_metric_history_excel(self, metric, uploaded_file):
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("File Excel kosong.")

        headers = [self._normalize_header(value) for value in rows[0]]
        aliases = {
            "periode": {"periode", "period", "bulan", "month", "kode_periode"},
            "tanggal_data": {"tanggal_data", "tanggal", "date", "tgl_data"},
            "metric_value": {"nilai_aktual", "actual", "realisasi", "metric_value", "nilai"},
            "target_value": {"target", "target_value", "target_rkap"},
            "keterangan": {"keterangan", "catatan", "note", "notes"},
        }
        column_map = {}
        for field, names in aliases.items():
            for idx, header in enumerate(headers):
                if header in names:
                    column_map[field] = idx
                    break
        if "metric_value" not in column_map:
            raise ValueError("Kolom nilai aktual tidak ditemukan.")

        imported = 0
        skipped = 0
        for raw_row in rows[1:]:
            if not raw_row or all(value in (None, "") for value in raw_row):
                continue

            def cell(field):
                idx = column_map.get(field)
                if idx is None or idx >= len(raw_row):
                    return None
                return raw_row[idx]

            tanggal_data = self._parse_excel_date(cell("tanggal_data"))
            periode = self._resolve_period(cell("periode"), tanggal_data=tanggal_data)
            metric_value = self._parse_decimal(cell("metric_value"))
            target_value = self._parse_decimal(cell("target_value"))
            keterangan = str(cell("keterangan") or "").strip()

            if not periode or metric_value is None:
                skipped += 1
                continue
            if tanggal_data is None:
                tanggal_data = periode.tanggal_selesai

            MonteCarloMetricHistory.objects.update_or_create(
                metric=metric,
                periode=periode,
                defaults={
                    "tanggal_data": tanggal_data,
                    "metric_value": metric_value,
                    "target_value": target_value,
                    "keterangan": keterangan,
                },
            )
            imported += 1
        return imported, skipped

    def pdf_text(self, text, style):
        escaped = escape(str(text or "")).replace("\n", "<br/>")
        return Paragraph(escaped, style)

    def pdf_label(self, value):
        return str(value) if value not in (None, "") else "-"

    def timeline_label(self, obj):
        months = [
            str(month)
            for month in range(1, 13)
            if getattr(obj, f"timeline_{month}", 0)
        ]
        return ", ".join(months) if months else "-"

    def pdf_view(self, request, summary_id):
        summary = get_object_or_404(ProfilRisikoKorporatSummary, pk=summary_id)
        items = list(
            summary.item
            .select_related(
                "bumn",
                "rkap_item",
                "sasaran_kbumn",
                "kategori_risiko",
                "taksonomi_t3",
                "matrix_cell_inheren__level_risiko",
                "matrix_cell_residual__level_risiko",
            )
            .prefetch_related(
                "daftar_penyebab__pemilik_risiko",
                "daftar_penyebab__jenis_existing_control",
                "daftar_penyebab__penilaian_efektivitas_kontrol",
                "daftar_penyebab__kategori_dampak",
                "sumber_risiko__reassessment_item__summary__unit_bisnis",
                "rencana_perlakuan_items__opsi_perlakuan_risiko",
                "rencana_perlakuan_items__pos_anggaran",
                "rencana_perlakuan_items__jenis_program_dalam_rkap",
                "rencana_perlakuan_items__jenis_rencana_perlakuan_risiko",
            )
            .order_by("no_item", "no_risiko")
        )

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="Profil_Risiko_Korporat_{summary.tahun}.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=18,
            leftMargin=18,
            topMargin=18,
            bottomMargin=18,
        )

        styles = getSampleStyleSheet()
        normal = ParagraphStyle(
            "PRKNormal",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.4,
            leading=7.6,
            alignment=TA_LEFT,
        )
        normal_center = ParagraphStyle(
            "PRKNormalCenter",
            parent=normal,
            alignment=TA_CENTER,
        )
        header = ParagraphStyle(
            "PRKHeader",
            parent=normal_center,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        section = ParagraphStyle(
            "PRKSection",
            parent=normal,
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=8,
        )
        title = ParagraphStyle(
            "PRKTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=16,
            alignment=TA_CENTER,
            spaceAfter=6,
        )
        subtitle = ParagraphStyle(
            "PRKSubtitle",
            parent=normal_center,
            fontSize=8,
            leading=10,
            spaceAfter=8,
        )

        elements = [
            Paragraph("LAPORAN PROFIL RISIKO KORPORAT", title),
            Paragraph(f"{escape(summary.nama_perusahaan)} - {summary.tahun}", subtitle),
        ]

        info_table = Table(
            [
                [
                    self.pdf_text("Judul", section),
                    self.pdf_text(summary.judul, normal),
                    self.pdf_text("Kode BUMN", section),
                    self.pdf_text(summary.kode_perusahaan, normal),
                ],
                [
                    self.pdf_text("Status", section),
                    self.pdf_text(summary.status or "-", normal),
                    self.pdf_text("Jumlah Risiko", section),
                    self.pdf_text(len(items), normal),
                ],
                [
                    self.pdf_text("Catatan", section),
                    self.pdf_text(summary.catatan or "-", normal),
                    "",
                    "",
                ],
            ],
            colWidths=[80, 300, 80, 300],
        )
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9CA3AF")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E5E7EB")),
            ("BACKGROUND", (2, 0), (2, 1), colors.HexColor("#E5E7EB")),
            ("SPAN", (1, 2), (3, 2)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        overview_data = [[
            self.pdf_text("No", header),
            self.pdf_text("Sasaran Korporat", header),
            self.pdf_text("Peristiwa Risiko", header),
            self.pdf_text("Kategori", header),
            self.pdf_text("T3", header),
            self.pdf_text("Inheren", header),
            self.pdf_text("Residual", header),
            self.pdf_text("Status", header),
        ]]
        for item in items:
            overview_data.append([
                self.pdf_text(item.no_risiko or item.no_item, normal_center),
                self.pdf_text(item.sasaran_korporat, normal),
                self.pdf_text(item.peristiwa_risiko, normal),
                self.pdf_text(item.kategori_risiko, normal),
                self.pdf_text(item.taksonomi_t3, normal),
                self.pdf_text(
                    f"D:{self.pdf_label(item.dampak)} K:{self.pdf_label(item.kemungkinan)} "
                    f"L:{self.pdf_label(item.level_risiko)} {item.get_level_name('inheren') or ''}",
                    normal_center,
                ),
                self.pdf_text(
                    f"D:{self.pdf_label(item.residual_dampak)} K:{self.pdf_label(item.residual_kemungkinan)} "
                    f"L:{self.pdf_label(item.residual_level_risiko)} {item.get_level_name('residual') or ''}",
                    normal_center,
                ),
                self.pdf_text(item.status or "-", normal_center),
            ])

        overview_table = Table(
            overview_data,
            colWidths=[28, 140, 180, 100, 110, 82, 82, 58],
            repeatRows=1,
        )
        overview_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#6B7280")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0070C0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(self.pdf_text("Ringkasan Risiko", section))
        elements.append(overview_table)
        elements.append(Spacer(1, 10))

        for item in items:
            elements.append(self.pdf_text(f"Detail Risiko {item.no_risiko or item.no_item}", section))
            detail_table = Table(
                [
                    [
                        self.pdf_text("Sasaran", section),
                        self.pdf_text(item.sasaran_korporat, normal),
                        self.pdf_text("Sasaran KBUMN", section),
                        self.pdf_text(item.sasaran_kbumn, normal),
                    ],
                    [
                        self.pdf_text("Peristiwa Risiko", section),
                        self.pdf_text(item.peristiwa_risiko, normal),
                        self.pdf_text("Deskripsi", section),
                        self.pdf_text(item.deskripsi_peristiwa_risiko or "-", normal),
                    ],
                    [
                        self.pdf_text("RKAP", section),
                        self.pdf_text(item.rkap_item or "-", normal),
                        self.pdf_text("BUMN", section),
                        self.pdf_text(f"{item.nama_bumn} ({item.kode_bumn})", normal),
                    ],
                ],
                colWidths=[75, 315, 75, 315],
            )
            detail_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9CA3AF")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F3F4F6")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(detail_table)
            elements.append(Spacer(1, 4))

            penyebab_rows = [[
                self.pdf_text("No", header),
                self.pdf_text("Pemilik", header),
                self.pdf_text("Penyebab", header),
                self.pdf_text("KRI / Threshold", header),
                self.pdf_text("Existing Control", header),
                self.pdf_text("Dampak", header),
            ]]
            for penyebab in item.daftar_penyebab.all():
                penyebab_rows.append([
                    self.pdf_text(penyebab.no_penyebab_risiko or penyebab.urutan, normal_center),
                    self.pdf_text(penyebab.pemilik_risiko or "-", normal),
                    self.pdf_text(penyebab.penyebab_risiko or "-", normal),
                    self.pdf_text(
                        f"{penyebab.key_risk_indicators or '-'}\n"
                        f"Aman: {self.pdf_label(penyebab.threshold_aman)}; "
                        f"Hati-hati: {self.pdf_label(penyebab.threshold_hati_hati)}; "
                        f"Bahaya: {self.pdf_label(penyebab.threshold_bahaya)}",
                        normal,
                    ),
                    self.pdf_text(
                        f"{self.pdf_label(penyebab.jenis_existing_control)}\n"
                        f"{self.pdf_label(penyebab.existing_control)}\n"
                        f"Efektivitas: {self.pdf_label(penyebab.penilaian_efektivitas_kontrol)}",
                        normal,
                    ),
                    self.pdf_text(
                        f"{self.pdf_label(penyebab.kategori_dampak)}\n"
                        f"{self.pdf_label(penyebab.deskripsi_dampak)}",
                        normal,
                    ),
                ])
            if len(penyebab_rows) == 1:
                penyebab_rows.append(["", "", self.pdf_text("Belum ada penyebab risiko.", normal), "", "", ""])

            penyebab_table = Table(
                penyebab_rows,
                colWidths=[28, 90, 170, 170, 170, 152],
                repeatRows=1,
            )
            penyebab_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#6B7280")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0070C0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(penyebab_table)
            elements.append(Spacer(1, 4))

            sumber_rows = [[
                self.pdf_text("Unit/Bidang", header),
                self.pdf_text("Risiko Bidang / Unit", header),
                self.pdf_text("Kode Penyebab", header),
                self.pdf_text("Penyebab", header),
                self.pdf_text("Keterangan", header),
            ]]
            for sumber in item.sumber_risiko.all():
                reassessment = sumber.reassessment_item
                unit = reassessment.summary.unit_bisnis if reassessment and reassessment.summary_id else "-"
                sumber_rows.append([
                    self.pdf_text(unit, normal),
                    self.pdf_text(reassessment.peristiwa_risiko if reassessment else "-", normal),
                    self.pdf_text(sumber.kode_penyebab_risiko or "-", normal_center),
                    self.pdf_text(sumber.penyebab_risiko or "-", normal),
                    self.pdf_text(sumber.keterangan or "-", normal),
                ])
            if len(sumber_rows) == 1:
                sumber_rows.append(["", self.pdf_text("Belum ada sumber risiko bidang/unit.", normal), "", "", ""])

            sumber_table = Table(
                sumber_rows,
                colWidths=[100, 230, 90, 220, 140],
                repeatRows=1,
            )
            sumber_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#6B7280")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0070C0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(sumber_table)
            elements.append(Spacer(1, 4))

            rencana_rows = [[
                self.pdf_text("No", header),
                self.pdf_text("Opsi / Jenis", header),
                self.pdf_text("Rencana", header),
                self.pdf_text("Output", header),
                self.pdf_text("Biaya / Pos / PRK", header),
                self.pdf_text("PIC / Timeline", header),
                self.pdf_text("Status", header),
            ]]
            for rencana in item.rencana_perlakuan_items.all():
                jenis = ", ".join(
                    str(jenis_rencana)
                    for jenis_rencana in rencana.jenis_rencana_perlakuan_risiko.all()
                )
                rencana_rows.append([
                    self.pdf_text(rencana.urutan, normal_center),
                    self.pdf_text(f"{self.pdf_label(rencana.opsi_perlakuan_risiko)}\n{jenis or '-'}", normal),
                    self.pdf_text(rencana.rencana_perlakuan_risiko or "-", normal),
                    self.pdf_text(rencana.output_perlakuan_risiko or "-", normal),
                    self.pdf_text(
                        f"{self.pdf_label(rencana.biaya_perlakuan_risiko)}\n"
                        f"{self.pdf_label(rencana.pos_anggaran)}\n"
                        f"PRK: {self.pdf_label(rencana.prk)}",
                        normal,
                    ),
                    self.pdf_text(f"{self.pdf_label(rencana.pic)}\nBulan: {self.timeline_label(rencana)}", normal),
                    self.pdf_text(rencana.status or "-", normal_center),
                ])
            if len(rencana_rows) == 1:
                rencana_rows.append(["", "", self.pdf_text("Belum ada rencana perlakuan risiko.", normal), "", "", "", ""])

            rencana_table = Table(
                rencana_rows,
                colWidths=[28, 118, 180, 145, 115, 115, 79],
                repeatRows=1,
            )
            rencana_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#6B7280")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0070C0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(rencana_table)
            elements.append(Spacer(1, 8))

        doc.build(elements)
        return response


class ProfilRisikoKorporatPenyebabInline(admin.StackedInline):
    model = ProfilRisikoKorporatPenyebab
    extra = 1
    ordering = ("urutan",)

    fields = (
        "urutan",
        "pemilik_risiko",
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
        "penyebab_risiko",
        "key_risk_indicators",
        "unit_satuan_kri",
        "threshold_aman",
        "threshold_hati_hati",
        "threshold_bahaya",
        "jenis_existing_control",
        "existing_control",
        "penilaian_efektivitas_kontrol",
        "kategori_dampak",
        "deskripsi_dampak",
        "perkiraan_waktu_terpapar",
    )

    readonly_fields = (
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
    )

    autocomplete_fields = (
        "pemilik_risiko",
        "jenis_existing_control",
        "penilaian_efektivitas_kontrol",
        "kategori_dampak",
    )   


class RisikoInherenKuantitatifInline(admin.StackedInline):
    model = RisikoInherenKuantitatif
    extra = 0
    max_num = 1

    fields = (
        "nilai_dampak",
        "probabilitas",
        "eksposur",
        "keterangan",
    )

    readonly_fields = (
        "eksposur",
    )


class RencanaPerlakuanRisikoKorporatInline(admin.StackedInline):
    model = RencanaPerlakuanRisikoKorporat
    extra = 1
    ordering = ("urutan",)

    fields = (
        "urutan",
        "opsi_perlakuan_risiko",
        "jenis_rencana_perlakuan_risiko",
        "rencana_perlakuan_risiko",
        "output_perlakuan_risiko",
        "biaya_perlakuan_risiko",
        "pos_anggaran",
        "prk",
        "jenis_program_dalam_rkap",
        "pic",
        "timeline_1",
        "timeline_2",
        "timeline_3",
        "timeline_4",
        "timeline_5",
        "timeline_6",
        "timeline_7",
        "timeline_8",
        "timeline_9",
        "timeline_10",
        "timeline_11",
        "timeline_12",
        "status",
        "keterangan",
    )

    autocomplete_fields = (
        "opsi_perlakuan_risiko",
        "pos_anggaran",
        "jenis_program_dalam_rkap",
    )

    filter_horizontal = (
        "jenis_rencana_perlakuan_risiko",
    )

@admin.register(ProfilRisikoKorporatItem)
class ProfilRisikoKorporatItemAdmin(admin.ModelAdmin):
    inlines = [
        ProfilRisikoKorporatPenyebabInline,
    ]

    list_display = (
        "no_item",
        "item_risiko_display",
        "summary",
        "tahun",
        "kategori_risiko",
        "risk_owner",
        "status",
    )

    search_fields = (
        "no_risiko",
        "no_item",
        "peristiwa_risiko",
        "deskripsi_peristiwa_risiko",
        "summary__judul",
        "daftar_penyebab__pemilik_risiko__username",
        "daftar_penyebab__pemilik_risiko__first_name",
        "daftar_penyebab__pemilik_risiko__last_name",
    )

    list_filter = (
        "summary__tahun",
        "summary",
        "kategori_risiko",
        "status",
    )

    ordering = (
        "summary",
        "no_item",
    )

    fieldsets = (
        ("Header Risiko", {
            "fields": (
                "summary",
                "no_item",
                "no_risiko",
                "bumn",
                "rkap_item",
            )
        }),
        ("Klasifikasi", {
            "fields": (
                "sasaran_korporat",
                "sasaran_kbumn",
                "kategori_risiko",
                "taksonomi_t3",
            )
        }),
        ("Peristiwa Risiko", {
            "fields": (
                "peristiwa_risiko",
                "deskripsi_peristiwa_risiko",
            )
        }),
        ("Penilaian Inheren", {
            "fields": (
                "dampak",
                "kemungkinan",
                "level_risiko",
                "matrix_cell_inheren",
            )
        }),
        ("Penilaian Residual", {
            "fields": (
                "residual_dampak",
                "residual_kemungkinan",
                "residual_level_risiko",
                "matrix_cell_residual",
            )
        }),
        ("Info Tambahan", {
            "fields": (
                "status",
            )
        }),
    )

    readonly_fields = (
        "level_risiko",
        "matrix_cell_inheren",
        "residual_level_risiko",
        "matrix_cell_residual",
    )

    autocomplete_fields = (
        "summary",
        "bumn",
        "rkap_item",
        "sasaran_kbumn",
        "kategori_risiko",
        "taksonomi_t3",
    )

    @admin.display(description="Peristiwa Risiko", ordering="peristiwa_risiko")
    def item_risiko_display(self, obj):
        return format_html(
            '<span title="{}">{}</span>',
            obj.get_display_label(),
            obj.short_label,
        )

    @admin.display(description="Tahun", ordering="summary__tahun")
    def tahun(self, obj):
        return obj.summary.tahun if obj.summary_id else "-"

    @admin.display(description="Risk Owner")
    def risk_owner(self, obj):
        owner = obj.daftar_penyebab.select_related("pemilik_risiko").filter(
            pemilik_risiko__isnull=False
        ).first()
        return owner.pemilik_risiko if owner else "-"

@admin.register(ProfilRisikoKorporatSumber)
class ProfilRisikoKorporatSumberAdmin(admin.ModelAdmin):
    list_display = (
        "risiko_korporat",
        "reassessment_item",
        "unit_bisnis_reassessment",
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
    )
    search_fields = (
        "risiko_korporat__sasaran_korporat",
        "risiko_korporat__peristiwa_risiko",
        "reassessment_item__peristiwa_risiko",
        "reassessment_item__penyebab_risiko",
        "kode_penyebab_risiko",
        "reassessment_item__summary__unit_bisnis__name",
    )
    list_filter = (
        "risiko_korporat__summary__tahun",
        "risiko_korporat__summary",
        "reassessment_item__summary__unit_bisnis",
    )
    ordering = (
        "risiko_korporat",
        "no_penyebab_risiko",
    )
    readonly_fields = (
        "no_penyebab_risiko",
        "kode_penyebab_risiko",
        "penyebab_risiko",
    )
    autocomplete_fields = (
        "risiko_korporat",
        "reassessment_item",
    )

    def unit_bisnis_reassessment(self, obj):
        if obj.reassessment_item_id and obj.reassessment_item.summary_id:
            return obj.reassessment_item.summary.unit_bisnis
        return "-"
    unit_bisnis_reassessment.short_description = "Unit/Bidang"

class TahunBukuAdmin(admin.ModelAdmin):
    list_display = ("tahun", "aktif")
    list_filter = ("aktif",)
    ordering = ("-tahun",)

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        try:
            if search_term:
                queryset |= self.model.objects.filter(tahun=int(search_term))
        except ValueError:
            pass
        return queryset, use_distinct


class PeriodeLaporanAdmin(admin.ModelAdmin):
    list_display = (
        "nama_periode",
        "tahun_buku",
        "kode_periode",
        "jenis_periode",
        "tanggal_mulai",
        "tanggal_selesai",
        "is_locked",
    )
    list_filter = ("jenis_periode", "tahun_buku", "is_locked")
    search_fields = ("nama_periode", "kode_periode", "tahun_buku__tahun")
    ordering = ("tahun_buku__tahun", "tanggal_mulai")

# =========================================================
# CUSTOM ADMIN SITE
# =========================================================

try:
    risk_admin_site.register(User, CustomUserAdmin)
except Exception:
    pass

try:
    risk_admin_site.register(Group, CustomGroupAdmin)
except Exception:
    pass

try:
    risk_admin_site.register(AppSetting, AppSettingAdmin)
except admin.sites.AlreadyRegistered:
    pass

risk_admin_site.register(BagianKontrakManajemen, BagianKontrakManajemenAdmin)
risk_admin_site.register(ItemKontrakManajemen, ItemKontrakManajemenAdmin)
risk_admin_site.register(RKMSummary, RKMSummaryAdmin)
risk_admin_site.register(RKMItem, RKMItemAdmin)
risk_admin_site.register(ReAssessmentSummary, ReAssessmentSummaryAdmin)
risk_admin_site.register(ReAssessmentItem, ReAssessmentItemAdmin)
risk_admin_site.register(KPMRSummary, KPMRSummaryAdmin)
risk_admin_site.register(RiskManagementReview, RiskManagementReviewAdmin)
risk_admin_site.register(ProfilRisikoKorporatSummary, ProfilRisikoKorporatSummaryAdmin)
risk_admin_site.register(ProfilRisikoKorporatItem, ProfilRisikoKorporatItemAdmin)
risk_admin_site.register(TaksonomiT3, TaksonomiT3Admin)
risk_admin_site.register(KategoriRisiko, KategoriRisikoAdmin)
risk_admin_site.register(SasaranKBUMN, SasaranKBUMNAdmin)
risk_admin_site.register(MasterBUMN, MasterBUMNAdmin)
risk_admin_site.register(MasterJenisExistingControl, MasterJenisExistingControlAdmin)
risk_admin_site.register(MasterEfektivitasKontrol, MasterEfektivitasKontrolAdmin)
risk_admin_site.register(MasterKategoriDampak, MasterKategoriDampakAdmin)
risk_admin_site.register(MasterSkalaDampak, MasterSkalaDampakAdmin)
risk_admin_site.register(MasterSkalaProbabilitas, MasterSkalaProbabilitasAdmin)
risk_admin_site.register(MasterOpsiPerlakuanRisiko, MasterOpsiPerlakuanRisikoAdmin)
risk_admin_site.register(MasterJenisRencanaPerlakuanRisiko, MasterJenisRencanaPerlakuanRisikoAdmin)
risk_admin_site.register(MasterPosAnggaran, MasterPosAnggaranAdmin)
risk_admin_site.register(MasterJenisProgramRKAP, MasterJenisProgramRKAPAdmin)
risk_admin_site.register(MasterLevelRisiko, MasterLevelRisikoAdmin)
risk_admin_site.register(RiskMatrix, RiskMatrixAdmin)
risk_admin_site.register(KPMRPeriode, KPMRPeriodeAdmin)
risk_admin_site.register(KPMRIndikatorResmi, KPMRIndikatorResmiAdmin)
risk_admin_site.register(KPMRSubIndikatorResmi, KPMRSubIndikatorResmiAdmin)
risk_admin_site.register(KinerjaPeriode, KinerjaPeriodeAdmin)
risk_admin_site.register(KinerjaIndikator, KinerjaIndikatorAdmin)
risk_admin_site.register(KompositRisikoTriwulan, KompositRisikoTriwulanAdmin)
risk_admin_site.register(RoadmapProgram, RoadmapProgramAdmin)
risk_admin_site.register(RoadmapPenilaianSemester, RoadmapPenilaianSemesterAdmin)
risk_admin_site.register(RKAPItem, RKAPItemAdmin)

try:
    risk_admin_site.register(KnowledgeBaseCategory, KnowledgeBaseCategoryAdmin)
except admin.sites.AlreadyRegistered:
    pass

try:
    risk_admin_site.register(KnowledgeBaseArticle, KnowledgeBaseArticleAdmin)
except admin.sites.AlreadyRegistered:
    pass

try:
    risk_admin_site.register(KontrakManajemen, KontrakManajemenAdmin)
except admin.sites.AlreadyRegistered:
    pass

try:
    risk_admin_site.register(MasterTemplateKM, MasterTemplateKMAdmin)
except admin.sites.AlreadyRegistered:
    pass

try:
    risk_admin_site.register(TahunBuku, TahunBukuAdmin)
except Exception:
    pass

try:
    risk_admin_site.register(PeriodeLaporan, PeriodeLaporanAdmin)
except Exception:
    pass

try:
    risk_admin_site.register(MasterBagianKM, MasterBagianKMAdmin)
except admin.sites.AlreadyRegistered:
    pass

try:
    risk_admin_site.register(RiwayatJabatanUser, RiwayatJabatanUserAdmin)
except admin.sites.AlreadyRegistered:
    pass
