from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import os
from pathlib import Path
import re
import sys

PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django

django.setup()

from django.contrib.auth.models import Group
from django.db import transaction
from openpyxl import load_workbook

from risk.models import ItemKontrakManajemen, KontrakManajemen, RKMItem, RKMSummary


DEFAULT_SOURCE_FILE = "/Users/armeizir/Downloads/NKO SM UBINFRA MEI 2026.xlsx"
SOURCE_FILE = Path(
    os.environ.get("RKM_SOURCE_FILE")
    or (sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SOURCE_FILE)
)
SHEET_NAME = "DetailN"
YEAR = 2026
MONTH = 5
UNIT_NAME = "UB INFRA"
RKM_TITLE = "RKM UB INFRA Mei 2026"


def clean(value):
    if value in (None, ""):
        return None
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value).strip()


def decimal_or_none(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def normalize_code(value):
    text = clean(value)
    if not text:
        return None
    text = re.sub(r"\s+", "", text.upper())
    match = re.fullmatch(r"([A-Z])(\d+)", text)
    if not match:
        return None
    return match.group(1), int(match.group(2)), f"{match.group(1)}{int(match.group(2))}"


def category_for(code_letter):
    return code_letter if code_letter in {"A", "B", "C", "D"} else None


def calculate_achievement(target, actual, polarity):
    target_dec = decimal_or_none(target)
    actual_dec = decimal_or_none(actual)
    if not target_dec or not actual_dec:
        return None
    if target_dec == 0 or actual_dec == 0:
        return None
    if polarity == "negatif":
        score = target_dec / actual_dec * Decimal("100")
    else:
        score = actual_dec / target_dec * Decimal("100")
    return score.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def read_detail_rows():
    workbook = load_workbook(SOURCE_FILE, data_only=True, read_only=True, keep_links=False)
    worksheet = workbook[SHEET_NAME]
    code_columns = []
    for column in range(3, worksheet.max_column + 1, 2):
        normalized = normalize_code(worksheet.cell(3, column).value)
        if normalized:
            code_columns.append((column, column + 1, *normalized))

    rows_by_month = {}
    for row in range(6, 18):
        month = worksheet.cell(row, 2).value
        if isinstance(month, (int, float)):
            rows_by_month[int(month)] = row

    items = []
    for target_col, actual_col, letter, number, code in code_columns:
        monthly_actuals = {}
        monthly_targets = {}
        for month, row in rows_by_month.items():
            monthly_targets[month] = clean(worksheet.cell(row, target_col).value)
            monthly_actuals[month] = clean(worksheet.cell(row, actual_col).value)
        items.append(
            {
                "letter": letter,
                "number": number,
                "code": code,
                "monthly_targets": monthly_targets,
                "monthly_actuals": monthly_actuals,
            }
        )
    return items


def run():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(SOURCE_FILE)

    detail_items = read_detail_rows()
    unit = Group.objects.get(name=UNIT_NAME)
    kontrak = KontrakManajemen.objects.get(tahun=YEAR, unit_bisnis=unit)
    km_items = {
        (item.master_bagian.kode_bagian.strip().upper(), item.no_urut): item
        for item in ItemKontrakManajemen.objects.filter(kontrak=kontrak).select_related("master_bagian")
    }

    missing = [item["code"] for item in detail_items if (item["letter"], item["number"]) not in km_items]
    if missing:
        raise RuntimeError(f"Item KM tidak ditemukan untuk kode DetailN: {', '.join(missing)}")

    with transaction.atomic():
        summary, _ = RKMSummary.objects.update_or_create(
            tahun=YEAR,
            bulan=MONTH,
            unit_bisnis=unit,
            kontrak_manajemen=kontrak,
            defaults={
                "judul": RKM_TITLE,
                "tanggal_mulai": date(YEAR, MONTH, 1),
                "tanggal_selesai": date(YEAR, MONTH, 31),
                "status": "Draft",
                "pic": UNIT_NAME,
                "tanggal_pengajuan": date(YEAR, MONTH, 31),
            },
        )

        imported = []
        for detail in detail_items:
            km_item = km_items[(detail["letter"], detail["number"])]
            targets = detail["monthly_targets"]
            actuals = detail["monthly_actuals"]
            target_mei = targets.get(MONTH)
            actual_mei = actuals.get(MONTH)
            achievement = calculate_achievement(target_mei, actual_mei, km_item.polaritas)

            rkm_item, _ = RKMItem.objects.update_or_create(
                summary=summary,
                km_item=km_item,
                defaults={
                    "no_item": detail["number"],
                    "kategori_rkm": category_for(detail["letter"]),
                    "sasaran": km_item.indikator_kinerja_kunci,
                    "kpi_indikator": km_item.indikator_kinerja_kunci,
                    "kpi_satuan": km_item.satuan,
                    "kpi_target": km_item.target,
                    "target_akumulasi": target_mei,
                    "target_akumulasi_satuan": km_item.satuan,
                    "target_januari": targets.get(1),
                    "target_februari": targets.get(2),
                    "target_maret": targets.get(3),
                    "target_april": targets.get(4),
                    "target_mei": target_mei,
                    "target_juni": targets.get(6),
                    "target_juli": targets.get(7),
                    "target_agustus": targets.get(8),
                    "target_september": targets.get(9),
                    "target_oktober": targets.get(10),
                    "target_november": targets.get(11),
                    "target_desember": targets.get(12),
                    "realisasi_januari": actuals.get(1),
                    "realisasi_februari": actuals.get(2),
                    "realisasi_maret": actuals.get(3),
                    "realisasi_april": actuals.get(4),
                    "realisasi_mei": actual_mei,
                    "jumlah_realisasi": actual_mei,
                    "persen_capaian": achievement,
                    "pic_rkm": UNIT_NAME,
                    "target_bulanan": f"Mei 2026 target: {target_mei or '-'} {km_item.satuan or ''}".strip(),
                    "realisasi": f"Mei 2026 realisasi: {actual_mei or '-'} {km_item.satuan or ''}".strip(),
                    "deviasi": f"Capaian Mei 2026: {achievement}%" if achievement is not None else "",
                    "keterangan": (
                        f"Sumber import: {SOURCE_FILE.name} sheet {SHEET_NAME}, kode {detail['code']}. "
                        f"Target/Realisasi Jan-Mei: "
                        f"Jan {targets.get(1) or '-'} / {actuals.get(1) or '-'}; "
                        f"Feb {targets.get(2) or '-'} / {actuals.get(2) or '-'}; "
                        f"Mar {targets.get(3) or '-'} / {actuals.get(3) or '-'}; "
                        f"Apr {targets.get(4) or '-'} / {actuals.get(4) or '-'}; "
                        f"Mei {target_mei or '-'} / {actual_mei or '-'}."
                    ),
                },
            )
            imported.append((detail["code"], rkm_item, target_mei, actual_mei, achievement))

    print(f"RKM: {summary.id} {summary.judul}")
    print(f"Imported/updated RKM items: {len(imported)}")
    for code, rkm_item, target, actual, achievement in imported:
        print(
            f"- {code}: RKMItem {rkm_item.id}; target Mei={target or '-'}; "
            f"realisasi Mei={actual or '-'}; capaian={achievement if achievement is not None else '-'}"
        )


if __name__ == "__main__":
    run()
