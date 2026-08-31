#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import re
import sqlite3
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.prod")

import django
django.setup()

from django.conf import settings
from django.db import connection, transaction

from openpyxl import load_workbook

from risk.models import (
    KontrakManajemen,
    ItemKontrakManajemen,
    RKMSummary,
    RKMItem,
)

YEAR = 2026
MONTH = 7

KM_ID = 10
UNIT_ID = 2

TITLE = "RKM UB KITRAN Juli 2026"

SOURCE_DEFAULT = Path("/tmp/RKM UBKITRANS JULI 2026.xlsx")
EXPECTED_SHA256 = (
    "1c31f5f98396f7cfa77c9d7b6cfe829f1cec4c5afc1305274ceee2439dbe760e"
)
SOURCE_SHEET = "Usulan RKM 2026"

# no_item RKM mengikuti canonical KM no_urut seperti RKM April.
GROUPS = [
    # no_item, km_item, section, parent row, first row, last row, expected KPI token
    (1,  122, "A", 8,   8,   15,  "Biaya Pokok Penyediaan"),
    (2,  131, "A", 23,  23,  27,  "Specific gas"),
    (3,  133, "A", 16,  16,  22,  "Equivalent Avability Factor"),
    (5,  138, "C", 34,  34,  59,  "SAIDI"),
    (6,  139, "C", 60,  60,  90,  "SAIFI"),
    (10, 146, "C", 91,  91,  106, "Susut Jaringan"),
    (13, 150, "D", 109, 109, 109, "Pengendalian penggunaan Anggaran"),
    (15, 153, "E", 112, 112, 118, "Peningkatan gap kompetensi"),
    (17, 156, "F", 121, 121, 121, "Compliance"),
]

KM_GUARDS = {
    122: ("Optimalisasi Biaya Pemeliharaan", 1),
    131: ("Specific Gas Consumption", 2),
    133: ("EAF Non MPP", 3),
    138: ("SAIDI KITRANS", 5),
    139: ("SAIFI KITRANS", 6),
    146: ("Susut KITRANS", 10),
    150: ("Pengendalian Penggunaan Anggaran Investasi", 13),
    153: ("Pengelolaan Human Capital", 15),
    156: ("Compliance", 17),
}

MONTH_COLS = {
    "januari": 13,
    "februari": 14,
    "maret": 15,
    "april": 16,
    "mei": 17,
    "juni": 18,
    "juli": 19,
}


class DryRunRollback(Exception):
    pass


def banner(text):
    print()
    print("=" * 160)
    print(text)
    print("=" * 160)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm(v):
    s = str(v or "").casefold().replace("\xa0", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def source_string(v):
    if v is None:
        return None

    if isinstance(v, bool):
        return "1" if v else "0"

    if isinstance(v, int):
        return str(v)

    if isinstance(v, float):
        # Avoid scientific notation / unnecessary .0
        s = format(v, ".15g")
        return s

    return str(v).strip() or None


def decimal_value(v):
    if v in (None, "", "-"):
        return None

    if isinstance(v, Decimal):
        return v

    if isinstance(v, (int, float)):
        return Decimal(str(v))

    s = str(v).strip()
    if not s or s in {"-", "AO", "AI", "AI/AO"}:
        return None

    s = s.replace("Rp", "").replace("rp", "").replace(" ", "")

    # Indonesian simple decimal.
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")

    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def join_unique(values):
    out = []
    seen = set()

    for v in values:
        s = text(v)
        if not s:
            continue

        k = norm(s)
        if not k or k in seen:
            continue

        seen.add(k)
        out.append(s)

    return "\n\n".join(out) if out else None


def percent_db(v):
    """
    Normalize kolom source '% Capaian' ke format RKMItem.persen_capaian.

    Source UBKITRANS mencampur:
      0.5833 -> 58.33%
      1      -> 100.00%
      57.619 -> 57.62%
      100    -> 100.00%

    Database RKMItem menyimpan persen pada skala 0..100,
    DecimalField(max_digits=8, decimal_places=2).
    """
    d = decimal_value(v)

    if d is None:
        return None

    if abs(d) <= Decimal("1"):
        d *= Decimal("100")

    return d.quantize(Decimal("0.01"))


def pct_display(v):
    d = percent_db(v)

    if d is None:
        return None

    return f"{d}%"


def validate_source(source: Path):
    banner("SOURCE GUARD")

    if not source.exists():
        raise RuntimeError(f"STOP: source tidak ditemukan: {source}")

    actual_hash = sha256(source)

    print("SOURCE :", source)
    print("SHA256 :", actual_hash)

    if actual_hash != EXPECTED_SHA256:
        raise RuntimeError(
            f"STOP: SHA256 mismatch; expected={EXPECTED_SHA256}"
        )

    print("HASH   : MATCH")

    wb = load_workbook(
        source,
        data_only=True,
        read_only=False,
        keep_links=False,
    )

    if SOURCE_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"STOP: sheet {SOURCE_SHEET!r} tidak ditemukan."
        )

    ws = wb[SOURCE_SHEET]

    if "2026" not in str(ws["A3"].value or ""):
        raise RuntimeError(
            f"STOP: A3 bukan tahun 2026: {ws['A3'].value!r}"
        )

    sign = str(ws["AB123"].value or "")
    if "05 Agustus 2026" not in sign:
        raise RuntimeError(
            f"STOP: signature/date guard berubah: AB123={sign!r}"
        )

    for _, _, _, parent, _, _, token in GROUPS:
        indicator = str(ws.cell(parent, 2).value or "")
        if norm(token) not in norm(indicator):
            raise RuntimeError(
                f"STOP: source KPI row {parent} berubah; "
                f"expected token={token!r}, actual={indicator!r}"
            )

    print("SHEET  :", SOURCE_SHEET)
    print("YEAR   : 2026")
    print("DATE   :", sign)
    print("KPI    : 9 consolidated groups")
    print("SOURCE : VALID")

    return wb, ws


def validate_master():
    banner("TARGET MASTER")

    km = (
        KontrakManajemen.objects
        .select_related("unit_bisnis")
        .get(pk=KM_ID)
    )

    print(
        f"KM={km.id} | {km.judul!r} | tahun={km.tahun} | "
        f"unit={km.unit_bisnis_id} {km.unit_bisnis} | status={km.status!r}"
    )

    if km.tahun != YEAR:
        raise RuntimeError("STOP: KM bukan tahun 2026.")

    if km.unit_bisnis_id != UNIT_ID:
        raise RuntimeError(
            f"STOP: unit KM={km.unit_bisnis_id}, expected={UNIT_ID}"
        )

    if km.status != "Final":
        raise RuntimeError(
            f"STOP: KM id={KM_ID} tidak Final: {km.status!r}"
        )

    items = {
        x.id: x
        for x in ItemKontrakManajemen.objects
        .filter(kontrak=km)
        .select_related("master_bagian")
        .order_by("no_urut", "id")
    }

    if len(items) != 20:
        raise RuntimeError(
            f"STOP: KM item count={len(items)}, expected=20"
        )

    for km_id, (token, expected_no) in KM_GUARDS.items():
        obj = items.get(km_id)

        if obj is None:
            raise RuntimeError(
                f"STOP: KMItem {km_id} tidak ditemukan pada KM {KM_ID}."
            )

        if norm(token) not in norm(obj.indikator_kinerja_kunci):
            raise RuntimeError(
                f"STOP: KMItem {km_id} label berubah: "
                f"{obj.indikator_kinerja_kunci!r}"
            )

        if obj.no_urut != expected_no:
            raise RuntimeError(
                f"STOP: KMItem {km_id} no_urut={obj.no_urut}, "
                f"expected={expected_no}"
            )

        print(
            f"KMITEM={obj.id:<4} | no={obj.no_urut:<3} | "
            f"{obj.indikator_kinerja_kunci}"
        )

    existing = RKMSummary.objects.filter(
        kontrak_manajemen=km,
        unit_bisnis_id=UNIT_ID,
        tahun=YEAR,
        bulan=MONTH,
    )

    print("Existing RKM July:", existing.count())

    if existing.exists():
        raise RuntimeError(
            "STOP: RKM UB KITRAN Juli 2026 sudah ada; "
            "V1 tidak overwrite data existing."
        )

    return km, items


def collect_group(ws, spec):
    no_item, km_id, section, parent, first, last, token = spec

    def col(row, idx):
        return ws.cell(row, idx).value

    # Source item-level fields follow the parent KPI row.
    indicator = text(col(parent, 2))
    unit = source_string(col(parent, 3))
    target = source_string(col(parent, 4))

    monthly = {
        name: source_string(col(parent, c))
        for name, c in MONTH_COLS.items()
    }

    jumlah = source_string(col(parent, 25))
    pct_raw = col(parent, 26)
    pct = percent_db(pct_raw)

    # Detail rows are consolidated just like the April importer.
    initiatives = join_unique(col(r, 5) for r in range(first, last + 1))
    programs = join_unique(col(r, 6) for r in range(first, last + 1))
    risks = join_unique(col(r, 7) for r in range(first, last + 1))
    mitigations = join_unique(col(r, 8) for r in range(first, last + 1))
    actions = join_unique(col(r, 9) for r in range(first, last + 1))

    budget = decimal_value(col(parent, 10))
    target_accum = source_string(col(parent, 11))
    target_accum_unit = source_string(col(parent, 12))
    actual_budget = decimal_value(col(parent, 27))
    pic = source_string(col(parent, 28))
    analysis = text(col(parent, 29))

    detail_blocks = []

    for r in range(first, last + 1):
        meaningful = any(
            col(r, c) not in (None, "")
            for c in range(5, 30)
        )

        if not meaningful:
            continue

        jan_jul = [
            source_string(col(r, c)) or "-"
            for c in range(13, 20)
        ]

        detail_blocks.append(
            "\n".join(
                [
                    f"Excel row {r}",
                    f"Program: {source_string(col(r,6)) or '-'}",
                    f"Risiko: {source_string(col(r,7)) or '-'}",
                    f"Mitigasi: {source_string(col(r,8)) or '-'}",
                    f"Rencana aksi: {source_string(col(r,9)) or '-'}",
                    f"Anggaran: {source_string(col(r,10)) or '-'}",
                    f"Realisasi Jan-Jul: {' | '.join(jan_jul)}",
                    f"PIC: {source_string(col(r,28)) or '-'}",
                    f"Analisa: {source_string(col(r,29)) or '-'}",
                ]
            )
        )

    keterangan = (
        f"Sumber: RKM UBKITRANS JULI 2026.xlsx | "
        f"sheet {SOURCE_SHEET} | parent row {parent}\n\n"
        + "\n\n".join(detail_blocks)
    )

    return {
        "no_item": no_item,
        "km_item_id": km_id,
        "section": section,
        "indicator": indicator,
        "unit": unit,
        "target": target,
        "initiative": initiatives,
        "program": programs,
        "risk": risks,
        "mitigation": mitigations,
        "action": actions,
        "budget": budget,
        "target_accum": target_accum,
        "target_accum_unit": target_accum_unit,
        "monthly": monthly,
        "jumlah": jumlah,
        "pct": pct,
        "actual_budget": actual_budget,
        "pic": pic,
        "analysis": analysis,
        "keterangan": keterangan,
        "parent_row": parent,
    }


def audit_rows(ws, km_items):
    banner("SOURCE -> KM MAPPING")

    rows = []

    for spec in GROUPS:
        src = collect_group(ws, spec)
        km = km_items[src["km_item_id"]]

        rows.append(src)

        print(
            f"no_item={src['no_item']:<3} "
            f"| row={src['parent_row']:<3} "
            f"| KM={km.id:<4} "
            f"| Jul={src['monthly']['juli']!r:<12} "
            f"| Jumlah={src['jumlah']!r:<12} "
            f"| Capaian={src['pct']!r}"
        )
        print("   SOURCE:", src["indicator"])
        print("   KM    :", km.indikator_kinerja_kunci)

    if len(rows) != 9:
        raise RuntimeError("STOP: consolidated source bukan 9 item.")

    if len({x["km_item_id"] for x in rows}) != 9:
        raise RuntimeError("STOP: KM mapping tidak unique 9/9.")

    return rows


def snapshot_km(km):
    return list(
        ItemKontrakManajemen.objects
        .filter(kontrak=km)
        .values_list(
            "id",
            "master_bagian_id",
            "no_urut",
            "indikator_kinerja_kunci",
            "satuan",
            "target",
            "bobot",
        )
        .order_by("id")
    )


def backup_sqlite():
    engine = connection.settings_dict.get("ENGINE", "")

    if "sqlite" not in engine:
        print("BACKUP: skipped, database bukan SQLite.")
        return None

    src = Path(str(connection.settings_dict["NAME"])).resolve()

    backup_dir = Path("/home/adminsvr/backup")
    backup_dir.mkdir(parents=True, exist_ok=True)

    dst = backup_dir / (
        "db_before_import_rkm_ubkitrans_juli_2026_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".sqlite3"
    )

    with sqlite3.connect(str(src)) as s:
        with sqlite3.connect(str(dst)) as d:
            s.backup(d)

    with sqlite3.connect(str(dst)) as db:
        integrity = db.execute(
            "PRAGMA integrity_check"
        ).fetchone()[0]

    print("BACKUP   :", dst)
    print("INTEGRITY:", integrity)

    if integrity != "ok":
        raise RuntimeError("STOP: backup integrity_check gagal.")

    return dst


def create_summary(km):
    obj = RKMSummary(
        judul=TITLE,
        tahun=YEAR,
        bulan=MONTH,
        unit_bisnis_id=UNIT_ID,
        kontrak_manajemen=km,
        tanggal_mulai=date(2026, 7, 1),
        tanggal_selesai=date(2026, 7, 31),
        status="Draft",
        status_pengajuan="Belum",
        pic="SM UBKITRANS",
    )

    obj.full_clean()
    obj.save()

    return obj


def create_item(summary, src, km_item):
    month_kwargs = {
        "realisasi_januari": src["monthly"]["januari"],
        "realisasi_februari": src["monthly"]["februari"],
        "realisasi_maret": src["monthly"]["maret"],
        "realisasi_april": src["monthly"]["april"],
        "realisasi_mei": src["monthly"]["mei"],
        "realisasi_juni": src["monthly"]["juni"],
        "realisasi_juli": src["monthly"]["juli"],
    }

    kwargs = {
        "summary": summary,
        "no_item": src["no_item"],
        "km_item": km_item,
        "kategori_rkm": src["section"],

        # Preserve source RKM KPI wording.
        "sasaran": src["indicator"],
        "kpi_indikator": src["indicator"],
        "kpi_satuan": src["unit"],
        "kpi_target": src["target"],

        "inisiatif_strategis": src["initiative"],
        "program_kerja_utama": src["program"],
        "risiko": src["risk"],
        "mitigasi_risiko": src["mitigation"],
        "rencana_aksi": src["action"],

        "anggaran_rp_ribu": src["budget"],
        "target_akumulasi": src["target_accum"],
        "target_akumulasi_satuan": src["target_accum_unit"],

        # Source has no explicit monthly target columns.
        "target_januari": None,
        "target_februari": None,
        "target_maret": None,
        "target_april": None,
        "target_mei": None,
        "target_juni": None,
        "target_juli": None,
        "target_agustus": None,
        "target_september": None,
        "target_oktober": None,
        "target_november": None,
        "target_desember": None,

        **month_kwargs,

        "realisasi_agustus": None,
        "realisasi_september": None,
        "realisasi_oktober": None,
        "realisasi_november": None,
        "realisasi_desember": None,

        "jumlah_realisasi": src["jumlah"],
        "persen_capaian": src["pct"],
        "realisasi_anggaran": src["actual_budget"],

        "pic_rkm": src["pic"],
        "hasil_analisa_program_kerja": src["analysis"],

        "target_bulanan": (
            f"Target KPI: {src['target'] or '-'} {src['unit'] or ''}; "
            f"Target akumulasi: {src['target_accum'] or '-'} "
            f"{src['target_accum_unit'] or ''}"
        ).strip(),

        "realisasi": (
            f"Juli: {src['monthly']['juli'] or '-'}; "
            f"Jumlah: {src['jumlah'] or '-'}"
        ),

        "deviasi": (
            f"Capaian: {pct_display(src['pct'])}"
            if src["pct"] is not None
            else "Capaian: -"
        ),

        "keterangan": src["keterangan"],
    }

    fields = {f.name for f in RKMItem._meta.fields}
    kwargs = {
        k: v
        for k, v in kwargs.items()
        if k in fields or k in {"summary", "km_item"}
    }

    item = RKMItem(**kwargs)
    item.full_clean()
    item.save()

    # Model.save() may recalculate achievement using generic logic.
    # Restore source-document values explicitly.
    RKMItem.objects.filter(pk=item.pk).update(
        jumlah_realisasi=src["jumlah"],
        persen_capaian=src["pct"],
        realisasi_anggaran=src["actual_budget"],
    )

    item.refresh_from_db()

    return item


def verify(summary, rows):
    banner("VERIFY IN TRANSACTION")

    items = list(
        RKMItem.objects
        .filter(summary=summary)
        .select_related("km_item")
        .order_by("no_item", "id")
    )

    expected_nos = [1, 2, 3, 5, 6, 10, 13, 15, 17]
    expected_km = [122, 131, 133, 138, 139, 146, 150, 153, 156]

    if len(items) != 9:
        raise RuntimeError(
            f"STOP VERIFY: items={len(items)}, expected=9"
        )

    if [x.no_item for x in items] != expected_nos:
        raise RuntimeError(
            f"STOP VERIFY: no_item={[x.no_item for x in items]}"
        )

    if [x.km_item_id for x in items] != expected_km:
        raise RuntimeError(
            f"STOP VERIFY: km mapping={[x.km_item_id for x in items]}"
        )

    by_no = {x.no_item: x for x in items}

    for src in rows:
        item = by_no[src["no_item"]]

        if item.realisasi_juli != src["monthly"]["juli"]:
            raise RuntimeError(
                f"STOP VERIFY item {item.no_item}: "
                f"Juli DB={item.realisasi_juli!r}, "
                f"source={src['monthly']['juli']!r}"
            )

        if item.jumlah_realisasi != src["jumlah"]:
            raise RuntimeError(
                f"STOP VERIFY item {item.no_item}: "
                f"jumlah DB={item.jumlah_realisasi!r}, "
                f"source={src['jumlah']!r}"
            )

        print(
            f"{item.no_item:02d} | "
            f"RKMItem={item.id:<4} | "
            f"KM={item.km_item_id:<4} | "
            f"Jul={item.realisasi_juli!r:<12} | "
            f"Jumlah={item.jumlah_realisasi!r:<12} | "
            f"Capaian={item.persen_capaian!r} | "
            f"{item.kpi_indikator}"
        )

    print("VERIFY 9/9: PASS")

    return items


def db_health(label):
    if "sqlite" not in connection.settings_dict.get("ENGINE", ""):
        return

    with connection.cursor() as cur:
        cur.execute("PRAGMA integrity_check")
        integrity = cur.fetchone()[0]

        cur.execute("PRAGMA foreign_key_check")
        fk = cur.fetchall()

    print(f"{label} integrity_check :", integrity)
    print(f"{label} foreign_key_check:", len(fk), "error")

    if integrity != "ok" or fk:
        raise RuntimeError(
            f"STOP: database health gagal pada {label}"
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        default=str(SOURCE_DEFAULT),
    )
    parser.add_argument(
        "--apply",
        action="store_true",
    )
    args = parser.parse_args()

    mode = "APPLY" if args.apply else "DRY RUN"

    banner(f"IMPORT RKM UBKITRANS JULI 2026 V1 SAFE — {mode}")

    source = Path(args.source)

    wb, ws = validate_source(source)
    km, km_items = validate_master()
    rows = audit_rows(ws, km_items)

    db_health("PRE")

    before_km = snapshot_km(km)

    if args.apply:
        backup_sqlite()

    try:
        with transaction.atomic():
            # Race-condition duplicate guard.
            existing = RKMSummary.objects.select_for_update().filter(
                kontrak_manajemen_id=KM_ID,
                unit_bisnis_id=UNIT_ID,
                tahun=YEAR,
                bulan=MONTH,
            )

            if existing.exists():
                raise RuntimeError(
                    "STOP: RKM Juli muncul setelah preflight."
                )

            summary = create_summary(km)

            created = []

            for src in rows:
                created.append(
                    create_item(
                        summary,
                        src,
                        km_items[src["km_item_id"]],
                    )
                )

            verify(summary, rows)

            after_km = snapshot_km(km)

            if before_km != after_km:
                raise RuntimeError(
                    "STOP: KM master berubah saat import."
                )

            if not args.apply:
                raise DryRunRollback()

    except DryRunRollback:
        banner("RESULT — DRY RUN")
        print("RKM Summary       : 1 (ROLLBACK)")
        print("RKM Items         : 9 (ROLLBACK)")
        print("KM mapping        : 9/9")
        print("KM master modified: 0")
        print("Database          : TIDAK DIUBAH")
        print("STATUS            : DRY RUN BERHASIL")
        return

    db_health("POST")

    banner("RESULT — APPLY")
    print("RKM ID            :", summary.id)
    print("Judul             :", summary.judul)
    print("Unit              :", summary.unit_bisnis)
    print("Periode           : Juli 2026")
    print("Status            :", summary.status)
    print("Status Pengajuan  :", summary.status_pengajuan)
    print("RKM Items         :", len(created))
    print("KM mapping        : 9/9")
    print("KM master modified: 0")
    print("Database          : TERSIMPAN")
    print("STATUS            : APPLY BERHASIL")


if __name__ == "__main__":
    main()
