from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
import os
from pathlib import Path
import sys

PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django

django.setup()

from django.contrib.auth.models import Group
from django.db import transaction

from risk.models import ItemKontrakManajemen, KontrakManajemen, RKMItem, RKMSummary


DEFAULT_SOURCE_FILE = "/Users/armeizir/Downloads/Draft RKM UBKITRANS APRIL (1).xlsx"
SOURCE_FILE = Path(
    os.environ.get("RKM_SOURCE_FILE")
    or (sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SOURCE_FILE)
)
SHEET_NAME = "Usulan RKM 2026"
YEAR = 2026
MONTH = 4
UNIT_NAME = "UB KITRAN"
CATEGORY_LABELS = {
    "A": "Nilai Ekonomi dan Sosial untuk Indonesia",
    "B": "Inovasi Model Bisnis",
    "C": "Kepemimpinan Teknologi",
    "D": "Peningkatan Investasi",
}


def clean(value):
    if value in (None, ""):
        return None
    return str(value).strip()


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("=") or text in {"-", "AO"}:
            return None
        text = text.replace("%", "").replace(",", ".")
    else:
        text = value
    try:
        return Decimal(str(text))
    except (InvalidOperation, ValueError):
        return None


def fmt(value):
    if value in (None, ""):
        return "-"
    return str(value).strip()


def as_percent(value):
    if value in (None, ""):
        return "-"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if -1 <= number <= 1:
        return f"{number * 100:.2f}%"
    return f"{number:.2f}"


def km_number_for_kpi(kpi):
    text = (kpi or "").lower()
    if "biaya pokok penyediaan" in text or "bpp" in text:
        return 1
    if "specific gas" in text:
        return 2
    if "eaf" in text or "avability" in text or "availability" in text:
        return 3
    if "efor" in text:
        return 4
    if "saidi" in text:
        return 5
    if "saifi" in text:
        return 6
    if "sustainability" in text:
        return 7
    if "manajemen risiko" in text or "kpmr" in text:
        return 8
    if "kesiapan pasokan" in text:
        return 9
    if "susut" in text:
        return 10
    if "tata kelola pembangkit" in text:
        return 11
    if "sistem manajemen terintegr" in text:
        return 12
    if "anggaran investasi" in text:
        return 13
    if "icofr" in text:
        return 14
    if "compliance" in text or "gcg" in text or "zero fatality" in text:
        return 17
    if "kompetensi" in text or "training" in text or "pelatihan" in text or "magang" in text:
        return 15
    if "k3l" in text:
        return 16
    return None


def read_rows():
    from openpyxl import load_workbook

    workbook = load_workbook(
        SOURCE_FILE,
        data_only=True,
        read_only=True,
        keep_links=False,
    )
    worksheet = workbook[SHEET_NAME]
    current = {}
    current_category = None
    rows = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=7, values_only=True), start=7):
        values = list(row)
        no = values[0] if len(values) > 0 else None
        kpi = clean(values[1] if len(values) > 1 else None)
        if isinstance(no, str) and no.strip() in CATEGORY_LABELS:
            current_category = no.strip()
            continue
        if isinstance(no, (int, float)) and kpi:
            current = {
                "excel_row": row_idx,
                "excel_no": int(no),
                "kategori_rkm": current_category,
                "kpi": kpi,
                "satuan_kpi": clean(values[2] if len(values) > 2 else None),
                "target_kpi": clean(values[3] if len(values) > 3 else None),
            }
        if not current:
            continue
        if not any(value not in (None, "") for value in values[5:29]):
            continue
        km_no = km_number_for_kpi(current["kpi"])
        if not km_no:
            continue
        rows.append(
            {
                **current,
                "km_no": km_no,
                "inisiatif": clean(values[4] if len(values) > 4 else None),
                "program": clean(values[5] if len(values) > 5 else None),
                "risiko": clean(values[6] if len(values) > 6 else None),
                "mitigasi": clean(values[7] if len(values) > 7 else None),
                "rencana_aksi": clean(values[8] if len(values) > 8 else None),
                "anggaran": clean(values[9] if len(values) > 9 else None),
                "target_akumulasi": clean(values[10] if len(values) > 10 else None),
                "satuan_target": clean(values[11] if len(values) > 11 else None),
                "realisasi_jan": clean(values[12] if len(values) > 12 else None),
                "realisasi_feb": clean(values[13] if len(values) > 13 else None),
                "realisasi_mar": clean(values[14] if len(values) > 14 else None),
                "realisasi_apr": clean(values[15] if len(values) > 15 else None),
                "realisasi_mei": clean(values[16] if len(values) > 16 else None),
                "realisasi_jun": clean(values[17] if len(values) > 17 else None),
                "realisasi_jul": clean(values[18] if len(values) > 18 else None),
                "realisasi_agu": clean(values[19] if len(values) > 19 else None),
                "realisasi_sep": clean(values[20] if len(values) > 20 else None),
                "realisasi_okt": clean(values[21] if len(values) > 21 else None),
                "realisasi_nov": clean(values[22] if len(values) > 22 else None),
                "realisasi_des": clean(values[23] if len(values) > 23 else None),
                "jumlah": clean(values[24] if len(values) > 24 else None),
                "capaian": clean(values[25] if len(values) > 25 else None),
                "realisasi_anggaran": clean(values[26] if len(values) > 26 else None),
                "pic": clean(values[27] if len(values) > 27 else None),
                "analisa": clean(values[28] if len(values) > 28 else None),
            }
        )
    return rows


def summarize_rows(rows):
    detail_lines = []
    for index, row in enumerate(rows, start=1):
        detail_lines.append(
            "\n".join(
                [
                    f"{index}. Baris Excel {row['excel_row']}",
                    f"   Program: {fmt(row['program'])}",
                    f"   Risiko: {fmt(row['risiko'])}",
                    f"   Mitigasi: {fmt(row['mitigasi'])}",
                    f"   Rencana aksi: {fmt(row['rencana_aksi'])}",
                    f"   Anggaran: {fmt(row['anggaran'])}; Realisasi anggaran: {fmt(row['realisasi_anggaran'])}",
                    f"   Realisasi Jan-Apr: {fmt(row['realisasi_jan'])} | {fmt(row['realisasi_feb'])} | {fmt(row['realisasi_mar'])} | {fmt(row['realisasi_apr'])}",
                    f"   PIC: {fmt(row['pic'])}",
                    f"   Analisa: {fmt(row['analisa'])}",
                ]
            )
        )
    return "\n\n".join(detail_lines)


def run():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(SOURCE_FILE)

    source_rows = read_rows()
    grouped = defaultdict(list)
    for row in source_rows:
        grouped[row["km_no"]].append(row)

    unit = Group.objects.get(name=UNIT_NAME)
    kontrak = KontrakManajemen.objects.get(tahun=YEAR, unit_bisnis=unit)
    km_items = {
        item.no_urut: item
        for item in ItemKontrakManajemen.objects.filter(kontrak=kontrak).order_by("no_urut")
    }

    missing = sorted(set(grouped) - set(km_items))
    if missing:
        raise RuntimeError(f"Item KM tidak ditemukan untuk no_urut: {missing}")

    with transaction.atomic():
        summary, _ = RKMSummary.objects.update_or_create(
            tahun=YEAR,
            bulan=MONTH,
            unit_bisnis=unit,
            kontrak_manajemen=kontrak,
            defaults={
                "judul": "RKM UB KITRAN April 2026",
                "tanggal_mulai": date(YEAR, MONTH, 1),
                "tanggal_selesai": date(YEAR, MONTH, 30),
                    "status": "Draft",
                    "pic": "UB KITRAN",
                    "tanggal_pengajuan": date(YEAR, MONTH, 30),
                },
        )

        imported = 0
        for no_urut in sorted(grouped):
            rows = grouped[no_urut]
            first = rows[0]
            km_item = km_items[no_urut]
            kpi_names = []
            for row in rows:
                if row["kpi"] not in kpi_names:
                    kpi_names.append(row["kpi"])

            item, _ = RKMItem.objects.update_or_create(
                summary=summary,
                km_item=km_item,
                defaults={
                    "no_item": no_urut,
                    "sasaran": " / ".join(kpi_names),
                    "kategori_rkm": first["kategori_rkm"],
                    "kpi_indikator": first["kpi"],
                    "kpi_satuan": first["satuan_kpi"],
                    "kpi_target": first["target_kpi"],
                    "inisiatif_strategis": first["inisiatif"],
                    "program_kerja_utama": "\n\n".join(filter(None, [row["program"] for row in rows])),
                    "risiko": "\n\n".join(filter(None, [row["risiko"] for row in rows])),
                    "mitigasi_risiko": "\n\n".join(filter(None, [row["mitigasi"] for row in rows])),
                    "rencana_aksi": "\n\n".join(filter(None, [row["rencana_aksi"] for row in rows])),
                    "anggaran_rp_ribu": decimal_or_none(first["anggaran"]),
                    "target_akumulasi": first["target_akumulasi"],
                    "target_akumulasi_satuan": first["satuan_target"],
                    "realisasi_januari": first["realisasi_jan"],
                    "realisasi_februari": first["realisasi_feb"],
                    "realisasi_maret": first["realisasi_mar"],
                    "realisasi_april": first["realisasi_apr"],
                    "realisasi_mei": first["realisasi_mei"],
                    "realisasi_juni": first["realisasi_jun"],
                    "realisasi_juli": first["realisasi_jul"],
                    "realisasi_agustus": first["realisasi_agu"],
                    "realisasi_september": first["realisasi_sep"],
                    "realisasi_oktober": first["realisasi_okt"],
                    "realisasi_november": first["realisasi_nov"],
                    "realisasi_desember": first["realisasi_des"],
                    "jumlah_realisasi": first["jumlah"],
                    "persen_capaian": decimal_or_none(first["capaian"]),
                    "realisasi_anggaran": decimal_or_none(first["realisasi_anggaran"]),
                    "pic_rkm": first["pic"],
                    "hasil_analisa_program_kerja": first["analisa"],
                    "target_bulanan": (
                        f"Target KPI: {fmt(first['target_kpi'])} {fmt(first['satuan_kpi'])}; "
                        f"Target akumulasi: {fmt(first['target_akumulasi'])} {fmt(first['satuan_target'])}"
                    ),
                    "realisasi": f"April: {fmt(first['realisasi_apr'])}; Jumlah: {fmt(first['jumlah'])}",
                    "deviasi": f"Capaian: {as_percent(first['capaian'])}",
                    "keterangan": summarize_rows(rows),
                },
            )
            imported += 1

        RKMItem.objects.filter(summary=summary).exclude(
            km_item__no_urut__in=grouped.keys()
        ).delete()

        summary.save()

    print(f"RKM: {summary.id} {summary.judul}")
    print(f"Source rows: {len(source_rows)}")
    print(f"Grouped KM items: {len(grouped)}")
    print(f"Imported/updated RKM items: {imported}")
    print("Mapping:")
    for no_urut in sorted(grouped):
        print(
            f"- KM {no_urut}: {km_items[no_urut].indikator_kinerja_kunci} "
            f"<- {len(grouped[no_urut])} baris Excel"
        )


if __name__ == "__main__":
    run()
