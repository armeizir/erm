"""Import official BID BIS KPMR TW II 2026 from the Hasil KPMR worksheet.

Audit only:
    python risk/scripts/import_kpmr_bis_tw2_2026.py /path/to/file.xlsx

Apply:
    python risk/scripts/import_kpmr_bis_tw2_2026.py /path/to/file.xlsx --apply
"""

from __future__ import annotations

import argparse
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django

django.setup()

from django.contrib.auth.models import Group
from django.db import transaction

from risk.models import KPMRIndikatorResmi, KPMRPeriode, KPMRSubIndikatorResmi
from risk.services.kpmr_automation import rating_for_score


YEAR = 2026
QUARTER = 2
UNIT_NAME = "BID BIS"
SHEET = "Hasil KPMR"

INDICATORS = {
    "I1": (5, "Pencapaian Nilai Eksposur Risiko dibandingkan target Risiko Residual", Decimal("30")),
    "I2": (6, "Pencapaian output pelaksanaan perlakuan Risiko dibandingkan target total output", Decimal("20")),
    "I3": (7, "Realisasi biaya pelaksanaan perlakuan Risiko dibandingkan anggaran", Decimal("20")),
    "I4": (8, "Ketepatan penilaian Risiko", Decimal("30")),
}

SUBINDICATORS = {
    "IDENTIFIKASI": (35, "Ketepatan identifikasi Risiko"),
    "KUANTIFIKASI": (36, "Ketepatan kuantifikasi Risiko"),
    "RENCANA": (37, "Ketepatan rencana perlakuan Risiko"),
    "PRIORITISASI": (38, "Ketepatan prioritisasi Risiko"),
}


def decimal(value) -> Decimal:
    if value is None:
        raise ValueError("Nilai hasil formula Excel kosong; buka dan simpan ulang file di Excel.")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def answer_for(code: str, result: Decimal) -> str:
    choices = {
        "I1": {Decimal("90"): "a", Decimal("60"): "b", Decimal("40"): "c"},
        "I2": {Decimal("100"): "a", Decimal("80"): "b", Decimal("60"): "c", Decimal("40"): "d", Decimal("20"): "e"},
        "I3": {Decimal("80"): "a", Decimal("40"): "b"},
    }
    try:
        return choices[code][result]
    except KeyError as exc:
        raise ValueError(f"Hasil {code} tidak dikenali: {result}") from exc


def main(path: Path, apply: bool) -> None:
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, data_only=True, read_only=True)
    if SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet {SHEET!r} tidak ditemukan.")
    sheet = workbook[SHEET]

    indicator_rows = []
    for code, (row, name, weight) in INDICATORS.items():
        result = decimal(sheet[f"N{row}"].value)
        score = decimal(sheet[f"O{row}"].value)
        answer = answer_for(code, result) if code != "I4" else "a,a,a,a"
        expected_score = decimal(result * weight / Decimal("100"))
        if score != expected_score:
            raise ValueError(f"Skor {code} tidak konsisten: Excel={score}, hitung={expected_score}")
        indicator_rows.append((code, name, weight, answer, result, score))

    sub_rows = []
    for code, (row, name) in SUBINDICATORS.items():
        result = decimal(sheet[f"N{row}"].value)
        score = decimal(sheet[f"O{row}"].value)
        if result != Decimal("90.00") or score != Decimal("22.50"):
            raise ValueError(f"Subindikator {code} harus a=90 dan skor=22.50, diperoleh {result}/{score}")
        sub_rows.append((code, name, result, score))

    total = decimal(sheet["O9"].value)
    calculated_total = decimal(sum(row[5] for row in indicator_rows))
    if total != calculated_total:
        raise ValueError(f"Total Excel={total} tidak sama dengan jumlah indikator={calculated_total}")
    rating = rating_for_score(total)

    print(f"Sumber: {path.name}")
    for code, _, _, answer, result, score in indicator_rows:
        print(f"{code}: jawaban={answer}, hasil={result}, skor={score}")
    print(f"TOTAL: {total} ({rating})")
    print("Mode:", "APPLY" if apply else "AUDIT SAJA")

    if not apply:
        print("Tidak ada data diubah. Jalankan kembali dengan --apply.")
        return

    unit = Group.objects.get(name=UNIT_NAME)
    with transaction.atomic():
        period, _ = KPMRPeriode.objects.update_or_create(
            tahun=YEAR,
            triwulan=QUARTER,
            unit_bisnis=unit,
            defaults={
                "skor_total": total,
                "rating": rating,
                "catatan": f"Diimpor dari {path.name}; mengikuti sheet {SHEET} TW II 2026.",
            },
        )
        indicators = {}
        for code, name, weight, answer, result, score in indicator_rows:
            indicator, _ = KPMRIndikatorResmi.objects.update_or_create(
                periode=period,
                kode=code,
                defaults={
                    "nama": name,
                    "bobot": weight,
                    "jawaban": answer,
                    "hasil": result,
                    "skor": score,
                    "dokumen_referensi": path.name,
                    "keterangan": f"Nilai resmi dari sheet {SHEET}.",
                },
            )
            indicators[code] = indicator

        for code, name, result, score in sub_rows:
            KPMRSubIndikatorResmi.objects.update_or_create(
                indikator=indicators["I4"],
                kode=code,
                defaults={
                    "nama": name,
                    "bobot": Decimal("25.00"),
                    "jawaban": "a",
                    "hasil": result,
                    "skor": score,
                    "keterangan": f"Nilai resmi dari sheet {SHEET}.",
                },
            )

    print(f"TERSIMPAN: {period} = {period.skor_total} ({period.rating})")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("file", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    main(args.file, args.apply)
