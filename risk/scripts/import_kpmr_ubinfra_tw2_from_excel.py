from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import os
import sys

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "riskproject.settings.dev")

import django

django.setup()

from django.contrib.auth.models import Group
from django.db import transaction

from risk.models import KPMRIndikatorResmi, KPMRPeriode, KPMRSubIndikatorResmi
from risk.services.kpmr_automation import rating_for_score


DEFAULT_SOURCE_FILE = "/Users/armeizir/Downloads/Kertas Kerja KPMR UB INFRA TW2.xlsx"
DEFAULT_YEAR = 2026
DEFAULT_QUARTER = 2
DEFAULT_UNIT_NAME = "UB INFRA"

INDICATORS = {
    "I1": {
        "row": 8,
        "nama": "Pencapaian Nilai Eksposur Risiko dibandingkan target Risiko Residual",
        "bobot": Decimal("30.00"),
        "options": {"a": Decimal("90"), "b": Decimal("60"), "c": Decimal("40")},
        "reference_cell": "J8",
        "note_cell": "K8",
    },
    "I2": {
        "row": 12,
        "nama": "Pencapaian output pelaksanaan perlakuan Risiko dibandingkan target total output",
        "bobot": Decimal("20.00"),
        "options": {
            "a": Decimal("100"),
            "b": Decimal("80"),
            "c": Decimal("60"),
            "d": Decimal("40"),
            "e": Decimal("20"),
        },
        "reference_cell": "J12",
        "note_cell": "K12",
    },
    "I3": {
        "row": 18,
        "nama": "Realisasi biaya pelaksanaan perlakuan Risiko dibandingkan anggaran",
        "bobot": Decimal("20.00"),
        "options": {"a": Decimal("80"), "b": Decimal("40")},
        "reference_cell": "J18",
        "note_cell": "K18",
    },
}

SUBINDICATORS = {
    "IDENTIFIKASI": {
        "row": 22,
        "nama": "Ketepatan identifikasi Risiko",
        "options": {"a": Decimal("90"), "b": Decimal("50")},
        "reference_cell": "J22",
        "note_cell": "K22",
    },
    "KUANTIFIKASI": {
        "row": 25,
        "nama": "Ketepatan kuantifikasi Risiko",
        "options": {"a": Decimal("90"), "b": Decimal("50")},
        "reference_cell": "J25",
        "note_cell": "K25",
    },
    "RENCANA": {
        "row": 28,
        "nama": "Ketepatan rencana perlakuan Risiko",
        "options": {"a": Decimal("90"), "b": Decimal("50")},
        "reference_cell": "J28",
        "note_cell": "K28",
    },
    "PRIORITISASI": {
        "row": 31,
        "nama": "Ketepatan prioritisasi Risiko",
        "options": {"a": Decimal("90"), "b": Decimal("50")},
        "reference_cell": "J31",
        "note_cell": "K31",
    },
}


def _source_file():
    return Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SOURCE_FILE)


def _arg(index, default):
    return sys.argv[index] if len(sys.argv) > index else default


def _decimal(value):
    return Decimal(value or 0).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _text(value):
    return str(value or "").strip()


def _choice(ws, row):
    return _text(ws.cell(row=row, column=4).value).lower()


def _score_from_choice(choice, options, label):
    if choice not in options:
        raise ValueError(f"Jawaban {label} tidak valid: {choice!r}")
    return options[choice]


def _weighted(raw_score, weight):
    return _decimal(raw_score * weight / Decimal("100"))


def run():
    path = _source_file()
    if not path.exists():
        raise FileNotFoundError(path)

    year = int(_arg(2, DEFAULT_YEAR))
    quarter = int(_arg(3, DEFAULT_QUARTER))
    unit_name = _arg(4, DEFAULT_UNIT_NAME)
    unit = Group.objects.get(name=unit_name)

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook["KPMR"]

    indicator_rows = []
    for code, definition in INDICATORS.items():
        answer = _choice(worksheet, definition["row"])
        raw_score = _score_from_choice(answer, definition["options"], code)
        score = _weighted(raw_score, definition["bobot"])
        indicator_rows.append(
            {
                "kode": code,
                "nama": definition["nama"],
                "bobot": definition["bobot"],
                "jawaban": answer,
                "hasil": raw_score,
                "skor": score,
                "dokumen_referensi": _text(worksheet[definition["reference_cell"]].value),
                "keterangan": _text(worksheet[definition["note_cell"]].value),
            }
        )

    sub_rows = []
    for code, definition in SUBINDICATORS.items():
        answer = _choice(worksheet, definition["row"])
        raw_score = _score_from_choice(answer, definition["options"], code)
        score = _weighted(raw_score, Decimal("25.00"))
        sub_rows.append(
            {
                "kode": code,
                "nama": definition["nama"],
                "bobot": Decimal("25.00"),
                "jawaban": answer,
                "hasil": raw_score,
                "skor": score,
                "dokumen_referensi": _text(worksheet[definition["reference_cell"]].value),
                "keterangan": _text(worksheet[definition["note_cell"]].value),
            }
        )

    i4_raw = _decimal(sum(row["skor"] for row in sub_rows))
    i4_score = _weighted(i4_raw, Decimal("30.00"))
    indicator_rows.append(
        {
            "kode": "I4",
            "nama": "Ketepatan penilaian Risiko",
            "bobot": Decimal("30.00"),
            "jawaban": ",".join(row["jawaban"] for row in sub_rows),
            "hasil": i4_raw,
            "skor": i4_score,
            "dokumen_referensi": "Kertas Kerja KPMR",
            "keterangan": "Dihitung dari 4 sub indikator sesuai kertas kerja Excel.",
        }
    )

    total_score = _decimal(sum(row["skor"] for row in indicator_rows))
    rating = rating_for_score(total_score)

    with transaction.atomic():
        period, _ = KPMRPeriode.objects.update_or_create(
            tahun=year,
            triwulan=quarter,
            unit_bisnis=unit,
            defaults={
                "skor_total": total_score,
                "rating": rating,
                "catatan": f"Diimpor dari {path.name}. Perhitungan mengikuti jawaban parameter di sheet KPMR.",
            },
        )
        indicator_by_code = {}
        for row in indicator_rows:
            indicator, _ = KPMRIndikatorResmi.objects.update_or_create(
                periode=period,
                kode=row["kode"],
                defaults={
                    "nama": row["nama"],
                    "bobot": row["bobot"],
                    "jawaban": row["jawaban"],
                    "hasil": row["hasil"],
                    "skor": row["skor"],
                    "dokumen_referensi": row["dokumen_referensi"],
                    "keterangan": row["keterangan"],
                },
            )
            indicator_by_code[row["kode"]] = indicator
        i4 = indicator_by_code["I4"]
        for row in sub_rows:
            KPMRSubIndikatorResmi.objects.update_or_create(
                indikator=i4,
                kode=row["kode"],
                defaults={
                    "nama": row["nama"],
                    "bobot": row["bobot"],
                    "jawaban": row["jawaban"],
                    "hasil": row["hasil"],
                    "skor": row["skor"],
                    "keterangan": row["keterangan"],
                },
            )

    print("IMPORT KPMR SUMMARY")
    print(f"- Periode: {period}")
    print(f"- Skor total: {total_score}")
    print(f"- Rating: {rating}")
    for row in indicator_rows:
        print(f"- {row['kode']}: jawaban={row['jawaban']} hasil={row['hasil']} skor={row['skor']}")
    print("\nCATATAN")
    print("- Jika angka Excel berubah, pastikan pilihan jawaban di kolom D sheet KPMR sudah benar.")
    print("- Judul di file lampiran masih tertulis TW 1, tetapi script menyimpan ke TW sesuai argumen.")


if __name__ == "__main__":
    run()
