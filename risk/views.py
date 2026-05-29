from collections import defaultdict
import calendar
from decimal import Decimal, InvalidOperation
from io import BytesIO
import json
from statistics import mean, median, pstdev

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from monthly_report.models import MonthlyRiskReport
from corporate_risk.models import MonteCarloMetricHistory, MultiMetricMonteCarloResult, RiskMetric

from .services.kpmr_automation import kpmr_dashboard_rows, kpmr_dashboard_summary
from .models import (
    ItemKontrakManajemen,
    KPMRItem,
    KPMRSummary,
    KontrakManajemen,
    ProfilRisikoKorporatItem,
    ProfilRisikoKorporatPenyebab,
    ProfilRisikoKorporatSummary,
    RKMSummary,
    ReAssessmentSummary,
    RiskMatrix,
)


LEVEL_FALLBACKS = [
    {"name": "Low", "codes": ["LOW"], "default_color": "#5b8f3a"},
    {"name": "Low to Moderate", "codes": ["LOW_TO_MODERATE", "LOW-MODERATE", "LTM"], "default_color": "#b8d4a2"},
    {"name": "Moderate", "codes": ["MODERATE", "MEDIUM"], "default_color": "#f3ef19"},
    {"name": "Moderate to High", "codes": ["MODERATE_TO_HIGH", "MODERATE-HIGH", "MTH"], "default_color": "#f2b01e"},
    {"name": "High", "codes": ["HIGH"], "default_color": "#d50f0f"},
]

MODE_CHOICES = [
    ("inheren", "Inheren"),
    ("residual", "Residual"),
]

BULAN_CHOICES = [
    (1, "Januari"),
    (2, "Februari"),
    (3, "Maret"),
    (4, "April"),
    (5, "Mei"),
    (6, "Juni"),
    (7, "Juli"),
    (8, "Agustus"),
    (9, "September"),
    (10, "Oktober"),
    (11, "November"),
    (12, "Desember"),
]
BULAN_LABELS = dict(BULAN_CHOICES)


def _normalize(text):
    return (text or "").strip().lower()


def _get_mode_label(mode):
    return dict(MODE_CHOICES).get(mode, "Inheren")


def _resolve_level_bucket(level_name):
    normalized = _normalize(level_name)
    for bucket in LEVEL_FALLBACKS:
        if _normalize(bucket["name"]) == normalized:
            return bucket["name"]
    if "high" in normalized and "moderate" in normalized:
        return "Moderate to High"
    if "high" in normalized:
        return "High"
    if "moderate" in normalized and "low" in normalized:
        return "Low to Moderate"
    if "moderate" in normalized:
        return "Moderate"
    if "low" in normalized:
        return "Low"
    return level_name or "Tidak Terkategori"


def _fallback_level_from_score(score):
    if score >= 15:
        return "High", "#d00000"  # merah
    elif score >= 12:
        return "Moderate to High", "#f4a300"  # orange
    elif score >= 8:
        return "Moderate", "#fff200"  # kuning
    elif score >= 5:
        return "Low to Moderate", "#a9c98f"  # hijau muda
    else:
        return "Low", "#5a8f3a"  # hijau tua
    

def _default_matrix():
    return RiskMatrix.objects.filter(is_default=True, aktif=True).prefetch_related(
        "cells__skala_dampak", "cells__skala_probabilitas", "cells__level_risiko"
    ).first() or RiskMatrix.objects.filter(aktif=True).prefetch_related(
        "cells__skala_dampak", "cells__skala_probabilitas", "cells__level_risiko"
    ).first()


def _matrix_lookup(matrix):
    lookup = {}
    dampak_labels = {}
    kemungkinan_labels = {}
    legend_map = {}

    if not matrix:
        return lookup, dampak_labels, kemungkinan_labels, legend_map

    for cell in matrix.cells.all():
        impact = getattr(cell.skala_dampak, "urutan", None)
        likelihood = getattr(cell.skala_probabilitas, "urutan", None)
        if impact is None or likelihood is None:
            continue
        level_name = cell.level_risiko.nama if cell.level_risiko_id else None
        color = cell.warna_hex or getattr(cell.level_risiko, "warna_hex", None)
        lookup[(impact, likelihood)] = {
            "score": cell.skor,
            "level": level_name,
            "color": color,
            "cell_id": cell.id,
        }
        dampak_labels[impact] = cell.skala_dampak.nama
        kemungkinan_labels[likelihood] = cell.skala_probabilitas.nama
        if level_name:
            legend_map.setdefault(_resolve_level_bucket(level_name), color or "#d9d9d9")

    return lookup, dampak_labels, kemungkinan_labels, legend_map


def _selected_risk_fields(mode):
    if mode == "residual":
        return "residual_dampak", "residual_kemungkinan", "residual_level_risiko", "matrix_cell_residual"
    return "dampak", "kemungkinan", "level_risiko", "matrix_cell_inheren"


def _build_risk_entry(item, mode, matrix_lookup):
    dampak_field, kemungkinan_field, level_field, cell_field = _selected_risk_fields(mode)
    impact = getattr(item, dampak_field)
    likelihood = getattr(item, kemungkinan_field)
    stored_score = getattr(item, level_field)
    stored_cell = getattr(item, cell_field, None)

    score = stored_score
    level_name = None
    color = None
    matrix_source = "fallback"
    is_mappable = impact is not None and likelihood is not None

    cell_meta = matrix_lookup.get((impact, likelihood)) if is_mappable else None
    if cell_meta:
        score = cell_meta["score"]
        level_name = cell_meta["level"]
        color = cell_meta["color"]
        matrix_source = "default_matrix"
    elif stored_cell:
        level_name = getattr(getattr(stored_cell, "level_risiko", None), "nama", None)
        color = stored_cell.warna_hex or getattr(getattr(stored_cell, "level_risiko", None), "warna_hex", None)
        score = stored_cell.skor or score
        matrix_source = "stored_cell"

    if not is_mappable:
        level_name = "Belum Dipetakan"
        color = "#e5e7eb"
        score = None
        matrix_source = "not_mapped"
    elif not level_name:
        level_name, fallback_color = _fallback_level_from_score(score or (impact * likelihood))
        color = color or fallback_color
        matrix_source = "fallback"

    return {
        "id": item.id,
        "no_risiko": item.no_risiko,
        "peristiwa_risiko": item.peristiwa_risiko,
        "kategori_risiko": str(item.kategori_risiko) if item.kategori_risiko_id else "-",
        "pemilik_risiko": (
            str(item.daftar_penyebab.first().pemilik_risiko)
            if item.daftar_penyebab.exists()
            and item.daftar_penyebab.first().pemilik_risiko
            else "-"
        ),
        "status": item.status or "-",
        "summary": str(item.summary),
        "dampak": impact,
        "kemungkinan": likelihood,
        "level": level_name,
        "level_bucket": _resolve_level_bucket(level_name),
        "score": score or (impact * likelihood if is_mappable else None),
        "color": color or "#d9d9d9",
        "mode": mode,
        "matrix_source": matrix_source,
        "is_mappable": is_mappable,
    }


def _get_filtered_items(request):
    summary_id = request.GET.get("summary")
    year = request.GET.get("tahun") or _selected_year(request)
    status = request.GET.get("status")
    owner = request.GET.get("pemilik")
    category_id = request.GET.get("kategori")
    mode = request.GET.get("mode", "inheren")
    if mode not in dict(MODE_CHOICES):
        mode = "inheren"

    items = ProfilRisikoKorporatItem.objects.all().select_related(
        "summary", "kategori_risiko", "taksonomi_t3", "bumn", "sasaran_kbumn",
        "matrix_cell_inheren__level_risiko", "matrix_cell_residual__level_risiko"
    )
    selected_summary = None

    if summary_id:
        items = items.filter(summary_id=summary_id)
        selected_summary = ProfilRisikoKorporatSummary.objects.filter(pk=summary_id).first()
    if year:
        items = items.filter(summary__tahun=year)
    if status:
        items = items.filter(status=status)
    if owner:
        items = items.filter(daftar_penyebab__pemilik_risiko_id=owner).distinct()
    if category_id:
        items = items.filter(kategori_risiko_id=category_id) 

    return items.order_by("summary", "no_item"), selected_summary, mode


def _risk_matrix_context(items_qs, mode="inheren", selected_summary=None):
    matrix = _default_matrix()
    matrix_lookup, dampak_labels, kemungkinan_labels, legend_map = _matrix_lookup(matrix)
    items = list(items_qs)

    max_impact = max(dampak_labels.keys(), default=5)
    max_likelihood = max(kemungkinan_labels.keys(), default=5)
    size = max(getattr(matrix, "ukuran", 5), max_impact, max_likelihood)
    size = min(max(size, 5), 5)

    level_counts = defaultdict(int)
    cell_items = defaultdict(list)
    drilldown_items = []

    for item in items:
        entry = _build_risk_entry(item, mode, matrix_lookup)
        if not entry:
            continue
        if entry["is_mappable"]:
            level_counts[entry["level_bucket"]] += 1
            legend_map.setdefault(entry["level_bucket"], entry["color"])
            cell_items[(entry["dampak"], entry["kemungkinan"])].append(entry)
        drilldown_items.append(entry)

    grid = []
    for likelihood in range(size, 0, -1):
        row = {
            "value": likelihood,
            "label": kemungkinan_labels.get(likelihood, f"Skala {likelihood}"),
            "cells": [],
        }
        for impact in range(1, size + 1):
            cell_meta = matrix_lookup.get((impact, likelihood))
            cell_risks = sorted(
                cell_items.get((impact, likelihood), []),
                key=lambda risk: (risk["no_risiko"] or 0, risk["peristiwa_risiko"]),
            )
            score = impact * likelihood
            level_name, color = _fallback_level_from_score(score)

            row["cells"].append(
                {
                    "impact": impact,
                    "likelihood": likelihood,
                    "score": score,
                    "level": level_name,
                    "color": color,
                    "count": len(cell_risks),
                    "risks": cell_risks,
                }
            )
        grid.append(row)

    impact_axis = [
        {"value": impact, "label": dampak_labels.get(impact, f"Skala {impact}")}
        for impact in range(1, size + 1)
    ]

    summaries = ProfilRisikoKorporatSummary.objects.order_by("-tahun", "judul")
    years = list(
        ProfilRisikoKorporatSummary.objects.order_by("-tahun").values_list("tahun", flat=True).distinct()
    )
    statuses = list(
        ProfilRisikoKorporatItem.objects.exclude(status__isnull=True).exclude(status__exact="").order_by("status").values_list("status", flat=True).distinct()
    )
    owners = list(
        ProfilRisikoKorporatPenyebab.objects
        .exclude(pemilik_risiko__isnull=True)
        .order_by("pemilik_risiko__name")
        .values_list("pemilik_risiko_id", "pemilik_risiko__name")
        .distinct()
    )
    categories = list(
        ProfilRisikoKorporatItem.objects.filter(kategori_risiko__isnull=False).select_related("kategori_risiko").order_by("kategori_risiko__nama").values_list("kategori_risiko__id", "kategori_risiko__nama").distinct()
    )

    total_risks = len(drilldown_items)
    mapped_risks = len([item for item in drilldown_items if item["is_mappable"]])
    defaults = {bucket["name"]: bucket["default_color"] for bucket in LEVEL_FALLBACKS}
    legend = [
        {"name": "Low", "color": "#5a8f3a"},
        {"name": "Low to Moderate", "color": "#a9c98f"},
        {"name": "Moderate", "color": "#fff200"},
        {"name": "Moderate to High", "color": "#f4a300"},
        {"name": "High", "color": "#d00000"},
    ]

    return {
        "matrix": matrix,
        "grid": grid,
        "impact_axis": impact_axis,
        "legend": legend,
        "total_risks": total_risks,
        "mapped_risks": mapped_risks,
        "unmapped_risks": total_risks - mapped_risks,
        "level_counts": {
            "high": level_counts.get("High", 0),
            "moderate_to_high": level_counts.get("Moderate to High", 0),
            "moderate": level_counts.get("Moderate", 0),
            "low_to_moderate": level_counts.get("Low to Moderate", 0),
            "low": level_counts.get("Low", 0),
        },
        "mode": mode,
        "mode_label": _get_mode_label(mode),
        "matrix_source_label": "RiskMatrixCell default" if matrix else "Fallback skor standar",
        "drilldown_rows": sorted(
            drilldown_items,
            key=lambda row: (0 if row["score"] is None else row["score"] * -1, row["no_risiko"] or 0),
        ),
        "filters": {
            "summaries": summaries,
            "years": years,
            "statuses": statuses,
            "owners": owners,
            "categories": categories,
            "selected_summary": selected_summary,
            "modes": MODE_CHOICES,
            "months": BULAN_CHOICES,
        },
    }


def _selected_year(request):
    year = request.GET.get("tahun")
    if year:
        try:
            return int(year)
        except ValueError:
            pass
    latest = (
        ProfilRisikoKorporatSummary.objects.order_by("-tahun")
        .values_list("tahun", flat=True)
        .first()
    )
    return latest


def _selected_month(request):
    month = request.GET.get("bulan")
    if month:
        try:
            month_int = int(month)
            if 1 <= month_int <= 12:
                return month_int
        except ValueError:
            pass
    return None


def _selected_tab(request):
    tab = request.GET.get("tab", "profil")
    return tab if tab in {"profil", "prediksi", "kpmr", "km"} else "profil"


def _decimal_or_none(value):
    if value in (None, ""):
        return None
    cleaned = str(value).strip().replace("%", "").replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _format_decimal(value):
    if value is None:
        return ""
    normalized = value.quantize(Decimal("0.01"))
    if normalized == normalized.to_integral():
        return str(normalized.to_integral())
    return str(normalized)


def _format_percent(value):
    if value is None:
        return ""
    return f"{_format_decimal(value)}%"


def _calculate_km_score(item, target, realisasi):
    if target is None or realisasi is None or target == 0:
        return None, None
    if item.polaritas == "negatif":
        achievement = (target / realisasi * Decimal("100")) if realisasi else None
    else:
        achievement = realisasi / target * Decimal("100")
    if achievement is None:
        return None, None
    score = Decimal(str(item.bobot or 0)) * achievement / Decimal("100")
    return achievement, score


def _corporate_profile_rows(year):
    queryset = ProfilRisikoKorporatSummary.objects.all().order_by("-tahun", "judul")
    if year:
        queryset = queryset.filter(tahun=year)

    rows = []
    for summary in queryset:
        items = ProfilRisikoKorporatItem.objects.filter(summary=summary)
        mapped_inherent = items.filter(dampak__isnull=False, kemungkinan__isnull=False).count()
        mapped_residual = items.filter(
            residual_dampak__isnull=False,
            residual_kemungkinan__isnull=False,
        ).count()
        rows.append(
            {
                "id": summary.id,
                "judul": summary.judul,
                "tahun": summary.tahun,
                "status": summary.status,
                "total": items.count(),
                "mapped_inherent": mapped_inherent,
                "mapped_residual": mapped_residual,
            }
        )
    return rows


def _unit_profile_rows(year, month):
    summaries = ReAssessmentSummary.objects.select_related("unit_bisnis").order_by(
        "unit_bisnis__name",
        "judul",
    )
    if year:
        summaries = summaries.filter(tahun=year)

    rows = []
    for summary in summaries:
        reports = summary.monthly_reports.select_related("periode").order_by(
            "-periode__tanggal_mulai",
            "-versi",
        )
        if month:
            reports = reports.filter(periode__tanggal_mulai__month=month)
        report = reports.first()
        total_profile_items = summary.item.count()
        reported_items = report.items.count() if report else 0
        high_items = report.items.filter(realisasi_level_risiko__icontains="tinggi").count() if report else 0
        rows.append(
            {
                "id": summary.id,
                "unit": summary.unit_bisnis.name if summary.unit_bisnis_id else "-",
                "judul": summary.judul,
                "tahun": summary.tahun,
                "profile_items": total_profile_items,
                "report": report,
                "bulan": BULAN_LABELS.get(report.periode.tanggal_mulai.month) if report else "-",
                "reported_items": reported_items,
                "high_items": high_items,
                "coverage": round((reported_items / total_profile_items) * 100, 1)
                if total_profile_items
                else 0,
            }
        )
    return rows


def _selected_kontrak(request, year):
    contracts = KontrakManajemen.objects.select_related("unit_bisnis").order_by(
        "unit_bisnis__name",
        "judul",
    )
    if year:
        contracts = contracts.filter(tahun=year)

    selected_id = request.GET.get("km")
    selected = None
    if selected_id:
        selected = contracts.filter(pk=selected_id).first()
    if selected is None:
        selected = contracts.filter(unit_bisnis__name="BID MRK").first() or contracts.first()
    return selected, contracts


def _kontrak_manajemen_rows(year):
    contracts = KontrakManajemen.objects.select_related("unit_bisnis").order_by(
        "unit_bisnis__name",
        "-tahun",
        "judul",
    )
    if year:
        contracts = contracts.filter(tahun=year)

    rows = []
    for contract in contracts:
        reassessments = contract.reassessment_summary.all()
        monthly_reports = MonthlyRiskReport.objects.filter(
            reassessment__kontrak_manajemen=contract,
        ).distinct()
        rows.append(
            {
                "id": contract.id,
                "unit": contract.unit_bisnis.name if contract.unit_bisnis_id else "-",
                "judul": contract.judul,
                "tahun": contract.tahun,
                "status": contract.status,
                "total_items": ItemKontrakManajemen.objects.filter(kontrak=contract).count(),
                "rkm_count": contract.rkm_summary.count(),
                "profile_count": reassessments.count(),
                "monthly_report_count": monthly_reports.count(),
            }
        )
    return rows


def _kontrak_manajemen_detail(selected_contract, year, month):
    if selected_contract is None:
        return {"groups": [], "rkm": None, "total_bobot": 0, "total_nilai": ""}

    rkm_qs = RKMSummary.objects.filter(kontrak_manajemen=selected_contract).order_by("-tahun", "-bulan")
    if year:
        rkm_qs = rkm_qs.filter(tahun=year)
    if month:
        rkm_qs = rkm_qs.filter(bulan=month)
    rkm = rkm_qs.first()
    rkm_items = {
        item.km_item_id: item
        for item in rkm.item.select_related("km_item")
    } if rkm else {}

    items = ItemKontrakManajemen.objects.filter(kontrak=selected_contract).select_related(
        "master_bagian",
        "bagian",
    ).order_by("master_bagian__urutan", "master_bagian__kode_bagian", "no_urut")

    groups = []
    current_code = None
    current_group = None
    total_bobot = Decimal("0")
    total_nilai = Decimal("0")
    has_nilai = False

    for item in items:
        master = item.master_bagian
        code = master.kode_bagian if master else "-"
        name = master.nama_bagian if master else "Tanpa Bagian"
        if code != current_code:
            current_group = {
                "kode": code,
                "nama": name,
                "bobot": Decimal("0"),
                "nilai": Decimal("0"),
                "has_nilai": False,
                "items": [],
            }
            groups.append(current_group)
            current_code = code

        rkm_item = rkm_items.get(item.id)
        target_bulanan_raw = rkm_item.target_bulanan if rkm_item else ""
        realisasi_raw = rkm_item.realisasi if rkm_item else ""
        target_value = _decimal_or_none(target_bulanan_raw) or _decimal_or_none(item.target)
        realisasi_value = _decimal_or_none(realisasi_raw)
        achievement, score = _calculate_km_score(item, target_value, realisasi_value)

        bobot = Decimal(str(item.bobot or 0))
        total_bobot += bobot
        current_group["bobot"] += bobot
        if score is not None:
            has_nilai = True
            current_group["has_nilai"] = True
            current_group["nilai"] += score
            total_nilai += score

        current_group["items"].append(
            {
                "no": item.no_urut,
                "indikator": item.indikator_kinerja_kunci,
                "formula": item.formula or "",
                "satuan": item.satuan or "",
                "bobot": _format_decimal(bobot),
                "target_tahunan": item.target or "",
                "target_bulanan": target_bulanan_raw or item.target or "",
                "realisasi": realisasi_raw,
                "pencapaian": _format_percent(achievement),
                "nilai": _format_decimal(score),
                "indicator": "good" if score is not None and score >= bobot else "attention" if score is not None else "",
                "keterangan": rkm_item.keterangan if rkm_item else "",
            }
        )

    for group in groups:
        group["bobot_display"] = _format_decimal(group["bobot"])
        group["nilai_display"] = _format_decimal(group["nilai"]) if group["has_nilai"] else ""

    return {
        "groups": groups,
        "rkm": rkm,
        "total_bobot": _format_decimal(total_bobot),
        "total_nilai": _format_decimal(total_nilai) if has_nilai else "",
    }


def _selected_prediction_risk(request):
    risk_id = request.GET.get("prediksi_risiko")
    if not risk_id:
        return None
    try:
        return int(risk_id)
    except (TypeError, ValueError):
        return None


def _risk_prediction_options(year, selected_summary):
    items = ProfilRisikoKorporatItem.objects.select_related("summary").order_by("summary__tahun", "no_item")
    if year:
        items = items.filter(summary__tahun=year)
    if selected_summary:
        items = items.filter(summary=selected_summary)
    return items


def _risk_prediction_rows(year, selected_risk_id=None):
    queryset = (
        MultiMetricMonteCarloResult.objects
        .select_related("corporate_risk_item", "forecast_periode")
        .order_by("-created_at")
    )
    if year:
        queryset = queryset.filter(forecast_periode__tahun_buku__tahun=year)
    if selected_risk_id:
        queryset = queryset.filter(corporate_risk_item_id=selected_risk_id)

    rows = []
    for result in queryset[:25]:
        rows.append(
            {
                "id": result.id,
                "risk": result.corporate_risk_item,
                "periode": result.forecast_periode,
                "target": result.target_value,
                "actual_ytd": (result.simulation_snapshot or {}).get("target_analysis", {}).get("actual_total"),
                "forecast_total": result.forecast_total,
                "gap": result.target_gap,
                "potential_loss": result.potential_loss,
                "prob_achieve": result.probability_achieve_target,
                "prob_not_achieve": result.probability_not_achieve_target,
                "worst_case": result.worst_case_value,
                "baseline": result.baseline_value,
                "best_case": result.best_case_value,
                "var_95": result.var_95,
                "target_status": result.target_status or "-",
                "risk_status": result.risk_status or "-",
                "requires_mitigation": result.requires_mitigation,
            }
        )
    return rows


def _metric_statistics(values):
    values = [float(value) for value in values if value not in (None, "")]
    if not values:
        return None

    growth_rates = []
    for idx in range(1, len(values)):
        previous = values[idx - 1]
        current = values[idx]
        if previous:
            growth_rates.append((current - previous) / previous * 100)

    avg_value = mean(values)
    std_value = pstdev(values) if len(values) > 1 else 0
    coefficient_variation = (std_value / avg_value * 100) if avg_value else 0

    trend = "Stabil"
    if len(values) > 1:
        first_value = values[0]
        last_value = values[-1]
        if last_value > first_value * 1.05:
            trend = "Meningkat"
        elif last_value < first_value * 0.95:
            trend = "Menurun"

    volatility = "Rendah"
    if coefficient_variation > 20:
        volatility = "Tinggi"
    elif coefficient_variation > 10:
        volatility = "Sedang"

    return {
        "count": len(values),
        "min": min(values),
        "max": max(values),
        "mean": avg_value,
        "median": median(values),
        "std_dev": std_value,
        "coefficient_variation": coefficient_variation,
        "avg_growth": mean(growth_rates) if growth_rates else 0,
        "min_growth": min(growth_rates) if growth_rates else 0,
        "max_growth": max(growth_rates) if growth_rates else 0,
        "trend": trend,
        "volatility": volatility,
    }


def _risk_prediction_history(selected_risk_id):
    if not selected_risk_id:
        return {"selected_item": None, "metrics": [], "histories": []}

    selected_item = (
        ProfilRisikoKorporatItem.objects
        .select_related("summary")
        .filter(pk=selected_risk_id)
        .first()
    )
    if not selected_item:
        return {"selected_item": None, "metrics": [], "histories": []}

    metrics = list(
        RiskMetric.objects
        .filter(corporate_risk_item=selected_item, is_active=True)
        .order_by("-is_target_metric", "name")
    )
    histories = []
    for metric in metrics:
        metric_histories = (
            MonteCarloMetricHistory.objects
            .filter(metric=metric)
            .select_related("periode")
            .order_by("tanggal_data", "id")
        )
        history_rows = list(metric_histories)
        histories.append(
            {
                "metric": metric,
                "rows": history_rows,
                "count": len(history_rows),
                "statistics": _metric_statistics([row.metric_value for row in history_rows]),
            }
        )
    return {
        "selected_item": selected_item,
        "metrics": metrics,
        "histories": histories,
    }


def _risk_prediction_summary(rows):
    total = len(rows)
    need_mitigation = len([row for row in rows if row["requires_mitigation"]])
    not_achieved = len([row for row in rows if row["target_status"] == "Tidak Tercapai"])
    avg_probability = Decimal("0")
    probabilities = [
        Decimal(str(row["prob_not_achieve"]))
        for row in rows
        if row["prob_not_achieve"] not in (None, "")
    ]
    if probabilities:
        avg_probability = sum(probabilities, Decimal("0")) / Decimal(str(len(probabilities)))
    return {
        "total": total,
        "need_mitigation": need_mitigation,
        "not_achieved": not_achieved,
        "avg_probability_not_achieve": avg_probability,
    }


def _risk_prediction_detail(selected_risk_id, year):
    if not selected_risk_id:
        return {}

    queryset = (
        MultiMetricMonteCarloResult.objects
        .filter(corporate_risk_item_id=selected_risk_id)
        .select_related("corporate_risk_item", "forecast_periode")
        .order_by("-created_at")
    )
    if year:
        queryset = queryset.filter(forecast_periode__tahun_buku__tahun=year)

    result = queryset.first()
    if not result:
        return {}

    snapshot = result.simulation_snapshot or {}
    target_analysis = snapshot.get("target_analysis") or {}
    metrics = (result.metric_snapshot or {}).get("metrics", [])
    simulation_count = (
        target_analysis.get("total_simulation")
        or snapshot.get("n_simulations")
        or 0
    )
    projection_rows = snapshot.get("projection_rows") or []
    target_projection_rows = snapshot.get("target_projection_rows") or []
    chart_series = snapshot.get("chart_series") or {}
    distribution = sorted(float(value) for value in target_analysis.get("distribution_sample", []) if value is not None)
    target_value = float(result.target_value or target_analysis.get("target_value") or 0)

    return {
        "result": result,
        "target_analysis": target_analysis,
        "metrics": metrics,
        "simulation_count": simulation_count,
        "projection_rows": projection_rows,
        "target_projection_rows": target_projection_rows,
        "distribution_labels_json": json.dumps(list(range(1, len(distribution) + 1))),
        "distribution_values_json": json.dumps(distribution),
        "distribution_target_json": json.dumps([target_value] * len(distribution)),
        "multi_metric_labels_json": json.dumps(chart_series.get("labels", [])),
        "multi_metric_mean_json": json.dumps(chart_series.get("mean", [])),
        "multi_metric_p20_json": json.dumps(chart_series.get("p20", [])),
        "multi_metric_p40_json": json.dumps(chart_series.get("p40", [])),
        "multi_metric_p60_json": json.dumps(chart_series.get("p60", [])),
        "multi_metric_p80_json": json.dumps(chart_series.get("p80", [])),
    }


def _export_workbook(context, params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RCC Summary"

    dark_fill = PatternFill("solid", fgColor="0D2E5E")
    header_fill = PatternFill("solid", fgColor="D9E5F3")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    sheet["A1"] = "Risk Control Center (RCC)"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A2"] = f"Mode Monitoring: {context['mode_label']}"
    sheet["A3"] = f"Matriks Aktif: {context['matrix'].nama if context['matrix'] else 'Fallback Standard'}"
    sheet["A4"] = f"Sumber Perhitungan RCC: {context['matrix_source_label']}"

    filter_pairs = [
        ("Profil", params.get("summary") or "Semua"),
        ("Tahun", params.get("tahun") or "Semua"),
        ("Status", params.get("status") or "Semua"),
        ("Pemilik", params.get("pemilik") or "Semua"),
        ("Kategori", params.get("kategori") or "Semua"),
    ]
    start_row = 5
    for idx, (label, value) in enumerate(filter_pairs, start=start_row):
        sheet[f"A{idx}"] = label
        sheet[f"B{idx}"] = value
        sheet[f"A{idx}"].font = bold_font

    stats_row = 12
    stats = [
        ("Total Risiko", context["total_risks"]),
        ("High", context["level_counts"]["high"]),
        ("Moderate to High", context["level_counts"]["moderate_to_high"]),
        ("Moderate", context["level_counts"]["moderate"]),
        ("Low to Moderate", context["level_counts"]["low_to_moderate"]),
        ("Low", context["level_counts"]["low"]),
    ]
    for col_idx, (label, value) in enumerate(stats, start=1):
        cell = sheet.cell(stats_row, col_idx)
        cell.value = label
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        value_cell = sheet.cell(stats_row + 1, col_idx)
        value_cell.value = value
        value_cell.font = Font(size=14, bold=True)
        value_cell.alignment = Alignment(horizontal="center")

    matrix_sheet = workbook.create_sheet("Heatmap")
    matrix_sheet["A1"] = f"Heatmap RCC - {context['mode_label']}"
    matrix_sheet["A1"].font = Font(size=14, bold=True)
    matrix_sheet["A3"] = "Kemungkinan \\ Dampak"
    matrix_sheet["A3"].fill = dark_fill
    matrix_sheet["A3"].font = white_font

    for col_idx, impact in enumerate(context["impact_axis"], start=2):
        cell = matrix_sheet.cell(3, col_idx)
        cell.value = f"{impact['value']} - {impact['label']}"
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(context["grid"], start=4):
        label_cell = matrix_sheet.cell(row_idx, 1)
        label_cell.value = f"{row['value']} - {row['label']}"
        label_cell.fill = header_fill
        label_cell.font = bold_font
        label_cell.alignment = Alignment(wrap_text=True)
        for col_idx, cell_data in enumerate(row["cells"], start=2):
            cell = matrix_sheet.cell(row_idx, col_idx)
            cell.value = f"{cell_data['count']} risiko\n{cell_data['level']}\nSkor {cell_data['score']}"
            cell.fill = PatternFill("solid", fgColor=(cell_data["color"] or "#D9D9D9").replace("#", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    detail_sheet = workbook.create_sheet("Drilldown")
    headers = [
        "No Risiko",
        "Peristiwa Risiko",
        "Kategori Risiko",
        "Pemilik Risiko",
        "Status",
        "Mode",
        "Dampak",
        "Kemungkinan",
        "Level",
        "Skor",
        "Summary",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = detail_sheet.cell(1, col_idx)
        cell.value = header
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, risk in enumerate(context["drilldown_rows"], start=2):
        detail_sheet.cell(row_idx, 1).value = risk["no_risiko"]
        detail_sheet.cell(row_idx, 2).value = risk["peristiwa_risiko"]
        detail_sheet.cell(row_idx, 3).value = risk["kategori_risiko"]
        detail_sheet.cell(row_idx, 4).value = risk["pemilik_risiko"]
        detail_sheet.cell(row_idx, 5).value = risk["status"]
        detail_sheet.cell(row_idx, 6).value = context["mode_label"]
        detail_sheet.cell(row_idx, 7).value = risk["dampak"]
        detail_sheet.cell(row_idx, 8).value = risk["kemungkinan"]
        detail_sheet.cell(row_idx, 9).value = risk["level"]
        detail_sheet.cell(row_idx, 10).value = risk["score"]
        detail_sheet.cell(row_idx, 11).value = risk["summary"]

    widths = {
        "RCC Summary": {"A": 20, "B": 32},
        "Heatmap": {"A": 28, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22},
        "Drilldown": {"A": 12, "B": 50, "C": 26, "D": 26, "E": 18, "F": 14, "G": 12, "H": 14, "I": 22, "J": 10, "K": 30},
    }
    for ws in workbook.worksheets:
        for col, width in widths.get(ws.title, {}).items():
            ws.column_dimensions[col].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def dashboard(request):
    items, selected_summary, mode = _get_filtered_items(request)
    context = _risk_matrix_context(items, mode=mode, selected_summary=selected_summary)
    selected_year = _selected_year(request)
    selected_month = _selected_month(request)
    active_tab = _selected_tab(request)
    selected_prediction_risk_id = _selected_prediction_risk(request)
    selected_kontrak, kontrak_options = _selected_kontrak(request, selected_year)
    corporate_profile_rows = _corporate_profile_rows(selected_year)
    unit_profile_rows = _unit_profile_rows(selected_year, selected_month)
    kontrak_manajemen_rows = _kontrak_manajemen_rows(selected_year)
    kontrak_manajemen_detail = _kontrak_manajemen_detail(selected_kontrak, selected_year, selected_month)
    kpmr_rows = kpmr_dashboard_rows(selected_year, selected_month)
    kpmr_summary = kpmr_dashboard_summary(kpmr_rows)
    risk_prediction_options = _risk_prediction_options(selected_year, selected_summary)
    risk_prediction_history = _risk_prediction_history(selected_prediction_risk_id)
    risk_prediction_rows = _risk_prediction_rows(selected_year, selected_prediction_risk_id)
    risk_prediction_summary = _risk_prediction_summary(risk_prediction_rows)
    risk_prediction_detail = _risk_prediction_detail(selected_prediction_risk_id, selected_year)

    context["active_tab"] = active_tab
    context["selected_kontrak"] = selected_kontrak
    context["kontrak_options"] = kontrak_options
    context["filters"]["selected_year"] = selected_year
    context["filters"]["selected_month"] = selected_month
    context["corporate_profile_rows"] = corporate_profile_rows
    context["unit_profile_rows"] = unit_profile_rows
    context["kontrak_manajemen_rows"] = kontrak_manajemen_rows
    context["kontrak_manajemen_detail"] = kontrak_manajemen_detail
    context["corporate_profile_count"] = sum(row["total"] for row in corporate_profile_rows)
    context["unit_profile_count"] = sum(row["profile_items"] for row in unit_profile_rows)
    context["unit_report_count"] = len([row for row in unit_profile_rows if row["report"]])
    context["kontrak_manajemen_count"] = len(kontrak_manajemen_rows)
    context["kontrak_manajemen_item_count"] = sum(row["total_items"] for row in kontrak_manajemen_rows)
    context["kpmr_rows"] = kpmr_rows
    context["kpmr_summary"] = kpmr_summary
    context["risk_prediction_rows"] = risk_prediction_rows
    context["risk_prediction_summary"] = risk_prediction_summary
    context["risk_prediction_options"] = risk_prediction_options
    context["risk_prediction_history"] = risk_prediction_history
    context["risk_prediction_detail"] = risk_prediction_detail
    context["selected_prediction_risk_id"] = selected_prediction_risk_id
    context["selected_period_label"] = (
        f"{BULAN_LABELS[selected_month]} {selected_year}"
        if selected_month and selected_year
        else f"Tahun {selected_year}"
        if selected_year
        else "Semua periode"
    )
    context["page_title"] = "Risk Control Center (RCC)"
    return render(request, "dashboard.html", context)


def export_rcc_excel(request):
    items, selected_summary, mode = _get_filtered_items(request)
    context = _risk_matrix_context(items, mode=mode, selected_summary=selected_summary)
    file_bytes = _export_workbook(context, request.GET)

    response = HttpResponse(
        file_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="rcc_{mode}.xlsx"'
    return response


def kpmr_review_view(request, summary_id):
    summary = get_object_or_404(KPMRSummary, pk=summary_id)

    items = summary.item.select_related(
        "reassessment_item",
        "reassessment_item__km_item"
    ).order_by("no_item")

    context = {
        "summary": summary,
        "items": items,
    }

    return render(request, "risk/kpmr_review.html", context)


@csrf_exempt
def kpmr_update_item(request):
    if request.method == "POST":
        data = json.loads(request.body)

        item_id = data.get("id")
        field = data.get("field")
        value = data.get("value")

        try:
            item = KPMRItem.objects.get(id=item_id)

            if field == "perlakuan_risiko":
                item.perlakuan_risiko = value
            elif field == "bukti":
                item.bukti = value
            elif field == "nilai_kpmr":
                item.nilai_kpmr = int(value) if value else None
            elif field == "status_kpmr":
                item.status_kpmr = value
            elif field == "catatan":
                item.catatan = value

            item.save()

            return JsonResponse({"status": "success"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    return JsonResponse({"status": "invalid"})
